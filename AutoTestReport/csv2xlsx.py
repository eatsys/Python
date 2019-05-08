import openpyxl as xl
import csv

outputfile = xl.load_workbook('D:\\work\\Reamon\\Python\\template\\DVT-WIFI-2.4G-CONDUCTED TX.xlsx')

inpufile = csv.reader(open('D:\\work\\Reamon\\Python\\log\\'+ '2G-PowerLevel-11n-PORTA.csv'))
next(inpufile)
i = 5
for row in inpufile:
    outputfile["2.4G-DPD-ON"].cell(row=i, column=6).value = float(row[4])
    outputfile["2.4G-DPD-ON"].cell(row=i, column=13).value = float(row[7])
    i +=1

inpufile = csv.reader(open('D:\\work\\Reamon\\Python\\log\\'+ '2G-PowerLevel-11ac-PORTA.csv'))
next(inpufile)
i = 257
for row in inpufile:
    outputfile["2.4G-DPD-ON"].cell(row=i, column=6).value = float(row[4])
    outputfile["2.4G-DPD-ON"].cell(row=i, column=13).value = float(row[7])
    i += 1

inpufile = csv.reader(open('D:\\work\\Reamon\\Python\\log\\'+ '2G-PowerLevel-11n-PORTB.csv'))
next(inpufile)
i = 5
for row in inpufile:
    outputfile["2.4G-DPD-ON"].cell(row=i, column=9).value = float(row[4])
    outputfile["2.4G-DPD-ON"].cell(row=i, column=16).value = float(row[7])
    i += 1

inpufile = csv.reader(open('D:\\work\\Reamon\\Python\\log\\'+ '2G-PowerLevel-11ac-PORTB.csv'))
next(inpufile)
i = 257
for row in inpufile:
    outputfile["2.4G-DPD-ON"].cell(row=i, column=9).value = float(row[4])
    outputfile["2.4G-DPD-ON"].cell(row=i, column=16).value = float(row[7])
    i += 1


outputfile.save('D:\\work\\Reamon\\Python\\report\\TX.xlsx')


outputfile = xl.load_workbook('D:\\work\\Reamon\\Python\\template\\DVT-WIFI-WB62-2.4G-RX Dynamic-AM1.xlsx')

inpufile = csv.reader(open('D:\\work\\Reamon\\Python\\log\\'+ '2G-Dynamic-11g_n_ac-PORTA.csv'))
next(inpufile)
i = 5
for row in inpufile:
    outputfile["2.4G Receive Dynamic"].cell(row=i, column=3).value = float(row[4])
    outputfile["2.4G Receive Dynamic"].cell(row=i, column=5).value = float(row[7])
    i +=1

inpufile = csv.reader(open('D:\\work\\Reamon\\Python\\log\\'+ '2G-Dynamic-11g_n_ac-PORTB.csv'))
next(inpufile)
i = 5
for row in inpufile:
    #outputfile["2.4G Receive Dynamic"].cell(row=i, column=3).value = float(row[4])
    outputfile["2.4G Receive Dynamic"].cell(row=i, column=7).value = float(row[7])
    i += 1


outputfile.save('D:\\work\\Reamon\\Python\\report\\RX.xlsx')
