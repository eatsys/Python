__author__ = 'DVTRF'

from openpyxl import Workbook
from data.data import *
from data.parameters import AP_TYPE, RADIO, ANGLE_NUM, LINE_LOSS, RUN_TPYE
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import Series, LineChart, ScatterChart, RadarChart, Reference
from openpyxl.drawing.image import Image
import time

now_time = time.strftime('%Y%m%d%H%M%S', time.localtime())
stop_time = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime())
if RUN_TPYE == '0':
    test_type = '_OTA'
elif RUN_TPYE == '1':
    test_type = '_Conductive'
else:
    test_type = ''
filename = AP_TYPE + '_Rate_over_Range' + test_type + '_Test_Report_' + RADIO + '_' + now_time + '.xlsx'
logger.info('Report: ' + filename)
workbook = Workbook()
worksheet_cover = workbook.active
worksheet_environment = workbook.active
worksheet_range = workbook.active
worksheet_data = workbook.active
worksheet_cover = workbook.create_sheet('Overview', 0)
worksheet_cover.sheet_properties.tabColor = '1072BA'
worksheet_cover.sheet_view.showGridLines = False
worksheet_environment = workbook.create_sheet('Environment', 1)
worksheet_environment.sheet_view.showGridLines = False
worksheet_range = workbook.create_sheet('Rate_over_Range', 2)
worksheet_range.sheet_view.showGridLines = False
if ANGLE_NUM > 1:
    worksheet_angle = workbook.create_sheet('Rate_over_Angle', 3)
    worksheet_angle.sheet_view.showGridLines = False
    index_value = 4
else:
    index_value = 3
worksheet_data = workbook.create_sheet('Data', index_value)
worksheet_data.sheet_view.showGridLines = False

worksheet_cover.column_dimensions['B'].width = 15.0
worksheet_cover.column_dimensions['C'].width = 25.0
worksheet_cover.column_dimensions['D'].width = 35.0
worksheet_cover.row_dimensions[7].height = 70.0

worksheet_data.column_dimensions['A'].width = 12.0
worksheet_data.column_dimensions['B'].width = 21.0
worksheet_data.column_dimensions['C'].width = 10.0
worksheet_data.column_dimensions['D'].width = 22.0
worksheet_data.column_dimensions['E'].width = 29.0
worksheet_data.column_dimensions['F'].width = 22.0
worksheet_data.column_dimensions['G'].width = 29.0
worksheet_data.column_dimensions['H'].width = 13.0
worksheet_data.column_dimensions['I'].width = 20.0
worksheet_data.column_dimensions['J'].width = 13.0
worksheet_data.column_dimensions['K'].width = 20.0
worksheet_data.column_dimensions['L'].width = 13.0
worksheet_data.column_dimensions['M'].width = 20.0
worksheet_data.column_dimensions['N'].width = 13.0
worksheet_data.column_dimensions['O'].width = 20.0
worksheet_data.column_dimensions['P'].width = 8.0
worksheet_data.column_dimensions['Q'].width = 12.0
worksheet_data.column_dimensions['R'].width = 12.0
worksheet_data.column_dimensions['S'].width = 12.0
worksheet_data.column_dimensions['T'].width = 12.0
worksheet_data.column_dimensions['U'].width = 12.0
worksheet_data.column_dimensions['V'].width = 12.0
worksheet_data.column_dimensions['W'].width = 38.0
worksheet_data.column_dimensions['X'].width = 25.0
worksheet_data.column_dimensions['Y'].width = 38.0
worksheet_data.column_dimensions['Z'].width = 25.0
worksheet_data.row_dimensions[1].height = 26.0
for row in range(1, 100):
    worksheet_data.row_dimensions[row].height = 21.0

cover_base_font = Font(name='Arial Unicode MS',
                       size=11,
                       bold=False,
                       italic=False,
                       vertAlign=None,
                       underline='none',
                       strike=False,
                       color='EF000000')

cover_base_alignment = Alignment(vertical='center')

cover_report_font = Font(name='Verdana',
                         size=28,
                         bold=True,
                         vertAlign=None,
                         color='EF000000')

cover_info_font = Font(name='Times New Roman',
                       size=11,
                       bold=False,
                       italic=True,
                       vertAlign=None,
                       color='EF000000')

data_head_font = Font(name='Arial Unicode MS',
                      size=14,
                      bold=True,
                      color='000000')

data_font = Font(name='Times New Roman',
                 size=11,
                 bold=False,
                 italic=False,
                 color='000000')

data_alignment = Alignment(vertical='center',
                           horizontal='center')

data_border = Border(left=Side(border_style='dotted',
                               color='000000'),
                     right=Side(border_style='dotted',
                                color='000000'),
                     top=Side(border_style='dotted',
                              color='000000'),
                     bottom=Side(border_style='dotted',
                                 color='000000'),
                     diagonal=Side(border_style=None,
                                   color='000000'),
                     diagonal_direction=0,
                     outline=Side(border_style=None,
                                  color='000000'),
                     vertical=Side(border_style=None,
                                   color='000000'),
                     horizontal=Side(border_style=None,
                                     color='000000')
                     )


data_head_fill = PatternFill('solid', fgColor='FFCC66')

merge_font = Font(name='Arial Unicode MS',
                  size=11,
                  bold=False,
                  color='EF000000'
                  )


def write_row():
    # title
    if ANGLE_NUM > 1:
        title = ['Channel', 'Path_Loss(dB)', 'Angle', 'DS_Throughput', 'DS_Throughput_avg', 'US_Throughput',
                 'US_Throughput_avg', 'Sta_Rssi', 'Sta_Rssi_avg', 'AP_Rssi', 'AP_Rssi_avg', 'DS_Rate', 'DS_Rate_avg',
                 'US_Rate', 'US_Rate_avg', 'Time', 'BW(DS)', 'NSS(DS)', 'MCS(DS)', 'BW(US)', 'NSS(US)', 'MCS(US)',
                 'STA RSSI(per chain)', 'AP POWER', 'AP RSSI(per chain)', 'STA POWER']
    else:
        title = ['Channel', 'Path_Loss(dB)', 'Angle', 'DS_Throughput', 'US_Throughput', 'Sta_Rssi', 'AP_Rssi',
                 'DS_Rate',
                 'US_Rate', 'Time', 'BW(DS)', 'NSS(DS)', 'MCS(DS)', 'BW(US)', 'NSS(US)', 'MCS(US)',
                 'STA RSSI(per chain)', 'AP POWER', 'AP RSSI(per chain)', 'STA POWER']
    row_value = 1
    for t in title:
        worksheet_data.cell(1, row_value, value=t)
        row_value += 1
    for r in range(len(title)):
        r += 1
        worksheet_data.cell(1, r).font = data_head_font
        worksheet_data.cell(1, r).fill = data_head_fill
        border_range = ANGLE_NUM*int(len(Att_rep))
        for c in range(border_range + 1):
            c += 1
            worksheet_data.cell(c, r).border = data_border


def write_single(posX, single_result):
    posY = 2
    for st in single_result:
        if single_result == Att_rep:
            st = st + LINE_LOSS
        elif single_result in (Tx_Throught, Rx_Throught):
            st = float(st)
            data_format = '0.000'
        elif single_result in (Ap_Rssi, Sta_Rssi, Tx_Rate, Rx_Rate):
            st = int(st)
            data_format = '0'
        else:
            data_format = 'General'
        worksheet_data[str(chr(posX) + str(posY))] = st
        worksheet_data[str(chr(posX) + str(posY))].font = data_font
        worksheet_data[str(chr(posX) + str(posY))].alignment = data_alignment
        worksheet_data[str(chr(posX) + str(posY))].number_format = data_format
        posY += 1
    posX += 1
    return posX


def write_multi(posX, multi_result):
    posY = 2
    posZ = posY + int(ANGLE_NUM) - 1
    for mt in multi_result:
        if multi_result == Att_rep:
            mt = mt + LINE_LOSS
        worksheet_data.merge_cells(str(chr(posX) + str(posY) + ":" + chr(posX) + str(posZ)))
        worksheet_data[str(chr(posX) + str(posY))] = mt
        worksheet_data[str(chr(posX) + str(posY))].font = data_font
        worksheet_data[str(chr(posX) + str(posY))].alignment = data_alignment
        posY += int(ANGLE_NUM)
        posZ += int(ANGLE_NUM)
    posX += 1
    return posX


def write_avg(posX):
    posXX = posX - 1
    posY = 2
    posZ = posY + int(ANGLE_NUM) - 1
    for av in Att_rep:
        worksheet_data.merge_cells(str(chr(posX) + str(posY) + ":" + chr(posX) + str(posZ)))
        worksheet_data[str(chr(posX) + str(posY))] = '=AVERAGE(' + chr(posXX) + str(posY) + ':' + chr(posXX) + str(
            posZ) + ')'
        worksheet_data[str(chr(posX) + str(posY))].font = data_font
        worksheet_data[str(chr(posX) + str(posY))].alignment = data_alignment
        worksheet_data[str(chr(posX) + str(posY))].number_format = '0'
        worksheet_data[str(chr(posX) + str(posY))].border = data_border
        posY += int(ANGLE_NUM)
        posZ += int(ANGLE_NUM)
    posX += 1
    return posX


def write_range(posA, posB, posC, posD):
    # if mode == 'DS':
    #     name = 'DS Graph'
    #     tp_name = 'DS Throughput'
    #     chart = 'chart_tx'
    #     color_line = '#3399FF'
    # elif mode == 'US':
    #     name = 'US Graph'
    #     tp_name = 'US Throughput'
    #     chart = 'chart_rx'
    #     color_line = '#33CC66'
    chart1 = LineChart(smooth=True)
    chart1.title = 'Throughput'
    chart1.style = 13
    chart1.y_axis.title = 'Mbps'
    chart1.x_axis.title = 'RSSI(dBm)'
    tp_tx_list = []
    tp_rx_list = []
    rssi_sta_list = []
    rssi_ap_list = []
    row_value = 2
    # print(worksheet_data['B2'].value)
    # print(worksheet_data.cell(2,2).value)
    print(posA, posB, posC, posD)
    for att in Att_rep:
        # tp_tx = worksheet_data.cell(row_value, posA - 65).value
        # tp_tx_list.append(tp_tx)
        tp_tx = worksheet_data[str(chr(posA)+str(row_value))].value
        tp_tx_list.append(tp_tx)
        tp_rx = worksheet_data.cell(posB - 65, row_value).value
        tp_rx_list.append(tp_rx)
        rssi_sta = worksheet_data.cell(posC - 65, row_value).value
        rssi_sta_list.append(rssi_sta)
        rssi_ap = worksheet_data.cell(posD - 65, row_value).value
        rssi_ap_list.append(rssi_ap)
        row_value = row_value + ANGLE_NUM
    print(tp_tx_list, tp_rx_list, rssi_sta_list, rssi_sta_list)

    data = Reference(worksheet_data, min_col=posA-65, min_row=2, max_row=ANGLE_NUM*int(len(Att_rep)))
    print(data)
    chart1.add_data(data)
    cats = Reference(worksheet_data, min_col=posB-65, min_row=2, max_row=ANGLE_NUM*int(len(Att_rep)))
    print(cats)
    chart1.set_categories(cats)
    chart1.shape = 4
    worksheet_range.add_chart(chart1, 'B2')
    # if mode == 'DS':
    #     s1 = chart.series[0]
    #     s1.marker.symbol = "triangle"
    #     s1.marker.graphicalProperties.solidFill = "FF0000"  # Marker filling
    #     s1.marker.graphicalProperties.line.solidFill = "FF0000"  # Marker outline
    #
    #     s1.graphicalProperties.line.noFill = True
    #
    #     s2 = chart.series[1]
    #     s2.graphicalProperties.line.solidFill = "00AAAA"
    #     s2.graphicalProperties.line.dashStyle = "sysDot"
    #     s2.graphicalProperties.line.width = 100050  # width in EMUs
    #
    #     s2 = chart.series[2]
    #     s2.smooth = True  # Make the line smooth


def write_radar(posA, posE):
    posB = ord('C')
    posY = 2
    posD = 2
    if posE == 'B':
        tp_type = 'DS Throughput'
    elif posE == 'J':
        tp_type = 'US Throughput'
    else:
        print('Check postion')
        pass
    for att in Att_rep:
        # insert line graph
        radar = radar_name = 'radar' + str(att)
        radar = workbook.add_chart({"type": "radar"})
        # radar = eval(radar)
        radar.set_title({
            "name": 'Attenuation=' + str(att) + 'dB'
        })
        radar.set_y_axis({
            "name": "Mbps"
        })
        radar.set_x_axis(
            {
                'name': 'Angle'
            }
        )
        if posE == 'B':
            radar.set_chartarea({
                'border': {'none': True},
                'gradient': {'colors': ['#EEE7F8', '#DED2F1', '#CBB8E9']}
            })
        else:
            radar.set_chartarea({
                'border': {'none': True},
                'gradient': {'colors': ['#EEF8FF', '#C3F1FF', '#A2EBFF']}
            })
        if ANGLE_NUM > 1:
            radar.add_series({
                'name': tp_type,
                'categories': '=Data!$' + chr(posB) + '$' + str(posY) + ':$' + chr(posB) + '$' +
                              str(int(ANGLE_NUM) + posY - 1),
                'values': '=Data!$' + chr(posA) + '$' + str(posY) + ':$' + chr(posA) + '$' +
                          str(int(ANGLE_NUM) + posY - 1),
                'line': {'color': '#3399FF'},
                'gradient': {'colors': ['#FFE9D7', '#FFD4B2', '#FFC18B']},
                'marker': {
                    'type': 'circle',
                    'size': 3,
                    'border': {'color': 'red'},
                    'fill': {'color': 'yellow'},
                },
            })
        else:
            pass
        if posE == 'B':
            radar.set_plotarea({
                'border': {'none': True},
                'fill': {'none': True},
                # 'gradient': {'colors': ['#EEE7F8', '#DED2F1', '#CBB8E9']}
                'gradient': {'colors': ['#E6DCF4', '#DED2F1', '#CBB8E9']}
            })
        else:
            radar.set_plotarea({
                'border': {'none': True},
                'fill': {'none': True},
                # 'gradient': {'colors': ['#EEF8FF', '#C3F1FF', '#A2EBFF']}
                'gradient': {'colors': ['#DCF7FF', '#C3F1FF', '#A2EBFF']}
            })
        worksheet_angle.insert_chart(posE + str(posD), radar, {'x_scale': 1, 'y_scale': 1.4})
        posY += int(ANGLE_NUM)
        posD += 20


def Generate_Test_Report():
    img = Image('./images/CIG.png')
    worksheet_cover.add_image(img, 'B1')

    worksheet_cover.merge_cells('B4:I4')
    worksheet_cover['B4'] = 'Cambridge Industries Group (CIG)'
    worksheet_cover['B4'].font = cover_base_font
    worksheet_cover.merge_cells('B5:I5')
    worksheet_cover['B5'] = 'Partnership for the Next Generation Broadband Access'
    worksheet_cover['B5'].font = cover_base_font

    worksheet_cover.merge_cells('B7:F7')
    worksheet_cover['B7'] = 'WIFI Performance Test Report'
    worksheet_cover['B7'].font = cover_report_font
    worksheet_cover['B7'].alignment = cover_base_alignment

    worksheet_cover['B10'] = 'Finish Time'
    worksheet_cover['B10'].font = cover_base_font
    worksheet_cover['C10'] = stop_time
    worksheet_cover['C10'].font = cover_info_font

    worksheet_cover['B12'] = 'DUT(AP)'
    worksheet_cover['B12'].font = cover_base_font
    worksheet_cover['C13'] = 'Product'
    worksheet_cover['C13'].font = cover_base_font
    worksheet_cover['D13'] = AP_TYPE
    worksheet_cover['D13'].font = cover_info_font
    worksheet_cover['C14'] = 'Hardware Version'
    worksheet_cover['C14'].font = cover_base_font
    worksheet_cover['C15'] = 'Software Version'
    worksheet_cover['C15'].font = cover_base_font
    worksheet_cover['C16'] = 'Operating Band'
    worksheet_cover['C16'].font = cover_base_font
    worksheet_cover['C17'] = '2.4G Operation Mode'
    worksheet_cover['C17'].font = cover_base_font
    worksheet_cover['C18'] = '2.4G Antenna Configuration'
    worksheet_cover['C18'].font = cover_base_font
    worksheet_cover['C19'] = '5G Operation Mode'
    worksheet_cover['C19'].font = cover_base_font
    worksheet_cover['C20'] = '5G Antenna Configuration'
    worksheet_cover['C20'].font = cover_base_font
    worksheet_cover['B22'] = 'STATION'
    worksheet_cover['B22'].font = cover_base_font
    worksheet_cover['C23'] = 'Station Type'
    worksheet_cover['C23'].font = cover_base_font
    worksheet_cover['C24'] = 'Model'
    worksheet_cover['C24'].font = cover_base_font
    worksheet_cover['C25'] = 'Version'
    worksheet_cover['C25'].font = cover_base_font
    worksheet_cover['C26'] = 'Operating Band'
    worksheet_cover['C26'].font = cover_base_font
    worksheet_cover['C27'] = '2.4G Operation Mode'
    worksheet_cover['C27'].font = cover_base_font
    worksheet_cover['C28'] = '2.4G Antenna Configuration'
    worksheet_cover['C28'].font = cover_base_font
    worksheet_cover['C29'] = '5G Operation Mode'
    worksheet_cover['C29'].font = cover_base_font
    worksheet_cover['C30'] = '5G Antenna Configuration'
    worksheet_cover['C30'].font = cover_base_font
    worksheet_cover['B32'] = 'TEST TOOLS'
    worksheet_cover['B32'].font = cover_base_font
    worksheet_cover['C33'] = 'Test Software'
    worksheet_cover['C33'].font = cover_base_font
    worksheet_cover['D33'] = 'Ixchroit6.7'
    worksheet_cover['D33'].font = cover_info_font
    worksheet_cover['C34'] = 'Test Script'
    worksheet_cover['C34'].font = cover_base_font
    worksheet_cover['D34'] = 'High_Performance_Throughput.scr'
    worksheet_cover['D34'].font = cover_info_font

    worksheet_environment['B1'] = 'TEST DIAGRAM AND ENVIRONMENT'
    worksheet_environment['B1'].font = cover_base_font
    img = Image('./images/environment.png')
    worksheet_environment.add_image(img, 'B3')
    img = Image('./images/tp.png')
    worksheet_environment.add_image(img, 'M3')

    try:
        rep_to_excel = Reportdata_Get()
    except:
        logger.error('No file')
    else:
        try:
            rep_to_excel.Ch_get()
            if ANGLE_NUM > 1:
                posX = write_multi(65, Channel)
            else:
                posX = write_single(65, Channel)
        except:
            logger.error('No file')
        try:
            rep_to_excel.Att_get()
            if ANGLE_NUM > 1:
                posX = write_multi(posX, Att_rep)
            else:
                posX = write_single(posX, Att_rep)
        except:
            logger.error('No file')
        try:
            rep_to_excel.Angle_get()
            posX = write_single(posX, Angle)
        except:
            logger.error('No file')
        try:
            rep_to_excel.Tx_tp_get()
            posX = write_single(posX, Tx_Throught)
            posX_range_txtp = posX
        except:
            logger.error('No file')
        if ANGLE_NUM > 1:
            posX = write_avg(posX)
        try:
            rep_to_excel.Rx_tp_get()
            posX = write_single(posX, Rx_Throught)
            posX_range_rxtp = posX
        except:
            logger.error('No file')
        if ANGLE_NUM > 1:
            posX = write_avg(posX)
        try:
            rep_to_excel.Sta_rssi_get()
            posX = write_single(posX, Sta_Rssi)
            posX_range_starssi = posX
        except:
            logger.error('No file')
        if ANGLE_NUM > 1:
            posX = write_avg(posX)
        try:
            rep_to_excel.Ap_rssi_get()
            posX = write_single(posX, Ap_Rssi)
            posX_range_aprssi = posX
        except:
            logger.error('No file')
        if ANGLE_NUM > 1:
            posX = write_avg(posX)
        try:
            rep_to_excel.Tx_rate_get()
            posX = write_single(posX, Tx_Rate)
        except:
            logger.error('No file')
        if ANGLE_NUM > 1:
            posX = write_avg(posX)
        try:
            rep_to_excel.Rx_rate_get()
            posX = write_single(posX, Rx_Rate)
        except:
            logger.error('No file')
        if ANGLE_NUM > 1:
            posX = write_avg(posX)
        try:
            rep_to_excel.Dura_Time_get()
            posX = write_single(posX, Dura_Time)
        except:
            logger.error('No file')
        try:
            rep_to_excel.BW_TxRate_get()
            posX = write_single(posX, BW_Tx_Rate)
        except:
            logger.error('No file')
        try:
            rep_to_excel.NSS_TxRate_get()
            posX = write_single(posX, NSS_Tx_Rate)
        except:
            logger.error('No file')
        try:
            rep_to_excel.MCS_TxRate_get()
            posX = write_single(posX, MCS_Tx_Rate)
        except:
            logger.error('No file')
        try:
            rep_to_excel.BW_RxRate_get()
            posX = write_single(posX, BW_Rx_Rate)
        except:
            logger.error('No file')
        try:
            rep_to_excel.NSS_RxRate_get()
            posX = write_single(posX, NSS_Rx_Rate)
        except:
            logger.error('No file')
        try:
            rep_to_excel.MCS_RxRate_get()
            posX = write_single(posX, MCS_Rx_Rate)
        except:
            logger.error('No file')
        try:
            rep_to_excel.RSSI_TXANT_get()
            posX = write_single(posX, TX_RSSI_ANT)
        except:
            logger.error('No file')
        try:
            rep_to_excel.POWER_TXANT_get()
            posX = write_single(posX, TX_POWER_ANT)
        except:
            logger.error('No file')
        try:
            rep_to_excel.RSSI_RXANT_get()
            posX = write_single(posX, RX_RSSI_ANT)
        except:
            logger.error('No file')
        try:
            rep_to_excel.POWER_RXANT_get()
            write_single(posX, RX_POWER_ANT)
        except:
            logger.error('No file')
        try:
            write_range(posX_range_txtp, posX_range_rxtp, posX_range_starssi, posX_range_aprssi)
        except:
            logger.error('No data')
        if ANGLE_NUM > 1:
            pass
            # write_radar(posX_range_txtp, 'B')
            # write_radar(posX_range_rxtp, 'J')
        write_row()
    workbook.save('./Report/' + filename)


if __name__ == "__main__":
    print('AP is ', AP_TYPE)
    Generate_Test_Report()
