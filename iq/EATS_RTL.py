#!/user/bin/env python
# encoding: utf-8
# @time      : 2020/3/4 14:08

__author__ = 'Ethan'


from __future__ import division
from openpyxl import load_workbook
from colorama import init, Fore, Style
from parameters import IQ_IP, IQ_IP_INTERFERE, PATHLOSS_WANGTED, PATHLOSS_INTERFERE, DUT_IP, DUT_USERNAME, DUT_PASSWORD, \
    DUT_COM, DUT_BAUDRATE, EXT1, EXT2, EXT3, ID_2G, ID_5G_LOW, ID_5G_HIGH, LOG_ENABLE, CALI_2G, CALI_5G, \
    AUTO_ADJUST_POWER, accuracy_limit_left, accuracy_limit_right, RX_PACKETS, RX_DYNAMIC
from iq import IQxel
import os
import sys
import csv
from telnetlib import Telnet
import time
import re
import logging

logger = logging.getLogger()
# GEN TIME
now_time = time.strftime("%Y%m%d%H%M%S", time.localtime())
# logger.debug(now_time)
logger.setLevel(logging.DEBUG)  # logger的总开关，只有大于Debug的日志才能被logger对象处理

# 第二步，创建一个handler，用于写入日志文件
file_handler = logging.FileHandler('./log/log_' + now_time + '.txt', mode='w')
file_handler.setLevel(logging.DEBUG)  # 输出到file的log等级的开关
# 创建该handler的formatter
file_handler.setFormatter(
    logging.Formatter(
        fmt='%(asctime)s - %(filename)s[line:%(lineno)d] - %(levelname)s: %(message)s',
        datefmt='%Y-%m-%d %H:%M:%S')
)
# 添加handler到logger中
logger.addHandler(file_handler)

# 第三步，创建一个handler，用于输出到控制台
console_handler = logging.StreamHandler()
if LOG_ENABLE == '1':
    console_handler.setLevel(logging.DEBUG)  # 输出到控制台的log等级的开关
else:
    console_handler.setLevel(logging.INFO)  # 输出到控制台的log等级的开关
# 创建该handler的formatter
console_handler.setFormatter(
    logging.Formatter(
        fmt='%(asctime)s - %(levelname)s: %(message)s',
        datefmt='%Y-%m-%d %H:%M:%S')
)
logger.addHandler(console_handler)


class DUT:
    def __init__(self):
        self.tn = Telnet()
        self.tn.set_debuglevel(int(LOG_ENABLE))

    def close(self):
        if self.tn is not None:
            self.tn.close()
            self.tn = None

    def login(self, host, username, password):
        self.tn.open(host, port=23)
        time.sleep(1)
        command_result = self.tn.read_very_eager().decode('ascii')
        logger.debug(command_result)
        self.tn.read_until(b'Login as: ', timeout=1)
        self.tn.write(username.encode('ascii') + b'\n')
        self.tn.read_until(b'Password: ', timeout=1)
        self.tn.write(password.encode('ascii') + b'\n')

        time.sleep(1)
        command_result = self.tn.read_very_eager().decode('ascii')
        logger.debug(command_result)
        if 'wrong' not in command_result:
            logging.info('%s Sign up' % host)
            return True
        else:
            logging.warning('%s Login Fail' % host)
            return False

    def init(self, str1, str2, str3):
        self.tn.read_until(b'> ', timeout=1)
        self.tn.write(str1.encode('ascii') + b'\n')

        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(str2.encode('ascii') + b'\n')

        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(str3.encode('ascii') + b'\n')
        # self.tn.read_until(b'# ', timeout=1)
        # self.tn.write(b'\r\n')
        # time.sleep(5)

        # self.tn.read_until(b'#', timeout=1)
        # self.tn.write(str4.encode('ascii') + b'\n')
        # # self.tn.read_until(b'#', timeout=1)
        # # self.tn.write(b'\r\n')
        # # time.sleep(5)

    def ex_command(self, command):
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(command.encode('ascii') + b'\n')
        time.sleep(2)
        command_result = self.tn.read_very_eager().decode('ascii')
        logging.info('\n%s' % command_result)

    def tx(self):
        if int(channel) < 5000:
            band = ID_2G
            bands = '2'
            bandl = 'b'
        elif int(channel) < 5500:
            band = ID_5G_LOW
            bands = '5'
            bandl = 'a'
        else:
            band = ID_5G_HIGH
            bands = '5'
            bandl = 'a'
        # for 2.4g
        self.tn.write(b'wl -i wl%s ap\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s ssid ""\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s down\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s isup\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s mpc 0\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s ap\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s ssid ""\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s down\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s isup\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s country ALL\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s wsec 0\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s stbc_rx 1\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s scansuppress 1\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s ap\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s ssid ""\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s down\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s isup\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s band auto\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s channels\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s txbf 0\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s isup\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s up\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s isup\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s ap\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s ssid ""\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s down\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s isup\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s hw_txchain\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s hw_rxchain\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s spect 0\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s ibss_gmode -1\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s phytype\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s bw_cap 5g 255\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s bw_cap 2g 3\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s ap\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s ap\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s ssid ""\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s down\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s isup\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s mbss 0\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s frameburst 0\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s ap\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s ssid ""\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s down\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s isup\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s ampdu 0\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s gmode\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s gmode auto\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s isup\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s up\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s isup\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s PM 0\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s isup\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s longpkt 1\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s pkteng_maxlen\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s stbc_tx 0\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s ap\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s ssid ""\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s down\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s isup\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s bi 65535\n' % band.encode('ascii'))
        if int(bw) == 40:
            bws = '4'
        else:
            bws = '-1'
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s mimo_txbw %s\n' % (band.encode('ascii'), bws.encode('ascii')))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s ap\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s ssid ""\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s down\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s isup\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s ldpc_cap\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s 2g_rate auto\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s 5g_rate auto\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s ampdu 0\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s frameburst 1\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s hw_txchain\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s txchain 15\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s rxchain 15\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s hw_txchain\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s hw_rxchain\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s hw_txchain\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s hw_rxchain\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s hw_txchain\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s hw_rxchain\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s isup\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s spatial_policy 1\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s stbc_tx\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s txchain\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        # (11b:wl -i wl0 txcore  -k 2)
        # (11g:wl -i wl0 txcore  -o 1)
        # (11n:wl -i wl0 txcore  -s 1 -c 1)
        # (11ac:wl -i wl0 txcore  -s 1 -c 1)
        # (11ax:wl -i wl0 txcore  -s 1 -c 1)
        if mode == '11b':
            chains = 'k'
            addtional_chain = ''
        elif mode == '11g' or mode == '11a':
            chains = 'o'
            addtional_chain = ''
        elif mode == '11n' or mode == '11ac' or mode == '11ax':
            chains = 'c'
            addtional_chain = '-s 1'
        else:
            # logger.info('no mode')
            print('mode error', mode)
            exit(-1)
        if str(chain) == '0':
            chain_value = '1'
        elif str(chain) == '1':
            chain_value = '2'
        elif str(chain) == '2':
            chain_value = '4'
        elif str(chain) == '3':
            chain_value = '8'
        else:
            # logger.info('wrong chain')
            print('chain error')
            exit(-1)
        self.tn.write(b'wl -i wl%s txcore %s -%s %s\n' % (band.encode('ascii'), addtional_chain.encode('ascii'),
                                                          chains.encode('ascii'), chain_value.encode('ascii')))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s hw_txchain\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s hw_rxchain\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s phy_watchdog\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s phy_watchdog 0\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s chanspec\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s ap\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s ssid ""\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s down\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s isup\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s ap\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s ssid ""\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s down\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s isup\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s band %s\n' % (band.encode('ascii'), bandl.encode('ascii')))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s vht_features 3\n' % band.encode('ascii'))
        if int(channel) < 5000:
            channels = '{:.0f}'.format((int(channel) - 2407) / 5)
            if bw == '40':
                bw_channel = '40l'
            else:
                bw_channel = bw
        else:
            channels = '{:.0f}'.format((int(channel) - 5000) / 5)
            bw_channel = bw
        if bw == '40':
            channels = int(channels) - 2
        elif bw == '80':
            channels = int(channels) - 6
        elif bw == '160':
            channels = int(channels) - 14
        channels = str(channels)
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s chanspec %s/%s\n' % (band.encode('ascii'), channels.encode('ascii'),
                                                        bw_channel.encode('ascii')))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s isup\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s up\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s isup\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s phy_watchdog\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s phy_forcecal 1\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s phy_activecal\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s phy_watchdog\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s ap\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s ssid ""\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s isup\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s ssid test2\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s ap\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s ssid ""\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s he cap\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s ap\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s ssid ""\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s down\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s isup\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s he features 7\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s he cap\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s ap\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s ssid ""\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s down\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s isup\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s ampdu_rts 0\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s ldpc_cap\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s sgi_tx\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        # (11g:2g_rate -r 6 -b 20)
        # (11n:2g_rate -h 0 --ldpc -b 20)
        # (11ac:2g_rate -v 0 -s 1 --ldpc -b 20)
        # (11ax:2g_rate -e 0 -s 1 -i 1 --ldpc -b 20)
        if mode == '11b' or mode == '11g' or mode == '11a':
            ratess = 'r'
            addtional = ''
        elif mode == '11n':
            ratess = 'h'
            addtional = '--ldpc'
        elif mode == '11ac':
            ratess = 'v'
            addtional = '-s 1 --ldpc'
        elif mode == '11ax':
            ratess = 'e'
            addtional = '-s 1 -i 0 --ldpc'
        else:
            # logger.info('no mode')
            exit(-1)
        self.tn.write(b'wl -i wl%s %sg_rate -%s %s %s -b %s\n' % (band.encode('ascii'), bands.encode('ascii'),
                                                                  ratess.encode('ascii'), rates.encode('ascii'),
                                                                  addtional.encode('ascii'), bw.encode('ascii')))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s phy_txpwrctrl 1\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s isup\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s up\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s isup\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s pkteng_stop rx\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s pkteng_stop tx\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s txpwr1 -d -o %d\n' % (band.encode('ascii'), gain))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s isup\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s longpkt 1\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s isup\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s phy_watchdog\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s phy_forcecal 1\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s phy_activecal\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s cur_etheraddr\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s pkteng_stop rx\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s pkteng_stop tx\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s mpc 0\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s isup\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s interference\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s interference_override 0\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s pkteng_maxlen\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s ap\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s ssid ""\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(
            b'wl -i wl%s pkteng_start 00:11:22:33:44:55 tx 650 1500 0 00:90:4C:1F:D0:10\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        print('TX COMMANDS DONE')

    def get_paras(self):
        # self.tn.read_until(b'# ', timeout=1)
        # self.tn.write(b'hipriv.sh "vap0 set_tx_pow rf_reg_ctl 0"' + b'\n')
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'flash allhw' + b'\n')
        command_result = self.tn.read_until(b'HW_WLAN1', timeout=2)
        logger.debug(command_result)
        command_result = re.search(b'HW_wlan%s\w+=\w+', command_result)
        logger.debug(command_result.group())
        default_value = re.sub('^[0~9]', '', command_result.group().decode('ascii'))
        print('Default Value(HEX):', default_value)
        return default_value

    def tx_off(self):
        if int(channel) < 5000:
            band = ID_2G
        elif int(channel) < 5500:
            band = ID_5G_LOW
        else:
            band = ID_5G_HIGH
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s pkteng_stop rx\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s pkteng_stop tx\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s interference_override 0\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s phy_watchdog 1\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s isup\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s longpkt 0\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s ap\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s ssid ""\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s down\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s isup\n' % band.encode('ascii'))

    def adjust_power(self, gain):
        if int(channel) < 5000:
            band = ID_2G
        elif int(channel) > 5499:
            band = ID_5G_HIGH
        else:
            band = ID_5G_LOW
        self.tn.read_until(b'# ', timeout=1)
        # self.tn.write(b'iwpriv wlan%s mp_txpower patha=%d,pathb=%d' % (band.encode('ascii'), pwra_paras, pwrb_paras) + b'\n')
        self.tn.write(b'wl -i wl%s txpwr1 -q -o %d\n' % (band.encode('ascii'), gain))

    def rx(self):
        if int(channel) < 5000:
            band = ID_2G
            bands = '2'
            bandl = 'b'
        elif int(channel) < 5500:
            band = ID_5G_LOW
            bands = '5'
            bandl = 'a'
        else:
            band = ID_5G_HIGH
            bands = '5'
            bandl = 'a'
        # for 2.4g
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s mpc 0\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s ap\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s ssid ""\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s down\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s isup\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s country ALL\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s wsec 0\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s stbc_rx 1\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s scansuppress 1\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s ap\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s ssid ""\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s down\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s isup\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s band auto\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s channels\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s txbf 0\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s isup\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s up\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s isup\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s ap\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s ssid ""\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s down\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s isup\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s hw_txchain\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s hw_rxchain\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s spect 0\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s ibss_gmode -1\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s phytype\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s bw_cap 5g 255\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s bw_cap 2g 3\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s ap\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s ap\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s ssid ""\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s down\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s isup\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s mbss 0\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s frameburst 0\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s ap\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s ssid ""\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s down\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s isup\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s ampdu 0\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s gmode\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s gmode auto\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s isup\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s up\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s isup\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s PM 0\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s isup\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s longpkt 1\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s pkteng_maxlen\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s stbc_tx 0\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s ap\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s ssid ""\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s down\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s isup\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s bi 65535\n' % band.encode('ascii'))
        if int(bw) == 40:
            bws = '4'
        else:
            bws = '-1'
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s mimo_txbw %s\n' % (band.encode('ascii'), bws.encode('ascii')))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s ap\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s ssid ""\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s down\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s isup\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s ldpc_cap\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s 2g_rate auto\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s 5g_rate auto\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s ampdu 0\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s frameburst 1\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s hw_txchain\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s txchain 15\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s rxchain 15\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s hw_txchain\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s hw_rxchain\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s hw_txchain\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s hw_rxchain\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s hw_txchain\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s hw_rxchain\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s isup\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s spatial_policy 1\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s stbc_tx\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s txchain\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        # (11b:wl -i wl0 txcore  -k 2)
        # (11g:wl -i wl0 txcore  -o 1)
        # (11n:wl -i wl0 txcore  -s 1 -c 1)
        # (11ac:wl -i wl0 txcore  -s 1 -c 1)
        # (11ax:wl -i wl0 txcore  -s 1 -c 1)
        if mode == '11b':
            chains = 'k'
            addtional_chain = ''
        elif mode == '11g' or mode == '11a':
            chains = 'o'
            addtional_chain = ''
        elif mode == '11n' or mode == '11ac' or mode == '11ax':
            chains = 'c'
            addtional_chain = '-s 1'
        else:
            # logger.info('no mode')
            exit(-1)
        if str(chain) == '0':
            chain_value = '1'
        elif str(chain) == '1':
            chain_value = '2'
        elif str(chain) == '2':
            chain_value = '4'
        elif str(chain) == '3':
            chain_value = '8'
        else:
            # logger.info('wrong chain')
            exit(-1)
        self.tn.write(b'wl -i wl%s txcore %s -%s %s\n' % (band.encode('ascii'), addtional_chain.encode('ascii'),
                                                          chains.encode('ascii'), chain_value.encode('ascii')))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s rxchain %s\n' % (band.encode('ascii'), chain_value.encode('ascii')))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s hw_txchain\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s hw_rxchain\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s phy_watchdog\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s ap\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s ssid ""\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s down\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s isup\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s ap\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s ssid ""\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s down\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s isup\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s band %s\n' % (band.encode('ascii'), bandl.encode('ascii')))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s vht_features 3\n' % band.encode('ascii'))
        if int(channel) < 5000:
            channels = '{:.0f}'.format((int(channel) - 2407) / 5)
            if bw == '40':
                bw_channel = '40l'
            else:
                bw_channel = bw
        else:
            channels = '{:.0f}'.format((int(channel) - 5000) / 5)
            bw_channel = bw
        if bw == '40':
            channels = int(channels) - 2
        elif bw == '80':
            channels = int(channels) - 6
        elif bw == '160':
            channels = int(channels) - 14
        channels = str(channels)
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s chanspec %s/%s\n' % (band.encode('ascii'), channels.encode('ascii'),
                                                        bw_channel.encode('ascii')))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s isup\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s up\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s isup\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s phy_watchdog\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s phy_forcecal 1\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s phy_activecal\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s phy_watchdog\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s ap\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s ssid ""\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s isup\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s ssid test2\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s ap\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s ssid ""\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s ldpc_cap\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s sgi_tx\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s he cap\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s he cap\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        # (11g:2g_rate -r 6 -b 20)
        # (11n:2g_rate -h 0 --ldpc -b 20)
        # (11ac:2g_rate -v 0 -s 1 --ldpc -b 20)
        # (11ax:2g_rate -e 0 -s 1 -i 1 --ldpc -b 20)
        if mode == '11b' or mode == '11g' or mode == '11a':
            ratess = 'r'
            addtional = ''
        elif mode == '11n':
            ratess = 'h'
            addtional = '--ldpc'
        elif mode == '11ac':
            ratess = 'v'
            addtional = '-s 1 --ldpc'
        elif mode == '11ax':
            ratess = 'e'
            addtional = '-s 1 -i 1 --ldpc'
        else:
            # logger.info('no mode')
            exit(-1)
        self.tn.write(b'wl -i wl%s %sg_rate -%s %s %s -b %s\n' % (band.encode('ascii'), bands.encode('ascii'),
                                                                  ratess.encode('ascii'), rates.encode('ascii'),
                                                                  addtional.encode('ascii'), bw.encode('ascii')))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s counters --nz --rx --ucode\n' % band.encode('ascii'))
        print('RX COMMANDS DONE')

    # def get_statistics(self):
    #     if int(channel) < 5000:
    #         band = ID_2G
    #     elif int(channel) > 5499:
    #         band = ID_5G_HIGH
    #     else:
    #         band = ID_5G_LOW
    #     self.tn.read_until(b'# ', timeout=1)
    #     self.tn.write(b'wl -i wl%s counters --rx --ucode\n' % band.encode('ascii'))
    #     command_result = self.tn.read_until(b'# ', timeout=2)
    #     # print(command_result)
    #     time.sleep(1)
    #     if bw == '160':
    #         rx_counts = re.findall(b'rxstrt(.\d+)', command_result)[0].split()[0].decode('utf-8')
    #     else:
    #         rx_counts = re.findall(b'rxdtocast(.\d+)', command_result)[0].split()[0].decode('utf-8')
    #     logger.info('Packets: ' + RX_PACKETS + ' Packets Counts: ' + rx_counts)
    #     return rx_counts

    def get_statistics(self):
        if int(channel) < 5000:
            band = ID_2G
        elif int(channel) > 5499:
            band = ID_5G_HIGH
        else:
            band = ID_5G_LOW
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s counters --rx --ucode\n' % band.encode('ascii'))
        command_result = self.tn.read_until(b'# ', timeout=2)
        # print(command_result)
        time.sleep(1)
        rx_counts = re.findall(b'rxdtocast(.\d+)', command_result)[0].split()[0].decode('utf-8')
        logger.info('Packets: ' + RX_PACKETS + ' Packets Counts: ' + rx_counts)
        return rx_counts

    def rx_off(self):
        if int(channel) < 5000:
            band = ID_2G
        elif int(channel) < 5500:
            band = ID_5G_LOW
        else:
            band = ID_5G_HIGH
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s ap\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s ssid ""\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s down\n' % band.encode('ascii'))
        self.tn.read_until(b'# ', timeout=1)
        self.tn.write(b'wl -i wl%s isup\n' % band.encode('ascii'))

    def set_default(self):
        self.tn.write(b'wl -i wl0 down\r\n')
        self.tn.write(b'wl -i wl1 down\r\n')
        self.tn.write(b'wl -i wl2 down\r\n')


if __name__ == '__main__':
    init(autoreset=True)

    directory_r = os.path.exists(r'./Result')
    if directory_r is False:
        os.makedirs('Result')

    directory_l = os.path.exists(r'./log')
    if directory_l is False:
        os.makedirs('log')

    # input sn
    try:
        Number = sys.argv[1]
    except:
        Number = input("SN:")
    sn = "SN_" + Number

    try:
        filename = sys.argv[2]
    except:
        filename = 'TEST_FLOW.txt'

    # test define
    target_pwr_1M, target_pwr_2M, target_pwr_5_5M, target_pwr_11M, target_pwr_6M, target_pwr_9M, \
    target_pwr_12M, target_pwr_18M, target_pwr_24M, target_pwr_36M, target_pwr_48M, target_pwr_54M, \
    target_pwr_HT20_MCS0, target_pwr_HT20_MCS1, target_pwr_HT20_MCS2, \
    target_pwr_HT20_MCS3, target_pwr_HT20_MCS4, target_pwr_HT20_MCS5, target_pwr_HT20_MCS6, \
    target_pwr_HT20_MCS7, target_pwr_HT40_MCS0, target_pwr_HT40_MCS1, target_pwr_HT40_MCS2, \
    target_pwr_HT40_MCS3, target_pwr_HT40_MCS4, target_pwr_HT40_MCS5, target_pwr_HT40_MCS6, \
    target_pwr_HT40_MCS7, target_pwr_VHT20_MCS0, target_pwr_VHT20_MCS1, target_pwr_VHT20_MCS2, \
    target_pwr_VHT20_MCS3, target_pwr_VHT20_MCS4, target_pwr_VHT20_MCS5, target_pwr_VHT20_MCS6, \
    target_pwr_VHT20_MCS7, target_pwr_VHT20_MCS8, target_pwr_VHT40_MCS0, target_pwr_VHT40_MCS1, \
    target_pwr_VHT40_MCS2, target_pwr_VHT40_MCS3, target_pwr_VHT40_MCS4, target_pwr_VHT40_MCS5, \
    target_pwr_VHT40_MCS6, target_pwr_VHT40_MCS7, target_pwr_VHT40_MCS8, target_pwr_VHT40_MCS9, \
    target_pwr_VHT80_MCS0, target_pwr_VHT80_MCS1, target_pwr_VHT80_MCS2, target_pwr_VHT80_MCS3, \
    target_pwr_VHT80_MCS4, target_pwr_VHT80_MCS5, target_pwr_VHT80_MCS6, target_pwr_VHT80_MCS7, \
    target_pwr_VHT80_MCS8, target_pwr_VHT80_MCS9, \
    target_pwr_VHT160_MCS0, target_pwr_VHT160_MCS1, target_pwr_VHT160_MCS2, target_pwr_VHT160_MCS3, \
    target_pwr_VHT160_MCS4, target_pwr_VHT160_MCS5, target_pwr_VHT160_MCS6, target_pwr_VHT160_MCS7, \
    target_pwr_VHT160_MCS8, target_pwr_VHT160_MCS9, target_pwr_HE20_HE0, target_pwr_HE20_HE1, target_pwr_HE20_HE2, \
    target_pwr_HE20_HE3, target_pwr_HE20_HE4, target_pwr_HE20_HE5, \
    target_pwr_HE20_HE6, target_pwr_HE20_HE7, target_pwr_HE20_HE8, target_pwr_HE20_HE9, target_pwr_HE20_HE10, \
    target_pwr_HE20_HE11, target_pwr_HE40_HE0, target_pwr_HE40_HE1, target_pwr_HE40_HE2, target_pwr_HE40_HE3, \
    target_pwr_HE40_HE4, target_pwr_HE40_HE5, target_pwr_HE40_HE6, target_pwr_HE40_HE7, target_pwr_HE40_HE8, \
    target_pwr_HE40_HE9, target_pwr_HE40_HE10, target_pwr_HE40_HE11, target_pwr_HE80_HE0, target_pwr_HE80_HE1, \
    target_pwr_HE80_HE2, target_pwr_HE80_HE3, target_pwr_HE80_HE4, target_pwr_HE80_HE5, target_pwr_HE80_HE6, \
    target_pwr_HE80_HE7, target_pwr_HE80_HE8, target_pwr_HE80_HE9, target_pwr_HE80_HE10, target_pwr_HE80_HE11, \
    target_pwr_HE160_HE0, target_pwr_HE160_HE1, target_pwr_HE160_HE2, target_pwr_HE160_HE3, target_pwr_HE160_HE4, \
    target_pwr_HE160_HE5, target_pwr_HE160_HE6, target_pwr_HE160_HE7, target_pwr_HE160_HE8, target_pwr_HE160_HE9, \
    target_pwr_HE160_HE10, target_pwr_HE160_HE11 = [None] * 115

    target_EVM_1M, target_EVM_2M, target_EVM_5_5M, target_EVM_11M, target_EVM_6M, target_EVM_9M, target_EVM_12M, \
    target_EVM_18M, target_EVM_24M, target_EVM_36M, target_EVM_48M, target_EVM_54M, target_EVM_HT20_MCS0, \
    target_EVM_HT20_MCS1, target_EVM_HT20_MCS2, target_EVM_HT20_MCS3, target_EVM_HT20_MCS4, target_EVM_HT20_MCS5, \
    target_EVM_HT20_MCS6, target_EVM_HT20_MCS7, target_EVM_HT40_MCS0, target_EVM_HT40_MCS1, target_EVM_HT40_MCS2, \
    target_EVM_HT40_MCS3, target_EVM_HT40_MCS4, target_EVM_HT40_MCS5, target_EVM_HT40_MCS6, target_EVM_HT40_MCS7, \
    target_EVM_VHT20_MCS0, target_EVM_VHT20_MCS1, target_EVM_VHT20_MCS2, target_EVM_VHT20_MCS3, target_EVM_VHT20_MCS4, \
    target_EVM_VHT20_MCS5, target_EVM_VHT20_MCS6, target_EVM_VHT20_MCS7, target_EVM_VHT20_MCS8, target_EVM_VHT40_MCS0, \
    target_EVM_VHT40_MCS1, target_EVM_VHT40_MCS2, target_EVM_VHT40_MCS3, target_EVM_VHT40_MCS4, target_EVM_VHT40_MCS5, \
    target_EVM_VHT40_MCS6, target_EVM_VHT40_MCS7, target_EVM_VHT40_MCS8, target_EVM_VHT40_MCS9, target_EVM_VHT80_MCS0, \
    target_EVM_VHT80_MCS1, target_EVM_VHT80_MCS2, target_EVM_VHT80_MCS3, target_EVM_VHT80_MCS4, target_EVM_VHT80_MCS5, \
    target_EVM_VHT80_MCS6, target_EVM_VHT80_MCS7, target_EVM_VHT80_MCS8, target_EVM_VHT80_MCS9, \
    target_EVM_VHT160_MCS0, target_EVM_VHT160_MCS1, target_EVM_VHT160_MCS2, target_EVM_VHT160_MCS3, \
    target_EVM_VHT160_MCS4, target_EVM_VHT160_MCS5, target_EVM_VHT160_MCS6, target_EVM_VHT160_MCS7, \
    target_EVM_VHT160_MCS8, target_EVM_VHT160_MCS9, target_EVM_HE20_HE0, target_EVM_HE20_HE1, target_EVM_HE20_HE2, \
    target_EVM_HE20_HE3, target_EVM_HE20_HE4, target_EVM_HE20_HE5, \
    target_EVM_HE20_HE6, target_EVM_HE20_HE7, target_EVM_HE20_HE8, target_EVM_HE20_HE9, target_EVM_HE20_HE10, \
    target_EVM_HE20_HE11, target_EVM_HE40_HE0, target_EVM_HE40_HE1, target_EVM_HE40_HE2, target_EVM_HE40_HE3, \
    target_EVM_HE40_HE4, target_EVM_HE40_HE5, target_EVM_HE40_HE6, target_EVM_HE40_HE7, target_EVM_HE40_HE8, \
    target_EVM_HE40_HE9, target_EVM_HE40_HE10, target_EVM_HE40_HE11, target_EVM_HE80_HE0, target_EVM_HE80_HE1, \
    target_EVM_HE80_HE2, target_EVM_HE80_HE3, target_EVM_HE80_HE4, target_EVM_HE80_HE5, target_EVM_HE80_HE6, \
    target_EVM_HE80_HE7, target_EVM_HE80_HE8, target_EVM_HE80_HE9, target_EVM_HE80_HE10, target_EVM_HE80_HE11, \
    target_EVM_HE160_HE0, target_EVM_HE160_HE1, target_EVM_HE160_HE2, target_EVM_HE160_HE3, target_EVM_HE160_HE4, \
    target_EVM_HE160_HE5, target_EVM_HE160_HE6, target_EVM_HE160_HE7, target_EVM_HE160_HE8, target_EVM_HE160_HE9, \
    target_EVM_HE160_HE10, target_EVM_HE160_HE11 = [None] * 115

    target_sens_1M, target_sens_2M, target_sens_5_5M, target_sens_11M, target_sens_6M, target_sens_9M, target_sens_12M, \
    target_sens_18M, target_sens_24M, target_sens_36M, target_sens_48M, target_sens_54M, target_sens_HT20_MCS0, \
    target_sens_HT20_MCS1, target_sens_HT20_MCS2, target_sens_HT20_MCS3, target_sens_HT20_MCS4, target_sens_HT20_MCS5, \
    target_sens_HT20_MCS6, target_sens_HT20_MCS7, target_sens_HT40_MCS0, target_sens_HT40_MCS1, target_sens_HT40_MCS2, \
    target_sens_HT40_MCS3, target_sens_HT40_MCS4, target_sens_HT40_MCS5, target_sens_HT40_MCS6, target_sens_HT40_MCS7, \
    target_sens_VHT20_MCS0, target_sens_VHT20_MCS1, target_sens_VHT20_MCS2, target_sens_VHT20_MCS3, target_sens_VHT20_MCS4, \
    target_sens_VHT20_MCS5, target_sens_VHT20_MCS6, target_sens_VHT20_MCS7, target_sens_VHT20_MCS8, target_sens_VHT40_MCS0, \
    target_sens_VHT40_MCS1, target_sens_VHT40_MCS2, target_sens_VHT40_MCS3, target_sens_VHT40_MCS4, target_sens_VHT40_MCS5, \
    target_sens_VHT40_MCS6, target_sens_VHT40_MCS7, target_sens_VHT40_MCS8, target_sens_VHT40_MCS9, target_sens_VHT80_MCS0, \
    target_sens_VHT80_MCS1, target_sens_VHT80_MCS2, target_sens_VHT80_MCS3, target_sens_VHT80_MCS4, target_sens_VHT80_MCS5, \
    target_sens_VHT80_MCS6, target_sens_VHT80_MCS7, target_sens_VHT80_MCS8, target_sens_VHT80_MCS9, \
    target_sens_VHT160_MCS0, target_sens_VHT160_MCS1, target_sens_VHT160_MCS2, target_sens_VHT160_MCS3, \
    target_sens_VHT160_MCS4, target_sens_VHT160_MCS5, target_sens_VHT160_MCS6, target_sens_VHT160_MCS7, \
    target_sens_VHT160_MCS8, target_sens_VHT160_MCS9, target_sens_HE20_HE0, target_sens_HE20_HE1, target_sens_HE20_HE2, \
    target_sens_HE20_HE3, target_sens_HE20_HE4, target_sens_HE20_HE5, \
    target_sens_HE20_HE6, target_sens_HE20_HE7, target_sens_HE20_HE8, target_sens_HE20_HE9, target_sens_HE20_HE10, \
    target_sens_HE20_HE11, target_sens_HE40_HE0, target_sens_HE40_HE1, target_sens_HE40_HE2, target_sens_HE40_HE3, \
    target_sens_HE40_HE4, target_sens_HE40_HE5, target_sens_HE40_HE6, target_sens_HE40_HE7, target_sens_HE40_HE8, \
    target_sens_HE40_HE9, target_sens_HE40_HE10, target_sens_HE40_HE11, target_sens_HE80_HE0, target_sens_HE80_HE1, \
    target_sens_HE80_HE2, target_sens_HE80_HE3, target_sens_HE80_HE4, target_sens_HE80_HE5, target_sens_HE80_HE6, \
    target_sens_HE80_HE7, target_sens_HE80_HE8, target_sens_HE80_HE9, target_sens_HE80_HE10, target_sens_HE80_HE11, \
    target_sens_HE160_HE0, target_sens_HE160_HE1, target_sens_HE160_HE2, target_sens_HE160_HE3, target_sens_HE160_HE4, \
    target_sens_HE160_HE5, target_sens_HE160_HE6, target_sens_HE160_HE7, target_sens_HE160_HE8, target_sens_HE160_HE9, \
    target_sens_HE160_HE10, target_sens_HE160_HE11 = [None] * 115

    target_MAX_1M, target_MAX_2M, target_MAX_5_5M, target_MAX_11M, target_MAX_6M, target_MAX_9M, target_MAX_12M, \
    target_MAX_18M, target_MAX_24M, target_MAX_36M, target_MAX_48M, target_MAX_54M, target_MAX_HT20_MCS0, \
    target_MAX_HT20_MCS1, target_MAX_HT20_MCS2, target_MAX_HT20_MCS3, target_MAX_HT20_MCS4, target_MAX_HT20_MCS5, \
    target_MAX_HT20_MCS6, target_MAX_HT20_MCS7, target_MAX_HT40_MCS0, target_MAX_HT40_MCS1, target_MAX_HT40_MCS2, \
    target_MAX_HT40_MCS3, target_MAX_HT40_MCS4, target_MAX_HT40_MCS5, target_MAX_HT40_MCS6, target_MAX_HT40_MCS7, \
    target_MAX_VHT20_MCS0, target_MAX_VHT20_MCS1, target_MAX_VHT20_MCS2, target_MAX_VHT20_MCS3, target_MAX_VHT20_MCS4, \
    target_MAX_VHT20_MCS5, target_MAX_VHT20_MCS6, target_MAX_VHT20_MCS7, target_MAX_VHT20_MCS8, target_MAX_VHT40_MCS0, \
    target_MAX_VHT40_MCS1, target_MAX_VHT40_MCS2, target_MAX_VHT40_MCS3, target_MAX_VHT40_MCS4, target_MAX_VHT40_MCS5, \
    target_MAX_VHT40_MCS6, target_MAX_VHT40_MCS7, target_MAX_VHT40_MCS8, target_MAX_VHT40_MCS9, target_MAX_VHT80_MCS0, \
    target_MAX_VHT80_MCS1, target_MAX_VHT80_MCS2, target_MAX_VHT80_MCS3, target_MAX_VHT80_MCS4, target_MAX_VHT80_MCS5, \
    target_MAX_VHT80_MCS6, target_MAX_VHT80_MCS7, target_MAX_VHT80_MCS8, target_MAX_VHT80_MCS9, \
    target_MAX_VHT160_MCS0, target_MAX_VHT160_MCS1, target_MAX_VHT160_MCS2, target_MAX_VHT160_MCS3, \
    target_MAX_VHT160_MCS4, target_MAX_VHT160_MCS5, target_MAX_VHT160_MCS6, target_MAX_VHT160_MCS7, \
    target_MAX_VHT160_MCS8, target_MAX_VHT160_MCS9, target_MAX_HE20_HE0, target_MAX_HE20_HE1, target_MAX_HE20_HE2, \
    target_MAX_HE20_HE3, target_MAX_HE20_HE4, target_MAX_HE20_HE5, \
    target_MAX_HE20_HE6, target_MAX_HE20_HE7, target_MAX_HE20_HE8, target_MAX_HE20_HE9, target_MAX_HE20_HE10, \
    target_MAX_HE20_HE11, target_MAX_HE40_HE0, target_MAX_HE40_HE1, target_MAX_HE40_HE2, target_MAX_HE40_HE3, \
    target_MAX_HE40_HE4, target_MAX_HE40_HE5, target_MAX_HE40_HE6, target_MAX_HE40_HE7, target_MAX_HE40_HE8, \
    target_MAX_HE40_HE9, target_MAX_HE40_HE10, target_MAX_HE40_HE11, target_MAX_HE80_HE0, target_MAX_HE80_HE1, \
    target_MAX_HE80_HE2, target_MAX_HE80_HE3, target_MAX_HE80_HE4, target_MAX_HE80_HE5, target_MAX_HE80_HE6, \
    target_MAX_HE80_HE7, target_MAX_HE80_HE8, target_MAX_HE80_HE9, target_MAX_HE80_HE10, target_MAX_HE80_HE11, \
    target_MAX_HE160_HE0, target_MAX_HE160_HE1, target_MAX_HE160_HE2, target_MAX_HE160_HE3, target_MAX_HE160_HE4, \
    target_MAX_HE160_HE5, target_MAX_HE160_HE6, target_MAX_HE160_HE7, target_MAX_HE160_HE8, target_MAX_HE160_HE9, \
    target_MAX_HE160_HE10, target_MAX_HE160_HE11 = [None] * 115

    target_aj_1M, target_aj_2M, target_aj_5_5M, target_aj_11M, target_aj_6M, target_aj_9M, target_aj_12M, \
    target_aj_18M, target_aj_24M, target_aj_36M, target_aj_48M, target_aj_54M, target_aj_HT20_MCS0, \
    target_aj_HT20_MCS1, target_aj_HT20_MCS2, target_aj_HT20_MCS3, target_aj_HT20_MCS4, target_aj_HT20_MCS5, \
    target_aj_HT20_MCS6, target_aj_HT20_MCS7, target_aj_HT40_MCS0, target_aj_HT40_MCS1, target_aj_HT40_MCS2, \
    target_aj_HT40_MCS3, target_aj_HT40_MCS4, target_aj_HT40_MCS5, target_aj_HT40_MCS6, target_aj_HT40_MCS7, \
    target_aj_VHT20_MCS0, target_aj_VHT20_MCS1, target_aj_VHT20_MCS2, target_aj_VHT20_MCS3, target_aj_VHT20_MCS4, \
    target_aj_VHT20_MCS5, target_aj_VHT20_MCS6, target_aj_VHT20_MCS7, target_aj_VHT20_MCS8, target_aj_VHT40_MCS0, \
    target_aj_VHT40_MCS1, target_aj_VHT40_MCS2, target_aj_VHT40_MCS3, target_aj_VHT40_MCS4, target_aj_VHT40_MCS5, \
    target_aj_VHT40_MCS6, target_aj_VHT40_MCS7, target_aj_VHT40_MCS8, target_aj_VHT40_MCS9, target_aj_VHT80_MCS0, \
    target_aj_VHT80_MCS1, target_aj_VHT80_MCS2, target_aj_VHT80_MCS3, target_aj_VHT80_MCS4, target_aj_VHT80_MCS5, \
    target_aj_VHT80_MCS6, target_aj_VHT80_MCS7, target_aj_VHT80_MCS8, target_aj_VHT80_MCS9, \
    target_aj_VHT160_MCS0, target_aj_VHT160_MCS1, target_aj_VHT160_MCS2, target_aj_VHT160_MCS3, \
    target_aj_VHT160_MCS4, target_aj_VHT160_MCS5, target_aj_VHT160_MCS6, target_aj_VHT160_MCS7, \
    target_aj_VHT160_MCS8, target_aj_VHT160_MCS9, target_aj_HE20_HE0, target_aj_HE20_HE1, target_aj_HE20_HE2, \
    target_aj_HE20_HE3, target_aj_HE20_HE4, target_aj_HE20_HE5, \
    target_aj_HE20_HE6, target_aj_HE20_HE7, target_aj_HE20_HE8, target_aj_HE20_HE9, target_aj_HE20_HE10, \
    target_aj_HE20_HE11, target_aj_HE40_HE0, target_aj_HE40_HE1, target_aj_HE40_HE2, target_aj_HE40_HE3, \
    target_aj_HE40_HE4, target_aj_HE40_HE5, target_aj_HE40_HE6, target_aj_HE40_HE7, target_aj_HE40_HE8, \
    target_aj_HE40_HE9, target_aj_HE40_HE10, target_aj_HE40_HE11, target_aj_HE80_HE0, target_aj_HE80_HE1, \
    target_aj_HE80_HE2, target_aj_HE80_HE3, target_aj_HE80_HE4, target_aj_HE80_HE5, target_aj_HE80_HE6, \
    target_aj_HE80_HE7, target_aj_HE80_HE8, target_aj_HE80_HE9, target_aj_HE80_HE10, target_aj_HE80_HE11, \
    target_aj_HE160_HE0, target_aj_HE160_HE1, target_aj_HE160_HE2, target_aj_HE160_HE3, target_aj_HE160_HE4, \
    target_aj_HE160_HE5, target_aj_HE160_HE6, target_aj_HE160_HE7, target_aj_HE160_HE8, target_aj_HE160_HE9, \
    target_aj_HE160_HE10, target_aj_HE160_HE11 = [None] * 115

    target_naj_1M, target_naj_2M, target_naj_5_5M, target_naj_11M, target_naj_6M, target_naj_9M, target_naj_12M, \
    target_naj_18M, target_naj_24M, target_naj_36M, target_naj_48M, target_naj_54M, target_naj_HT20_MCS0, \
    target_naj_HT20_MCS1, target_naj_HT20_MCS2, target_naj_HT20_MCS3, target_naj_HT20_MCS4, target_naj_HT20_MCS5, \
    target_naj_HT20_MCS6, target_naj_HT20_MCS7, target_naj_HT40_MCS0, target_naj_HT40_MCS1, target_naj_HT40_MCS2, \
    target_naj_HT40_MCS3, target_naj_HT40_MCS4, target_naj_HT40_MCS5, target_naj_HT40_MCS6, target_naj_HT40_MCS7, \
    target_naj_VHT20_MCS0, target_naj_VHT20_MCS1, target_naj_VHT20_MCS2, target_naj_VHT20_MCS3, target_naj_VHT20_MCS4, \
    target_naj_VHT20_MCS5, target_naj_VHT20_MCS6, target_naj_VHT20_MCS7, target_naj_VHT20_MCS8, target_naj_VHT40_MCS0, \
    target_naj_VHT40_MCS1, target_naj_VHT40_MCS2, target_naj_VHT40_MCS3, target_naj_VHT40_MCS4, target_naj_VHT40_MCS5, \
    target_naj_VHT40_MCS6, target_naj_VHT40_MCS7, target_naj_VHT40_MCS8, target_naj_VHT40_MCS9, target_naj_VHT80_MCS0, \
    target_naj_VHT80_MCS1, target_naj_VHT80_MCS2, target_naj_VHT80_MCS3, target_naj_VHT80_MCS4, target_naj_VHT80_MCS5, \
    target_naj_VHT80_MCS6, target_naj_VHT80_MCS7, target_naj_VHT80_MCS8, target_naj_VHT80_MCS9, \
    target_naj_VHT160_MCS0, target_naj_VHT160_MCS1, target_naj_VHT160_MCS2, target_naj_VHT160_MCS3, \
    target_naj_VHT160_MCS4, target_naj_VHT160_MCS5, target_naj_VHT160_MCS6, target_naj_VHT160_MCS7, \
    target_naj_VHT160_MCS8, target_naj_VHT160_MCS9, target_naj_HE20_HE0, target_naj_HE20_HE1, target_naj_HE20_HE2, \
    target_naj_HE20_HE3, target_naj_HE20_HE4, target_naj_HE20_HE5, \
    target_naj_HE20_HE6, target_naj_HE20_HE7, target_naj_HE20_HE8, target_naj_HE20_HE9, target_naj_HE20_HE10, \
    target_naj_HE20_HE11, target_naj_HE40_HE0, target_naj_HE40_HE1, target_naj_HE40_HE2, target_naj_HE40_HE3, \
    target_naj_HE40_HE4, target_naj_HE40_HE5, target_naj_HE40_HE6, target_naj_HE40_HE7, target_naj_HE40_HE8, \
    target_naj_HE40_HE9, target_naj_HE40_HE10, target_naj_HE40_HE11, target_naj_HE80_HE0, target_naj_HE80_HE1, \
    target_naj_HE80_HE2, target_naj_HE80_HE3, target_naj_HE80_HE4, target_naj_HE80_HE5, target_naj_HE80_HE6, \
    target_naj_HE80_HE7, target_naj_HE80_HE8, target_naj_HE80_HE9, target_naj_HE80_HE10, target_naj_HE80_HE11, \
    target_naj_HE160_HE0, target_naj_HE160_HE1, target_naj_HE160_HE2, target_naj_HE160_HE3, target_naj_HE160_HE4, \
    target_naj_HE160_HE5, target_naj_HE160_HE6, target_naj_HE160_HE7, target_naj_HE160_HE8, target_naj_HE160_HE9, \
    target_naj_HE160_HE10, target_naj_HE160_HE11 = [None] * 115

    # report lable
    gen_tx_report = gen_txmax_report = gen_rx_report = gen_rxdynamic_report = gen_aj_report = gen_naj_report = \
        gen_txmimo_report = 1
    # connect equipments and dut
    try:
        iq_wanted = IQxel(IQ_IP)
    except Exception as err:
        logger.info(err)
        iq_wanted = None
        logger.info(Fore.RED + 'IQ connect fail!' + Style.RESET_ALL)
        exit(1)
    else:
        mw_iq = iq_wanted.read_idn
        iq_wanted.set_pathloss(PATHLOSS_WANGTED)
        logger.info('IQ connected!')
    try:
        iq_interfere = IQxel(IQ_IP_INTERFERE)
    except Exception as err:
        logger.info(err)
        iq_interfere = None
        logger.info(Fore.RED + 'INTERFERE IQ connect fail!' + Style.RESET_ALL)
    else:
        mw_iq_inter = iq_interfere.read_idn
        iq_interfere.set_pathloss(PATHLOSS_INTERFERE)
        logger.info('INTERFERE IQ connected!')
    try:
        dt = DUT()
    except Exception as err:
        logger.info(err + Fore.RED + 'DUT Open fail!' + Style.RESET_ALL)
    else:
        dt.login(DUT_IP, DUT_USERNAME, DUT_PASSWORD)
        dt.init(EXT1, EXT2, EXT3)
        logger.info('DUT connected!')

    # calibration
    if CALI_2G == '1' and CALI_5G == '1':
        cali_list = [1, 1]
        band_list = [1, 2]
    elif CALI_2G == '1':
        cali_list = [1, 0]
        band_list = [1, 0]
    elif CALI_5G == '1':
        cali_list = [0, 1]
        band_list = [0, 2]
    else:
        band_list = []
        cali_list = []
    if band_list and cali_list is not None:
        cali_para = [[0 for cali in cali_list] for band in band_list]
        cali_para[0][0] = cali_list[0]
        cali_para[0][1] = band_list[0]
        cali_para[1][0] = cali_list[1]
        cali_para[1][1] = band_list[1]
        logger.debug(cali_para)
        for cali, band in cali_para:
            cali = str(cali)
            band = str(band)
            logger.info('Band' + band)
            if cali == '1':
                if band == '1':
                    channel = 2412
                    cali_channel_list = [1, 7, 11]
                    cali_mode = '11g'
                    cali_rate = '0'
                    # radio_adress = '77'
                    # upc_length = '12'
                else:
                    channel = 5180
                    cali_channel_list = [36, 64, 100, 120, 140, 149, 161]
                    cali_mode = '11a'
                    cali_rate = '6'
                    # radio_adress = '155'
                    # upc_length = '28'
                # dt.init_cali()
                logger.debug(band)
                logger.info('Calibration...')
                # ppm cali
                logger.info('PPM Calibration...')
                chain = '0'
                mode = '11g/a'
                bw = '20'
                offset = dt.init_ppm()
                if iq_wanted is not None:
                    iq_wanted.use_pathloss(mw_iq)
                    logger.debug('Set pathloss')
                iq_wanted.vsa(mw_iq)
                iq_wanted.analysis()
                pwr_len, txq_len, data_pwr, data_txq = iq_wanted.get_status()
                symbol_clock_error = iq_wanted.get_ppm()
                ppm = int(float(symbol_clock_error))
                logger.debug('Default ppm: ' + str(ppm))
                ppm = int(offset) + ppm
                dt.adjust_ppm(ppm)
                time.sleep(3)
                iq_wanted.analysis()
                pwr_len, txq_len, data_pwr, data_txq = iq_wanted.get_status()
                symbol_clock_error = iq_wanted.get_ppm()
                cali_ppm = int(float(symbol_clock_error))
                logger.debug('Cali ppm: ' + str(cali_ppm))
                if cali_ppm > 5 or cali_ppm < -5:
                    logger.info(Fore.RED + 'PPM CALIBRATION FAIL' + Style.RESET_ALL)
                else:
                    logger.info(Fore.GREEN + 'PPM CALIBRATION SUCCESS' + Style.RESET_ALL)
                # power cali
                logger.info('Power Calibration...')
                cali_chain_list = ['0', '1']
                cali_power_list = [200, 160, 120]
                for cali_chain in cali_chain_list:
                    if cali_chain == '0':
                        chain = '01'
                    else:
                        chain = '10'
                    for cali_channel in cali_channel_list:
                        adjust_power_list = []
                        dt.cali_pwr(cali_channel, chain)
                        if cali_channel < 30:
                            channel = 2407 + cali_channel * 5
                        else:
                            channel = 5000 + cali_channel * 5
                        for cali_power in cali_power_list:
                            dt.adjust_pwr(cali_power)
                            iq_wanted.set_port(int(cali_power / 10))
                            target_power = int(cali_power / 10)
                            mode = cali_mode
                            rates = cali_rate
                            iq_wanted.analysis()
                            pwr_len, txq_len, data_pwr, data_txq = iq_wanted.get_status()
                            avg_power = iq_wanted.adjust_power()
                            adjust_power = int(float(avg_power) * 10.0)
                            adjust_power_list.append(adjust_power)
                        dt.adjust_pwr(300)
                        dt.tx_off()
                        logger.debug(adjust_power_list)
                        dt.cali_pwr_write(adjust_power_list)
                dt.cali_para_write()
                dt.upc_write()
                dt.crc()
                logger.info(Fore.GREEN + 'POWER CALIBRATION SUCCESS' + Style.RESET_ALL)
            else:
                logger.info(Fore.YELLOW + 'Calibration skip' + Style.RESET_ALL)
    else:
        logger.info(Fore.YELLOW + 'Calibration skip' + Style.RESET_ALL)

    # TEST FLOW
    # filename = 'TEST_FLOW.txt'
    f = open(filename)
    result = list()
    for line in f.readlines():
        # logger.debug(len(line))
        if len(line) < 30 or line.startswith('//') or line.isspace():
            continue
            pass
        else:
            line = line.strip()
            line = line.split()
            # logger.debug(line)
            # logger.debug('Channel:', line[1], 'Rate:', line[2], 'Chain:', line[3])
            item = line[0]
            item = re.sub('IQ_WIFI_TEST_', '', item)
            logger.debug(item)
            channel = line[1]
            rate = line[2]
            chain = line[3]

            # mode,bw,rates
            if rate == '1M' or rate == '2M' or rate == '5.5M' or rate == '11M':
                mode = '11b'
                bw = '20'
                rates = re.sub('M', '', rate)
                per_spec = 0.08
            elif rate == '6M' or rate == '9M' or rate == '12M' or rate == '18M' \
                    or rate == '24M' or rate == '36M' or rate == '48M' or rate == '54M':
                if int(channel) < 5000:
                    mode = '11g'
                else:
                    mode = '11a'
                bw = '20'
                rates = re.sub('M', '', rate)
                per_spec = 0.1
            elif rate == 'HT20-MCS0' or rate == 'HT20-MCS1' or rate == 'HT20-MCS2' or rate == 'HT20-MCS3' \
                    or rate == 'HT20-MCS4' or rate == 'HT20-MCS5' or rate == 'HT20-MCS6' or rate == 'HT20-MCS7':
                if int(channel) < 5000:
                    mode = '11n'
                else:
                    mode = '11n'
                bw = '20'
                rates = re.sub('HT20-MCS', '', rate)
                per_spec = 0.1
            elif rate == 'HT40-MCS0' or rate == 'HT40-MCS1' or rate == 'HT40-MCS2' or rate == 'HT40-MCS3' \
                    or rate == 'HT40-MCS4' or rate == 'HT40-MCS5' or rate == 'HT40-MCS6' or rate == 'HT40-MCS7':
                if int(channel) < 5000:
                    mode = '11n'
                else:
                    mode = '11n'
                bw = '40'
                rates = re.sub('HT40-MCS', '', rate)
                per_spec = 0.1
                # channel = int(channel) - 10
            elif rate == 'VHT20-MCS0' or rate == 'VHT20-MCS1' or rate == 'VHT20-MCS2' or rate == 'VHT20-MCS3' \
                    or rate == 'VHT20-MCS4' or rate == 'VHT20-MCS5' or rate == 'VHT20-MCS6' or rate == 'VHT20-MCS7' or \
                    rate == 'VHT20-MCS8':
                mode = '11ac'
                bw = '20'
                rates = re.sub('VHT20-MCS', '', rate)
                per_spec = 0.1
            elif rate == 'VHT40-MCS0' or rate == 'VHT40-MCS1' or rate == 'VHT40-MCS2' or rate == 'VHT40-MCS3' \
                    or rate == 'VHT40-MCS4' or rate == 'VHT40-MCS5' or rate == 'VHT40-MCS6' or rate == 'VHT40-MCS7' or \
                    rate == 'VHT40-MCS8' or rate == 'VHT40-MCS9':
                mode = '11ac'
                bw = '40'
                rates = re.sub('VHT40-MCS', '', rate)
                per_spec = 0.1
                # channel = int(channel) - 10
            elif rate == 'VHT80-MCS0' or rate == 'VHT80-MCS1' or rate == 'VHT80-MCS2' or rate == 'VHT80-MCS3' \
                    or rate == 'VHT80-MCS4' or rate == 'VHT80-MCS5' or rate == 'VHT80-MCS6' or rate == 'VHT80-MCS7' or \
                    rate == 'VHT80-MCS8' or rate == 'VHT80-MCS9':
                mode = '11ac'
                bw = '80'
                rates = re.sub('VHT80-MCS', '', rate)
                per_spec = 0.1
            elif rate == 'VHT160-MCS0' or rate == 'VHT160-MCS1' or rate == 'VHT160-MCS2' or rate == 'VHT160-MCS3' \
                    or rate == 'VHT160-MCS4' or rate == 'VHT160-MCS5' or rate == 'VHT160-MCS6' or rate == 'VHT160-MCS7' or \
                    rate == 'VHT160-MCS8' or rate == 'VHT160-MCS9':
                mode = '11ac'
                bw = '160'
                rates = re.sub('VHT160-MCS', '', rate)
                per_spec = 0.1
            elif rate == 'HE20-HE0' or rate == 'HE20-HE1' or rate == 'HE20-HE2' or rate == 'HE20-HE3' \
                    or rate == 'HE20-HE4' or rate == 'HE20-HE5' or rate == 'HE20-HE6' or rate == 'HE20-HE7' or \
                    rate == 'HE20-HE8' or rate == 'HE20-HE9' or rate == 'HE20-HE10' or rate == 'HE20-HE11':
                mode = '11ax'
                bw = '20'
                rates = re.sub('HE20-HE', '', rate)
                per_spec = 0.1
            elif rate == 'HE40-HE0' or rate == 'HE40-HE1' or rate == 'HE40-HE2' or rate == 'HE40-HE3' \
                    or rate == 'HE40-HE4' or rate == 'HE40-HE5' or rate == 'HE40-HE6' or rate == 'HE40-HE7' or \
                    rate == 'HE40-HE8' or rate == 'HE40-HE9' or rate == 'HE40-HE10' or rate == 'HE40-HE11':
                mode = '11ax'
                bw = '40'
                rates = re.sub('HE40-HE', '', rate)
                per_spec = 0.1
                # channel = int(channel) - 10
            elif rate == 'HE80-HE0' or rate == 'HE80-HE1' or rate == 'HE80-HE2' or rate == 'HE80-HE3' \
                    or rate == 'HE80-HE4' or rate == 'HE80-HE5' or rate == 'HE80-HE6' or rate == 'HE80-HE7' or \
                    rate == 'HE80-HE8' or rate == 'HE80-HE9' or rate == 'HE80-HE10' or rate == 'HE80-HE11':
                mode = '11ax'
                bw = '80'
                rates = re.sub('HE80-HE', '', rate)
                per_spec = 0.1
            elif rate == 'HE160-HE0' or rate == 'HE160-HE1' or rate == 'HE160-HE2' or rate == 'HE160-HE3' \
                    or rate == 'HE160-HE4' or rate == 'HE160-HE5' or rate == 'HE160-HE6' or rate == 'HE160-HE7' or \
                    rate == 'HE160-HE8' or rate == 'HE160-HE9' or rate == 'HE160-HE10' or rate == 'HE160-HE11':
                mode = '11ax'
                bw = '160'
                rates = re.sub('HE160-HE', '', rate)
                per_spec = 0.1
            # chain
            chain = re.sub('CHAIN', '', chain)
            path = chain
            # read spec
            if int(channel) < 5000:
                # INIT SPEC
                # 2.4g
                spec_file_2g = load_workbook('./spec_2g.xlsx')
                # logger.debug(spec_file.sheesnames)
                sheet_2g = spec_file_2g['Sheet1']
                rows_2g = []
                ratelist_2g = ['1M', '2M', '5_5M', '11M', '6M', '9M', '12M', '18M', '24M', '36M', '48M', '54M',
                               'HT20_MCS0', 'HT20_MCS1', 'HT20_MCS2', 'HT20_MCS3', 'HT20_MCS4', 'HT20_MCS5',
                               'HT20_MCS6', 'HT20_MCS7', 'HT40_MCS0', 'HT40_MCS1', 'HT40_MCS2', 'HT40_MCS3',
                               'HT40_MCS4', 'HT40_MCS5', 'HT40_MCS6', 'HT40_MCS7', 'VHT20_MCS0', 'VHT20_MCS1',
                               'VHT20_MCS2', 'VHT20_MCS3', 'VHT20_MCS4', 'VHT20_MCS5', 'VHT20_MCS6', 'VHT20_MCS7',
                               'VHT20_MCS8', 'VHT40_MCS0', 'VHT40_MCS1', 'VHT40_MCS2', 'VHT40_MCS3', 'VHT40_MCS4',
                               'VHT40_MCS5', 'VHT40_MCS6', 'VHT40_MCS7', 'VHT40_MCS8', 'VHT40_MCS9', 'HE20_HE0',
                               'HE20_HE1', 'HE20_HE2', 'HE20_HE3', 'HE20_HE4', 'HE20_HE5', 'HE20_HE6',
                               'HE20_HE7', 'HE20_HE8', 'HE20_HE9', 'HE20_HE10', 'HE20_HE11', 'HE40_HE0',
                               'HE40_HE1', 'HE40_HE2', 'HE40_HE3', 'HE40_HE4', 'HE40_HE5', 'HE40_HE6',
                               'HE40_HE7', 'HE40_HE8', 'HE40_HE9', 'HE40_HE10', 'HE40_HE11']
                for row_2g in sheet_2g:
                    rows_2g.append(row_2g)
                    # logger.debug(rows)
                for r in range(sheet_2g.max_row):
                    for c in range(sheet_2g.max_column):
                        # logger.debug(rows[r][c].value)
                        rows_2g[r][c].value = str(rows_2g[r][c].value).strip()
                        rs = r + 1
                        cs = c + 1
                        if rows_2g[r][c].value == 'POWER_ACCURACY':
                            spec_pwr = abs(rows_2g[r][cs].value)
                        elif rows_2g[r][c].value == 'Power_Gain_Index':
                            gain_24 = abs(rows_2g[r][cs].value)
                        elif rows_2g[r][c].value == 'EVM_MARGIN':
                            evm_margin = abs(rows_2g[r][cs].value)
                        elif rows_2g[r][c].value == 'Symbol_Clock_Error':
                            spec_symbol_clock_error = abs(rows_2g[r][cs].value)
                        elif rows_2g[r][c].value == 'XCAP':
                            xcap_24 = abs(rows_2g[r][cs].value)
                        elif rows_2g[r][c].value == 'LO_Leakage':
                            spec_lo_leakage = -abs(rows_2g[r][cs].value)
                        elif rows_2g[r][c].value == 'MASK':
                            spec_mask = abs(rows_2g[r][cs].value)
                        elif rows_2g[r][c].value == 'OBW_20M':
                            spec_obw_20M = rows_2g[r][cs].value
                        elif rows_2g[r][c].value == 'OBW_40M':
                            spec_obw_40M = rows_2g[r][cs].value
                        for x in ratelist_2g:
                            if rows_2g[r][c].value == x + '_power':
                                exec('target_pwr_%s=%d' % (x, rows_2g[rs][c].value))
                                break
                        for i in ratelist_2g:
                            if rows_2g[r][c].value == i + '_EVM':
                                exec('target_EVM_%s=%d' % (i, rows_2g[rs][c].value))
                                break
                        for j in ratelist_2g:
                            if rows_2g[r][c].value == j + '_sens':
                                exec('target_sens_%s=%d' % (j, rows_2g[rs][c].value))
                                break
                        for y in ratelist_2g:
                            if rows_2g[r][c].value == y + '_aj':
                                exec('target_aj_%s=%d' % (y, rows_2g[rs][c].value))
                                break
                        for k in ratelist_2g:
                            if rows_2g[r][c].value == k + '_naj':
                                exec('target_naj_%s=%d' % (k, rows_2g[rs][c].value))
                                break
                        for l in ratelist_2g:
                            if rows_2g[r][c].value == l + '_MAX':
                                exec('target_MAX_%s=%d' % (l, rows_2g[rs][c].value))
                                break
                        spec_obw_80M = spec_obw_160M = None
            else:
                # 5g
                spec_file_5g = load_workbook('./spec_5g.xlsx')
                # logger.debug(spec_file.sheesnames)
                sheet_5g = spec_file_5g['Sheet1']
                rows_5g = []
                ratelist_5g = ['6M', '9M', '12M', '18M', '24M', '36M', '48M', '54M', 'HT20_MCS0', 'HT20_MCS1',
                               'HT20_MCS2', 'HT20_MCS3', 'HT20_MCS4', 'HT20_MCS5', 'HT20_MCS6', 'HT20_MCS7',
                               'HT40_MCS0', 'HT40_MCS1', 'HT40_MCS2', 'HT40_MCS3', 'HT40_MCS4', 'HT40_MCS5',
                               'HT40_MCS6', 'HT40_MCS7', 'VHT20_MCS0', 'VHT20_MCS1', 'VHT20_MCS2', 'VHT20_MCS3',
                               'VHT20_MCS4', 'VHT20_MCS5', 'VHT20_MCS6', 'VHT20_MCS7', 'VHT20_MCS8', 'VHT40_MCS0',
                               'VHT40_MCS1', 'VHT40_MCS2', 'VHT40_MCS3', 'VHT40_MCS4', 'VHT40_MCS5', 'VHT40_MCS6',
                               'VHT40_MCS7', 'VHT40_MCS8', 'VHT40_MCS9', 'VHT80_MCS0', 'VHT80_MCS1', 'VHT80_MCS2',
                               'VHT80_MCS3', 'VHT80_MCS4', 'VHT80_MCS5', 'VHT80_MCS6', 'VHT80_MCS7', 'VHT80_MCS8',
                               'VHT80_MCS9', 'VHT160_MCS0', 'VHT160_MCS1', 'VHT160_MCS2',
                               'VHT160_MCS3', 'VHT160_MCS4', 'VHT160_MCS5', 'VHT160_MCS6', 'VHT160_MCS7', 'VHT160_MCS8',
                               'VHT160_MCS9', 'HE20_HE0', 'HE20_HE1', 'HE20_HE2', 'HE20_HE3', 'HE20_HE4',
                               'HE20_HE5', 'HE20_HE6', 'HE20_HE7', 'HE20_HE8', 'HE20_HE9', 'HE20_HE10',
                               'HE20_HE11', 'HE40_HE0', 'HE40_HE1', 'HE40_HE2', 'HE40_HE3', 'HE40_HE4',
                               'HE40_HE5', 'HE40_HE6', 'HE40_HE7', 'HE40_HE8', 'HE40_HE9', 'HE40_HE10',
                               'HE40_HE11', 'HE80_HE0', 'HE80_HE1', 'HE80_HE2', 'HE80_HE3', 'HE80_HE4',
                               'HE80_HE5', 'HE80_HE6', 'HE80_HE7', 'HE80_HE8', 'HE80_HE9', 'HE80_HE10',
                               'HE80_HE11', 'HE160_HE0', 'HE160_HE1', 'HE160_HE2', 'HE160_HE3', 'HE160_HE4',
                               'HE160_HE5', 'HE160_HE6', 'HE160_HE7', 'HE160_HE8', 'HE160_HE9', 'HE160_HE10',
                               'HE160_HE11']
                for row_5g in sheet_5g:
                    rows_5g.append(row_5g)
                    # logger.debug(rows)
                for rr in range(sheet_5g.max_row):
                    for cc in range(sheet_5g.max_column):
                        # logger.debug(rows[r][c].value)
                        rows_5g[rr][cc].value = str(rows_5g[rr][cc].value).strip()
                        rrs = rr + 1
                        ccs = cc + 1
                        if rows_5g[rr][cc].value == 'POWER_ACCURACY':
                            spec_pwr = abs(rows_5g[rr][ccs].value)
                        elif rows_5g[rr][cc].value == 'Power_Gain_Index':
                            gain_5 = abs(rows_5g[rr][ccs].value)
                        elif rows_5g[rr][cc].value == 'EVM_MARGIN':
                            evm_margin = abs(rows_5g[rr][ccs].value)
                        elif rows_5g[rr][cc].value == 'Symbol_Clock_Error':
                            spec_symbol_clock_error = abs(rows_5g[rr][ccs].value)
                        elif rows_5g[rr][cc].value == 'XCAP':
                            xcap_5 = abs(rows_5g[rr][ccs].value)
                        elif rows_5g[rr][cc].value == 'LO_Leakage':
                            spec_lo_leakage = -abs(rows_5g[rr][ccs].value)
                        elif rows_5g[rr][cc].value == 'MASK':
                            spec_mask = abs(rows_5g[rr][ccs].value)
                        elif rows_5g[rr][cc].value == 'OBW_20M':
                            spec_obw_20M = rows_5g[rr][ccs].value
                        elif rows_5g[rr][cc].value == 'OBW_40M':
                            spec_obw_40M = rows_5g[rr][ccs].value
                        elif rows_5g[rr][cc].value == 'OBW_80M':
                            spec_obw_80M = rows_5g[rr][ccs].value
                        elif rows_5g[rr][cc].value == 'OBW_160M':
                            spec_obw_160M = rows_5g[rr][ccs].value
                        for xx in ratelist_5g:
                            if rows_5g[rr][cc].value == xx + '_power':
                                exec('target_pwr_%s=%d' % (xx, rows_5g[rrs][cc].value))
                                break
                        for ii in ratelist_5g:
                            if rows_5g[rr][cc].value == ii + '_EVM':
                                exec('target_EVM_%s=%d' % (ii, rows_5g[rrs][cc].value))
                                break
                        for jj in ratelist_5g:
                            if rows_5g[rr][cc].value == jj + '_sens':
                                exec('target_sens_%s=%d' % (jj, rows_5g[rrs][cc].value))
                                break
                        for yy in ratelist_5g:
                            if rows_5g[rr][cc].value == yy + '_aj':
                                exec('target_aj_%s=%d' % (yy, rows_5g[rrs][cc].value))
                                break
                        for kk in ratelist_5g:
                            if rows_5g[rr][cc].value == kk + '_naj':
                                exec('target_naj_%s=%d' % (kk, rows_5g[rrs][cc].value))
                                break
                        for ll in ratelist_5g:
                            if rows_5g[rr][cc].value == ll + '_MAX':
                                exec('target_MAX_%s=%d' % (ll, rows_5g[rrs][cc].value))
                                break

            logger.info('*************************************************************')

            logger.info('Mode: ' + mode + ' Channel: ' + channel + ' BW: ' + bw + ' Rate: ' + rate + ' Chain: ' + chain)
            if iq_wanted is not None:
                iq_wanted.use_pathloss(mw_iq, chain)
                logger.debug('Set pathloss')
            if iq_interfere is not None:
                iq_interfere.use_pathloss(mw_iq_inter, chain)
            if item == 'TX':
                if gen_tx_report == 1:
                    # GEN REPORT title
                    tx_result_name = sn + '_' + 'TX_Result' + '_' + now_time + '.csv'
                    with open('./Result/' + tx_result_name, 'w', newline='') as f:
                        writer = csv.writer(f)
                        writer.writerow(['FREQ', 'DATA_RATE', 'CHAIN', 'TX_POWER', 'POWER', 'GAIN', 'LIMIT',
                                         'RESULT', 'EVM', 'LIMIT', 'RESULT', 'FREQ_ERROR', 'LIMIT', 'RESULT',
                                         'LOLEAKAGE', 'LIMIT', 'RESULT', 'OBW', 'LIMIT', 'RESULT',
                                         'MASK', 'LIMIT', 'RESULT', 'FLATNESS', 'LIMIT', 'RESULT',
                                         'RAMPONTIME', 'LIMIT', 'RESULT', 'RAMPOFFTIME', 'LIMIT', 'RESULT'])
                    with open('./log/' + 'log_' + tx_result_name, 'w', newline='') as f:
                        writer = csv.writer(f)
                        writer.writerow(['FREQ', 'DATA_RATE', 'CHAIN', 'TX_POWER', 'POWER', 'GAIN', 'LIMIT',
                                         'RESULT', 'EVM', 'LIMIT', 'RESULT', 'FREQ_ERROR', 'LIMIT', 'RESULT',
                                         'LOLEAKAGE', 'LIMIT', 'RESULT', 'OBW', 'LIMIT', 'RESULT',
                                         'MASK', 'LIMIT', 'RESULT', 'FLATNESS', 'LIMIT', 'RESULT',
                                         'RAMPONTIME', 'LIMIT', 'RESULT', 'RAMPOFFTIME', 'LIMIT', 'RESULT'])
                    gen_tx_report += 1
                avg_power = avg_evm = symbol_clock_error = lo_leakage = obw = mask = flatness = ramp_on_time = ramp_off_time = 0.0
                adjust_result = result_evm = result_symbol_clock_error = result_lo_leakage = result_mask = \
                    result_flasness = 'Pass'
                # gain = dt.get_paras()
                rate_t = re.sub('-', '_', rate)
                rate_t = re.sub('\.', '_', rate_t)
                targetpower = eval('target_pwr_' + rate_t)
                if len(line) > 4:
                    targetpower = int(line[4])
                else:
                    targetpower = int(targetpower)
                gain = int(targetpower)
                spec_evm = eval('target_EVM_' + rate_t)
                power_accuracy_left = float(targetpower) - float(spec_pwr) + float(accuracy_limit_left)
                power_accuracy_right = float(targetpower) + float(spec_pwr) - float(accuracy_limit_right)
                logger.debug('Power accuracy: ' + str(power_accuracy_left) + ' ' + str(power_accuracy_right))
                # dut tx enable
                dt.tx()
                # get RESULT
                iq_wanted.vsa(mw_iq, mode, bw, rates, channel, targetpower)
                iq_wanted.analysis(mode, rates)
                avg_power, result_pwr = iq_wanted.get_power('1', targetpower, spec_pwr)
                avg_evm, result_evm = iq_wanted.get_evm(mode, spec_evm, evm_margin)
                symbol_clock_error, result_symbol_clock_error = iq_wanted.get_ppm(mode, spec_symbol_clock_error)
                lo_leakage, result_lo_leakage = iq_wanted.get_lo_leakage(mode, spec_lo_leakage)
                obw, spec_obw, result_obw = iq_wanted.get_obw(bw, spec_obw_20M, spec_obw_40M, spec_obw_80M,
                                                              spec_obw_160M)
                mask, result_mask = iq_wanted.get_mask(spec_mask)
                flatness, spec_flatness, result_flatness = iq_wanted.get_flatness(mode)
                ramp_on_time, result_ramp_on_time, ramp_off_time, result_ramp_off_time = iq_wanted.get_ramp_time(mode)
                if mode == '11b':
                    spec_ramp_on_time = spec_ramp_off_time = 2.0
                else:
                    spec_ramp_on_time = spec_ramp_off_time = 'NA'
                if AUTO_ADJUST_POWER == '1':
                    # if avg_power == 'NA' or result_evm == 'NA':
                    if avg_power == 'NA' or float(avg_power) > 99.000:
                        logger.info(Fore.RED + 'Error!' + Style.RESET_ALL)
                    else:
                        # print(targetpower, spec_pwr, accuracy_limit_left, accuracy_limit_right)
                        max_power = float(targetpower + spec_pwr)
                        delta_power = float(avg_power) - float(targetpower)
                        delta_power = float('{:.3f}'.format(delta_power))
                        logger.debug('Default Measurer Power: ' + str(avg_power))
                        logger.debug('Target Power: ' + str(targetpower))
                        logger.debug('Delta Power: ' + str(delta_power))
                        # power is nomal but some is fail
                        if float(power_accuracy_left) <= float(avg_power) <= float(power_accuracy_right):
                            if result_evm == 'Fail' or result_symbol_clock_error == 'Fail' \
                                    or result_lo_leakage == 'Fail' or result_mask == 'Fail':
                                logger.info(Fore.RED + 'TX\'s quality are failed!' + Style.RESET_ALL)
                            else:
                                logger.info('Get Good Result!')
                        # power is not nomal
                        else:
                            power_counts = 0
                            setup = 2
                            setups = 2
                            logger.debug('Step:' + str(setup))
                            init_value = targetpower
                            gain = int(init_value) + setup
                            # gain = hex(gain)
                            # gain = re.sub('0x', '', gain)
                            get_power = []
                            get_delta_power = []
                            c = 1
                            get_power.append(avg_power)
                            logger.debug('Measurer Power List: ' + str(get_power))
                            get_delta_power.append(delta_power)
                            logger.debug('Power Deviation List: ' + str(get_delta_power))
                            logger.debug('Adjust Counts:' + str(c))
                            # ADJUST
                            while float(avg_power) < power_accuracy_left or float(avg_power) > power_accuracy_right:
                                logger.debug('L ' + str(power_accuracy_left) + str(avg_power) + ' R ' +
                                             str(power_accuracy_right))
                                logger.info(Fore.GREEN + 'Adjust...' + Style.RESET_ALL)
                                if 99.000 > float(avg_power) > power_accuracy_right:
                                    adj = -1
                                else:
                                    adj = 1
                                c = c + 1
                                logger.debug('NEW VALUE:' + str(gain))
                                dt.adjust_power(gain)
                                iq_wanted.vsa(mw_iq, mode, bw, rates, channel, targetpower)
                                iq_wanted.analysis(mode, rates)
                                avg_power, result_pwr = iq_wanted.get_power(1, targetpower, spec_pwr)
                                avg_evm, result_evm = iq_wanted.get_evm(mode, spec_evm, evm_margin)
                                symbol_clock_error, result_symbol_clock_error = iq_wanted.get_ppm(mode,
                                                                                                  spec_symbol_clock_error)
                                lo_leakage, result_lo_leakage = iq_wanted.get_lo_leakage(mode, spec_lo_leakage)
                                obw, spec_obw, result_obw = iq_wanted.get_obw(bw, spec_obw_20M, spec_obw_40M,
                                                                              spec_obw_80M,
                                                                              spec_obw_160M)
                                mask, result_mask = iq_wanted.get_mask(spec_mask)
                                flatness, result_flasness = iq_wanted.get_flatness(spec_flatness)
                                ramp_on_time, result_ramp_on_time, ramp_off_time, result_ramp_off_time = iq_wanted.get_ramp_time()
                                spec_ramp_on_time = spec_ramp_off_time = 2.0
                                get_power.append(avg_power)
                                logger.debug('Measurer Power List: ' + str(get_power))
                                delta_power = float(avg_power) - targetpower
                                delta_power = float('{:.3f}'.format(delta_power))
                                avg_power = float(avg_power)
                                min_adj = float(avg_power) - float(targetpower)
                                max_adj = float(avg_power) - float(max_power)
                                get_delta_power.append(delta_power)
                                logger.debug('Power Deviation List: ' + str(get_delta_power))
                                logger.debug('Adjust Counts: ' + str(c))
                                adjust_status = get_delta_power[c - 1] - get_delta_power[c - 2]
                                adjust_status = float('{:.3f}'.format(adjust_status))
                                logger.debug('Power Added: ' + str(adjust_status))
                                if adjust_status < 0:
                                    setup = 2 * setups * adj
                                elif adjust_status > 0:
                                    setup = 2 * adj
                                if abs(adjust_status) < 0.2:
                                    power_counts = power_counts + 1
                                # logger.debug(power_counts)
                                if power_counts > 5:
                                    logger.info(
                                        Fore.RED + 'Power added was too small, Power adjust stop' + Style.RESET_ALL)
                                    continue
                                else:
                                    logger.debug('Step: ' + str(setup))
                                    # gain = int(gain, 16)
                                    gain = int(gain) + setup
                                    # gain = hex(gain)
                                    # gain = re.sub('0x', '', gain)
                                    logger.debug(
                                        result_evm + result_symbol_clock_error + result_lo_leakage + result_mask)
                                    logger.info('*************************************************************')
                with open('./Result/' + tx_result_name, 'a+', newline='') as write_tx_result:
                    writer_file = csv.writer(write_tx_result)
                    writer_file.writerow(
                        [channel, rate, chain, targetpower, avg_power, gain, spec_pwr, result_pwr,
                         avg_evm, spec_evm, result_evm, symbol_clock_error, spec_symbol_clock_error,
                         result_symbol_clock_error, lo_leakage, spec_lo_leakage, result_lo_leakage, obw,
                         spec_obw, result_obw, mask, spec_mask, result_mask, flatness, spec_flatness,
                         result_flatness, ramp_on_time, spec_ramp_on_time, result_ramp_on_time,
                         ramp_off_time, spec_ramp_off_time, result_ramp_off_time])
                dt.tx_off()
            elif item == 'TX_MAX':
                if gen_txmax_report == 1:
                    # GEN REPORT
                    tx_result_name = sn + '_' + 'TX_MAX_Result' + '_' + now_time + '.csv'
                    with open('./Result/' + tx_result_name, 'w', newline='') as f:
                        writer = csv.writer(f)
                        writer.writerow(['RESULT', 'FREQ', 'DATA_RATE', 'CHAIN', 'TX_POWER', 'POWER', 'GAIN', 'LIMIT',
                                         'RESULT', 'EVM', 'LIMIT', 'RESULT', 'FREQ_ERROR', 'LIMIT', 'RESULT',
                                         'LOLEAKAGE', 'LIMIT', 'RESULT', 'OBW', 'LIMIT', 'RESULT',
                                         'MASK', 'LIMIT', 'RESULT', 'FLAsnESS', 'LIMIT', 'RESULT',
                                         'RAMPONTIME', 'LIMIT', 'RESULT', 'RAMPOFFTIME', 'LIMIT', 'RESULT'])
                    with open('./log/' + 'log_' + tx_result_name, 'w', newline='') as f:
                        writer = csv.writer(f)
                        writer.writerow(['FREQ', 'DATA_RATE', 'CHAIN', 'TX_POWER', 'POWER', 'GAIN', 'LIMIT',
                                         'RESULT', 'EVM', 'LIMIT', 'RESULT', 'FREQ_ERROR', 'LIMIT', 'RESULT',
                                         'LOLEAKAGE', 'LIMIT', 'RESULT', 'OBW', 'LIMIT', 'RESULT',
                                         'MASK', 'LIMIT', 'RESULT', 'FLAsnESS', 'LIMIT', 'RESULT',
                                         'RAMPONTIME', 'LIMIT', 'RESULT', 'RAMPOFFTIME', 'LIMIT', 'RESULT'])
                    gen_txmax_report += 1
                result_list = []
                adjust_result = result_evm = result_symbol_clock_error = result_lo_leakage = result_mask = \
                    result_flasness = 'Pass'
                gain = dt.get_paras()
                rate_t = re.sub('-', '_', rate)
                rate_t = re.sub('\.', '_', rate_t)
                targetpower = eval('target_pwr_' + rate_t)
                spec_evm = eval('target_EVM_' + rate_t)
                power_accuracy_left = float(targetpower) - float(spec_pwr) + float(accuracy_limit_left)
                power_accuracy_right = float(targetpower) + float(spec_pwr) - float(accuracy_limit_right)
                # dt.set_default()
                dt.tx()
                # RESULT
                iq_wanted.vsa(mw_iq)
                iq_wanted.analysis()
                pwr_len, txq_len, data_pwr, data_txq = iq_wanted.get_status()
                avg_power, result_evm, result_symbol_clock_error, result_lo_leakage, result_mask, result_list = \
                    iq_wanted.get_data(pwr_len, txq_len, data_pwr, data_txq, mode, channel, rate, chain,
                                       tx_result_name, targetpower, spec_pwr, gain, spec_evm, evm_margin,
                                       spec_symbol_clock_error, spec_lo_leakage, spec_mask, spec_obw_20M,
                                       spec_obw_40M, spec_obw_80M, spec_obw_160M, result_list)
                if avg_power == 'NA' or result_evm == 'NA':
                    # avg_power = avg_evm = symbol_clock_error = lo_leakage = result = result_pwr = result_evm = \
                    #    result_symbol_clock_error = result_lo_leakage = result_mask = 'NA'
                    logger.info(Fore.RED + 'Capture failed!' + Style.RESET_ALL)
                    continue
                # elif result_evm == 'Fail' or result_symbol_clock_error == 'Fail' or result_lo_leakage == 'Fail' \
                #         or result_mask == 'Fail':
                #     logger.info(Fore.RED + 'The rate\'s TX Quality is Fail, Adjust quit' + Style.RESET_ALL)
                #     continue
                else:
                    logger.info('Default Measurer Power: ' + str(avg_power) + ' Target Power: ' + str(targetpower))
                    logger.info('INIT_POWER_VALUE: ' + str(gain))
                    delta_power = float(avg_power) - float(targetpower)
                    delta_power = '{:.3f}'.format(delta_power)
                    adjust_init_power_counts = 0
                    # power is too low
                    while float(avg_power) < float(power_accuracy_left) or \
                            float(avg_power) > float(power_accuracy_right):
                        logger.debug(str(power_accuracy_left) + 'L' + avg_power + 'R' + str(power_accuracy_right))
                        logger.info(Fore.LIGHTRED_EX + 'The power need to adjust to target power!' + Style.RESET_ALL)
                        if float(avg_power) < float(power_accuracy_left):
                            delta_power = float(avg_power) - float(power_accuracy_left)
                            delta_setup = 1
                        else:
                            delta_power = float(avg_power) - float(power_accuracy_right)
                            delta_setup = -1
                        logger.debug(delta_power)
                        gain = int(gain) - int(delta_power * 2.0) + delta_setup
                        logger.debug(gain)
                        # gain = int(gain, 16)
                        # gain = int(gain) + setup
                        # gain = hex(gain)
                        dt.adjust_power(gain)
                        iq_wanted.vsa(mw_iq)
                        iq_wanted.analysis()
                        pwr_len, txq_len, data_pwr, data_txq = iq_wanted.get_status()
                        avg_power, result_evm, result_symbol_clock_error, result_lo_leakage, result_mask, result_list = \
                            iq_wanted.get_data(pwr_len, txq_len, data_pwr, data_txq, mode, channel, rate,
                                               chain, tx_result_name, targetpower, spec_pwr, gain, spec_evm, evm_margin,
                                               spec_symbol_clock_error, spec_lo_leakage, spec_mask, spec_obw_20M,
                                               spec_obw_40M, spec_obw_80M, spec_obw_160M, result_list)
                        adjust_init_power_counts = adjust_init_power_counts + 1
                        if adjust_init_power_counts > 10:
                            logger.error('Power Adjusted failed!')
                            continue
                    # power is ok, ADJUST
                    adjust_delta = 1
                    gain = int(gain) + adjust_delta
                    # gain = hex(gain)
                    get_power = []
                    get_delta_power = []
                    c = 1
                    get_power.append(avg_power)
                    logger.debug('Measurer Power List: ' + str(get_power))
                    delta_power = float(delta_power)
                    get_delta_power.append(delta_power)
                    logger.debug('Power Deviation List: ' + str(get_delta_power))
                    logger.debug('Adjust Counts: ' + str(c))
                    setup = power_counts = 0
                    while result_evm == 'Pass' and result_symbol_clock_error == 'Pass' and result_lo_leakage == 'Pass' \
                            and result_mask == 'Pass' and adjust_result == 'Pass' and 0 < gain <= 63:
                        c = c + 1
                        logger.info(Fore.GREEN + 'Adjust...' + Style.RESET_ALL)
                        logger.debug('NEW VALUE: ' + str(gain))
                        dt.adjust_power(gain)
                        iq_wanted.vsa(mw_iq)
                        iq_wanted.analysis()
                        pwr_len, txq_len, data_pwr, data_txq = iq_wanted.get_status()
                        avg_power, result_evm, result_symbol_clock_error, result_lo_leakage, result_mask, result_list = \
                            iq_wanted.get_data(pwr_len, txq_len, data_pwr, data_txq, mode, channel, rate,
                                               chain, tx_result_name, targetpower, spec_pwr, gain, spec_evm, evm_margin,
                                               spec_symbol_clock_error, spec_lo_leakage, spec_mask, spec_obw_20M,
                                               spec_obw_40M, spec_obw_80M, spec_obw_160M, result_list)
                        get_power.append(avg_power)
                        logger.debug('Measurer Power List: ' + str(get_power))
                        delta_power = float(avg_power) - targetpower
                        delta_power = '{:.3f}'.format(delta_power)
                        delta_power = float(delta_power)
                        get_delta_power.append(delta_power)
                        logger.debug('Power Deviation List: ' + str(get_delta_power))
                        logger.debug('Adjust Counts: ' + str(c))
                        adjust_status = get_delta_power[c - 1] - get_delta_power[c - 2]
                        adjust_status = float('{:.3f}'.format(adjust_status))
                        logger.debug('Power Added:' + str(adjust_status))
                        if adjust_status < 0.1:
                            setup = 2 * adjust_delta
                            power_counts = power_counts + 1
                        elif adjust_status > 0.1:
                            setup = 1
                        if power_counts > 10:
                            logger.info(Fore.RED + 'Power changed little, Power adjust stop' + Style.RESET_ALL)
                        else:
                            logger.info('Step: ' + str(setup))
                            # gain = int(gain, 16)
                            gain = int(gain) + setup
                            # gain = hex(gain)
                            # gain = re.sub('0x', '', gain)
                            logger.debug(result_evm + result_symbol_clock_error + result_lo_leakage + result_mask)
                            logger.info('*************************************************************')
                result_final = result_list[len(result_list) - 2]
                with open('./Result/' + tx_result_name, 'a+', newline='') as write_tx_result:
                    writer_file = csv.writer(write_tx_result)
                    writer_file.writerow(result_final)
            elif item == 'RX':
                if gen_rx_report == 1:
                    # GEN REPORT
                    rx_result_name = sn + '_' + 'RX_Result' + '_' + now_time + '.csv'
                    with open('./Result/' + rx_result_name, 'w', newline='') as f:
                        writer = csv.writer(f)
                        writer.writerow(
                            ['FREQ', 'DATA_RATE', 'CHAIN', 'SENSITIVITY SPEC', 'SENSITIVITY', 'RX PACKETS', 'RESULT'])
                    if RX_DYNAMIC == '1':
                        rx_dynamic_result = sn + '_' + 'RX_Dynamic' + '_' + now_time + '.csv'
                        with open('./Result/' + rx_dynamic_result, 'w', newline='') as f:
                            writer = csv.writer(f)
                            writer.writerow(['FREQ', 'DATA_RATE', 'CHAIN', 'VSG_POWER', 'VSG_Packets', 'PER', 'RESULT'])
                    gen_rx_report += 1
                start = int(line[4])
                stop = int(line[5])
                per_list = []
                sens_list = []
                logger.info('Start: ' + str(start) + ' Stop: ' + str(stop))
                # loss = iq.read_pathloss(channel, chain)
                # dt.set_default()
                rx_counts = []
                # dt.get_statistics()# init counts
                counts = dt.get_statistics()
                rx_counts.append(counts)
                # dt.rx_on()
                dt.rx()
                # logger.debug(channel)
                iq_wanted.vsg(mw_iq, mode, bw, rates, channel, start)
                counts = dt.get_statistics()
                rx_counts.append(counts)
                per = 1.0 - (float(rx_counts[-1]) - float(rx_counts[-2]))/float(RX_PACKETS)
                per = float(per)
                per = '{:.3f}'.format(per)
                logger.debug(per)
                if RX_DYNAMIC == '1':
                    if float(per) <= float(per_spec):
                        per_result = 'Pass'
                    else:
                        per_result = 'Fail'
                    with open('./Result/' + rx_dynamic_result, 'a+', newline='') as f2:
                        writer2 = csv.writer(f2)
                        writer2.writerow([channel, rate, chain, start, RX_PACKETS, per, per_result])
                per_list.append(per)
                sens_list.append(start)
                if start > stop:
                    while start > stop and float(per) <= float(per_spec):
                        start = start - 1
                        # dt.rx(mode, channel, bw, rates, chain)
                        # dt.rx()
                        iq_wanted.vsg(mw_iq,  mode, bw, rates, channel, start)
                        time.sleep(0.2)
                        counts = dt.get_statistics()
                        rx_counts.append(counts)
                        per = 1.0 - (float(rx_counts[-1]) - float(rx_counts[-2])) / float(RX_PACKETS)
                        per = float(per)
                        per = '{:.3f}'.format(per)
                        logger.debug(per)
                        per_list.append(per)
                        sens_list.append(start)
                        if RX_DYNAMIC == '1':
                            if float(per) <= float(per_spec):
                                per_result = 'Pass'
                            else:
                                per_result = 'Fail'
                            with open('./Result/' + rx_dynamic_result, 'a+', newline='') as f2:
                                writer2 = csv.writer(f2)
                                writer2.writerow([channel, rate, chain, start, RX_PACKETS, per, per_result])
                    logger.debug(str(len(per_list)) + str(per_list))
                    logger.debug(str(len(sens_list)) + str(sens_list))
                    per = per_list[len(per_list) - 2]
                    sens = sens_list[len(sens_list) - 2]
                    logger.info('Sensitivity: ' + str(sens))
                    rate_t = re.sub('-', '_', rate)
                    rate_t = re.sub('\.', '_', rate_t)
                    sens_spec = eval('target_sens_' + rate_t)
                    if sens <= sens_spec and float(per) <= float(per_spec):
                        result = 'Pass'
                    else:
                        result = 'Fail'
                    with open('./Result/' + rx_result_name, 'a+', newline='') as f2:
                        writer2 = csv.writer(f2)
                        writer2.writerow([channel, rate, chain, sens_spec, sens, RX_PACKETS, per, result])
                else:
                    while start < stop and float(per) <= float(per_spec):
                        start = start + 1
                        # dt.rx(mode, channel, bw, rates, chain)
                        # dt.rx()
                        iq_wanted.vsg(mw_iq,  mode, bw, rates, channel, start)
                        time.sleep(0.2)
                        counts = dt.get_statistics()
                        rx_counts.append(counts)
                        per = 1.0 - (float(rx_counts[-1]) - float(rx_counts[-2])) / float(RX_PACKETS)
                        per = float(per)
                        per = '{:.3f}'.format(per)
                        logger.debug(per)
                        per_list.append(per)
                        sens_list.append(start)
                        if RX_DYNAMIC == '1':
                            if float(per) <= float(per_spec):
                                per_result = 'Pass'
                            else:
                                per_result = 'Fail'
                            with open('./Result/' + rx_dynamic_result, 'a+', newline='') as f2:
                                writer2 = csv.writer(f2)
                                writer2.writerow([channel, rate, chain, start, RX_PACKETS, per, per_result])
                    logger.debug(str(len(per_list)) + str(per_list))
                    logger.debug(str(len(sens_list)) + str(sens_list))
                    per = per_list[len(per_list) - 2]
                    sens = sens_list[len(sens_list) - 2]
                    logger.info('Rx Level: ' + str(sens))
                    rate_t = re.sub('-', '_', rate)
                    rate_t = re.sub('\.', '_', rate_t)
                    max_spec = eval('target_MAX_' + rate_t)
                    if sens >= max_spec and float(per) <= float(per_spec):
                        result = 'Pass'
                    else:
                        result = 'Fail'
                    with open('./Result/' + rx_result_name, 'a+', newline='') as f2:
                        writer2 = csv.writer(f2)
                        writer2.writerow([channel, rate, chain, sens_spec, sens, RX_PACKETS, per, result])
                logger.info('*************************************************************')
                dt.rx_off()
            elif item == 'AJ':
                if gen_aj_report == 1:
                    # GEN REPORT
                    aj_result_name = sn + '_' + 'AJ_Result' + '_' + now_time + '.csv'
                    with open('./Result/' + aj_result_name, 'w', newline='') as f:
                        writer = csv.writer(f)
                        writer.writerow(['FREQ', 'DATA_RATE', 'CHAIN', 'Adjacent Channel Rejection SPEC',
                                         'Adjacent Channel Rejection',
                                         'RESULT'])
                    gen_aj_report += 1
                iq_interfere.vsg_off()
                # GET SENS
                start = int(line[4])
                sens_list = []
                per_list = []
                logger.info('Start: ' + str(start))
                dt.get_statistics()  # reset counts
                # dt.rx_on()
                dt.rx()
                # logger.debug(channel)
                iq_wanted.vsg(mw_iq, start)
                per = dt.get_statistics()
                per_list.append(per)
                sens_list.append(start)
                if float(per) > float(per_spec):
                    logger.info('Sensitivity fail!')
                else:
                    while float(per) <= float(per_spec):
                        start = start - 1
                        # dt.rx(mode, channel, bw, rates, chain)
                        # dt.rx()
                        iq_wanted.vsg(mw_iq, start)
                        per = dt.get_statistics()
                        per_list.append(per)
                        sens_list.append(start)
                    logger.debug(str(len(per_list)) + str(per_list))
                    logger.debug(str(len(sens_list)) + str(sens_list))
                    per = per_list[len(per_list) - 2]
                    sens = sens_list[len(sens_list) - 2]
                    logger.info('Sensitivity: ' + str(sens))
                    # AJ TEST
                    aj_list = []
                    aj_per_list = []
                    # AJ sens
                    if mode == '11b':
                        aj_sens = sens + 6
                    else:
                        aj_sens = sens + 3
                    # GET SENS SPEC
                    rate_t = re.sub('-', '_', rate)
                    rate_t = re.sub('\.', '_', rate_t)
                    aj_spec = eval('target_aj_' + rate_t)
                    logger.debug(aj_spec)
                    aj_start = aj_sens + aj_spec
                    logger.debug(aj_start)
                    # AJ aj_vsg
                    iq_interfere.vsg_aj(mw_iq_inter, aj_start)
                    iq_wanted.vsg(mw_iq, aj_sens)
                    aj_per = dt.get_statistics()
                    aj_per_list.append(per)
                    aj_list.append(aj_start)
                    if float(aj_per) > float(per_spec):
                        aj = aj_start
                        result = 'Fail'
                        iq_interfere.vsg_off()
                    else:
                        while float(aj_per) <= float(per_spec) and aj_start < -10:
                            aj_start = aj_start + 1
                            # dt.rx(mode, channel, bw, rates, chain)
                            # dt.rx()
                            iq_interfere.vsg_aj(mw_iq_inter, aj_start)
                            iq_wanted.vsg(mw_iq, aj_sens)
                            aj_per = dt.get_statistics()
                            aj_per_list.append(aj_per)
                            aj_list.append(aj_start)
                        iq_interfere.vsg_off()
                        logger.debug(str(len(aj_per_list)) + str(aj_per_list))
                        logger.debug(str(len(aj_list)) + str(aj_list))
                        aj_per = aj_per_list[len(aj_per_list) - 2]
                        aj = aj_list[len(aj_list) - 2] - aj_sens
                        logger.info('Adjacent Channel Rejection: ' + str(aj))
                        # RESULT
                        if aj < aj_spec:
                            result = 'Fail'
                        else:
                            result = 'Pass'
                        with open('./Result/' + aj_result_name, 'a+', newline='') as f3:
                            writer3 = csv.writer(f3)
                            writer3.writerow([channel, rate, chain, aj_spec, aj, result])
                logger.info('*************************************************************')
                dt.rx_off()
            elif item == 'NAJ':
                if gen_naj_report == 1:
                    # GEN REPORT
                    naj_result_name = sn + '_' + 'NAJ_Result' + '_' + now_time + '.csv'
                    with open('./Result/' + naj_result_name, 'w', newline='') as f:
                        writer = csv.writer(f)
                        writer.writerow(['FREQ', 'DATA_RATE', 'CHAIN', 'NonAdjacent Channel Rejection SPEC',
                                         'NonAdjacent Channel Rejection',
                                         'RESULT'])
                    gen_naj_report += 1
                if 2400 < int(channel) < 2483.5 and bw == '40':
                    logger.info('Test Channel is not right! No adjacent channel!')
                elif mode == '11b':
                    logger.info('Test mode is nor right!')
                else:
                    iq_interfere.vsg_off()
                    # GET SENS
                    start = int(line[4])
                    sens_list = []
                    per_list = []
                    logger.info('Start: ' + str(start))
                    dt.get_statistics()  # reset counts
                    # dt.rx_on()
                    dt.rx()
                    # logger.debug(channel)
                    iq_wanted.vsg(mw_iq, start)
                    per = dt.get_statistics()
                    per_list.append(per)
                    sens_list.append(start)
                    if float(per) > float(per_spec):
                        logger.info('Sensitivity fail!')
                    else:
                        while float(per) <= float(per_spec):
                            start = start - 1
                            # dt.rx(mode, channel, bw, rates, chain)
                            # dt.rx()
                            iq_wanted.vsg(mw_iq, start)
                            per = dt.get_statistics()
                            per_list.append(per)
                            sens_list.append(start)
                        logger.debug(str(len(per_list)) + str(per_list))
                        logger.debug(str(len(sens_list)) + str(sens_list))
                        per = per_list[len(per_list) - 2]
                        sens = sens_list[len(sens_list) - 2]
                        logger.info('Sensitivity: ' + str(sens))
                        # AJ TEST
                        naj_list = []
                        naj_per_list = []
                        # AJ sens
                        naj_sens = sens + 3
                        # GET SENS SPEC
                        rate_t = re.sub('-', '_', rate)
                        rate_t = re.sub('\.', '_', rate_t)
                        naj_spec = eval('target_naj_' + rate_t)
                        logger.debug(naj_spec)
                        naj_start = naj_sens + naj_spec
                        logger.debug(naj_start)
                        # AJ aj_vsg
                        iq_interfere.vsg_naj(mw_iq_inter, naj_start)
                        iq_wanted.vsg(mw_iq, naj_sens)
                        naj_per = dt.get_statistics()
                        naj_per_list.append(per)
                        naj_list.append(naj_start)
                        if float(naj_per) > float(per_spec):
                            naj = naj_start
                            result = 'Fail'
                            iq_interfere.vsg_off()
                        else:
                            while float(naj_per) <= float(per_spec) and naj_start < -10:
                                naj_start = naj_start + 1
                                # dt.rx(mode, channel, bw, rates, chain)
                                # dt.rx()
                                iq_interfere.vsg_naj(mw_iq_inter, naj_start)
                                iq_wanted.vsg(mw_iq, naj_sens)
                                naj_per = dt.get_statistics()
                                naj_per_list.append(naj_per)
                                naj_list.append(naj_start)
                            iq_interfere.vsg_off()
                            logger.debug(str(len(naj_per_list)) + str(naj_per_list))
                            logger.debug(str(len(naj_list)) + str(naj_list))
                            naj_per = naj_per_list[len(naj_per_list) - 2]
                            naj = naj_list[len(naj_list) - 2] - naj_sens
                            logger.info('NonAdjacent Channel Rejection: ' + str(naj))
                            # RESULT
                            if naj < naj_spec:
                                result = 'Fail'
                            else:
                                result = 'Pass'
                            with open('./Result/' + naj_result_name, 'a+', newline='') as f3:
                                writer3 = csv.writer(f3)
                                writer3.writerow([channel, rate, chain, naj_spec, naj, result])
                    logger.info('*************************************************************')
                    dt.rx_off()
            elif item == 'TX_MIMO':
                if gen_txmimo_report == 1:
                    # GEN REPORT
                    txmimo_result_name = sn + '_' + 'TX_MIMO_Result' + '_' + now_time + '.csv'
                    with open('./Result/' + txmimo_result_name, 'w', newline='') as f:
                        writer = csv.writer(f)
                        writer.writerow(['FREQ', 'DATA_RATE', 'CHAIN', 'TX_POWER', 'POWER', 'GAIN', 'LIMIT',
                                         'RESULT', 'EVM', 'LIMIT', 'RESULT', 'FREQ_ERROR', 'LIMIT', 'RESULT',
                                         'LOLEAKAGE', 'LIMIT', 'RESULT', 'OBW', 'LIMIT', 'RESULT',
                                         'MASK', 'LIMIT', 'RESULT', 'FLAsnESS', 'LIMIT', 'RESULT',
                                         'RAMPONTIME', 'LIMIT', 'RESULT', 'RAMPOFFTIME', 'LIMIT', 'RESULT'])
                    gen_txmimo_report += 1
                adjust_result = result_evm = result_symbol_clock_error = result_lo_leakage = result_mask = \
                    result_flasness = 'Pass'
                # dt.set_default()
                dt.tx_mimo()
                # RESULT
                rate_t = re.sub('-', '_', rate)
                rate_t = re.sub('\.', '_', rate_t)
                targetpower = eval('target_pwr_' + rate_t)
                spec_evm = eval('target_EVM_' + rate_t)
                iq_wanted.vsa_mimo(mw_iq)
                iq_wanted.analysis()
                pwr_len, txq_len, data_pwr, data_txq = iq_wanted.get_status()
                avg_power, result_evm, result_symbol_clock_error, result_lo_leakage, result_mask = \
                    iq_wanted.get_data(pwr_len, txq_len, data_pwr, data_txq, mode, channel, rate, chain,
                                       tx_result_name,
                                       targetpower, spec_pwr, gain, spec_evm, evm_margin,
                                       spec_symbol_clock_error,
                                       spec_lo_leakage, spec_mask, spec_obw_20M, spec_obw_40M, spec_obw_80M,
                                       spec_obw_160M)
                if AUTO_ADJUST_POWER == '1':
                    # if avg_power == 'NA' or result_evm == 'NA':
                    if avg_power == 'NA' or float(avg_power) > 99.000:
                        logger.info(Fore.RED + 'Error!' + Style.RESET_ALL)
                    else:
                        # accuracy_limit_left = 0.5
                        # accuracy_limit_right = 0.5
                        power_accuracy_left = float(targetpower) - float(spec_pwr) + float(accuracy_limit_left)
                        power_accuracy_right = float(targetpower) + float(spec_pwr) - float(
                            accuracy_limit_right)
                        max_power = float(targetpower + spec_pwr)
                        delta_power = float(avg_power) - float(targetpower)
                        delta_power = float('{:.3f}'.format(delta_power))
                        logger.debug('Default Measurer Power: ' + avg_power)
                        logger.debug('Target Power: ' + targetpower)
                        logger.debug('Delta Power: ' + delta_power)
                        # power is nomal but some is fail
                        if targetpower <= avg_power <= max_power:
                            if result_evm == 'Fail' or result_symbol_clock_error == 'Fail' \
                                    or result_lo_leakage == 'Fail' or result_mask == 'Fail':
                                logger.info(Fore.RED + 'TX\'s quality are failed!' + Style.RESET_ALL)
                            else:
                                logger.info('Get Good Result!')
                        # power is not nomal
                        else:
                            power_counts = 0
                            setup = 2
                            setups = 2
                            logger.debug('Step:' + setup)
                            init_value = dt.get_paras()
                            gain = init_value + setup
                            gain = hex(gain)
                            gain = re.sub('0x', '', gain)
                            get_power = []
                            get_delta_power = []
                            c = 1
                            get_power.append(avg_power)
                            logger.debug('Measurer Power List: ' + get_power)
                            get_delta_power.append(delta_power)
                            logger.debug('Power Deviation List: ' + get_delta_power)
                            logger.debug('Adjust Counts:' + c)
                            # ADJUST
                            min_adj = avg_power - targetpower
                            max_adj = avg_power - max_power
                            while min_adj < 0 or max_adj > 0:
                                logger.info(Fore.GREEN + 'Adjust...' + Style.RESET_ALL)
                                if avg_power > max_power:
                                    adj = -1
                                else:
                                    adj = 1
                                c = c + 1
                                logger.debug('NEW VALUE(HEX):' + gain)
                                gain = dt.adjust_power(gain)
                                iq_wanted.vsa(mw_iq)
                                iq_wanted.analysis()
                                pwr_len, txq_len, data_pwr, data_txq = iq_wanted.get_status()
                                avg_power, result_evm, result_symbol_clock_error, result_lo_leakage, result_mask = \
                                    iq_wanted.get_data(pwr_len, txq_len, data_pwr, data_txq, mode, channel,
                                                       rate, chain,
                                                       tx_result_name, targetpower, spec_pwr, gain, spec_evm,
                                                       evm_margin,
                                                       spec_symbol_clock_error, spec_lo_leakage, spec_mask,
                                                       spec_obw_20M,
                                                       spec_obw_40M, spec_obw_80M, spec_obw_160M)
                                get_power.append(avg_power)
                                logger.debug('Measurer Power List: ' + get_power)
                                delta_power = float(avg_power) - targetpower
                                delta_power = float('{:.3f}'.format(delta_power))
                                avg_power = float(avg_power)
                                min_adj = avg_power - targetpower
                                max_adj = avg_power - max_power
                                get_delta_power.append(delta_power)
                                logger.debug('Power Deviation List: ' + get_delta_power)
                                logger.debug('Adjust Counts: ' + c)
                                adjust_status = get_delta_power[c - 1] - get_delta_power[c - 2]
                                adjust_status = float('{:.3f}'.format(adjust_status))
                                logger.debug('Power Added: ' + adjust_status)
                                if adjust_status < 0.1:
                                    adjust_result = 'Pass'
                                    setup = 2 * setups * adj
                                    power_counts = power_counts + 1
                                elif adjust_status > 0.1:
                                    adjust_status = 'Pass'
                                    setup = 2 * adj
                                # logger.debug(power_counts)
                                if power_counts > 5:
                                    logger.info(
                                        Fore.RED + 'Power added was too small, Power adjust stop' + Style.RESET_ALL)
                                    adjust_result = 'Fail'
                                else:
                                    logger.debug('Step: ' + setup)
                                    gain = int(gain, 16)
                                    gain = int(gain) + setup
                                    gain = hex(gain)
                                    gain = re.sub('0x', '', gain)
                                    logger.debug(
                                        result_evm + result_symbol_clock_error + result_lo_leakage + result_mask)
                                    logger.info('*************************************************************')
                dt.tx_off()
    logger.info('************************TEST DONE****************************')
    dt.close()
    if iq_wanted is not None:
        iq_wanted.close()
    if iq_interfere is not None:
        iq_interfere.close()
