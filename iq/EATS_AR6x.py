#!/user/bin/env python
# encoding: utf-8
# @Author  : Ethan
# @time    : 2019/11/21 16:40

from __future__ import division
from colorama import init, Fore, Style
from openpyxl import load_workbook
from parameters import IQ_IP, IQ_PORT, IQ_ROUT, RX_PACKETS, EVM_AG, EVM_N, EVM_AC, EVM_AX, RL, IQ_IP_INTERFERE,\
    IQ_PORT_INTERFERE, IQ_ROUT_INTERFERE, DUT_IP, DUT_COM, DUT_BAUDRATE, DUT_USERNAME, DUT_PASSWORD, ID_2G, \
    ID_5G_LOW, ID_5G_HIGH, VAP_2G, VAP_5G, LOG_ENABLE, CALI_2G, CALI_5G, AUTO_ADJUST_POWER, accuracy_limit_left, \
    accuracy_limit_right, RX_DYNAMIC, PATHLOSS_WANGTED, PATHLOSS_INTERFERE, AUTO_ADJUST_POWER
import os
import sys
import time
import visa
import serial
import csv
import re
import logging
logger = logging.getLogger()


class IQxel():
    def __init__(self, ip, visaDLL=None, *args):
        self.ip = ip
        # self.visaDLL = 'C:\windows\system32\visa64.dll' if visaDLL is None else visaDLL
        self.address = 'TCPIP0::%s::hislip0::INSTR' % self.ip
        # self.resourceManager = visa.ResourceManager(self.visaDLL)
        self.resourceManager = visa.ResourceManager()
        self.instance = self.resourceManager.open_resource(self.address)

    def close(self):
        if self.instance is not None:
            self.resourceManager.close()
            self.instance = None

    def __formats__(self, format_spec: object) -> object:
        """
        format the result
        :param format_spec:
        :return:
        """
        self.format_spec = format_spec
        format_spec = float(format_spec)
        format_spec = '{:.3f}'.format(format_spec)
        return format_spec

    def save_image(self, imagname, fmt):
        """
        if save image
        :param imagname:
        :param fmt:
        :return:
        """
        assert fmt in ['jpg', 'png'], 'Invalid postfix of image'
        logger.debug('MMEM:STOR:IMAG "%s.%s"' % (imagname, fmt))
        self.instance.write('MMEM:STOR:IMAG "%s.%s"' % (imagname, fmt))

    def reset(self):
        self.instance.write('*RST')

    @property
    def read_idn(self):
        idn = self.instance.query('*IDN?')
        logger.debug(idn)
        iq_model = idn.split(',')[1]
        logger.debug(iq_model)
        iq_model = iq_model[1]
        logger.debug(iq_model)
        IQ_PORT_model = IQ_PORT.isdigit()
        logger.debug(IQ_PORT_model)
        if iq_model == 'IQXEL' and IQ_PORT_model is True:
            mw = ''
        elif iq_model == 'IQXEL-M' and IQ_PORT_model is False:
            mw = ''
        elif iq_model == 'IQXEL-MW' and IQ_PORT_model is False:
            mw = 'M'
        elif iq_model == 'IQXEL-M2W' and IQ_PORT_model is False:
            mw = ''
        else:
            mw = ''
        logger.debug(mw)
        return mw

    def set_pathloss(self, pathloss):
        """
        set pathloss, needs to change to support to 8x8 or any path
        2019-11-15 support for any path
        :return:
        """
        pathloss_list = []
        with open(pathloss) as f:
            pathloss = csv.reader(f)
            for path_loss in pathloss:
                logger.debug(path_loss)
                pathloss_list.append(path_loss)
            # write pathloss to iq
            rows = [row for row in pathloss_list]
            path_num = len(rows[0]) - 1
            for path in range(path_num):
                logger.debug(path)
                self.instance.write('MEM:TABLE "' + str(path) + '"; MEM:TABLE:DEFINE "FREQ,LOSS"')
                for path_loss in pathloss_list:
                    channel = path_loss[0]
                    pathloss = path_loss[path+1]
                    logger.debug(pathloss)
                    self.instance.write('MEM:TABLE "' + str(path) + '";MEMory:TABLe:INSert:POINt %s MHz,%s' %
                                            (channel, pathloss))
                self.instance.write('MEMory:TABLe "' + str(path) + '";MEMory:TABLe:STORe')

    def use_pathloss(self, mw):
        self.instance.write('%sVSA%s;RFC:USE "%s",RF%s' % (mw, IQ_ROUT, chain, IQ_PORT))
        self.instance.write('%sVSA%s;RFC:STAT  ON,RF%s' % (mw, IQ_ROUT, IQ_PORT))
        self.instance.write('%sVSG%s;RFC:USE "%s",RF%s' % (mw, IQ_ROUT, chain, IQ_PORT))
        self.instance.write('%sVSG%s;RFC:STAT  ON,RF%s' % (mw, IQ_ROUT, IQ_PORT))

    def vsa(self, mw):
        # print(mw, IQ_ROUT, IQ_PORT, IQ_ROUT)
        self.instance.write('%sROUT%s;PORT:RES RF%s,VSA%s' % (mw, IQ_ROUT, IQ_PORT, IQ_ROUT))
        self.instance.write('CHAN1;WIFI')
        self.instance.write('VSA1;TRIG:SOUR VIDeo')
        logger.debug(channel)
        channels = int(channel) * 1000000
        logger.debug(channels)
        self.instance.write('%sVSA%s;FREQ:cent %s' % (mw, IQ_ROUT, str(channels)))
        if RL == '1':
            self.instance.write('%sVSA%s;RLEVel:AUTO' % (mw, IQ_ROUT))
        else:
            rlevel = target_power + 12
            logger.debug(str(rlevel))
            self.instance.write('%sVSA%s;RLEV %d;*wai;*opc?' % (mw, IQ_ROUT, rlevel))
        if int(bw) > 80:
            sampling_rate = 240000000
        else:
            sampling_rate = 160000000
        self.instance.write('%sVSA%s;SRAT %d' % (mw, IQ_ROUT, sampling_rate))
        if mode == '11b':
            capture_time = (32 - float(rates) * 2) / 1000
            standard = 'DSSS'
            mask_limit = 'AUTO'
        elif mode == '11g' or mode == '11a':
            capture_time = (13.2 - float(rates) * 3 / 16) / 1000
            standard = 'OFDM'
            mask_limit = 'AUTO'
        elif mode == '11n':
            capture_time = (12 - float(rates) * 10 / 7) / 1000
            standard = 'OFDM'
            mask_limit = 'AUTO'
        elif mode == '11ac':
            capture_time = (8 - float(rates) * 2 / 3) / 1000
            standard = 'OFDM'
            mask_limit = 'AUTO'
        elif mode == '11ax':
            capture_time = (9 - float(rates) * 2 / 3) / 1000
            standard = 'OFDM'
            mask_limit = '11AX'
        else:
            capture_time = 10
            standard = 'OFDM'
            mask_limit = 'AUTO'
        if mode == '11a' and EVM_AG == '1':
            channel_estimation = 'DATA'
        elif mode == '11g' and EVM_AG == '1':
            channel_estimation = 'DATA'
        elif mode == '11n' and EVM_N == '1':
            channel_estimation = 'DATA'
        elif mode == '11ac' and EVM_AC == '1':
            channel_estimation = 'DATA'
        elif mode == '11ax' and EVM_AX == '1':
            channel_estimation = 'DATA'
        else:
            channel_estimation = 'LTF'
        logger.debug('Captime: ' + str(capture_time))
        self.instance.write('CHAN1;WIFI;CONF:STAN %s' % standard)
        self.instance.write('CHAN1;WIFI;CONF:OFDM:CEST %s' % channel_estimation)
        self.instance.write('CHAN1;WIFI;CONF:SPEC:HLIM:TYPE %s' % mask_limit)
        self.instance.write('VSA%s;CAPT:TIME %s' % (IQ_ROUT, capture_time))

    def analysis(self):
        """
        let iq get the data and analysis
        :return:
        """
        if mode == '11b':
            logger.debug('DSSS')
            self.instance.write('CHAN1')
            self.instance.write('VSA%s ;init' % IQ_ROUT)
            self.instance.write('WIFI')
            self.instance.write('calc:pow 0, 2')
            time.sleep(1)
            self.instance.write('calc:txq 0, 2')
            time.sleep(1)
            self.instance.write('calc:ccdf 0, 2')
            self.instance.write('calc:ramp 0, 2')
            self.instance.write('calc:spec 0, 2')
            self.instance.query('WIFI;FETC:SEGM:POW:AVER?')
        else:
            logger.debug('OFDM')
            self.instance.write('CHAN1')
            self.instance.write('VSA%s ;init' % IQ_ROUT)
            self.instance.write('WIFI')
            self.instance.write('calc:pow 0, 2')
            time.sleep(1)
            self.instance.write('calc:txq 0, 2')
            time.sleep(1)
            self.instance.write('calc:ccdf 0, 2')
            self.instance.write('calc:ramp 0, 2')
            self.instance.write('calc:spec 0, 2')
            self.instance.query('WIFI;FETC:SEGM:POW:AVER?')

    def get_status(self):
        """
        get the data and check
        :return:
        """
        time.sleep(1)
        data_pwr = self.instance.query_ascii_values('WIFI;FETC:SEGM:POW:AVER?')
        if mode == '11b':
            data_txq = self.instance.query_ascii_values('WIFI;FETC:SEGM:TXQ:DSSS:AVER?')
            txqlen = 10
        else:
            data_txq = self.instance.query_ascii_values('WIFI;FETC:SEGM:TXQ:OFDM:AVER?')
            txqlen = 8
        pwr_len = len(data_pwr)
        txq_len = len(data_txq)
        logger.debug(data_pwr)
        logger.debug(pwr_len)
        logger.debug(data_txq)
        logger.debug(txq_len)
        status = 0
        while pwr_len != 2 or txq_len != txqlen:
            self.analysis()
            data_pwr = self.instance.query_ascii_values('WIFI;FETC:SEGM:POW:AVER?')
            if mode == '11b':
                data_txq = self.instance.query_ascii_values('WIFI;FETC:SEGM:TXQ:DSSS:AVER?')
            else:
                data_txq = self.instance.query_ascii_values('WIFI;FETC:SEGM:TXQ:OFDM:AVER?')
            pwr_len = len(data_pwr)
            txq_len = len(data_txq)
            logger.debug(data_pwr)
            logger.debug(pwr_len)
            logger.debug(data_txq)
            logger.debug(txq_len)
            status += 1
            if status > 2:
                break
        return pwr_len, txq_len, data_pwr, data_txq

    def get_data(self, pwr_len, txq_len, data_pwr, data_txq, mode, channel, rate, chain, tx_result_name, target_power,
                 spec_pwr, pwra_paras, spec_evm, evm_margin, spec_symbol_clock_error, spec_lo_leakage, spec_mask,
                 spec_obw_20M, spec_obw_40M,spec_obw_80M):
        """
        get the test data and write to report
        :param pwr_len:
        :param txq_len:
        :param data_pwr:
        :param data_txq:
        :param mode:
        :param channel:
        :param rate:
        :param chain:
        :param tx_result_name:
        :param target_power:
        :param spec_pwr:
        :param pwra_paras:
        :param spec_evm:
        :param evm_margin:
        :param spec_symbol_clock_error:
        :param spec_lo_leakage:
        :param spec_mask:
        :param spec_obw_20M:
        :param spec_obw_40M:
        :param spec_obw_80M:
        :return:
        """
        if mode == '11b':
            if pwr_len < 2:
                logger.error('Error: ' + str(data_pwr))
                avg_power = 'NA'
                result_pwr = 'NA'
            else:
                # logger.debug(data_pwr)
                avg_power = data_pwr[1]
                logger.debug(avg_power)
                avg_power = self.__formats__(avg_power)
                delta_pwr = abs(float(avg_power) - target_power)
                logger.debug(delta_pwr)
                # spec_pwr = 2
                if delta_pwr > spec_pwr:
                    logger.info('Power:               ' + Fore.RED + avg_power + Style.RESET_ALL + 'dBm')
                    result_pwr = 'Fail'
                else:
                    logger.info('Power:               ' + Fore.BLUE + avg_power + Style.RESET_ALL + 'dBm')
                    result_pwr = 'Pass'
            # txquality
            if txq_len < 10:
                logger.error('Error:' + str(data_txq))
                avg_evm = 'NA'
                result_evm = 'NA'
                symbol_clock_error = 'NA'
                result_symbol_clock_error = 'NA'
                lo_leakage = 'NA'
                result_lo_leakage = 'NA'
            else:
                avg_evm = data_txq[1]
                avg_evm = self.__formats__(avg_evm)
                peak_evm = data_txq[2]
                peak_evm = self.__formats__(peak_evm)
                freq_error = data_txq[4]
                freq_error = self.__formats__(freq_error)
                peak_freq_error = data_txq[5]
                peak_freq_error = self.__formats__(peak_freq_error)
                symbol_clock_error = data_txq[6]
                symbol_clock_error = self.__formats__(symbol_clock_error)
                lo_leakage = data_txq[7]
                lo_leakage = self.__formats__(lo_leakage)
                # rate = rate + 'EVM'
                spec_evm = spec_evm - evm_margin
                if float(avg_evm) > spec_evm:
                    logger.info('EVM:                 ' + Fore.RED + avg_evm + Style.RESET_ALL + 'dB')
                    result_evm = 'Fail'
                else:
                    logger.info('EVM:                 ' + Fore.BLUE + avg_evm + Style.RESET_ALL + 'dB')
                    result_evm = 'Pass'
                logger.info('EVM Peak:            ' + peak_evm + 'dB')
                logger.info('Frequency Error:     ' + freq_error + 'kHz')
                logger.info('Frequency Error Peak:' + peak_freq_error + 'kHz')
                # spec_symbol_clock_error = eval('target_evm_'+rate)
                if abs(float(symbol_clock_error)) > abs(float(spec_symbol_clock_error)):
                    logger.info('Symbol Clock Error:  ' + Fore.RED + symbol_clock_error + Style.RESET_ALL + 'ppm')
                    result_symbol_clock_error = 'Fail'
                else:
                    logger.info('Symbol Clock Error:  ' + Fore.BLUE + symbol_clock_error + Style.RESET_ALL + 'ppm')
                    result_symbol_clock_error = 'Pass'
                # spec_lo_leakage = -15
                if abs(float(lo_leakage)) < abs(float(spec_lo_leakage)):
                    logger.info('LO Leakage:          ' + Fore.RED + lo_leakage + Style.RESET_ALL + 'dB')
                    result_lo_leakage = 'Fail'
                else:
                    logger.info('LO Leakage:          ' + Fore.BLUE + lo_leakage + Style.RESET_ALL + 'dB')
                    result_lo_leakage = 'Pass'
            # mask
            data_mask = self.instance.query_ascii_values('WIFI;FETC:SEGM:SPEC:AVER:VIOL?')
            time.sleep(0.2)
            logger.debug(data_mask)
            data_mask_len = len(data_mask)
            logger.debug(data_mask_len)
            if data_mask_len != 2:
                logger.error('Error: ' + str(data_mask))
                mask = 'NA'
                result_mask = 'NA'
            else:
                # logger.debug(data_mask)
                mask = data_mask[1]
                mask = self.__formats__(mask)
                # spec_mask = 5.12
                if abs(float(mask)) > spec_mask:
                    logger.info('Mask:                ' + Fore.RED + mask + Style.RESET_ALL + '%')
                    result_mask = 'Fail'
                else:
                    logger.info('Mask:                ' + Fore.BLUE + mask + Style.RESET_ALL + '%')
                    result_mask = 'Pass'
            # OBW
            data_obw = self.instance.query_ascii_values('WIFI;FETC:SEGM:SPEC:AVER:OBW?')
            time.sleep(0.2)
            logger.debug(data_obw)
            spec_obw = spec_obw_20M
            data_obw_len = len(data_obw)
            logger.debug(data_obw_len)
            if data_obw_len != 4:
                logger.error('Error: ' + str(data_obw))
                obw = 'NA'
                result_obw = 'NA'
            else:
                # logger.debug(data_obw)
                obw = data_obw[1]
                obw = self.__formats__(obw)
                obw = round(float(obw) / 1000000, 3)
                if obw > spec_obw:
                    logger.info('OBW:                 ' + Fore.RED + str(obw) + Style.RESET_ALL + 'MHz')
                    result_obw = 'Fail'
                else:
                    logger.info('OBW:                 ' + Fore.BLUE + str(obw) + Style.RESET_ALL + 'MHz')
                    result_obw = 'Pass'
            # RAMP
            # on time
            data_ontime = self.instance.query_ascii_values('WIFI;FETC:SEGM:RAMP:ON:TRIS?')
            time.sleep(0.2)
            logger.debug(data_ontime)
            spec_ramp_on_time = 2.0
            data_ontime_len = len(data_ontime)
            logger.debug(data_ontime_len)
            if data_ontime_len < 4:
                logger.error('Error: ' + str(data_ontime))
                ramp_on_time = 'NA'
                result_ramp_on_time = 'NA'
            else:
                # logger.debug(data_ontime)
                ramp_on_time = data_ontime[1]
                ramp_on_time = float(ramp_on_time) * 1000000
                ramp_on_time = self.__formats__(ramp_on_time)
                if float(ramp_on_time) > spec_ramp_on_time:
                    logger.info('Ramp On Time:        ' + Fore.RED + ramp_on_time + Style.RESET_ALL + 'us')
                    result_ramp_on_time = 'Fail'
                else:
                    logger.info('Ramp On Time:        ' + Fore.BLUE + ramp_on_time + Style.RESET_ALL + 'us')
                    result_ramp_on_time = 'Pass'
            # off time
            data_offtime = self.instance.query_ascii_values('WIFI;FETC:SEGM:RAMP:OFF:TRIS?')
            time.sleep(0.2)
            logger.debug(data_offtime)
            spec_ramp_off_time = 2.0
            data_offtime_len = len(data_offtime)
            logger.debug(data_offtime_len)
            if data_offtime_len < 4:
                logger.error('Error: ' + str(data_offtime))
                ramp_off_time = 'NA'
                result_ramp_off_time = 'NA'
            else:
                # logger.debug(data_offtime)
                ramp_off_time = data_offtime[1]
                ramp_off_time = float(ramp_off_time) * 1000000
                ramp_off_time = self.__formats__(ramp_off_time)
                if float(ramp_off_time) > spec_ramp_off_time:
                    logger.info('Ramp Off Time:       ' + Fore.RED + ramp_off_time + Style.RESET_ALL + 'us')
                    result_ramp_off_time = 'Fail'
                else:
                    logger.info('Ramp Off Time:       ' + Fore.BLUE + ramp_off_time + Style.RESET_ALL + 'us')
                    result_ramp_off_time = 'Pass'
            flatness = 'NA'
            spec_flatness = 'NA'
            result_flatness = 'NA'

        else:
            logger.debug('OFDM')
            if pwr_len < 2:
                logger.error('Error: ' + str(data_pwr))
                avg_power = 'NA'
                result_pwr = 'NA'
            else:
                # logger.debug(data)
                avg_power = data_pwr[1]
                avg_power = self.__formats__(avg_power)
                delta_pwr = abs(float(avg_power) - target_power)
                logger.debug(delta_pwr)
                ##spec_pwr = 2
                if delta_pwr > spec_pwr:
                    logger.info('Power:               ' + Fore.RED + avg_power + Style.RESET_ALL + 'dBm')
                    result_pwr = 'Fail'
                else:
                    logger.info('Power:               ' + Fore.BLUE + avg_power + Style.RESET_ALL + 'dBm')
                    result_pwr = 'Pass'
            # txquality
            if txq_len < 8:
                logger.error('Error: ' + str(data_txq))
                avg_evm = 'NA'
                result_evm = 'NA'
                symbol_clock_error = 'NA'
                result_symbol_clock_error = 'NA'
                lo_leakage = 'NA'
                result_lo_leakage = 'NA'
            else:
                logger.debug(data_txq)
                avg_evm = data_txq[1]
                avg_evm = self.__formats__(avg_evm)
                freq_error = data_txq[3]
                freq_error = self.__formats__(freq_error)
                symbol_clock_error = data_txq[4]
                symbol_clock_error = self.__formats__(symbol_clock_error)
                lo_leakage = data_txq[5]
                lo_leakage = self.__formats__(lo_leakage)
                spec_evm = spec_evm - abs(evm_margin)
                # logger.info(spec_evm)
                if float(avg_evm) > spec_evm:
                    logger.info('EVM:                 ' + Fore.RED + avg_evm + Style.RESET_ALL + 'dB')
                    result_evm = 'Fail'
                else:
                    logger.info('EVM:                 ' + Fore.BLUE + avg_evm + Style.RESET_ALL + 'dB')
                    result_evm = 'Pass'
                logger.info('Frequency Error:     ' + freq_error + 'kHz')
                # spec_symbol_clock_error = 10.0
                if abs(float(symbol_clock_error)) > abs(float(spec_symbol_clock_error)):
                    logger.info('Symbol Clock Error:  ' + Fore.RED + symbol_clock_error + Style.RESET_ALL + 'ppm')
                    result_symbol_clock_error = 'Fail'
                else:
                    logger.info('Symbol Clock Error:  ' + Fore.BLUE + symbol_clock_error + Style.RESET_ALL + 'ppm')
                    result_symbol_clock_error = 'Pass'
                # spec_lo_leakage = -15
                if abs(float(lo_leakage)) < abs(float(spec_lo_leakage)):
                    logger.info('LO Leakage:          ' + Fore.RED + lo_leakage + Style.RESET_ALL + 'dB')
                    result_lo_leakage = 'Fail'
                else:
                    logger.info('LO Leakage:          ' + Fore.BLUE + lo_leakage + Style.RESET_ALL + 'dB')
                    result_lo_leakage = 'Pass'
            # mask
            data_mask = self.instance.query_ascii_values('WIFI;FETC:SEGM:SPEC:AVER:VIOL?')
            time.sleep(0.2)
            logger.debug(data_mask)
            data_mask_len = len(data_mask)
            logger.debug(data_mask_len)
            if data_mask_len != 2:
                logger.error('Error: ' + str(data_mask))
                mask = 'NA'
                result_mask = 'NA'
            else:
                # logger.debug(data_mask)
                mask = data_mask[1]
                mask = self.__formats__(mask)
                # spec_mask = 5.12
                if abs(float(mask)) > spec_mask:
                    logger.info('Mask:                ' + Fore.RED + mask + Style.RESET_ALL + '%')
                    result_mask = 'Fail'
                else:
                    logger.info('Mask:                ' + Fore.BLUE + mask + Style.RESET_ALL + '%')
                    result_mask = 'Pass'
            # OBW
            data_obw = self.instance.query_ascii_values('WIFI;FETC:SEGM:SPEC:AVER:OBW?')
            time.sleep(0.2)
            logger.debug(data_obw)
            if bw == '80':
                spec_obw = spec_obw_80M
            elif bw == '40':
                spec_obw = spec_obw_40M
            else:
                spec_obw = spec_obw_20M
            data_obw_len = len(data_obw)
            logger.debug(data_obw_len)
            if data_obw_len != 4:
                logger.error('Error: ' + str(data_obw))
                obw = 'NA'
                result_obw = 'NA'
            else:
                # logger.debug(data_obw)
                obw = data_obw[1]
                obw = self.__formats__(obw)
                obw = round(float(obw) / 1000000, 3)
                if obw > spec_obw:
                    logger.info('OBW:                 ' + Fore.RED + str(obw) + Style.RESET_ALL + 'MHz')
                    result_obw = 'Fail'
                else:
                    logger.info('OBW:                 ' + Fore.BLUE + str(obw) + Style.RESET_ALL + 'MHz')
                    result_obw = 'Pass'
            # flatness
            # datas = self.instance.query('FETC:SEGM1:OFDM:SFL:SIGN1:AVER?')
            # time.sleep(0.2)
            # time.sleep(10)
            # logger.info('flatness', datas)
            data_flatness = self.instance.query('WIFI;FETC:SEGM:OFDM:SFL:AVER:CHEC?')
            # data_flatness = self.query('WIFI;FETC:SEGM:OFDM:SFL:AVER?')
            time.sleep(0.2)
            logger.debug(data_flatness)
            data_flatness = data_flatness.split(',')
            spec_flatness = 0
            data_flatness_len = len(data_flatness)
            logger.debug(data_flatness_len)
            if data_flatness_len < 2:
                logger.error('Error: ' + str(data_flatness))
                flatness = 'NA'
                result_flatness = 'NA'
            else:
                # logger.debug(data_flatness)
                flatness = data_flatness[0]
                if int(flatness) == spec_flatness:
                    logger.info('flatness:            ' + Fore.BLUE + flatness + Style.RESET_ALL)
                    result_flatness = 'Pass'
                else:
                    logger.info('flatness:            ' + Fore.RED + flatness + Style.RESET_ALL)
                    result_flatness = 'Fail'
            ramp_on_time = 'NA'
            spec_ramp_on_time = 'NA'
            result_ramp_on_time = 'NA'
            ramp_off_time = 'NA'
            spec_ramp_off_time = 'NA'
            result_ramp_off_time = 'NA'

        if AUTO_ADJUST_POWER == '1' and avg_power != 'NA':
            if float(avg_power) < float(power_accuracy_left) or float(avg_power) > float(power_accuracy_right):
                logger.info('Adjust power')
            else:
                with open('./Result/' + tx_result_name, 'a+', newline='') as write_result:
                    writer_file = csv.writer(write_result)
                    writer_file.writerow(
                        [channel, rate, chain, target_power, avg_power, pwra_paras, spec_pwr, result_pwr,
                         avg_evm, spec_evm, result_evm, symbol_clock_error, spec_symbol_clock_error,
                         result_symbol_clock_error, lo_leakage, spec_lo_leakage, result_lo_leakage, obw,
                         spec_obw, result_obw, mask, spec_mask, result_mask, flatness, spec_flatness,
                         result_flatness, ramp_on_time, spec_ramp_on_time, result_ramp_on_time,
                         ramp_off_time, spec_ramp_off_time, result_ramp_off_time])
        else:
            with open('./Result/' + tx_result_name, 'a+', newline='') as write_result:
                writer_file = csv.writer(write_result)
                writer_file.writerow([channel, rate, chain, target_power, avg_power, pwra_paras, spec_pwr, result_pwr,
                                      avg_evm, spec_evm, result_evm, symbol_clock_error, spec_symbol_clock_error,
                                      result_symbol_clock_error, lo_leakage, spec_lo_leakage, result_lo_leakage, obw,
                                      spec_obw, result_obw, mask, spec_mask, result_mask, flatness, spec_flatness,
                                      result_flatness, ramp_on_time, spec_ramp_on_time, result_ramp_on_time,
                                      ramp_off_time, spec_ramp_off_time, result_ramp_off_time])

        return avg_power, result_evm, result_symbol_clock_error, result_lo_leakage, result_mask

    def get_power(self):
        """
        just for calibration, only for power calibration
        :return:
        """
        if pwr_len < 2:
            logger.error('Error: ' + str(data_pwr))
            avg_power = 'NA'
        else:
            # logger.debug(data_pwr)
            avg_power = data_pwr[1]
            logger.debug(avg_power)
            avg_power = self.__formats__(avg_power)
            delta_pwr = abs(float(avg_power) - target_power)
            logger.debug(delta_pwr)
            spec_pwr = 2
            if delta_pwr > spec_pwr:
                logger.info('Power:               ' + Fore.RED + avg_power + Style.RESET_ALL + 'dBm')
            else:
                logger.info('Power:               ' + Fore.BLUE + avg_power + Style.RESET_ALL + 'dBm')
        return avg_power

    def get_ppm(self):
        """
        just for calibration, for ppm calibration only
        :return:
        """
        # txquality
        if txq_len < 8:
            logger.error('Error: ' + str(data_txq))
            avg_evm = 'NA'
            result_evm = 'NA'
            symbol_clock_error = 'NA'
            result_symbol_clock_error = 'NA'
            lo_leakage = 'NA'
            result_lo_leakage = 'NA'
        else:
            logger.debug(data_txq)
            symbol_clock_error = data_txq[4]
            symbol_clock_error = self.__formats__(symbol_clock_error)
            spec_symbol_clock_error = 5.0
            if abs(float(symbol_clock_error)) > abs(float(spec_symbol_clock_error)):
                logger.info('Symbol Clock Error:  ' + Fore.RED + symbol_clock_error + Style.RESET_ALL + 'ppm')
            else:
                logger.info('Symbol Clock Error:  ' + Fore.BLUE + symbol_clock_error + Style.RESET_ALL + 'ppm')
        return symbol_clock_error

    def vsg(self, mw, rlevel):
        # rint(mw, IQ_ROUT, IQ_PORT, IQ_ROUT)
        self.instance.write('%sROUT%s;PORT:RES RF%s,VSG%s' % (mw, IQ_ROUT, IQ_PORT, IQ_ROUT))
        self.instance.write('CHAN1;WIFI')
        channels = int(channel) * 1000000
        self.instance.write('%sVSG%s;FREQ:cent %d' % (mw, IQ_ROUT, channels))
        if int(bw) > 80:
            sampling_rate = 240000000
        else:
            sampling_rate = 160000000
        self.instance.write('%sVSA%s;SRAT %d' % (mw, IQ_ROUT, sampling_rate))
        # rlevel = start + loss
        logger.debug(rlevel)
        self.instance.write('%sVSG%s;POW:lev %d' % (mw, IQ_ROUT, rlevel))
        self.instance.write('VSG1;POW:STAT ON')
        self.instance.write('VSG1;MOD:STAT ON')
        self.instance.write('VSG1;WAVE:EXEC OFF')
        self.instance.write('VSG1;WLIS:COUN %d' % int(RX_PACKETS))
        # wave = 'OFDM-6'
        if mode == '11b' and rates == '1':
            wave = 'DSSS-1'
            vsg_delay = int(RX_PACKETS) / 100 + 5 - int(rates) * int(RX_PACKETS) / 1000
        elif mode == '11b' and rates == '2':
            wave = 'DSSS-2L'
            vsg_delay = int(RX_PACKETS) / 100 - int(rates) * int(RX_PACKETS) / 1000
        elif mode == '11b' and rates == '5.5':
            wave = 'CCK-5_5S'
            vsg_delay = int(RX_PACKETS) / 100 - int(float(rates) * int(RX_PACKETS) / 1000)
        elif mode == '11b' and rates == '11':
            wave = 'CCK-11S'
            vsg_delay = int(RX_PACKETS) / 100 - int(rates) * int(RX_PACKETS) / 1800
        elif mode == '11g' or mode == '11a':
            wave = 'OFDM-' + rates
            vsg_delay = int(RX_PACKETS) / 100 - int(rates) * int(RX_PACKETS) / 6000
        elif mode == '11ng' or mode == '11na':
            wave = 'HT' + bw + '_MCS' + rates
            vsg_delay = int(RX_PACKETS) / 100 - int(rates) * int(RX_PACKETS) / 1000
        elif mode == '11ac':
            wave = '11AC_VHT' + bw + '_S1_MCS' + rates
            vsg_delay = int(RX_PACKETS) / 100 - int(rates) * int(RX_PACKETS) / 1000
        elif mode == '11ax':
            wave = '11AX_HE' + bw + '_S1_HE' + rates
            vsg_delay = int(RX_PACKETS) / 100 - int(rates) * int(RX_PACKETS) / 1200
        # logger.info(wave)
        logger.info('Delay Time:' + str(vsg_delay))
        self.instance.write('VSG1; WAVE:LOAD "/user/WiFi_%s.iqvsg"' % wave)
        self.instance.write('VSG1 ;wave:exec off')
        self.instance.write('WLIST:WSEG1:DATA "/user/WiFi_%s.iqvsg"' % wave)
        self.instance.write('wlist:wseg1:save')
        self.instance.write('WLIST:COUNT:ENABLE WSEG1')
        self.instance.write('WAVE:EXEC ON, WSEG1')
        time.sleep(vsg_delay)
        self.instance.write('VSG1 ;WAVE:EXEC OFF')
        self.instance.write('WLIST:COUNT:DISABLE WSEG1')

    def vsg_aj(self, mw, rlevel):
        self.instance.write('%sROUT%s;PORT:RES RF%s,VSG%s' % (mw, IQ_ROUT_INTERFERE, IQ_PORT_INTERFERE, IQ_ROUT_INTERFERE))
        self.instance.write('CHAN1;WIFI')
        if int(bw) == 20 and 2400 < int(channel) < 2440:
            channel_aj = int(channel) + 25
        elif int(bw) == 20 and 2440 < int(channel) < 2483.5:
            channel_aj = int(channel) - 25
        elif int(bw) == 40 and int(channel) == 2422:
            channel_aj = int(channel) + 50
        elif int(bw) == 40 and int(channel) == 2452:
            channel_aj = int(channel) - 50
        elif int(bw) == 20:
            if 5100 < int(channel) < 5250 or 5500 <= int(channel) <= 5600 or 5745 <= int(channel) <= 5785:
                channel_aj = int(channel) + 20
            elif 5250 < int(channel) <= 5320 or 5600 < int(channel) < 5745 or 5785 < int(channel) <= 5825:
                channel_aj = int(channel) - 20
        elif int(bw) == 40:
            if 5100 < int(channel) < 5250 or 5500 <= int(channel) <= 5600 or 5745 <= int(channel) <= 5785:
                channel_aj = int(channel) + 40
            elif 5250 < int(channel) <= 5320 or 5600 < int(channel) < 5745 or 5785 < int(channel) <= 5825:
                channel_aj = int(channel) - 40
        elif int(bw) == 80:
            if 5100 < int(channel) < 5250 or 5500 <= int(channel) <= 5600:
                channel_aj = int(channel) + 80
            elif 5250 < int(channel) <= 5320 or 5600 < int(channel) < 5745 or 5745 < int(channel) <= 5825:
                channel_aj = int(channel) - 80
        elif int(bw) == 160:
            if 5100 < int(channel) < 5500:
                channel_aj = int(channel) + 160
            elif 5500 < int(channel) <= 5825:
                channel_aj = int(channel) - 160
        else:
            logger.info('Test Channel is not right! No adjacent channel!')
            return False
        channels = channel_aj * 1000000
        self.instance.write('%sVSG%s;FREQ:cent %d' % (mw, IQ_ROUT, channels))
        if int(bw) > 80:
            sampling_rate = 240000000
        else:
            sampling_rate = 160000000
        self.instance.write('%sVSA%s;SRAT %d' % (mw, IQ_ROUT, sampling_rate))
        # rlevel = start + loss
        logger.debug(rlevel)
        self.instance.write('%sVSG%s;POW:lev %d' % (mw, IQ_ROUT, rlevel))
        # wave = 'OFDM-6'
        if mode == '11b' and rates == '1':
            wave = 'DSSS-1'
        elif mode == '11b' and rates == '2':
            wave = 'DSSS-2L'
        elif mode == '11b' and rates == '5.5':
            wave = 'CCK-5_5S'
        elif mode == '11b' and rates == '11':
            wave = 'CCK-11S'
        elif mode == '11g' or mode == '11a':
            wave = 'OFDM-' + rates
        elif mode == '11ng' or mode == '11na':
            wave = 'HT' + bw + '_MCS' + rates
        elif mode == '11ac':
            wave = '11AC_VHT' + bw + '_S1_MCS' + rates
        elif mode == '11ax':
            wave = '11AX_HE' + bw + '_S1_HE' + rates
        # logger.info(wave)
        self.instance.write('VSG1; WAVE:LOAD "/user/WiFi_%s.iqvsg"' % wave)
        self.instance.write('VSG1;WAVE:EXEC ON')
        self.instance.write('VSG1;POW:STAT ON')
        self.instance.write('VSG1;MOD:STAT ON')

    def vsg_naj(self, mw, rlevel):
        self.instance.write('%sROUT%s;PORT:RES RF%s,VSG%s' % (mw, IQ_ROUT_INTERFERE, IQ_PORT_INTERFERE, IQ_ROUT_INTERFERE))
        self.instance.write('CHAN1;WIFI')
        if int(bw) == 20 and 2400 < int(channel) <= 2422:
            channel_naj = int(channel) + 50
        elif int(bw) == 20 and 2462 <= int(channel) < 2483.5:
            channel_naj = int(channel) - 50
        elif int(bw) == 20:
            if 5100 < int(channel) < 5250 or 5500 <= int(channel) <= 5600 or 5745 <= int(channel) <= 5785:
                channel_naj = int(channel) + 40
            elif 5250 < int(channel) <= 5320 or 5600 < int(channel) < 5745 or 5785 < int(channel) <= 5825:
                channel_naj = int(channel) - 40
        elif int(bw) == 40:
            if 5100 < int(channel) < 5250 or 5500 <= int(channel) <= 5600 or 5745 <= int(channel) <= 5785:
                channel_naj = int(channel) + 80
            elif 5250 < int(channel) <= 5320 or 5600 < int(channel) < 5745 or 5785 < int(channel) <= 5825:
                channel_naj = int(channel) - 80
        elif int(bw) == 80:
            if 5100 < int(channel) < 5250 or 5500 <= int(channel) <= 5600:
                channel_naj = int(channel) + 160
            elif 5250 < int(channel) <= 5320 or 5600 < int(channel) < 5745 or 5745 < int(channel) <= 5825:
                channel_naj = int(channel) - 160
        elif int(bw) == 160:
            if 5100 < int(channel) < 5500:
                channel_naj = int(channel) + 320
            elif 5500 < int(channel) <= 5825:
                channel_naj = int(channel) - 320
        else:
            logger.info('Test Channel is not right! No adjacent channel!')
            return False
        channels = channel_naj * 1000000
        self.instance.write('%sVSG%s;FREQ:cent %d' % (mw, IQ_ROUT, channels))
        if int(bw) > 80:
            sampling_rate = 240000000
        else:
            sampling_rate = 160000000
        self.instance.write('%sVSA%s;SRAT %d' % (mw, IQ_ROUT, sampling_rate))
        # rlevel = start + loss
        logger.debug(rlevel)
        self.instance.write('%sVSG%s;POW:lev %d' % (mw, IQ_ROUT, rlevel))
        # wave = 'OFDM-6'
        if mode == '11b' and rates == '1':
            wave = 'DSSS-1'
        elif mode == '11b' and rates == '2':
            wave = 'DSSS-2L'
        elif mode == '11b' and rates == '5.5':
            wave = 'CCK-5_5S'
        elif mode == '11b' and rates == '11':
            wave = 'CCK-11S'
        elif mode == '11g' or mode == '11a':
            wave = 'OFDM-' + rates
        elif mode == '11ng' or mode == '11na':
            wave = 'HT' + bw + '_MCS' + rates
        elif mode == '11ac':
            wave = '11AC_VHT' + bw + '_S1_MCS' + rates
        elif mode == '11ax':
            wave = '11AX_HE' + bw + '_S1_HE' + rates
        # logger.info(wave)
        self.instance.write('VSG1; WAVE:LOAD "/user/WiFi_%s.iqvsg"' % wave)
        self.instance.write('VSG1;WAVE:EXEC ON')
        self.instance.write('VSG1;POW:STAT ON')
        self.instance.write('VSG1;MOD:STAT ON')

    def vsg_off(self):
        self.instance.write('VSG1 ;WAVE:EXEC OFF')


class DUT():
    """
    for serial connect
    """

    def __init__(self, com, baudrate):
        """
        param, timeout must set if use readline() or readlines()
        :param com:
        :param baudrate:
        """
        self.com = com
        self.baudrate = baudrate
        self.sn = serial.Serial(self.com, self.baudrate)
        self.sn.timeout = 0.5

    def close(self):
        if self.sn is not None:
            self.sn.close()
            self.sn = None

    def login(self, username, password):
        """

        :param username:
        :param password:
        :return:
        """
        if username is not None:
            self.sn.read_until(b'Login: ')
            self.sn.write(username.encode('ascii') + b'\n')
        if password is not None:
            self.sn.read_until(b'Password:')
            self.sn.write(password.encode('ascii') + b'\n')

        time.sleep(1)
        command_result = self.sn.readlines()
        for result in command_result:
            logger.debug(result.strip().decode('utf-8'))
        if 'wrong' not in command_result:
            logging.info('%s Sign up' % username)
            return True
        else:
            logging.warning('%s Login Fail' % username)
            return False

    def init(self, str1, str2, str3, str4):
        """
        additional command if you need to excite
        :param str1:
        :param str2:
        :param str3:
        :return:
        """
        self.sn.write(b'\r\n')
        self.sn.write(str1.encode('ascii') + b'\r\n')
        self.sn.write(str2.encode('ascii') + b'\r\n')
        self.sn.write(str3.encode('ascii') + b'\r\n')
        self.sn.write(str4.encode('ascii') + b'\r\n')
        command_result = self.sn.readlines()
        for result in command_result:
            logger.debug(result.strip().decode('utf-8'))

    def init_cali(self):
        """
        init for calibration, cancel it and use init() or manual
        :return:
        """
        logger.info('Initialization for calibration...')
        self.sn.write(b'shell sh\r\n')
        time.sleep(1)
        self.sn.write(b'stty echo\r\n')
        time.sleep(1)
        self.sn.write(b'cd /mnt/caldata\r\n')
        time.sleep(1)
        self.sn.write(b'rm mtdblock0\r\n')
        time.sleep(1)
        self.sn.write(b'rm cfg_ont_hisi.ini\r\n')
        time.sleep(1)
        self.sn.write(b'exit\r\n')
        time.sleep(1)
        self.sn.write(b'wifi calibrate test parameter init\r\n')
        time.sleep(2)
        self.sn.write(b'shell WifiChipInit.sh \r\n')
        time.sleep(5)
        command_result = self.sn.readlines()
        for result in command_result:
            logger.debug(result.strip().decode('utf-8'))

    def init_ppm(self):
        """
        output a tx signal for ppm cali
        :return:
        """
        if band == '1':
            self.sn.write(b'shell iwpriv Hisilicon0 adjust_ppm 0\r\n')
            logger.debug('shell iwpriv Hisilicon0 adjust_ppm 0')
            self.sn.write(b'shell WifiTxInit.sh 1 11g 20 1 54 0001\r\n')
            logger.debug('shell WifiTxInit.sh 1 11g 20 1 54 0001')
            time.sleep(3)
        else:
            self.sn.write(b'shell iwpriv Hisilicon1 adjust_ppm 0\r\n')
            logger.debug('shell iwpriv Hisilicon1 adjust_ppm 0')
            self.sn.write(b'shell WifiTxInit.sh 2 11a 20 36 54 0001\r\n')
            logger.debug('shell WifiTxInit.sh 2 11a 20 36 54 0001')
            time.sleep(3)
        command_result = self.sn.readlines()
        for result in command_result:
            logger.debug(result.strip().decode('utf-8'))

    def adjust_ppm(self, ppm):
        """
        adjust ppm to 0
        :param ppm:
        :return:
        """
        if band == '1':
            hi_id = '0'
        else:
            hi_id = '1'
        self.sn.write(b'shell iwpriv Hisilicon%s adjust_ppm %s\r\n' % (hi_id.encode('ascii'), str(ppm).encode('ascii')))
        command_result = self.sn.readlines()
        for result in command_result:
            logger.debug(result.strip().decode('utf-8'))

    def write_ppm(self, ppm):
        """
        write ppm to flash
        :param ppm:
        :return:
        """
        self.sn.write(b'wifi calibrate test parameter write wifichip %s type freqoffset len 4 value %s\r\n'
                      % (band.encode('ascii'), ppm.encode('ascii')))
        command_result = self.sn.readlines()
        for result in command_result:
            logger.debug(result.strip().decode('utf-8'))

    def cali_pwr(self, cali_channel, chain):
        """
        send a tx signal for cali
        :param cali_channel:
        :param chain:
        :return:
        """
        if band == '1':
            cali_mode = '11b'
            cali_rate = '11'
        else:
            cali_mode = '11a'
            cali_rate = '6'
        if chain == '0':
            chains = '01'
        else:
            chains = '10'
        self.sn.write(b'shell WifiTxPowerInit.sh %s %s 20 %s %s 00%s\r\n' %
                      (band.encode('ascii'), cali_mode.encode('ascii'), str(cali_channel).encode('ascii'),
                       cali_rate.encode('ascii'), str(chain).encode('ascii')) + b'\r\n')
        time.sleep(2)
        logger.debug(f'shell WifiTxPowerInit.sh {band} {cali_mode} 20 {cali_channel} {cali_rate} 00{chains}')
        command_result = self.sn.readlines()
        for result in command_result:
            logger.debug(result.strip().decode('utf-8'))

    def adjust_pwr(self, adjust_power):
        """
        adjust the cali signal power, 200,160,120
        :param adjust_power:
        :return:
        """
        if int(channel) < 5000:
            vap = VAP_2G
        else:
            vap = VAP_5G
        self.sn.write(b'shell iwpriv vap%s txpower %s\r\n' % (vap.encode('ascii'), str(adjust_power).encode('ascii')))
        logger.debug(f'shell iwpriv vap{vap} txpower {adjust_power}')
        command_result = self.sn.readlines()
        for result in command_result:
            logger.debug(result.strip().decode('utf-8'))

    def cali_pwr_write(self, adjust_power_list):
        """
        write the test power to cali power
        :param adjust_power_list:
        :return:
        """
        if band == '1':
            hi_id = '0'
        else:
            hi_id = '1'
        if cali_channel < 5:
            subband = '0'
        elif cali_channel == 8 or cali_channel == 40:
            subband = '1'
        elif cali_channel == 12 or cali_channel == 56:
            subband = '2'
        elif cali_channel == 104:
            subband = '3'
        elif cali_channel == 120:
            subband = '4'
        elif cali_channel == 136:
            subband = '5'
        elif cali_channel == 157:
            subband = '6'
        else:
            subband = '0'
        self.sn.write(b'shell iwpriv Hisilicon%s cali_power "%s %s %s %s %s"\r\n' %
                      (hi_id.encode('ascii'), str(cali_chain).encode('ascii'), subband.encode('ascii'),
                       str(adjust_power_list[0]).encode('ascii'), str(adjust_power_list[1]).encode('ascii'),
                       str(adjust_power_list[2]).encode('ascii')) + b'\r\n')
        logger.debug(f'shell iwpriv Hisilicon{hi_id} cali_power "{cali_chain} {subband} {adjust_power_list[0]} '
                     f'{adjust_power_list[1]} {adjust_power_list[2]}"')
        command_result = self.sn.readlines()
        for result in command_result:
            logger.debug(result.strip().decode('utf-8'))

    def cali_para_write(self):
        """
        get the cali param and write to flash
        :return:
        """
        if band == '1':
            hi_id = '0'
        else:
            hi_id = '1'
        self.sn.write(b'\r\n')
        self.sn.write(b'shell iwpriv Hisilicon%s get_power_param\r\n' % hi_id.encode('ascii') + b'\r\n')
        logger.debug(f'shell iwpriv Hisilicon{hi_id} get_power_param')
        self.sn.write(b'ls\r\n')
        command_result = self.sn.read_until(b'Enabled')
        logger.debug(command_result)
        pwrcali_para = re.findall(b'get_power_param:\r+\n+(.+)\r+', command_result)[0].decode('utf-8')
        logger.debug(command_result)
        self.sn.write(b'wifi calibrate test parameter write wifichip %s type power len %s value %s\r\n' %
                      (band.encode('ascii'), radio_adress.encode('ascii'), pwrcali_para.encode('ascii')) + b'\r\n')
        logger.debug(
            f'wifi calibrate test parameter write wifichip {band} type power len {radio_adress} value {pwrcali_para}')
        command_result = self.sn.readlines()
        for result in command_result:
            logger.debug(result.strip().decode('utf-8'))
        self.sn.write(b'wifi calibrate test parameter read wifichip %s type power len %s\r\n' %
                      (band.encode('ascii'), radio_adress.encode('ascii')) + b'\r\n')
        logger.debug((f'wifi calibrate test parameter read wifichip %s type power len {radio_adress}'))
        pwrcali_result = self.sn.read_until(b'success!')
        logger.debug(pwrcali_result)

    def crc(self):
        """
        crc cali
        :return:
        """
        self.sn.write(b'wifi calibrate parameter crc calc\r\n')
        self.sn.write(b'wifi calibrate parameter crc check\r\n')
        command_result = self.sn.readlines()
        for result in command_result:
            logger.debug(result.strip().decode('utf-8'))

    def ex_command(self, command):
        """
        not use
        :param command:
        :return:
        """
        self.sn.write(command.encode('ascii') + b'\n')
        time.sleep(2)
        command_result = self.sn.read_very_eager().decode('ascii')
        logging.info('\n%s' % command_result)

    def tx(self):
        """
        for tx test, set a tx signal
        :return:
        """
        if int(channel) < 5000:
            band = ID_2G
        elif int(channel) < 5500:
            band = ID_5G_LOW
        else:
            band = ID_5G_HIGH
        self.sn.write(b'shell WifiTxRxSwitch.sh %s tx off\r\n' % band.encode('ascii'))
        logger.debug(channel)
        if bw == '40':
            channels = int(channel) - 10
        elif bw == '80':
            channels = int(channel) - 30
        else:
            channels = channel
        if int(channel) < 5000:
            channels = '{:.0f}'.format((int(channels) - 2407) / 5)  # 2407+5*1=2412
        else:
            channels = '{:.0f}'.format((int(channels) - 5000) / 5)  # 2407+5*1=2412
        if chain == '0':
            chains = '01'
        else:
            chains = '10'
        # tx
        self.sn.write(b'\r\n')
        self.sn.write(b'shell WifiTxInit.sh %s %s %s %s %s 00%s' %
                      (band.encode('ascii'), mode.encode('ascii'), bw.encode('ascii'), str(channels).encode('ascii'),
                       rates.encode('ascii'), chains.encode('ascii')) + b'\r\n')
        logger.debug(f'shell WifiTxInit.sh {band} {mode} {bw} {channels} {rates} 00{chains}')
        command_result = self.sn.readlines()
        for result in command_result:
            logger.debug(result.strip().decode('utf-8'))
        logger.info('TX COMMANDS DONE')

    def tx_off(self):
        """
        off the tx signal
        :return:
        """
        if int(channel) < 5000:
            band = ID_2G
        elif int(channel) < 5500:
            band = ID_5G_LOW
        else:
            band = ID_5G_HIGH
        self.sn.write(b'shell WifiTxRxSwitch.sh %s tx off\r\n' % band.encode('ascii'))
        logger.debug(f'shell WifiTxRxSwitch.sh {band} tx off')
        command_result = self.sn.readlines()
        for result in command_result:
            logger.debug(result.strip().decode('utf-8'))

    def get_paras(self):
        """
        not use
        :return:
        """
        if int(channel) < 5000:
            gain = gain_24
        else:
            gain = gain_5
        if int(channel) < 5000:
            band = ID_2G
        elif int(channel) < 5500:
            band = ID_5G_LOW
        else:
            band = ID_5G_HIGH
        # power para
        if int(channel) < 2437 and chain == '0':
            channel_groups = '98'
        elif int(channel) < 2437 and chain == '1':
            channel_groups = 'a8'
        elif 2437 < int(channel) < 2462 and chain == '0':
            channel_groups = '94'
        elif 2437 < int(channel) < 2462 and chain == '1':
            channel_groups = 'a4'
        elif int(channel) >= 2462 and chain == '0':
            channel_groups = '90'
        elif int(channel) >= 2462 and chain == '1':
            channel_groups = 'a0'
        else:
            logger.info('5g need add')
        self.sn.write(b'hipriv.sh "vap%s set_tx_pow rf_reg_ctl 1"' % band.encode('ascii') + b'\n')

        self.sn.write(b'hipriv.sh "vap%s reginfo soc 0x200380%s 0x200380%s";dmesg -c'
                      % (band.encode('ascii'), channel_groups.encode('ascii'), channel_groups.encode('ascii')) + b'\n')
        # command_result = self.sn.read_very_eager().decode('ascii')
        command_result = self.sn.read_until(b'value=0x\w+\r\r\nWAP(Dopra Linux) # ', timeout=2)
        command_result = re.search(b'value=0x\w+', command_result)
        default_value = re.sub('value=0x', '', command_result.group().decode('ascii'))
        logger.info('Default Value(HEX): ' + default_value)
        default_value = int(default_value, 16)
        return default_value

    def adjust_power(self):
        """
        not use
        :return:
        """
        if int(channel) < 5000:
            band = ID_2G
        elif int(channel) < 5500:
            band = ID_5G_LOW
        else:
            band = ID_5G_HIGH
        # power para
        if int(channel) < 2437 and chain == '0':
            channel_groups = '98'
        elif int(channel) < 2437 and chain == '1':
            channel_groups = 'a8'
        elif 2437 < int(channel) < 2462 and chain == '0':
            channel_groups = '94'
        elif 2437 < int(channel) < 2462 and chain == '1':
            channel_groups = 'a4'
        elif int(channel) >= 2462 and chain == '0':
            channel_groups = '90'
        elif int(channel) >= 2462 and chain == '1':
            channel_groups = 'a0'
        else:
            logger.info('5g need add')
        self.sn.write(b'hipriv.sh "vap%s regwrite soc 0x200380%s 0x%s"'
                      % (band.encode('ascii'), channel_groups.encode('ascii'), pwr_paras.encode('ascii')) + b'\n')

        self.sn.write(b'hipriv.sh "vap%s reginfo soc 0x200380%s 0x200380%s";dmesg -c'
                      % (band.encode('ascii'), channel_groups.encode('ascii'), channel_groups.encode('ascii')) + b'\n')
        command_result = self.sn.read_until(b'value=0x\w+\r\r\nWAP(Dopra Linux) # ', timeout=1)
        command_result = re.search(b'value=0x\w+', command_result)
        default_value = re.sub('value=0x', '', command_result.group().decode('ascii'))
        logger.info('Check Value(HEX): ' + default_value)
        gain = default_value
        return gain

    def set_default(self):
        """
        off all the tx and rx
        :return:
        """
        self.sn.write(b'shell WifiTxRxSwitch.sh 1 tx off\r\n')
        self.sn.write(b'shell WifiTxRxSwitch.sh 2 tx off\r\n')
        self.sn.write(b'shell WifiTxRxSwitch.sh 1 rx off\r\n')
        self.sn.write(b'shell WifiTxRxSwitch.sh 2 rx off\r\n')
        # self.sn.write(b'shell WifiRxReset.sh %s' % band.encode('ascii') + b'\r\n')
        command_result = self.sn.readlines()
        for result in command_result:
            # logger.debug(result.decode('utf-8'))
            logger.debug(result.strip().decode('utf-8'))

    def rx_on(self):
        """
        turn on the rx
        :return:
        """
        if int(channel) < 5000:
            band = ID_2G
        elif int(channel) < 5500:
            band = ID_5G_LOW
        else:
            band = ID_5G_HIGH
        self.sn.write(b'shell WifiTxRxSwitch.sh %s rx on' % band.encode('ascii') + b'\r\n')
        command_result = self.sn.readlines()
        for result in command_result:
            logger.debug(result.strip().decode('utf-8'))

    def rx_off(self):
        """
        turn off the rx
        :return:
        """
        if int(channel) < 5000:
            band = ID_2G
        elif int(channel) < 5500:
            band = ID_5G_LOW
        else:
            band = ID_5G_HIGH
        self.sn.write(b'shell WifiTxRxSwitch.sh %s rx off' % band.encode('ascii') + b'\r\n')
        command_result = self.sn.readlines()
        for result in command_result:
            logger.debug(result.strip().decode('utf-8'))

    def rx(self):
        """
        set for rx test
        :return:
        """
        if int(channel) < 5000:
            band = ID_2G
        elif int(channel) < 5500:
            band = ID_5G_LOW
        else:
            band = ID_5G_HIGH
        logger.debug(band)
        logger.debug(mode)
        logger.debug(bw)
        logger.debug(channel)
        logger.debug(chain)
        # self.sn.write(b'shell WifiRxReset.sh %s\r\n' % band.encode('ascii'))
        # self.sn.write(b'shell WifiTxRxSwitch.sh %s tx off\r\n' % band.encode('ascii'))
        # self.sn.write(b'shell WifiTxRxSwitch.sh %s rx on' % band.encode('ascii') + b'\r\n')
        # self.sn.write(b'\r\n')
        logger.debug(channel)
        if bw == '40':
            channels = int(channel) - 10
        elif bw == '80':
            channels = int(channel) - 30
        else:
            channels = int(channel)

        if int(channel) < 5000:
            channels = '{:.0f}'.format((int(channels) - 2407) / 5)
        else:
            channels = '{:.0f}'.format((int(channels) - 5000) / 5)
        logger.debug(channels)
        if chain == '0':
            chains = '01'
        else:
            chains = '10'
        # if mode == '11b':
        #     modes = ''
        #     rates = 'rate'
        # elif mode == '11g':
        #     modes = '11ng'
        #     rates = 'mcs'
        # logger.debug('chain set',chain,type(chain),chains)
        self.sn.write(b'shell WifiRxInit.sh %s %s %s %s 00%s' % (band.encode('ascii'), mode.encode('ascii'),
                                                                 bw.encode('ascii'), str(channels).encode('ascii'),
                                                                 chains.encode('ascii')) + b'\r\n')
        logger.debug(f'shell WifiRxInit.sh {band} {mode} {bw} {channels} 00{chains}')
        command_result = self.sn.readlines()
        for result in command_result:
            logger.debug(result.strip().decode('utf-8'))
        logger.info('RX COMMANDS DONE')

    def get_statistics(self):
        """
        get the per result
        :return:
        """
        if int(channel) < 5000:
            vap = VAP_2G
        else:
            vap = VAP_5G
        self.sn.write(b'shell iwpriv vap%s get_rx_info' % vap.encode('ascii') + b'\r\n')
        time.sleep(3)
        command_result = self.sn.readlines()
        logger.debug(command_result)
        for result in command_result:
            logger.debug(result.strip().decode('utf-8'))
        logger.debug(command_result[2])
        rx_good = re.findall(b'RX OK:(\d+)', command_result[2])[0].decode('utf-8')
        logger.debug(rx_good)
        # command_result = self.sn.read_until(b'RX ERROR')
        # command_result = re.search(b'RX OK:\w+', command_result)
        # logger.debug(command_result)
        # logger.debug(command_result.group())
        # rx_good = re.sub('RX OK:', '', command_result.group().decode('ascii'))
        PER_value = (int(RX_PACKETS) - int(rx_good)) / int(RX_PACKETS)
        logger.info('Packets: ' + str(RX_PACKETS) + 'PER: ' + str(PER_value))
        # self.sn.write(b'shell WifiTxRxSwitch.sh %s rx off' % band.encode('ascii') + b'\r\n')
        return PER_value


if __name__ == '__main__':
    init(autoreset=True)

    # generate result and log
    directory_r = os.path.exists(r'./Result')
    if directory_r is False:
        os.makedirs('Result')

    directory_l = os.path.exists(r'./log')
    if directory_l is False:
        os.makedirs('log')

    # GEN TIME
    now_time = time.strftime("%Y%m%d%H%M%S", time.localtime())
    # logger.debug(now_time)
    logger.setLevel(logging.DEBUG)  # logger的总开关，只有大于Debug的日志才能被logger对象处理

    # 第二步，创建一个handler，用于写入日志文件
    file_handler = logging.FileHandler('./log/log_' + now_time + '.txt', mode='w')
    file_handler.setLevel(logging.DEBUG)  # 输出到file的log等级的开关
    # 创建该handler的formatter
    file_handler.setFormatter(
        logging.Formatter(
            fmt='%(asctime)s - %(filename)s[line:%(lineno)d] - %(levelname)s: %(message)s',
            datefmt='%Y-%m-%d %H:%M:%S')
    )
    # 添加handler到logger中
    logger.addHandler(file_handler)

    # 第三步，创建一个handler，用于输出到控制台
    console_handler = logging.StreamHandler()
    if LOG_ENABLE == '1':
        console_handler.setLevel(logging.DEBUG)  # 输出到控制台的log等级的开关
    else:
        console_handler.setLevel(logging.INFO)  # 输出到控制台的log等级的开关
    # 创建该handler的formatter
    console_handler.setFormatter(
        logging.Formatter(
            fmt='%(asctime)s - %(levelname)s: %(message)s',
            datefmt='%Y-%m-%d %H:%M:%S')
    )
    logger.addHandler(console_handler)

    # input sn
    try:
        Number = sys.argv[1]
    except:
        Number = input("SN:")
    sn = "SN_" + Number

    # test define
    target_pwr_1M, target_pwr_2M, target_pwr_5_5M, target_pwr_11M, target_pwr_6M, target_pwr_9M, \
    target_pwr_12M, target_pwr_18M, target_pwr_24M, target_pwr_36M, target_pwr_48M, target_pwr_54M, \
    target_pwr_HT20_MCS0, target_pwr_HT20_MCS1, target_pwr_HT20_MCS2, \
    target_pwr_HT20_MCS3, target_pwr_HT20_MCS4, target_pwr_HT20_MCS5, target_pwr_HT20_MCS6, \
    target_pwr_HT20_MCS7, target_pwr_HT40_MCS0, target_pwr_HT40_MCS1, target_pwr_HT40_MCS2, \
    target_pwr_HT40_MCS3, target_pwr_HT40_MCS4, target_pwr_HT40_MCS5, target_pwr_HT40_MCS6, \
    target_pwr_HT40_MCS7, target_pwr_VHT20_MCS0, target_pwr_VHT20_MCS1, target_pwr_VHT20_MCS2, \
    target_pwr_VHT20_MCS3, target_pwr_VHT20_MCS4, target_pwr_VHT20_MCS5, target_pwr_VHT20_MCS6, \
    target_pwr_VHT20_MCS7, target_pwr_VHT20_MCS8, target_pwr_VHT40_MCS0, target_pwr_VHT40_MCS1, \
    target_pwr_VHT40_MCS2, target_pwr_VHT40_MCS3, target_pwr_VHT40_MCS4, target_pwr_VHT40_MCS5, \
    target_pwr_VHT40_MCS6, target_pwr_VHT40_MCS7, target_pwr_VHT40_MCS8, target_pwr_VHT40_MCS9, \
    target_pwr_VHT80_MCS0, target_pwr_VHT80_MCS1, target_pwr_VHT80_MCS2, target_pwr_VHT80_MCS3, \
    target_pwr_VHT80_MCS4, target_pwr_VHT80_MCS5, target_pwr_VHT80_MCS6, target_pwr_VHT80_MCS7, \
    target_pwr_VHT80_MCS8, target_pwr_VHT80_MCS9, \
    target_pwr_VHT160_MCS0, target_pwr_VHT160_MCS1, target_pwr_VHT160_MCS2, target_pwr_VHT160_MCS3, \
    target_pwr_VHT160_MCS4, target_pwr_VHT160_MCS5, target_pwr_VHT160_MCS6, target_pwr_VHT160_MCS7, \
    target_pwr_VHT160_MCS8, target_pwr_VHT160_MCS9, target_pwr_HE20_HE0, target_pwr_HE20_HE1, target_pwr_HE20_HE2, \
    target_pwr_HE20_HE3, target_pwr_HE20_HE4, target_pwr_HE20_HE5, \
    target_pwr_HE20_HE6, target_pwr_HE20_HE7, target_pwr_HE20_HE8, target_pwr_HE20_HE9, target_pwr_HE20_HE10, \
    target_pwr_HE20_HE11, target_pwr_HE40_HE0, target_pwr_HE40_HE1, target_pwr_HE40_HE2, target_pwr_HE40_HE3, \
    target_pwr_HE40_HE4, target_pwr_HE40_HE5, target_pwr_HE40_HE6, target_pwr_HE40_HE7, target_pwr_HE40_HE8, \
    target_pwr_HE40_HE9, target_pwr_HE40_HE10, target_pwr_HE40_HE11, target_pwr_HE80_HE0, target_pwr_HE80_HE1, \
    target_pwr_HE80_HE2, target_pwr_HE80_HE3, target_pwr_HE80_HE4, target_pwr_HE80_HE5, target_pwr_HE80_HE6, \
    target_pwr_HE80_HE7, target_pwr_HE80_HE8, target_pwr_HE80_HE9, target_pwr_HE80_HE10, target_pwr_HE80_HE11, \
    target_pwr_HE160_HE0, target_pwr_HE160_HE1, target_pwr_HE160_HE2, target_pwr_HE160_HE3, target_pwr_HE160_HE4, \
    target_pwr_HE160_HE5, target_pwr_HE160_HE6, target_pwr_HE160_HE7, target_pwr_HE160_HE8, target_pwr_HE160_HE9, \
    target_pwr_HE160_HE10, target_pwr_HE160_HE11 = [None] * 115

    target_EVM_1M, target_EVM_2M, target_EVM_5_5M, target_EVM_11M, target_EVM_6M, target_EVM_9M, target_EVM_12M, \
    target_EVM_18M, target_EVM_24M, target_EVM_36M, target_EVM_48M, target_EVM_54M, target_EVM_HT20_MCS0, \
    target_EVM_HT20_MCS1, target_EVM_HT20_MCS2, target_EVM_HT20_MCS3, target_EVM_HT20_MCS4, target_EVM_HT20_MCS5, \
    target_EVM_HT20_MCS6, target_EVM_HT20_MCS7, target_EVM_HT40_MCS0, target_EVM_HT40_MCS1, target_EVM_HT40_MCS2, \
    target_EVM_HT40_MCS3, target_EVM_HT40_MCS4, target_EVM_HT40_MCS5, target_EVM_HT40_MCS6, target_EVM_HT40_MCS7, \
    target_EVM_VHT20_MCS0, target_EVM_VHT20_MCS1, target_EVM_VHT20_MCS2, target_EVM_VHT20_MCS3, target_EVM_VHT20_MCS4, \
    target_EVM_VHT20_MCS5, target_EVM_VHT20_MCS6, target_EVM_VHT20_MCS7, target_EVM_VHT20_MCS8, target_EVM_VHT40_MCS0, \
    target_EVM_VHT40_MCS1, target_EVM_VHT40_MCS2, target_EVM_VHT40_MCS3, target_EVM_VHT40_MCS4, target_EVM_VHT40_MCS5, \
    target_EVM_VHT40_MCS6, target_EVM_VHT40_MCS7, target_EVM_VHT40_MCS8, target_EVM_VHT40_MCS9, target_EVM_VHT80_MCS0, \
    target_EVM_VHT80_MCS1, target_EVM_VHT80_MCS2, target_EVM_VHT80_MCS3, target_EVM_VHT80_MCS4, target_EVM_VHT80_MCS5, \
    target_EVM_VHT80_MCS6, target_EVM_VHT80_MCS7, target_EVM_VHT80_MCS8, target_EVM_VHT80_MCS9, \
    target_EVM_VHT160_MCS0, target_EVM_VHT160_MCS1, target_EVM_VHT160_MCS2, target_EVM_VHT160_MCS3, \
    target_EVM_VHT160_MCS4, target_EVM_VHT160_MCS5, target_EVM_VHT160_MCS6, target_EVM_VHT160_MCS7, \
    target_EVM_VHT160_MCS8, target_EVM_VHT160_MCS9, target_EVM_HE20_HE0, target_EVM_HE20_HE1, target_EVM_HE20_HE2, \
    target_EVM_HE20_HE3, target_EVM_HE20_HE4, target_EVM_HE20_HE5, \
    target_EVM_HE20_HE6, target_EVM_HE20_HE7, target_EVM_HE20_HE8, target_EVM_HE20_HE9, target_EVM_HE20_HE10, \
    target_EVM_HE20_HE11, target_EVM_HE40_HE0, target_EVM_HE40_HE1, target_EVM_HE40_HE2, target_EVM_HE40_HE3, \
    target_EVM_HE40_HE4, target_EVM_HE40_HE5, target_EVM_HE40_HE6, target_EVM_HE40_HE7, target_EVM_HE40_HE8, \
    target_EVM_HE40_HE9, target_EVM_HE40_HE10, target_EVM_HE40_HE11, target_EVM_HE80_HE0, target_EVM_HE80_HE1, \
    target_EVM_HE80_HE2, target_EVM_HE80_HE3, target_EVM_HE80_HE4, target_EVM_HE80_HE5, target_EVM_HE80_HE6, \
    target_EVM_HE80_HE7, target_EVM_HE80_HE8, target_EVM_HE80_HE9, target_EVM_HE80_HE10, target_EVM_HE80_HE11, \
    target_EVM_HE160_HE0, target_EVM_HE160_HE1, target_EVM_HE160_HE2, target_EVM_HE160_HE3, target_EVM_HE160_HE4, \
    target_EVM_HE160_HE5, target_EVM_HE160_HE6, target_EVM_HE160_HE7, target_EVM_HE160_HE8, target_EVM_HE160_HE9, \
    target_EVM_HE160_HE10, target_EVM_HE160_HE11 = [None] * 115

    target_sens_1M, target_sens_2M, target_sens_5_5M, target_sens_11M, target_sens_6M, target_sens_9M, target_sens_12M, \
    target_sens_18M, target_sens_24M, target_sens_36M, target_sens_48M, target_sens_54M, target_sens_HT20_MCS0, \
    target_sens_HT20_MCS1, target_sens_HT20_MCS2, target_sens_HT20_MCS3, target_sens_HT20_MCS4, target_sens_HT20_MCS5, \
    target_sens_HT20_MCS6, target_sens_HT20_MCS7, target_sens_HT40_MCS0, target_sens_HT40_MCS1, target_sens_HT40_MCS2, \
    target_sens_HT40_MCS3, target_sens_HT40_MCS4, target_sens_HT40_MCS5, target_sens_HT40_MCS6, target_sens_HT40_MCS7, \
    target_sens_VHT20_MCS0, target_sens_VHT20_MCS1, target_sens_VHT20_MCS2, target_sens_VHT20_MCS3, target_sens_VHT20_MCS4, \
    target_sens_VHT20_MCS5, target_sens_VHT20_MCS6, target_sens_VHT20_MCS7, target_sens_VHT20_MCS8, target_sens_VHT40_MCS0, \
    target_sens_VHT40_MCS1, target_sens_VHT40_MCS2, target_sens_VHT40_MCS3, target_sens_VHT40_MCS4, target_sens_VHT40_MCS5, \
    target_sens_VHT40_MCS6, target_sens_VHT40_MCS7, target_sens_VHT40_MCS8, target_sens_VHT40_MCS9, target_sens_VHT80_MCS0, \
    target_sens_VHT80_MCS1, target_sens_VHT80_MCS2, target_sens_VHT80_MCS3, target_sens_VHT80_MCS4, target_sens_VHT80_MCS5, \
    target_sens_VHT80_MCS6, target_sens_VHT80_MCS7, target_sens_VHT80_MCS8, target_sens_VHT80_MCS9, \
    target_sens_VHT160_MCS0, target_sens_VHT160_MCS1, target_sens_VHT160_MCS2, target_sens_VHT160_MCS3, \
    target_sens_VHT160_MCS4, target_sens_VHT160_MCS5, target_sens_VHT160_MCS6, target_sens_VHT160_MCS7, \
    target_sens_VHT160_MCS8, target_sens_VHT160_MCS9, target_sens_HE20_HE0, target_sens_HE20_HE1, target_sens_HE20_HE2, \
    target_sens_HE20_HE3, target_sens_HE20_HE4, target_sens_HE20_HE5, \
    target_sens_HE20_HE6, target_sens_HE20_HE7, target_sens_HE20_HE8, target_sens_HE20_HE9, target_sens_HE20_HE10, \
    target_sens_HE20_HE11, target_sens_HE40_HE0, target_sens_HE40_HE1, target_sens_HE40_HE2, target_sens_HE40_HE3, \
    target_sens_HE40_HE4, target_sens_HE40_HE5, target_sens_HE40_HE6, target_sens_HE40_HE7, target_sens_HE40_HE8, \
    target_sens_HE40_HE9, target_sens_HE40_HE10, target_sens_HE40_HE11, target_sens_HE80_HE0, target_sens_HE80_HE1, \
    target_sens_HE80_HE2, target_sens_HE80_HE3, target_sens_HE80_HE4, target_sens_HE80_HE5, target_sens_HE80_HE6, \
    target_sens_HE80_HE7, target_sens_HE80_HE8, target_sens_HE80_HE9, target_sens_HE80_HE10, target_sens_HE80_HE11, \
    target_sens_HE160_HE0, target_sens_HE160_HE1, target_sens_HE160_HE2, target_sens_HE160_HE3, target_sens_HE160_HE4, \
    target_sens_HE160_HE5, target_sens_HE160_HE6, target_sens_HE160_HE7, target_sens_HE160_HE8, target_sens_HE160_HE9, \
    target_sens_HE160_HE10, target_sens_HE160_HE11 = [None] * 115

    target_aj_1M, target_aj_2M, target_aj_5_5M, target_aj_11M, target_aj_6M, target_aj_9M, target_aj_12M, \
    target_aj_18M, target_aj_24M, target_aj_36M, target_aj_48M, target_aj_54M, target_aj_HT20_MCS0, \
    target_aj_HT20_MCS1, target_aj_HT20_MCS2, target_aj_HT20_MCS3, target_aj_HT20_MCS4, target_aj_HT20_MCS5, \
    target_aj_HT20_MCS6, target_aj_HT20_MCS7, target_aj_HT40_MCS0, target_aj_HT40_MCS1, target_aj_HT40_MCS2, \
    target_aj_HT40_MCS3, target_aj_HT40_MCS4, target_aj_HT40_MCS5, target_aj_HT40_MCS6, target_aj_HT40_MCS7, \
    target_aj_VHT20_MCS0, target_aj_VHT20_MCS1, target_aj_VHT20_MCS2, target_aj_VHT20_MCS3, target_aj_VHT20_MCS4, \
    target_aj_VHT20_MCS5, target_aj_VHT20_MCS6, target_aj_VHT20_MCS7, target_aj_VHT20_MCS8, target_aj_VHT40_MCS0, \
    target_aj_VHT40_MCS1, target_aj_VHT40_MCS2, target_aj_VHT40_MCS3, target_aj_VHT40_MCS4, target_aj_VHT40_MCS5, \
    target_aj_VHT40_MCS6, target_aj_VHT40_MCS7, target_aj_VHT40_MCS8, target_aj_VHT40_MCS9, target_aj_VHT80_MCS0, \
    target_aj_VHT80_MCS1, target_aj_VHT80_MCS2, target_aj_VHT80_MCS3, target_aj_VHT80_MCS4, target_aj_VHT80_MCS5, \
    target_aj_VHT80_MCS6, target_aj_VHT80_MCS7, target_aj_VHT80_MCS8, target_aj_VHT80_MCS9, \
    target_aj_VHT160_MCS0, target_aj_VHT160_MCS1, target_aj_VHT160_MCS2, target_aj_VHT160_MCS3, \
    target_aj_VHT160_MCS4, target_aj_VHT160_MCS5, target_aj_VHT160_MCS6, target_aj_VHT160_MCS7, \
    target_aj_VHT160_MCS8, target_aj_VHT160_MCS9, target_aj_HE20_HE0, target_aj_HE20_HE1, target_aj_HE20_HE2, \
    target_aj_HE20_HE3, target_aj_HE20_HE4, target_aj_HE20_HE5, \
    target_aj_HE20_HE6, target_aj_HE20_HE7, target_aj_HE20_HE8, target_aj_HE20_HE9, target_aj_HE20_HE10, \
    target_aj_HE20_HE11, target_aj_HE40_HE0, target_aj_HE40_HE1, target_aj_HE40_HE2, target_aj_HE40_HE3, \
    target_aj_HE40_HE4, target_aj_HE40_HE5, target_aj_HE40_HE6, target_aj_HE40_HE7, target_aj_HE40_HE8, \
    target_aj_HE40_HE9, target_aj_HE40_HE10, target_aj_HE40_HE11, target_aj_HE80_HE0, target_aj_HE80_HE1, \
    target_aj_HE80_HE2, target_aj_HE80_HE3, target_aj_HE80_HE4, target_aj_HE80_HE5, target_aj_HE80_HE6, \
    target_aj_HE80_HE7, target_aj_HE80_HE8, target_aj_HE80_HE9, target_aj_HE80_HE10, target_aj_HE80_HE11, \
    target_aj_HE160_HE0, target_aj_HE160_HE1, target_aj_HE160_HE2, target_aj_HE160_HE3, target_aj_HE160_HE4, \
    target_aj_HE160_HE5, target_aj_HE160_HE6, target_aj_HE160_HE7, target_aj_HE160_HE8, target_aj_HE160_HE9, \
    target_aj_HE160_HE10, target_aj_HE160_HE11 = [None] * 115

    target_naj_1M, target_naj_2M, target_naj_5_5M, target_naj_11M, target_naj_6M, target_naj_9M, target_naj_12M, \
    target_naj_18M, target_naj_24M, target_naj_36M, target_naj_48M, target_naj_54M, target_naj_HT20_MCS0, \
    target_naj_HT20_MCS1, target_naj_HT20_MCS2, target_naj_HT20_MCS3, target_naj_HT20_MCS4, target_naj_HT20_MCS5, \
    target_naj_HT20_MCS6, target_naj_HT20_MCS7, target_naj_HT40_MCS0, target_naj_HT40_MCS1, target_naj_HT40_MCS2, \
    target_naj_HT40_MCS3, target_naj_HT40_MCS4, target_naj_HT40_MCS5, target_naj_HT40_MCS6, target_naj_HT40_MCS7, \
    target_naj_VHT20_MCS0, target_naj_VHT20_MCS1, target_naj_VHT20_MCS2, target_naj_VHT20_MCS3, target_naj_VHT20_MCS4, \
    target_naj_VHT20_MCS5, target_naj_VHT20_MCS6, target_naj_VHT20_MCS7, target_naj_VHT20_MCS8, target_naj_VHT40_MCS0, \
    target_naj_VHT40_MCS1, target_naj_VHT40_MCS2, target_naj_VHT40_MCS3, target_naj_VHT40_MCS4, target_naj_VHT40_MCS5, \
    target_naj_VHT40_MCS6, target_naj_VHT40_MCS7, target_naj_VHT40_MCS8, target_naj_VHT40_MCS9, target_naj_VHT80_MCS0, \
    target_naj_VHT80_MCS1, target_naj_VHT80_MCS2, target_naj_VHT80_MCS3, target_naj_VHT80_MCS4, target_naj_VHT80_MCS5, \
    target_naj_VHT80_MCS6, target_naj_VHT80_MCS7, target_naj_VHT80_MCS8, target_naj_VHT80_MCS9, \
    target_naj_VHT160_MCS0, target_naj_VHT160_MCS1, target_naj_VHT160_MCS2, target_naj_VHT160_MCS3, \
    target_naj_VHT160_MCS4, target_naj_VHT160_MCS5, target_naj_VHT160_MCS6, target_naj_VHT160_MCS7, \
    target_naj_VHT160_MCS8, target_naj_VHT160_MCS9, target_naj_HE20_HE0, target_naj_HE20_HE1, target_naj_HE20_HE2, \
    target_naj_HE20_HE3, target_naj_HE20_HE4, target_naj_HE20_HE5, \
    target_naj_HE20_HE6, target_naj_HE20_HE7, target_naj_HE20_HE8, target_naj_HE20_HE9, target_naj_HE20_HE10, \
    target_naj_HE20_HE11, target_naj_HE40_HE0, target_naj_HE40_HE1, target_naj_HE40_HE2, target_naj_HE40_HE3, \
    target_naj_HE40_HE4, target_naj_HE40_HE5, target_naj_HE40_HE6, target_naj_HE40_HE7, target_naj_HE40_HE8, \
    target_naj_HE40_HE9, target_naj_HE40_HE10, target_naj_HE40_HE11, target_naj_HE80_HE0, target_naj_HE80_HE1, \
    target_naj_HE80_HE2, target_naj_HE80_HE3, target_naj_HE80_HE4, target_naj_HE80_HE5, target_naj_HE80_HE6, \
    target_naj_HE80_HE7, target_naj_HE80_HE8, target_naj_HE80_HE9, target_naj_HE80_HE10, target_naj_HE80_HE11, \
    target_naj_HE160_HE0, target_naj_HE160_HE1, target_naj_HE160_HE2, target_naj_HE160_HE3, target_naj_HE160_HE4, \
    target_naj_HE160_HE5, target_naj_HE160_HE6, target_naj_HE160_HE7, target_naj_HE160_HE8, target_naj_HE160_HE9, \
    target_naj_HE160_HE10, target_naj_HE160_HE11 = [None] * 115

    #report lable
    gen_tx_report = gen_rx_report = gen_rxdynamic_report = gen_aj_report = gen_naj_report = 1
    # connect equipments and dut
    try:
        iq_wanted = IQxel(IQ_IP)
    except Exception as err:
        logger.info(err)
        iq_wanted = None
        logger.info(Fore.RED + 'IQ connect fail!' + Style.RESET_ALL)
        exit(1)
    else:
        mw_iq = iq_wanted.read_idn
        iq_wanted.set_pathloss(PATHLOSS_WANGTED)
        logger.info('IQ connected!')
    try:
        iq_interfere = IQxel(IQ_IP_INTERFERE)
    except Exception as err:
        logger.info(err)
        iq_interfere = None
        logger.info(Fore.RED + 'INTERFERE IQ connect fail!' + Style.RESET_ALL)
    else:
        mw_iq_inter = iq_interfere.read_idn
        iq_interfere.set_pathloss(PATHLOSS_INTERFERE)
        logger.info('INTERFERE IQ connected!')
    try:
        dt = DUT('COM' + DUT_COM, DUT_BAUDRATE)
    except Exception as err:
        logger.info(err + Fore.RED + 'DUT Open fail!' + Style.RESET_ALL)
    else:
        logger.info('DUT connected!')

    # calibration
    if CALI_2G == '1' and CALI_5G == '1':
        cali_list = [1, 1]
        band_list = [1, 2]
    elif CALI_2G == '1':
        cali_list = [1, 0]
        band_list = [1, 0]
    elif CALI_5G == '1':
        cali_list = [0, 1]
        band_list = [0, 2]
    else:
        band_list = []
        cali_list = []
    if band_list and cali_list is not None:
        cali_para = [[0 for cali in cali_list] for band in band_list]
        cali_para[0][0] = cali_list[0]
        cali_para[0][1] = band_list[0]
        cali_para[1][0] = cali_list[1]
        cali_para[1][1] = band_list[1]
        logger.debug(cali_para)
        for cali, band in cali_para:
            cali = str(cali)
            band = str(band)
            logger.info('Band' + band)
            if cali == '1':
                if band == '1':
                    channel = 2412
                    cali_channel_list = [3, 8, 12]
                    cali_mode = '11b'
                    cali_rate = '11'
                    radio_adress = '77'
                else:
                    channel = 5180
                    cali_channel_list = [40, 56, 104, 120, 136, 157]
                    cali_mode = '11a'
                    cali_rate = '6'
                    radio_adress = '155'
                # dt.init_cali()
                logger.debug(band)
                logger.info('Calibration...')
                # ppm cali
                logger.info('PPM Calibration...')
                chain = '0'
                mode = '11g/a'
                bw = '20'
                dt.init_ppm()
                iq_wanted.set_port(20)
                iq_wanted.analysis()
                pwr_len, txq_len, data_pwr, data_txq = iq_wanted.get_status()
                symbol_clock_error = iq_wanted.adjust_ppm()
                init_ppm = int(float(symbol_clock_error))
                logger.debug('Default ppm: ' + str(init_ppm))
                dt.adjust_ppm(init_ppm)
                time.sleep(3)
                iq_wanted.analysis()
                pwr_len, txq_len, data_pwr, data_txq = iq_wanted.get_status()
                symbol_clock_error = iq_wanted.adjust_ppm()
                cali_ppm = int(float(symbol_clock_error))
                logger.debug('Cali ppm: ' + str(cali_ppm))
                if cali_ppm > 5 or cali_ppm < -5:
                    logger.info(Fore.RED + 'PPM CALIBRATION FAIL' + Style.RESET_ALL)
                else:
                    logger.info(Fore.GREEN + 'PPM CALIBRATION SUCCESS' + Style.RESET_ALL)
                # power cali
                logger.info('Power Calibration...')
                cali_chain_list = ['0', '1']
                cali_power_list = [200, 160, 120]
                for cali_chain in cali_chain_list:
                    if cali_chain == '0':
                        chain = '01'
                    else:
                        chain = '10'
                    for cali_channel in cali_channel_list:
                        adjust_power_list = []
                        dt.cali_pwr(cali_channel, chain)
                        if cali_channel < 30:
                            channel = 2407 + cali_channel * 5
                        else:
                            channel = 5000 + cali_channel * 5
                        for cali_power in cali_power_list:
                            dt.adjust_pwr(cali_power)
                            iq_wanted.set_port(int(cali_power / 10))
                            target_power = int(cali_power / 10)
                            mode = cali_mode
                            rates = cali_rate
                            iq_wanted.analysis()
                            pwr_len, txq_len, data_pwr, data_txq = iq_wanted.get_status()
                            avg_power = iq_wanted.adjust_power()
                            adjust_power = int(float(avg_power) * 10.0)
                            adjust_power_list.append(adjust_power)
                        dt.adjust_pwr(300)
                        dt.tx_off()
                        logger.debug(adjust_power_list)
                        dt.cali_pwr_write(adjust_power_list)
                dt.cali_para_write()
                dt.crc()
                logger.info(Fore.GREEN + 'POWER CALIBRATION SUCCESS' + Style.RESET_ALL)
            else:
                logger.info(Fore.YELLOW + 'Calibration skip' + Style.RESET_ALL)
    else:
        logger.info(Fore.YELLOW + 'Calibration skip' + Style.RESET_ALL)

    # TEST FLOW
    filename = 'TEST_FLOW.txt'
    f = open(filename)
    result = list()
    for line in f.readlines():
        # logger.debug(len(line))
        line = line.strip()
        if len(line) < 30 or line.startswith('//') or line.isspace():
            continue
            pass
        else:
            line = line.split()
            # logger.debug(line)
            # logger.debug('Channel:', line[1], 'Rate:', line[2], 'Chain:', line[3])
            item = line[0]
            item = re.sub('IQ_WIFI_TEST_', '', item)
            logger.debug(item)
            channel = line[1]
            rate = line[2]
            chain = line[3]

            # mode,bw,rates
            if rate == '1M' or rate == '2M' or rate == '5.5M' or rate == '11M':
                mode = '11b'
                bw = '20'
                rates = re.sub('M', '', rate)
                per_spec = 0.08
            elif rate == '6M' or rate == '9M' or rate == '12M' or rate == '18M' \
                    or rate == '24M' or rate == '36M' or rate == '48M' or rate == '54M':
                if int(channel) < 5000:
                    mode = '11g'
                else:
                    mode = '11a'
                bw = '20'
                rates = re.sub('M', '', rate)
                per_spec = 0.1
            elif rate == 'HT20-MCS0' or rate == 'HT20-MCS1' or rate == 'HT20-MCS2' or rate == 'HT20-MCS3' \
                    or rate == 'HT20-MCS4' or rate == 'HT20-MCS5' or rate == 'HT20-MCS6' or rate == 'HT20-MCS7':
                if int(channel) < 5000:
                    mode = '11ng'
                else:
                    mode = '11na'
                bw = '20'
                rates = re.sub('HT20-MCS', '', rate)
                per_spec = 0.1
            elif rate == 'HT40-MCS0' or rate == 'HT40-MCS1' or rate == 'HT40-MCS2' or rate == 'HT40-MCS3' \
                    or rate == 'HT40-MCS4' or rate == 'HT40-MCS5' or rate == 'HT40-MCS6' or rate == 'HT40-MCS7':
                if int(channel) < 5000:
                    mode = '11ng'
                else:
                    mode = '11na'
                bw = '40'
                rates = re.sub('HT40-MCS', '', rate)
                per_spec = 0.1
                # channel = int(channel) - 10
            elif rate == 'VHT20-MCS0' or rate == 'VHT20-MCS1' or rate == 'VHT20-MCS2' or rate == 'VHT20-MCS3' \
                    or rate == 'VHT20-MCS4' or rate == 'VHT20-MCS5' or rate == 'VHT20-MCS6' or rate == 'VHT20-MCS7' or \
                    rate == 'VHT20-MCS8':
                mode = '11ac'
                bw = '20'
                rates = re.sub('VHT20-MCS', '', rate)
                per_spec = 0.1
            elif rate == 'VHT40-MCS0' or rate == 'VHT40-MCS1' or rate == 'VHT40-MCS2' or rate == 'VHT40-MCS3' \
                    or rate == 'VHT40-MCS4' or rate == 'VHT40-MCS5' or rate == 'VHT40-MCS6' or rate == 'VHT40-MCS7' or \
                    rate == 'VHT40-MCS8' or rate == 'VHT40-MCS9':
                mode = '11ac'
                bw = '40'
                rates = re.sub('VHT40-MCS', '', rate)
                per_spec = 0.1
                # channel = int(channel) - 10
            elif rate == 'VHT80-MCS0' or rate == 'VHT80-MCS1' or rate == 'VHT80-MCS2' or rate == 'VHT80-MCS3' \
                    or rate == 'VHT80-MCS4' or rate == 'VHT80-MCS5' or rate == 'VHT80-MCS6' or rate == 'VHT80-MCS7' or \
                    rate == 'VHT80-MCS8' or rate == 'VHT80-MCS9':
                mode = '11ac'
                bw = '80'
                rates = re.sub('VHT80-MCS', '', rate)
                per_spec = 0.1
            elif rate == 'VHT160-MCS0' or rate == 'VHT160-MCS1' or rate == 'VHT160-MCS2' or rate == 'VHT160-MCS3' \
                    or rate == 'VHT160-MCS4' or rate == 'VHT160-MCS5' or rate == 'VHT160-MCS6' or rate == 'VHT160-MCS7' or \
                    rate == 'VHT160-MCS8' or rate == 'VHT160-MCS9':
                mode = '11ac'
                bw = '160'
                rates = re.sub('VHT160-MCS', '', rate)
                per_spec = 0.1
            elif rate == 'HE20-HE0' or rate == 'HE20-HE1' or rate == 'HE20-HE2' or rate == 'HE20-HE3' \
                    or rate == 'HE20-HE4' or rate == 'HE20-HE5' or rate == 'HE20-HE6' or rate == 'HE20-HE7' or \
                    rate == 'HE20-HE8' or rate == 'HE20-HE9' or rate == 'HE20-HE10' or rate == 'HE20-HE11':
                mode = '11ax'
                bw = '20'
                rates = re.sub('HE20-HE', '', rate)
                per_spec = 0.1
            elif rate == 'HE40-HE0' or rate == 'HE40-HE1' or rate == 'HE40-HE2' or rate == 'HE40-HE3' \
                    or rate == 'HE40-HE4' or rate == 'HE40-HE5' or rate == 'HE40-HE6' or rate == 'HE40-HE7' or \
                    rate == 'HE40-HE8' or rate == 'HE40-HE9' or rate == 'HE40-HE10' or rate == 'HE40-HE11':
                mode = '11ax'
                bw = '40'
                rates = re.sub('HE40-HE', '', rate)
                per_spec = 0.1
                # channel = int(channel) - 10
            elif rate == 'HE80-HE0' or rate == 'HE80-HE1' or rate == 'HE80-HE2' or rate == 'HE80-HE3' \
                    or rate == 'HE80-HE4' or rate == 'HE80-HE5' or rate == 'HE80-HE6' or rate == 'HE80-HE7' or \
                    rate == 'HE80-HE8' or rate == 'HE80-HE9' or rate == 'HE80-HE10' or rate == 'HE80-HE11':
                mode = '11ax'
                bw = '80'
                rates = re.sub('HE80-HE', '', rate)
                per_spec = 0.1
            elif rate == 'HE160-HE0' or rate == 'HE160-HE1' or rate == 'HE160-HE2' or rate == 'HE160-HE3' \
                    or rate == 'HE160-HE4' or rate == 'HE160-HE5' or rate == 'HE160-HE6' or rate == 'HE160-HE7' or \
                    rate == 'HE160-HE8' or rate == 'HE160-HE9' or rate == 'HE160-HE10' or rate == 'HE160-HE11':
                mode = '11ax'
                bw = '160'
                rates = re.sub('HE160-HE', '', rate)
                per_spec = 0.1
            # chain
            chain = re.sub('CHAIN', '', chain)
            path = chain
            # read spec
            if int(channel) < 5000:
                # INIT SPEC
                # 2.4g
                spec_file_2g = load_workbook('./spec_2g.xlsx')
                # logger.debug(spec_file.sheesnames)
                sheet_2g = spec_file_2g['Sheet1']
                rows_2g = []
                ratelist_2g = ['1M', '2M', '5_5M', '11M', '6M', '9M', '12M', '18M', '24M', '36M', '48M', '54M',
                               'HT20_MCS0', 'HT20_MCS1', 'HT20_MCS2', 'HT20_MCS3', 'HT20_MCS4', 'HT20_MCS5',
                               'HT20_MCS6', 'HT20_MCS7', 'HT40_MCS0', 'HT40_MCS1', 'HT40_MCS2', 'HT40_MCS3',
                               'HT40_MCS4', 'HT40_MCS5', 'HT40_MCS6', 'HT40_MCS7', 'VHT20_MCS0', 'VHT20_MCS1',
                               'VHT20_MCS2', 'VHT20_MCS3', 'VHT20_MCS4', 'VHT20_MCS5', 'VHT20_MCS6', 'VHT20_MCS7',
                               'VHT20_MCS8', 'VHT40_MCS0', 'VHT40_MCS1', 'VHT40_MCS2', 'VHT40_MCS3', 'VHT40_MCS4',
                               'VHT40_MCS5', 'VHT40_MCS6', 'VHT40_MCS7', 'VHT40_MCS8', 'VHT40_MCS9', 'HE20_HE0',
                               'HE20_HE1', 'HE20_HE2', 'HE20_HE3', 'HE20_HE4', 'HE20_HE5', 'HE20_HE6',
                               'HE20_HE7', 'HE20_HE8', 'HE20_HE9', 'HE20_HE10', 'HE20_HE11', 'HE40_HE0',
                               'HE40_HE1', 'HE40_HE2', 'HE40_HE3', 'HE40_HE4', 'HE40_HE5', 'HE40_HE6',
                               'HE40_HE7', 'HE40_HE8', 'HE40_HE9', 'HE40_HE10', 'HE40_HE11']
                for row_2g in sheet_2g:
                    rows_2g.append(row_2g)
                    # logger.debug(rows)
                for r in range(sheet_2g.max_row):
                    for c in range(sheet_2g.max_column):
                        # logger.debug(rows[r][c].value)
                        rows_2g[r][c].value = str(rows_2g[r][c].value).strip()
                        rs = r + 1
                        cs = c + 1
                        if rows_2g[r][c].value == 'POWER_ACCURACY':
                            spec_pwr = abs(rows_2g[r][cs].value)
                        elif rows_2g[r][c].value == 'Power_Gain_Index':
                            gain_24 = abs(rows_2g[r][cs].value)
                        elif rows_2g[r][c].value == 'EVM_MARGIN':
                            evm_margin = abs(rows_2g[r][cs].value)
                        elif rows_2g[r][c].value == 'Symbol_Clock_Error':
                            spec_symbol_clock_error = abs(rows_2g[r][cs].value)
                        elif rows_2g[r][c].value == 'XCAP':
                            xcap = abs(rows_2g[r][cs].value)
                        elif rows_2g[r][c].value == 'LO_Leakage':
                            spec_lo_leakage = -abs(rows_2g[r][cs].value)
                        elif rows_2g[r][c].value == 'MASK':
                            spec_mask = abs(rows_2g[r][cs].value)
                        elif rows_2g[r][c].value == 'OBW_20M':
                            spec_obw_20M = rows_2g[r][cs].value
                        elif rows_2g[r][c].value == 'OBW_40M':
                            spec_obw_40M = rows_2g[r][cs].value
                        for x in ratelist_2g:
                            if rows_2g[r][c].value == x + '_power':
                                exec('target_pwr_%s=%d' % (x, rows_2g[rs][c].value))
                                break
                        for i in ratelist_2g:
                            if rows_2g[r][c].value == i + '_EVM':
                                exec('target_EVM_%s=%d' % (i, rows_2g[rs][c].value))
                                break
                        for j in ratelist_2g:
                            if rows_2g[r][c].value == j + '_sens':
                                exec('target_sens_%s=%d' % (j, rows_2g[rs][c].value))
                                break
                        for y in ratelist_2g:
                            if rows_2g[r][c].value == y + '_aj':
                                exec('target_aj_%s=%d' % (y, rows_2g[rs][c].value))
                                break
                        for k in ratelist_2g:
                            if rows_2g[r][c].value == k + '_naj':
                                exec('target_naj_%s=%d' % (k, rows_2g[rs][c].value))
                                break
                        spec_obw_80M = spec_obw_160M = None
            else:
                # 5g
                spec_file_5g = load_workbook('./spec_5g.xlsx')
                # logger.debug(spec_file.sheesnames)
                sheet_5g = spec_file_5g['Sheet1']
                rows_5g = []
                ratelist_5g = ['6M', '9M', '12M', '18M', '24M', '36M', '48M', '54M', 'HT20_MCS0', 'HT20_MCS1',
                               'HT20_MCS2', 'HT20_MCS3', 'HT20_MCS4', 'HT20_MCS5', 'HT20_MCS6', 'HT20_MCS7',
                               'HT40_MCS0', 'HT40_MCS1', 'HT40_MCS2', 'HT40_MCS3', 'HT40_MCS4', 'HT40_MCS5',
                               'HT40_MCS6', 'HT40_MCS7', 'VHT20_MCS0', 'VHT20_MCS1', 'VHT20_MCS2', 'VHT20_MCS3',
                               'VHT20_MCS4', 'VHT20_MCS5', 'VHT20_MCS6', 'VHT20_MCS7', 'VHT20_MCS8', 'VHT40_MCS0',
                               'VHT40_MCS1', 'VHT40_MCS2', 'VHT40_MCS3', 'VHT40_MCS4', 'VHT40_MCS5', 'VHT40_MCS6',
                               'VHT40_MCS7', 'VHT40_MCS8', 'VHT40_MCS9', 'VHT80_MCS0', 'VHT80_MCS1', 'VHT80_MCS2',
                               'VHT80_MCS3', 'VHT80_MCS4', 'VHT80_MCS5', 'VHT80_MCS6', 'VHT80_MCS7', 'VHT80_MCS8',
                               'VHT80_MCS9', 'VHT160_MCS0', 'VHT160_MCS1', 'VHT160_MCS2',
                               'VHT160_MCS3', 'VHT160_MCS4', 'VHT160_MCS5', 'VHT160_MCS6', 'VHT160_MCS7', 'VHT160_MCS8',
                               'VHT160_MCS9', 'HE20_HE0', 'HE20_HE1', 'HE20_HE2', 'HE20_HE3', 'HE20_HE4',
                               'HE20_HE5', 'HE20_HE6', 'HE20_HE7', 'HE20_HE8', 'HE20_HE9', 'HE20_HE10',
                               'HE20_HE11', 'HE40_HE0', 'HE40_HE1', 'HE40_HE2', 'HE40_HE3', 'HE40_HE4',
                               'HE40_HE5', 'HE40_HE6', 'HE40_HE7', 'HE40_HE8', 'HE40_HE9', 'HE40_HE10',
                               'HE40_HE11', 'HE80_HE0', 'HE80_HE1', 'HE80_HE2', 'HE80_HE3', 'HE80_HE4',
                               'HE80_HE5', 'HE80_HE6', 'HE80_HE7', 'HE80_HE8', 'HE80_HE9', 'HE80_HE10',
                               'HE80_HE11', 'HE160_HE0', 'HE160_HE1', 'HE160_HE2', 'HE160_HE3', 'HE160_HE4',
                               'HE160_HE5', 'HE160_HE6', 'HE160_HE7', 'HE160_HE8', 'HE160_HE9', 'HE160_HE10',
                               'HE160_HE11']
                for row_5g in sheet_5g:
                    rows_5g.append(row_5g)
                    # logger.debug(rows)
                for rr in range(sheet_5g.max_row):
                    for cc in range(sheet_5g.max_column):
                        # logger.debug(rows[r][c].value)
                        rows_5g[rr][cc].value = str(rows_5g[rr][cc].value).strip()
                        rrs = rr + 1
                        ccs = cc + 1
                        if rows_5g[rr][cc].value == 'POWER_ACCURACY':
                            spec_pwr = abs(rows_5g[rr][ccs].value)
                        elif rows_5g[rr][cc].value == 'Power_Gain_Index':
                            gain_5 = abs(rows_5g[rr][ccs].value)
                        elif rows_5g[rr][cc].value == 'EVM_MARGIN':
                            evm_margin = abs(rows_5g[rr][ccs].value)
                        elif rows_5g[rr][cc].value == 'Symbol_Clock_Error':
                            spec_symbol_clock_error = abs(rows_5g[rr][ccs].value)
                        elif rows_5g[rr][cc].value == 'XCAP':
                            xcap = abs(rows_5g[rr][ccs].value)
                        elif rows_5g[rr][cc].value == 'LO_Leakage':
                            spec_lo_leakage = -abs(rows_5g[rr][ccs].value)
                        elif rows_5g[rr][cc].value == 'MASK':
                            spec_mask = abs(rows_5g[rr][ccs].value)
                        elif rows_5g[rr][cc].value == 'OBW_20M':
                            spec_obw_20M = rows_5g[rr][ccs].value
                        elif rows_5g[rr][cc].value == 'OBW_40M':
                            spec_obw_40M = rows_5g[rr][ccs].value
                        elif rows_5g[rr][cc].value == 'OBW_80M':
                            spec_obw_80M = rows_5g[rr][ccs].value
                        elif rows_5g[rr][cc].value == 'OBW_160M':
                            spec_obw_160M = rows_5g[rr][ccs].value
                        for xx in ratelist_5g:
                            if rows_5g[rr][cc].value == xx + '_target':
                                exec('target_pwr_%s=%d' % (xx, rows_5g[rrs][cc].value))
                                break
                        for ii in ratelist_5g:
                            if rows_5g[rr][cc].value == ii + '_EVM':
                                exec('target_evm_%s=%d' % (ii, rows_5g[rrs][cc].value))
                                break
                        for jj in ratelist_5g:
                            if rows_5g[rr][cc].value == jj + '_sens':
                                exec('target_sens_%s=%d' % (jj, rows_5g[rrs][cc].value))
                                break
                        for yy in ratelist_5g:
                            if rows_5g[rr][cc].value == yy + '_aj':
                                exec('target_aj_%s=%d' % (yy, rows_5g[rrs][cc].value))
                                break
                        for kk in ratelist_5g:
                            if rows_5g[rr][cc].value == kk + '_naj':
                                exec('target_naj_%s=%d' % (kk, rows_5g[rrs][cc].value))
                                break

            logger.info('*************************************************************')

            logger.info('Mode: ' + mode + ' Channel: ' + channel + ' BW: ' + bw + ' Rate: ' + rate + ' Chain: ' + chain)
            if iq_wanted is not None:
                iq_wanted.use_pathloss(mw_iq)
            if iq_interfere is not None:
                iq_interfere.use_pathloss(mw_iq_inter)
            if int(channel) < 5000:
                gain = gain_24
            else:
                gain = gain_5
            if item == 'TX':
                if gen_tx_report == 1:
                    # GEN REPORT
                    tx_result_name = sn + '_' + 'TX_Result' + '_' + now_time + '.csv'
                    with open('./Result/' + tx_result_name, 'w', newline='') as f:
                        writer = csv.writer(f)
                        writer.writerow(['FREQ', 'DATA_RATE', 'CHAIN', 'TX_POWER', 'POWER', 'GAIN', 'LIMIT',
                                         'RESULT', 'EVM', 'LIMIT', 'RESULT', 'FREQ_ERROR', 'LIMIT', 'RESULT',
                                         'LOLEAKAGE', 'LIMIT', 'RESULT', 'OBW', 'LIMIT', 'RESULT',
                                         'MASK', 'LIMIT', 'RESULT', 'flatness', 'LIMIT', 'RESULT',
                                         'RAMPONTIME', 'LIMIT', 'RESULT', 'RAMPOFFTIME', 'LIMIT', 'RESULT'])
                    gen_tx_report += 1
                adjust_result = result_evm = result_symbol_clock_error = result_lo_leakage = result_mask = \
                    result_flatness = 'Pass'
                # dt.set_default()
                dt.tx()
                # RESULT
                rate_t = re.sub('-', '_', rate)
                rate_t = re.sub('\.', '_', rate_t)
                target_power = eval('target_pwr_' + rate_t)
                spec_evm = eval('target_EVM_' + rate_t)
                iq_wanted.vsa(mw_iq)
                iq_wanted.analysis()
                pwr_len, txq_len, data_pwr, data_txq = iq_wanted.get_status()
                avg_power, result_evm, result_symbol_clock_error, result_lo_leakage, result_mask = \
                    iq_wanted.get_data(pwr_len, txq_len, data_pwr, data_txq, mode, channel, rate, chain, tx_result_name,
                                target_power, spec_pwr, gain, spec_evm, evm_margin, spec_symbol_clock_error,
                                spec_lo_leakage, spec_mask, spec_obw_20M, spec_obw_40M, spec_obw_80M)
                if AUTO_ADJUST_POWER == '1':
                    # if avg_power == 'NA' or result_evm == 'NA':
                    if avg_power == 'NA' or float(avg_power) > 99.000:
                        logger.info(Fore.RED + 'Error!' + Style.RESET_ALL)
                    else:
                        # accuracy_limit_left = 0.5
                        # accuracy_limit_right = 0.5
                        power_accuracy_left = float(target_power) - float(spec_pwr) + float(accuracy_limit_left)
                        power_accuracy_right = float(target_power) + float(spec_pwr) - float(accuracy_limit_right)
                        max_power = float(target_power + spec_pwr)
                        delta_power = float(avg_power) - float(target_power)
                        delta_power = float('{:.3f}'.format(delta_power))
                        logger.debug('Default Measurer Power: ' + avg_power)
                        logger.debug('Target Power: ' + target_power)
                        logger.debug('Delta Power: ' + delta_power)
                        # power is nomal but some is fail
                        if target_power <= avg_power <= max_power:
                            if result_evm == 'Fail' or result_symbol_clock_error == 'Fail' \
                                    or result_lo_leakage == 'Fail' or result_mask == 'Fail':
                                logger.info(Fore.RED + 'TX\'s quality are failed!' + Style.RESET_ALL)
                            else:
                                logger.info('Get Good Result!')
                        # power is not nomal
                        else:
                            low_power_counts = 0
                            setup = 2
                            setups = 2
                            logger.debug('Step:' + setup)
                            init_value = dt.get_paras()
                            pwr_paras = init_value + setup
                            pwr_paras = hex(pwr_paras)
                            pwr_paras = re.sub('0x', '', pwr_paras)
                            get_power = []
                            get_delta_power = []
                            c = 1
                            get_power.append(avg_power)
                            logger.debug('Measurer Power List: ' + get_power)
                            get_delta_power.append(delta_power)
                            logger.debug('Power Deviation List: ' + get_delta_power)
                            logger.debug('Adjust Counts:' + c)
                            # ADJUST
                            min_adj = avg_power - target_power
                            max_adj = avg_power - max_power
                            while min_adj < 0 or max_adj > 0:
                                logger.info(Fore.GREEN + 'Adjust...' + Style.RESET_ALL)
                                if avg_power > max_power:
                                    adj = -1
                                else:
                                    adj = 1
                                c = c + 1
                                logger.debug('NEW VALUE(HEX):' + pwr_paras)
                                gain = dt.adjust_power()
                                iq_wanted.vsa(mw_iq)
                                iq_wanted.analysis()
                                pwr_len, txq_len, data_pwr, data_txq = iq_wanted.get_status()
                                avg_power, result_evm, result_symbol_clock_error, result_lo_leakage, result_mask = \
                                    iq_wanted.get_data(pwr_len, txq_len, data_pwr, data_txq, mode, channel, rate, chain,
                                                tx_result_name, target_power, spec_pwr, gain, spec_evm, evm_margin,
                                                spec_symbol_clock_error, spec_lo_leakage, spec_mask, spec_obw_20M,
                                                spec_obw_40M, spec_obw_80M)
                                get_power.append(avg_power)
                                logger.debug('Measurer Power List: ' + get_power)
                                delta_power = float(avg_power) - target_power
                                delta_power = float('{:.3f}'.format(delta_power))
                                avg_power = float(avg_power)
                                min_adj = avg_power - target_power
                                max_adj = avg_power - max_power
                                get_delta_power.append(delta_power)
                                logger.debug('Power Deviation List: ' + get_delta_power)
                                logger.debug('Adjust Counts: ' + c)
                                adjust_status = get_delta_power[c - 1] - get_delta_power[c - 2]
                                adjust_status = float('{:.3f}'.format(adjust_status))
                                logger.debug('Power Added: ' + adjust_status)
                                if adjust_status < 0.1:
                                    adjust_result = 'Pass'
                                    setup = 2 * setups * adj
                                    low_power_counts = low_power_counts + 1
                                elif adjust_status > 0.1:
                                    adjust_status = 'Pass'
                                    setup = 2 * adj
                                # logger.debug(low_power_counts)
                                if low_power_counts > 5:
                                    logger.info(
                                        Fore.RED + 'Power added was too small, Power adjust stop' + Style.RESET_ALL)
                                    adjust_result = 'Fail'
                                else:
                                    logger.debug('Step: ' + setup)
                                    pwr_paras = int(pwr_paras, 16)
                                    pwr_paras = int(pwr_paras) + setup
                                    pwr_paras = hex(pwr_paras)
                                    pwr_paras = re.sub('0x', '', pwr_paras)
                                    logger.debug(
                                        result_evm + result_symbol_clock_error + result_lo_leakage + result_mask)
                                    logger.info('*************************************************************')
                dt.tx_off()
            elif item == 'RX':
                if gen_rx_report == 1:
                    # GEN REPORT
                    rx_result_name = sn + '_' + 'RX_Result' + '_' + now_time + '.csv'
                    with open('./Result/' + rx_result_name, 'w', newline='') as f:
                        writer = csv.writer(f)
                        writer.writerow(
                            ['FREQ', 'DATA_RATE', 'CHAIN', 'SENSITIVITY SPEC', 'SENSITIVITY', 'RX PACKETS', 'PER', 'RESULT'])
                    if RX_DYNAMIC == '1':
                        rx_dynamic_result = sn + '_' + 'RX_Dynamic' + '_' + now_time + '.csv'
                        with open('./Result/' + rx_dynamic_result, 'w', newline='') as f:
                            writer = csv.writer(f)
                            writer.writerow(['FREQ', 'DATA_RATE', 'CHAIN', 'VSG_POWER', 'VSG_Packets', 'PER', 'RESULT'])
                    gen_rx_report += 1
                start = int(line[4])
                stop = int(line[5])
                per_list = []
                sens_list = []
                logger.info('Start: ' + str(start) + ' Stop: ' + str(stop))
                # loss = iq.read_pathloss(channel, chain)
                # dt.set_default()
                dt.get_statistics()  # reset counts
                dt.rx_on()
                dt.rx()
                # logger.debug(channel)
                iq_wanted.vsg(mw_iq, start)
                per = dt.get_statistics()
                per = int(per)
                per_list.append(per)
                sens_list.append(start)
                while start > stop and per <= int(per_spec):
                    start = start - 1
                    # dt.rx(mode, channel, bw, rates, chain)
                    # dt.rx()
                    iq_wanted.vsg(mw_iq, start)
                    time.sleep(0.2)
                    per = dt.get_statistics()
                    per = int(per)
                    per_list.append(per)
                    sens_list.append(start)
                    if RX_DYNAMIC == '1':
                        if per <= per_spec:
                            per_result = 'Pass'
                        else:
                            per_result = 'Fail'
                        with open('./Result/' + rx_dynamic_result, 'a+', newline='') as f2:
                            writer2 = csv.writer(f2)
                            writer2.writerow([channel, rate, chain, start, RX_PACKETS, per, per_result])
                logger.debug(str(len(per_list)) + str(per_list))
                logger.debug(str(len(sens_list)) + str(sens_list))
                per = per_list[len(per_list) - 2]
                sens = sens_list[len(sens_list) - 2]
                logger.info('Sensitivity: ' + str(sens))
                rate_t = re.sub('-', '_', rate)
                rate_t = re.sub('\.', '_', rate_t)
                sens_spec = eval('target_sens_' + rate_t)
                if sens < sens_spec and per <= per_spec:
                    result = 'Pass'
                else:
                    result = 'Fail'
                with open('./Result/' + rx_result_name, 'a+', newline='') as f2:
                    writer2 = csv.writer(f2)
                    writer2.writerow([channel, rate, chain, sens_spec, sens, RX_PACKETS, per, result])
                logger.info('*************************************************************')
                dt.rx_off()
            elif item == 'AJ':
                if gen_aj_report == 1:
                    # GEN REPORT
                    aj_result_name = sn + '_' + 'AJ_Result' + '_' + now_time + '.csv'
                    with open('./Result/' + aj_result_name, 'w', newline='') as f:
                        writer = csv.writer(f)
                        writer.writerow(['FREQ', 'DATA_RATE', 'CHAIN', 'Adjacent Channel Rejection SPEC',
                                         'Adjacent Channel Rejection',
                                         'RESULT'])
                    gen_aj_report += 1
                if 2400 < int(channel) < 2483.5 and bw == '40':
                    logger.info('Test Channel is not right! No adjacent channel!')
                else:
                    iq_interfere.vsg_off()
                    # GET SENS
                    start = int(line[4])
                    sens_list = []
                    per_list = []
                    logger.info('Start: ' + str(start))
                    dt.get_statistics()  # reset counts
                    dt.rx_on()
                    dt.rx()
                    # logger.debug(channel)
                    iq_wanted.vsg(mw_iq, start)
                    per = dt.get_statistics()
                    per_list.append(per)
                    sens_list.append(start)
                    if float(per) > float(per_spec):
                        logger.info('Sensitivity fail!')
                    else:
                        while float(per) <= float(per_spec):
                            start = start - 1
                            # dt.rx(mode, channel, bw, rates, chain)
                            # dt.rx()
                            iq_wanted.vsg(mw_iq, start)
                            per = dt.get_statistics()
                            per_list.append(per)
                            sens_list.append(start)
                        logger.debug(str(len(per_list)) + str(per_list))
                        logger.debug(str(len(sens_list)) + str(sens_list))
                        per = per_list[len(per_list)-2]
                        sens = sens_list[len(sens_list)-2]
                        logger.info('Sensitivity: ' + str(sens))
                        # AJ TEST
                        aj_list = []
                        aj_per_list = []
                        # AJ sens
                        if mode == '11b':
                            aj_sens = sens + 6
                        else:
                            aj_sens = sens + 3
                        # GET SENS SPEC
                        rate_t = re.sub('-', '_', rate)
                        rate_t = re.sub('\.', '_', rate_t)
                        aj_spec = eval('target_aj_' + rate_t)
                        logger.debug(aj_spec)
                        aj_start = aj_sens + aj_spec
                        logger.debug(aj_start)
                        # AJ aj_vsg
                        iq_interfere.vsg_aj(mw_iq_inter, aj_start)
                        iq_wanted.vsg(mw_iq, aj_sens)
                        aj_per = dt.get_statistics()
                        aj_per_list.append(per)
                        aj_list.append(aj_start)
                        if float(aj_per) > float(per_spec):
                            aj = aj_start
                            result = 'Fail'
                            iq_interfere.vsg_off()
                        else:
                            while float(aj_per) <= float(per_spec) and aj_start < -10:
                                aj_start = aj_start + 1
                                # dt.rx(mode, channel, bw, rates, chain)
                                # dt.rx()
                                iq_interfere.vsg_aj(mw_iq_inter, aj_start)
                                iq_wanted.vsg(mw_iq, aj_sens)
                                aj_per = dt.get_statistics()
                                aj_per_list.append(aj_per)
                                aj_list.append(aj_start)
                            iq_interfere.vsg_off()
                            logger.debug(str(len(aj_per_list)) + str(aj_per_list))
                            logger.debug(str(len(aj_list)) + str(aj_list))
                            aj_per = aj_per_list[len(aj_per_list)-2]
                            aj = aj_list[len(aj_list)-2] - aj_sens
                            logger.info('Adjacent Channel Rejection: ' + str(aj))
                            # RESULT
                            if aj < aj_spec:
                                result = 'Fail'
                            else:
                                result = 'Pass'
                            with open('./Result/' + aj_result_name, 'a+', newline='') as f3:
                                writer3 = csv.writer(f3)
                                writer3.writerow([channel, rate, chain, aj_spec, aj, result])
                    logger.info('*************************************************************')
                    dt.rx_off()
            elif item == 'NAJ':
                if gen_naj_report == 1:
                    # GEN REPORT
                    naj_result_name = sn + '_' + 'NAJ_Result' + '_' + now_time + '.csv'
                    with open('./Result/' + naj_result_name, 'w', newline='') as f:
                        writer = csv.writer(f)
                        writer.writerow(['FREQ', 'DATA_RATE', 'CHAIN', 'NonAdjacent Channel Rejection SPEC',
                                         'NonAdjacent Channel Rejection',
                                         'RESULT'])
                    gen_naj_report += 1
                if 2400 < int(channel) < 2483.5 and bw == '40':
                    logger.info('Test Channel is not right! No adjacent channel!')
                elif mode == '11b':
                    logger.info('Test mode is nor right!')
                else:
                    iq_interfere.vsg_off()
                    # GET SENS
                    start = int(line[4])
                    sens_list = []
                    per_list = []
                    logger.info('Start: ' + str(start))
                    dt.get_statistics()  # reset counts
                    dt.rx_on()
                    dt.rx()
                    # logger.debug(channel)
                    iq_wanted.vsg(mw_iq, start)
                    per = dt.get_statistics()
                    per_list.append(per)
                    sens_list.append(start)
                    if float(per) > float(per_spec):
                        logger.info('Sensitivity fail!')
                    else:
                        while float(per) <= float(per_spec):
                            start = start - 1
                            # dt.rx(mode, channel, bw, rates, chain)
                            # dt.rx()
                            iq_wanted.vsg(mw_iq, start)
                            per = dt.get_statistics()
                            per_list.append(per)
                            sens_list.append(start)
                        logger.debug(str(len(per_list)) + str(per_list))
                        logger.debug(str(len(sens_list)) + str(sens_list))
                        per = per_list[len(per_list)-2]
                        sens = sens_list[len(sens_list)-2]
                        logger.info('Sensitivity: ' + str(sens))
                        # AJ TEST
                        naj_list = []
                        naj_per_list = []
                        # AJ sens
                        naj_sens = sens + 3
                        # GET SENS SPEC
                        rate_t = re.sub('-', '_', rate)
                        rate_t = re.sub('\.', '_', rate_t)
                        naj_spec = eval('target_naj_' + rate_t)
                        logger.debug(naj_spec)
                        naj_start = naj_sens + naj_spec
                        logger.debug(naj_start)
                        # AJ aj_vsg
                        iq_interfere.vsg_naj(mw_iq_inter, naj_start)
                        iq_wanted.vsg(mw_iq, naj_sens)
                        naj_per = dt.get_statistics()
                        naj_per_list.append(per)
                        naj_list.append(naj_start)
                        if float(naj_per) > float(per_spec):
                            naj = naj_start
                            result = 'Fail'
                            iq_interfere.vsg_off()
                        else:
                            while float(naj_per) <= float(per_spec) and naj_start < -10:
                                naj_start = naj_start + 1
                                # dt.rx(mode, channel, bw, rates, chain)
                                # dt.rx()
                                iq_interfere.vsg_naj(mw_iq_inter, naj_start)
                                iq_wanted.vsg(mw_iq, naj_sens)
                                naj_per = dt.get_statistics()
                                naj_per_list.append(naj_per)
                                naj_list.append(naj_start)
                            iq_interfere.vsg_off()
                            logger.debug(str(len(naj_per_list)) + str(naj_per_list))
                            logger.debug(str(len(naj_list)) + str(naj_list))
                            naj_per = naj_per_list[len(naj_per_list)-2]
                            naj = naj_list[len(naj_list)-2] - naj_sens
                            logger.info('NonAdjacent Channel Rejection: ' + str(naj))
                            # RESULT
                            if naj < naj_spec:
                                result = 'Fail'
                            else:
                                result = 'Pass'
                            with open('./Result/' + naj_result_name, 'a+', newline='') as f3:
                                writer3 = csv.writer(f3)
                                writer3.writerow([channel, rate, chain, naj_spec, naj, result])
                    logger.info('*************************************************************')
                    dt.rx_off()
    logger.info('************************TEST DONE****************************')
    dt.close()
    iq_wanted.close()
    iq_interfere.close()
