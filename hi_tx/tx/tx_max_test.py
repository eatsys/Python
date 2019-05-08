# -*- coding: utf-8 -*-
# @Time    : 2019/1/18 13:23
# @Author  : Ethan
# @FileName: tx_test.py

from iq import IQxel
from hi_tx_max import dut
import csv
import datetime
import time
import re
from openpyxl import load_workbook

try:
    Number = input("SN:")
    sn = "SN_"+Number

    #INIT IQ
    iq = IQxel('192.168.100.254')
    iq.open()
    iq.read_idn()
    iq.reset()
    iq.set_pathloss()

    #INIT DUT
    dut_file = csv.reader(open('./config.csv'))
    for rows in dut_file:
        if rows[0] == 'ip':
            ip = rows[1]
            print('DUT IP: ', ip)
        elif rows[0] == 'username':
            user = rows[1]
            print('DUT Username: ', user)
        elif rows[0] == 'password':
            pwd = rows[1]
            print('DUT Password: ', pwd)
        elif rows[0] == 'addition1':
            add1 = rows[1]
            print('Addition Parameters: ', add1)
        elif rows[0] == 'addition2':
            add2 = rows[1]
            print('Addition Parameters: ', add2)
        elif rows[0] == 'addition3':
            add3 = rows[1]
            print('Addition Parameters: ', add3)

    #INIT DUT
    dt = dut()
    dt.login(ip, user, pwd)
    dt.init(add1, add2, add3)
    #dt.ex_command('ls')

    #INIT SPEC
    spec_file = load_workbook('./spec.xlsx')
    #print(spec_file.sheetnames)
    sheet = spec_file['Sheet1']
    rows = []
    ratelist_pwr = ['1M', '2M', '5_5M', '11M','6M','9M','12M','18M','24M','36M','48M','54M','HT20_MCS0','HT20_MCS1',
                'HT20_MCS2','HT20_MCS3','HT20_MCS4','HT20_MCS5','HT20_MCS6','HT20_MCS7','HT40_MCS0','HT40_MCS1',
                'HT40_MCS2','HT40_MCS3','HT40_MCS4','HT40_MCS5','HT40_MCS6','HT40_MCS7']
    ratelist_evm = ['1M_EVM', '2M_EVM', '5_5M_EVM', '11M_EVM', '6M_EVM', '9M_EVM', '12M_EVM', '18M_EVM', '24M_EVM',
                '36M_EVM', '48M_EVM', '54M_EVM', 'HT20_MCS0_EVM', 'HT20_MCS1_EVM', 'HT20_MCS2_EVM', 'HT20_MCS3_EVM',
                'HT20_MCS4_EVM', 'HT20_MCS5_EVM', 'HT20_MCS6_EVM', 'HT20_MCS7_EVM', 'HT40_MCS0_EVM', 'HT40_MCS1_EVM',
                'HT40_MCS2_EVM', 'HT40_MCS3_EVM', 'HT40_MCS4_EVM', 'HT40_MCS5_EVM', 'HT40_MCS6_EVM', 'HT40_MCS7_EVM']
    target_pwr_1M, target_pwr_2M,  target_pwr_5_5M, target_pwr_11M, target_pwr_6M, target_pwr_9M, target_pwr_12M,\
        target_pwr_18M, target_pwr_24M, target_pwr_36M, target_pwr_48M, target_pwr_54M, target_pwr_HT20_MCS0,\
        target_pwr_HT20_MCS1, target_pwr_HT20_MCS2, target_pwr_HT20_MCS3, target_pwr_HT20_MCS4, target_pwr_HT20_MCS5,\
        target_pwr_HT20_MCS6, target_pwr_HT20_MCS7, target_pwr_HT40_MCS0, target_pwr_HT40_MCS1, target_pwr_HT40_MCS2,\
        target_pwr_HT40_MCS3, target_pwr_HT40_MCS4, target_pwr_HT40_MCS5, target_pwr_HT40_MCS6, target_pwr_HT40_MCS7\
        = [None] * 28
    target_1M_EVM, target_2M_EVM, target_5_5M_EVM, target_11M_EVM, target_6M_EVM, target_9M_EVM, target_12M_EVM,\
        target_18M_EVM, target_24M_EVM, target_36M_EVM, target_48M_EVM, target_54M_EVM, target_HT20_MCS0_EVM, \
        target_HT20_MCS1_EVM, target_HT20_MCS2_EVM, target_HT20_MCS3_EVM, target_HT20_MCS4_EVM, target_HT20_MCS5_EVM,\
        target_HT20_MCS6_EVM, target_HT20_MCS7_EVM, target_HT40_MCS0_EVM, target_HT40_MCS1_EVM, target_HT40_MCS2_EVM, \
        target_HT40_MCS3_EVM, target_HT40_MCS4_EVM, target_HT40_MCS5_EVM, target_HT40_MCS6_EVM, target_HT40_MCS7_EVM \
        = [None] * 28
    for row in sheet:
        rows.append(row)
        # print(rows)
    for r in range(sheet.max_row):
        for c in range(sheet.max_column):
            # print(rows[r][c].value)
            rows[r][c].value = str(rows[r][c].value).strip()
            rs = r + 1
            cs = c + 1
            if rows[r][c].value == 'POWER_ACCURACY':
                spec_pwr = abs(rows[r][cs].value)
            elif rows[r][c].value == 'EVM_MARGIN':
                evm_margin = abs(rows[r][cs].value)
            elif rows[r][c].value == 'Symbol_Clock_Error':
                spec_symbol_clock_error = abs(rows[r][cs].value)
            elif rows[r][c].value == 'LO_Leakage':
                spec_lo_leakage = abs(rows[r][cs].value)
            elif rows[r][c].value == 'MASK':
                spec_mask = abs(rows[r][cs].value)
            elif rows[r][c].value == 'OBW_20M':
                spec_obw_20M = rows[r][cs].value
            elif rows[r][c].value == 'OBW_40M':
                spec_obw_40M = rows[r][cs].value
            for x in ratelist_pwr:
                if rows[r][c].value == x + '_target':
                    exec('target_pwr_%s=%d' % (x, rows[rs][c].value))
                    break
            for i in ratelist_evm:
                if rows[r][c].value == i + '_target':
                    exec('target_%s=%d' % (i, rows[rs][c].value))
                    break
    targetpower = 0
    spec_evm = 0
    #GEN Report
    now_time = datetime.datetime.now()
    # print(now_time)
    now_time = str(now_time)
    now_time = now_time.split()
    # print(now_time)
    day_time = re.sub('-', '', now_time[0])
    now_time = now_time[1].split('.')
    # print(day_time)
    # print(now_time)
    now_time = re.sub(':', '', now_time[0])
    now_time = day_time + now_time
    # print(now_time)
    result_name = sn + '_' + 'TX_Result' + '_' + now_time + '.csv'
    with open('./Result/' + result_name, 'w', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(['FREQ', 'DATA_RATE', 'CHAIN', 'TX_POWER', 'POWER', 'LIMIT', 'RESULT',
                         'EVM', 'LIMIT', 'RESULT', 'FREQ_ERROR', 'LIMIT', 'RESULT',
                         'LOLEAKAGE', 'LIMIT', 'RESULT', 'OBW', 'LIMIT', 'RESULT',
                         'MASK', 'LIMIT', 'RESULT', 'FLATNESS', 'LIMIT', 'RESULT',
                         'RAMPONTIME', 'LIMIT', 'RESULT', 'RAMPOFFTIME', 'LIMIT', 'RESULT'])

    #TEST FLOW
    filename = 'TEST_FLOW.txt'
    f = open(filename)
    result = list()
    for line in f.readlines():
        if not len(line) or line.startswith('//'):
            continue
            pass
        else:
            line = line.strip()
            line = line.split()
            #print(line)
            #print('Channel:',line[1],'Rate:',line[2],'Chain:',line[3])
            channel = line[1]
            rate = line[2]
            chain = line[3]
            if rate == '1M' or rate == '2M' or rate == '5.5M' or rate == '11M':
                mode = '11b'
                bw = '20'
                rates = re.sub('M','',rate)
            elif rate == '6M' or rate == '9M' or rate == '12M' or rate == '18M' \
                    or rate == '24M' or rate == '36M' or rate == '48M' or rate == '54M':
                mode = '11g'
                bw = '20'
                rates = re.sub('M','',rate)
            elif rate == 'HT20-MCS0' or rate == 'HT20-MCS1' or rate == 'HT20-MCS2' or rate == 'HT20-MCS3' \
                    or rate == 'HT20-MCS4' or rate == 'HT20-MCS5' or rate == 'HT20-MCS6' or rate == 'HT20-MCS7':
                mode = '11ng20'
                bw = '20'
                rates = re.sub('HT20-MCS', '', rate)
            elif rate == 'HT40-MCS0' or rate == 'HT40-MCS1' or rate == 'HT40-MCS2' or rate == 'HT40-MCS3' \
                    or rate == 'HT40-MCS4' or rate == 'HT40-MCS5' or rate == 'HT40-MCS6' or rate == 'HT40-MCS7':
                mode = '11ng40plus'
                bw = '40'
                rates = re.sub('HT40-MCS', '', rate)

            if chain == 'CHAIN0':
                chain = re.sub('CHAIN', '', chain)
            elif chain == 'CHAIN1':
                chain = re.sub('CHAIN', '', chain)
            elif chain == 'CHAIN2':
                chain = re.sub('CHAIN', '', chain)
            else:
                chain = re.sub('CHAIN', '', chain)

            print('*************************************************************')
            print('Mode:',mode,'Channel:',channel,'BW:',bw,'Rate:',rate,'Chain:',chain)
            #channel = int(channel)
            #bw = str(bw)
            #rates = str(rates)
            #dt.tx(mode, channel, bw, rates, chain)
            #rate_t = re.sub('-', '_', rate)
            #rate_t = re.sub('\.', '_', rate_t)
            #targetpower = eval('target_pwr_' + rate_t)
            #rate_e = rate_t + '_EVM'
            #spec_evm = eval('target_' + rate_e)
            #channel = int(channel)
            #chain = int(chain)
            #iq.read_data(targetpower, mode, channel, chain)
            #iq.get_data(mode, rate, channel, chain, result_name, targetpower, spec_pwr, spec_evm, evm_margin,
            #            spec_symbol_clock_error, spec_lo_leakage, spec_mask, spec_obw_20M, spec_obw_40M)
            result_evm = result_symbol_clock_error = result_lo_leakage = result_mask = result_flatness = 'Pass'
            while result_evm or result_symbol_clock_error or result_lo_leakage or result_mask or result_flatness == 'Pass':
                channel = int(channel)
                bw = str(bw)
                rates = str(rates)
                dt.adjust_power(dt.tx(mode, channel, bw, rates, chain))

                rate_t = re.sub('-', '_', rate)
                rate_t = re.sub('\.', '_', rate_t)
                targetpower = eval('target_pwr_' + rate_t)
                rate_e = rate_t + '_EVM'
                spec_evm = eval('target_' + rate_e)
                channel = int(channel)
                chain = int(chain)
                iq.read_data(targetpower, mode, channel, chain)
                iq.get_data(mode, rate, channel, chain, result_name, targetpower, spec_pwr, spec_evm, evm_margin,
                            spec_symbol_clock_error, spec_lo_leakage, spec_mask, spec_obw_20M, spec_obw_40M)
                #from iq import result_evm, result_symbol_clock_error, result_lo_leakage, result_mask, result_flatness
            print('*************************************************************')

    dt.close()
    iq.close()

finally:
    time.sleep(1)