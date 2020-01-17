# -*- coding: utf-8 -*-
# @Time    : 2019/1/18 13:23
# @Author  : Ethan
# @FileName: tx_test.py
"""add 5g and RX """

from __future__ import division
from openpyxl import load_workbook
import os
import visa
import telnetlib
import pandas as pd
import time
import datetime
from colorama import init, Fore, Style
import csv
import re
import logging
logger = logging.getLogger()


def __isset__(v: object) -> object:
    try:
        type(eval(v))
    except Exception as error:
        return 0
    else:
        return 1


class IQxel():
    def __init__(self, ip, visaDLL=None, *args):
        self.ip = ip
        # self.visaDLL = 'C:\windows\system32\visa64.dll' if visaDLL is None else visaDLL
        self.address = 'TCPIP0::%s::hislip0::INSTR' % self.ip
        # self.resourceManager = visa.ResourceManager(self.visaDLL)
        self.resourceManager = visa.ResourceManager()
        self.instance = self.resourceManager.open_resource(self.address)

    def close(self):
        if self.instance is not None:
            self.resourceManager.close()
            self.instance = None

    def __formats__(self, format_spec: object) -> object:
        self.format_spec = format_spec
        format_spec = float(format_spec)
        format_spec = '{:.3f}'.format(format_spec)
        return format_spec

    def save_image(self, imagname, fmt):
        assert fmt in ['jpg', 'png'], 'Invalid postfix of image'
        logger.debug('MMEM:STOR:IMAG "%s.%s"' % (imagname, fmt))
        self.instance.write('MMEM:STOR:IMAG "%s.%s"' % (imagname, fmt))

    def reset(self):
        self.instance.write('*RST')

    def read_idn(self):
        idn = self.instance.query('*IDN?')
        iq_model = idn.split(',')[1]
        logger.debug(idn)
        return iq_model

    def init(self):
        self.instance.write('ROUT1;PORT:RES RF1,OFF')
        self.instance.write('ROUT1;PORT:RES RF2,OFF')
        logger.debug('iq init...')

    def read_pathloss(self, get_freqs, get_chains):
        self.get_freqs = get_freqs
        self.get_chains = get_chains
        logger.debug(type(get_chains))
        get_chains = int(get_chains) + 1
        pathloss = pd.read_csv('./pathloss.csv')
        logger.debug(pathloss)
        freqlist = pathloss['Frequency']
        logger.debug(freqlist)
        lens = len(freqlist)
        logger.debug(lens)
        for x in range(lens):
            logger.debug(x)
            logger.debug(pathloss.loc[x, 'Frequency'])
            if get_freqs == pathloss.loc[x, 'Frequency']:
                loss = pathloss.loc[x]  # get frequency loss list
                logger.debug(loss)
                loss = loss[get_chains]  # get chain loss value
                logger.debug('Channel:', get_freqs, 'Chain:', get_chains, 'pathloss:', loss)
                break
            else:
                pass
        return loss
        # freq = freqlist[0]
        # logger.debug (freq)
        # chaina = pathloss['1']
        # logger.info(chaina)
        # chaina_2412 = chaina[0]
        # logger.debug (chaina_2412)
        # chainb = pathloss['2']
        # logger.info(chainb)
        # chainb_2412 = chaina[0]
        # logger.info(chaina_2412)
        # logger.info(pathloss.loc[1, ['Frequency', '1','2','3','4']])
        # logger.info(pathloss.loc[1, 'Frequency'])
        # logger.info(pathloss.loc[1,'1'])

    def set_pathloss(self):
        pathloss = pd.read_csv('./pathloss.csv')
        logger.debug(pathloss)
        freqlist = pathloss['Frequency']
        logger.debug(freqlist)
        lens = len(freqlist)
        logger.debug(lens)
        # creat loss table
        self.instance.write('MEM:TABLE "1"; MEM:TABLE:DEFINE "FREQ,LOSS"')
        for x in range(lens):
            channel = pathloss.loc[x, 'Frequency']
            loss = pathloss.loc[x, '1']
            logger.debug(loss)
            self.instance.write('MEM:TABLE "1";MEMory:TABLe:INSert:POINt %f MHz,%f' % (channel, loss))
        self.instance.write('MEMory:TABLe "1";MEMory:TABLe:STORe')
        self.instance.write('MEM:TABLE "2"; MEM:TABLE:DEFINE "FREQ,LOSS"')
        for x in range(lens):
            channel = pathloss.loc[x, 'Frequency']
            loss = pathloss.loc[x, '2']
            logger.debug(loss)
            self.instance.write('MEM:TABLE "2";MEMory:TABLe:INSert:POINt %f MHz,%f' % (channel, loss))
        self.instance.write('MEMory:TABLe "2";MEMory:TABLe:STORe')
        self.instance.write('MEM:TABLE "3"; MEM:TABLE:DEFINE "FREQ,LOSS"')
        for x in range(lens):
            channel = pathloss.loc[x, 'Frequency']
            loss = pathloss.loc[x, '3']
            logger.debug(loss)
            self.instance.write('MEM:TABLE "3";MEMory:TABLe:INSert:POINt %f MHz,%f' % (channel, loss))
        self.instance.write('MEMory:TABLe "3";MEMory:TABLe:STORe')
        self.instance.write('MEM:TABLE "4"; MEM:TABLE:DEFINE "FREQ,LOSS"')
        for x in range(lens):
            channel = pathloss.loc[x, 'Frequency']
            loss = pathloss.loc[x, '4']
            logger.debug(loss)
            self.instance.write('MEM:TABLE "4";MEMory:TABLe:INSert:POINt %f MHz,%f' % (channel, loss))
        self.instance.write('MEMory:TABLe "4";MEMory:TABLe:STORe')

    def set_port(self, target_power):
        # loss_path = 1
        # idns = self.instance.query('*IDN?')
        # logger.debug(idns)
        # iq_model = idns.split(',')
        # logger.debug(iq_model)
        # iq_model = iq_model[1]
        logger.debug(iq_model)
        iq_port_model = iq_port.isdigit()
        logger.debug(iq_port_model)
        if iq_model == 'IQXEL' and iq_port_model is True:
            mw = ''
        elif iq_model == 'IQXEL-M' and iq_port_model is False:
            mw = ''
        elif iq_model == 'IQXEL-MW' and iq_port_model is False:
            mw = 'M'
        elif iq_model == 'IQXEL-M2W' and iq_port_model is False:
            mw = ''
        else:
            mw = ''
        if chain == '0':
            loss_path = '1'
        elif chain == '1':
            loss_path = '2'
        elif chain == '2':
            loss_path = '3'
        elif chain == '3':
            loss_path = '4'
        self.instance.write('%sVSA%s;RFC:USE "%s",RF%s' % (mw, iq_rout, loss_path, iq_port))
        self.instance.write('%sVSA%s;RFC:STAT  ON,RF%s' % (mw, iq_rout, iq_port))
        # rint(mw, iq_rout, iq_port, iq_rout)
        self.instance.write('%sROUT%s;PORT:RES RF%s,VSA%s' % (mw, iq_rout, iq_port, iq_rout))
        self.instance.write('VSA1;TRIG:SOUR VIDeo')
        self.instance.write('CHAN1;WIFI')
        channels = int(channel) * 1000000
        self.instance.write('%sVSA%s;FREQ:cent %d' % (mw, iq_rout, str(channels)))
        self.instance.write('%sVSA%s;SRAT 160000000' % (mw, iq_rout))
        # self.instance.write('%sVSA%s ;RLEVel:AUTO' % (mw, iq_rout))
        rlevel = target_power + 12
        logger.debug(str(rlevel))
        time.sleep(sleep_time)
        self.instance.write('%sVSA%s;RLEV %d;*wai;*opc?' % (mw, iq_rout, rlevel))

    def analysis(self):
        if mode == '11b':
            capture_time = (32 - float(rates) * 2) / 1000
        elif mode == '11g' or mode == '11a':
            capture_time = (13.2 - float(rates) * 3 / 16) / 1000
        elif mode == '11n':
            capture_time = (12 - float(rates) * 10 / 7) / 1000
        elif mode == '11ac':
            capture_time = (8 - float(rates) * 2 / 3) / 1000
        logger.debug('Captime: ' + str(capture_time))
        if mode == '11b':
            logger.info('DSSS')
            self.instance.write('CHAN1;WIFI;CONF:STAN DSSS')
            self.instance.write('VSA%s;CAPT:TIME %s' % (iq_rout, capture_time))
            self.instance.write('CHAN1')
            self.instance.write('VSA%s ;init' % iq_rout)
            self.instance.write('WIFI')
            self.instance.write('calc:pow 0, 2')
            time.sleep(1)
            self.instance.write('calc:txq 0, 2')
            time.sleep(1)
            # self.instance.write('calc:txq 0, 2')
            # time.sleep(1)
            self.instance.write('calc:ccdf 0, 2')
            self.instance.write('calc:ramp 0, 2')
            self.instance.write('calc:spec 0, 2')
            self.instance.query('WIFI;FETC:SEGM:POW:AVER?')
        else:
            logger.info('OFDM')
            self.instance.write('CHAN1;WIFI;CONF:STAN OFDM')
            self.instance.write('CHAN1;WIFI;CONF:OFDM:CEST DATA')
            self.instance.write('VSA%s;CAPT:TIME %s' % (iq_rout, capture_time))
            self.instance.write('CHAN1')
            self.instance.write('VSA%s ;init' % iq_rout)
            self.instance.write('WIFI')
            self.instance.write('calc:pow 0, 2')
            time.sleep(1)
            self.instance.write('calc:txq 0, 2')
            time.sleep(1)
            # self.instance.write('calc:txq 0, 2')
            # time.sleep(1)
            self.instance.write('calc:ccdf 0, 2')
            self.instance.write('calc:ramp 0, 2')
            self.instance.write('calc:spec 0, 2')
            self.instance.query('WIFI;FETC:SEGM:POW:AVER?')

    def get_status(self):
        time.sleep(2)
        data_pwr = self.instance.query_ascii_values('WIFI;FETC:SEGM:POW:AVER?')
        if mode == '11b':
            data_txq = self.instance.query_ascii_values('WIFI;FETC:SEGM:TXQ:DSSS:AVER?')
            txqlen = 10
        else:
            data_txq = self.instance.query_ascii_values('WIFI;FETC:SEGM:TXQ:OFDM:AVER?')
            txqlen = 8
        pwr_len = len(data_pwr)
        txq_len = len(data_txq)
        logger.debug(data_pwr)
        logger.debug(pwr_len)
        logger.debug(data_txq)
        logger.debug(txq_len)
        status = 0
        while pwr_len != 2 or txq_len != txqlen:
            self.analysis()
            data_pwr = self.instance.query_ascii_values('WIFI;FETC:SEGM:POW:AVER?')
            if mode == '11b':
                data_txq = self.instance.query_ascii_values('WIFI;FETC:SEGM:TXQ:DSSS:AVER?')
            else:
                data_txq = self.instance.query_ascii_values('WIFI;FETC:SEGM:TXQ:OFDM:AVER?')
            pwr_len = len(data_pwr)
            txq_len = len(data_txq)
            logger.debug(data_pwr)
            logger.debug(pwr_len)
            logger.debug(data_txq)
            logger.debug(txq_len)
            status += 1
            if status > 2:
                continue
        return pwr_len, txq_len, data_pwr, data_txq

    def get_data(self, pwr_len, txq_len, data_pwr, data_txq, mode, channel, rate, chain, tx_result_name, target_power,
                 spec_pwr, pwra_paras, spec_evm, evm_margin, spec_symbol_clock_error, spec_lo_leakage, spec_mask,
                 spec_obw_20M, spec_obw_40M,spec_obw_80M):
        if mode == '11b':
            if pwr_len < 2:
                logger.error('Error: ' + str(data_pwr))
                avg_power = 'NA'
                result_pwr = 'NA'
            else:
                # logger.debug(data_pwr)
                avg_power = data_pwr[1]
                logger.debug(avg_power)
                avg_power = self.__formats__(avg_power)
                delta_pwr = abs(float(avg_power) - target_power)
                logger.debug(delta_pwr)
                # spec_pwr = 2
                if delta_pwr > spec_pwr:
                    logger.info('Power:               ' + Fore.RED + avg_power + Style.RESET_ALL + 'dBm')
                    result_pwr = 'Fail'
                else:
                    logger.info('Power:               ' + Fore.BLUE + avg_power + Style.RESET_ALL + 'dBm')
                    result_pwr = 'Pass'
            # txquality
            if txq_len < 10:
                logger.error('Error:' + str(data_txq))
                avg_evm = 'NA'
                result_evm = 'NA'
                symbol_clock_error = 'NA'
                result_symbol_clock_error = 'NA'
                lo_leakage = 'NA'
                result_lo_leakage = 'NA'
            else:
                avg_evm = data_txq[1]
                avg_evm = self.__formats__(avg_evm)
                peak_evm = data_txq[2]
                peak_evm = self.__formats__(peak_evm)
                freq_error = data_txq[4]
                freq_error = self.__formats__(freq_error)
                peak_freq_error = data_txq[5]
                peak_freq_error = self.__formats__(peak_freq_error)
                symbol_clock_error = data_txq[6]
                symbol_clock_error = self.__formats__(symbol_clock_error)
                lo_leakage = data_txq[7]
                lo_leakage = self.__formats__(lo_leakage)
                # rate = rate + 'EVM'
                spec_evm = spec_evm - evm_margin
                if float(avg_evm) > spec_evm:
                    logger.info('EVM:                 ' + Fore.RED + avg_evm + Style.RESET_ALL + 'dB')
                    result_evm = 'Fail'
                else:
                    logger.info('EVM:                 ' + Fore.BLUE + avg_evm + Style.RESET_ALL + 'dB')
                    result_evm = 'Pass'
                logger.info('EVM Peak:            ' + peak_evm + 'dB')
                logger.info('Frequency Error:     ' + freq_error + 'kHz')
                logger.info('Frequency Error Peak:' + peak_freq_error + 'kHz')
                # spec_symbol_clock_error = eval('target_evm_'+rate)
                if abs(float(symbol_clock_error)) > abs(float(spec_symbol_clock_error)):
                    logger.info('Symbol Clock Error:  ' + Fore.RED + symbol_clock_error + Style.RESET_ALL + 'ppm')
                    result_symbol_clock_error = 'Fail'
                else:
                    logger.info('Symbol Clock Error:  ' + Fore.BLUE + symbol_clock_error + Style.RESET_ALL + 'ppm')
                    result_symbol_clock_error = 'Pass'
                # spec_lo_leakage = -15
                if abs(float(lo_leakage)) < abs(float(spec_lo_leakage)):
                    logger.info('LO Leakage:          ' + Fore.RED + lo_leakage + Style.RESET_ALL + 'dB')
                    result_lo_leakage = 'Fail'
                else:
                    logger.info('LO Leakage:          ' + Fore.BLUE + lo_leakage + Style.RESET_ALL + 'dB')
                    result_lo_leakage = 'Pass'
            # mask
            data_mask = self.instance.query_ascii_values('WIFI;FETC:SEGM:SPEC:AVER:VIOL?')
            time.sleep(sleep_time)
            logger.debug(data_mask)
            data_mask_len = len(data_mask)
            logger.debug(data_mask_len)
            if data_mask_len != 2:
                logger.error('Error: ' + str(data_mask))
                mask = 'NA'
                result_mask = 'NA'
            else:
                # logger.debug(data_mask)
                mask = data_mask[1]
                mask = self.__formats__(mask)
                # spec_mask = 5.12
                if abs(float(mask)) > spec_mask:
                    logger.info('Mask:                ' + Fore.RED + mask + Style.RESET_ALL + '%')
                    result_mask = 'Fail'
                else:
                    logger.info('Mask:                ' + Fore.BLUE + mask + Style.RESET_ALL + '%')
                    result_mask = 'Pass'
            # OBW
            data_obw = self.instance.query_ascii_values('WIFI;FETC:SEGM:SPEC:AVER:OBW?')
            time.sleep(sleep_time)
            logger.debug(data_obw)
            spec_obw = spec_obw_20M
            data_obw_len = len(data_obw)
            logger.debug(data_obw_len)
            if data_obw_len != 4:
                logger.error('Error: ' + str(data_obw))
                obw = 'NA'
                result_obw = 'NA'
            else:
                # logger.debug(data_obw)
                obw = data_obw[1]
                obw = self.__formats__(obw)
                obw = round(float(obw) / 1000000, 3)
                if obw > spec_obw:
                    logger.info('OBW:                 ' + Fore.RED + str(obw) + Style.RESET_ALL + 'MHz')
                    result_obw = 'Fail'
                else:
                    logger.info('OBW:                 ' + Fore.BLUE + str(obw) + Style.RESET_ALL + 'MHz')
                    result_obw = 'Pass'
            # RAMP
            # on time
            data_ontime = self.instance.query_ascii_values('WIFI;FETC:SEGM:RAMP:ON:TRIS?')
            time.sleep(sleep_time)
            logger.debug(data_ontime)
            spec_ramp_on_time = 2.0
            data_ontime_len = len(data_ontime)
            logger.debug(data_ontime_len)
            if data_ontime_len < 4:
                logger.error('Error: ' + str(data_ontime))
                ramp_on_time = 'NA'
                result_ramp_on_time = 'NA'
            else:
                # logger.debug(data_ontime)
                ramp_on_time = data_ontime[1]
                ramp_on_time = float(ramp_on_time) * 1000000
                ramp_on_time = self.__formats__(ramp_on_time)
                if float(ramp_on_time) > spec_ramp_on_time:
                    logger.info('Ramp On Time:        ' + Fore.RED + ramp_on_time + Style.RESET_ALL + 'us')
                    result_ramp_on_time = 'Fail'
                else:
                    logger.info('Ramp On Time:        ' + Fore.BLUE + ramp_on_time + Style.RESET_ALL + 'us')
                    result_ramp_on_time = 'Pass'
            # off time
            data_offtime = self.instance.query_ascii_values('WIFI;FETC:SEGM:RAMP:OFF:TRIS?')
            time.sleep(sleep_time)
            logger.debug(data_offtime)
            spec_ramp_off_time = 2.0
            data_offtime_len = len(data_offtime)
            logger.debug(data_offtime_len)
            if data_offtime_len < 4:
                logger.error('Error: ' + str(data_offtime))
                ramp_off_time = 'NA'
                result_ramp_off_time = 'NA'
            else:
                # logger.debug(data_offtime)
                ramp_off_time = data_offtime[1]
                ramp_off_time = float(ramp_off_time) * 1000000
                ramp_off_time = self.__formats__(ramp_off_time)
                if float(ramp_off_time) > spec_ramp_off_time:
                    logger.info('Ramp Off Time:       ' + Fore.RED + ramp_off_time + Style.RESET_ALL + 'us')
                    result_ramp_off_time = 'Fail'
                else:
                    logger.info('Ramp Off Time:       ' + Fore.BLUE + ramp_off_time + Style.RESET_ALL + 'us')
                    result_ramp_off_time = 'Pass'
            flatness = 'NA'
            spec_flatness = 'NA'
            result_flatness = 'NA'

        else:
            logger.debug('OFDM')
            if pwr_len < 2:
                logger.error('Error: ' + str(data_pwr))
                avg_power = 'NA'
                result_pwr = 'NA'
            else:
                # logger.debug(data)
                avg_power = data_pwr[1]
                avg_power = self.__formats__(avg_power)
                delta_pwr = abs(float(avg_power) - target_power)
                logger.debug(delta_pwr)
                ##spec_pwr = 2
                if delta_pwr > spec_pwr:
                    logger.info('Power:               ' + Fore.RED + avg_power + Style.RESET_ALL + 'dBm')
                    result_pwr = 'Fail'
                else:
                    logger.info('Power:               ' + Fore.BLUE + avg_power + Style.RESET_ALL + 'dBm')
                    result_pwr = 'Pass'
            # txquality
            if txq_len < 8:
                logger.error('Error: ' + str(data_txq))
                avg_evm = 'NA'
                result_evm = 'NA'
                symbol_clock_error = 'NA'
                result_symbol_clock_error = 'NA'
                lo_leakage = 'NA'
                result_lo_leakage = 'NA'
            else:
                logger.debug(data_txq)
                avg_evm = data_txq[1]
                avg_evm = self.__formats__(avg_evm)
                freq_error = data_txq[3]
                freq_error = self.__formats__(freq_error)
                symbol_clock_error = data_txq[4]
                symbol_clock_error = self.__formats__(symbol_clock_error)
                lo_leakage = data_txq[5]
                lo_leakage = self.__formats__(lo_leakage)
                spec_evm = spec_evm - abs(evm_margin)
                # logger.info(spec_evm)
                if float(avg_evm) > spec_evm:
                    logger.info('EVM:                 ' + Fore.RED + avg_evm + Style.RESET_ALL + 'dB')
                    result_evm = 'Fail'
                else:
                    logger.info('EVM:                 ' + Fore.BLUE + avg_evm + Style.RESET_ALL + 'dB')
                    result_evm = 'Pass'
                logger.info('Frequency Error:     ' + freq_error + 'kHz')
                # spec_symbol_clock_error = 10.0
                if abs(float(symbol_clock_error)) > abs(float(spec_symbol_clock_error)):
                    logger.info('Symbol Clock Error:  ' + Fore.RED + symbol_clock_error + Style.RESET_ALL + 'ppm')
                    result_symbol_clock_error = 'Fail'
                else:
                    logger.info('Symbol Clock Error:  ' + Fore.BLUE + symbol_clock_error + Style.RESET_ALL + 'ppm')
                    result_symbol_clock_error = 'Pass'
                # spec_lo_leakage = -15
                if abs(float(lo_leakage)) < abs(float(spec_lo_leakage)):
                    logger.info('LO Leakage:          ' + Fore.RED + lo_leakage + Style.RESET_ALL + 'dB')
                    result_lo_leakage = 'Fail'
                else:
                    logger.info('LO Leakage:          ' + Fore.BLUE + lo_leakage + Style.RESET_ALL + 'dB')
                    result_lo_leakage = 'Pass'
            # mask
            data_mask = self.instance.query_ascii_values('WIFI;FETC:SEGM:SPEC:AVER:VIOL?')
            time.sleep(sleep_time)
            logger.debug(data_mask)
            data_mask_len = len(data_mask)
            logger.debug(data_mask_len)
            if data_mask_len != 2:
                logger.error('Error: ' + str(data_mask))
                mask = 'NA'
                result_mask = 'NA'
            else:
                # logger.debug(data_mask)
                mask = data_mask[1]
                mask = self.__formats__(mask)
                # spec_mask = 5.12
                if abs(float(mask)) > spec_mask:
                    logger.info('Mask:                ' + Fore.RED + mask + Style.RESET_ALL + '%')
                    result_mask = 'Fail'
                else:
                    logger.info('Mask:                ' + Fore.BLUE + mask + Style.RESET_ALL + '%')
                    result_mask = 'Pass'
            # OBW
            data_obw = self.instance.query_ascii_values('WIFI;FETC:SEGM:SPEC:AVER:OBW?')
            time.sleep(sleep_time)
            logger.debug(data_obw)
            if bw == '80':
                spec_obw = spec_obw_80M
            elif bw == '40':
                spec_obw = spec_obw_40M
            else:
                spec_obw = spec_obw_20M
            data_obw_len = len(data_obw)
            logger.debug(data_obw_len)
            if data_obw_len != 4:
                logger.error('Error: ' + str(data_obw))
                obw = 'NA'
                result_obw = 'NA'
            else:
                # logger.debug(data_obw)
                obw = data_obw[1]
                obw = self.__formats__(obw)
                obw = round(float(obw) / 1000000, 3)
                if obw > spec_obw:
                    logger.info('OBW:                 ' + Fore.RED + str(obw) + Style.RESET_ALL + 'MHz')
                    result_obw = 'Fail'
                else:
                    logger.info('OBW:                 ' + Fore.BLUE + str(obw) + Style.RESET_ALL + 'MHz')
                    result_obw = 'Pass'
            # flatness
            # datas = self.instance.query('FETC:SEGM1:OFDM:SFL:SIGN1:AVER?')
            # time.sleep(sleep_time)
            # time.sleep(10)
            # logger.info('flatness', datas)
            data_flatness = self.instance.query('WIFI;FETC:SEGM:OFDM:SFL:AVER:CHEC?')
            # data_flatness = self.query('WIFI;FETC:SEGM:OFDM:SFL:AVER?')
            time.sleep(sleep_time)
            logger.debug(data_flatness)
            data_flatness = data_flatness.split(',')
            spec_flatness = 0
            data_flatness_len = len(data_flatness)
            logger.debug(data_flatness_len)
            if data_flatness_len < 2:
                logger.error('Error: ' + str(data_flatness))
                flatness = 'NA'
                result_flatness = 'NA'
            else:
                # logger.debug(data_flatness)
                flatness = data_flatness[0]
                if int(flatness) == spec_flatness:
                    logger.info('Flatness:            ' + Fore.BLUE + flatness + Style.RESET_ALL)
                    result_flatness = 'Pass'
                else:
                    logger.info('Flatness:            ' + Fore.RED + flatness + Style.RESET_ALL)
                    result_flatness = 'Fail'
            ramp_on_time = 'NA'
            spec_ramp_on_time = 'NA'
            result_ramp_on_time = 'NA'
            ramp_off_time = 'NA'
            spec_ramp_off_time = 'NA'
            result_ramp_off_time = 'NA'

        if super_mode == '1' and avg_power != 'NA':
            if float(avg_power) < float(power_accuracy_left) or float(avg_power) > float(power_accuracy_right):
                logger.info('Adjust power')
            else:
                with open('./Result/' + tx_result_name, 'a+', newline='') as write_result:
                    writer_file = csv.writer(write_result)
                    writer_file.writerow(
                        [channel, rate, chain, target_power, avg_power, pwra_paras, spec_pwr, result_pwr,
                         avg_evm, spec_evm, result_evm, symbol_clock_error, spec_symbol_clock_error,
                         result_symbol_clock_error, lo_leakage, spec_lo_leakage, result_lo_leakage, obw,
                         spec_obw, result_obw, mask, spec_mask, result_mask, flatness, spec_flatness,
                         result_flatness, ramp_on_time, spec_ramp_on_time, result_ramp_on_time,
                         ramp_off_time, spec_ramp_off_time, result_ramp_off_time])
        else:
            with open('./Result/' + tx_result_name, 'a+', newline='') as write_result:
                writer_file = csv.writer(write_result)
                writer_file.writerow([channel, rate, chain, target_power, avg_power, pwra_paras, spec_pwr, result_pwr,
                                      avg_evm, spec_evm, result_evm, symbol_clock_error, spec_symbol_clock_error,
                                      result_symbol_clock_error, lo_leakage, spec_lo_leakage, result_lo_leakage, obw,
                                      spec_obw, result_obw, mask, spec_mask, result_mask, flatness, spec_flatness,
                                      result_flatness, ramp_on_time, spec_ramp_on_time, result_ramp_on_time,
                                      ramp_off_time, spec_ramp_off_time, result_ramp_off_time])

        return avg_power, result_evm, result_symbol_clock_error, result_lo_leakage, result_mask

    def vsg(self):
        loss_path = 1
        idns = self.instance.query('*IDN?')
        logger.debug(idns)
        iq_port_model = iq_port.isdigit()
        logger.debug(iq_port_model)
        iq_model = idns.split(',')
        logger.debug(iq_model)
        iq_model = iq_model[1]
        logger.debug(iq_model)
        if iq_model == 'IQXEL' and iq_port_model is True:
            mw = ''
        elif iq_model == 'IQXEL-M' and iq_port_model is False:
            mw = ''
        elif iq_model == 'IQXEL-MW' and iq_port_model is False:
            mw = 'M'
        elif iq_model == 'IQXEL-M2W' and iq_port_model is False:
            mw = ''
        else:
            mw = ''
        if chain == '0':
            loss_path = '1'
        elif chain == '1':
            loss_path = '2'
        elif chain == '2':
            loss_path = '3'
        elif chain == '3':
            loss_path = '4'
        self.instance.write('%sVSG%s;RFC:USE "%s",RF%s' % (mw, iq_rout, loss_path, iq_port))
        self.instance.write('%sVSG%s;RFC:STAT  ON,RF%s' % (mw, iq_rout, iq_port))
        # rint(mw, iq_rout, iq_port, iq_rout)
        self.instance.write('%sROUT%s;PORT:RES RF%s,VSG%s' % (mw, iq_rout, iq_port, iq_rout))
        self.instance.write('CHAN1;WIFI')
        channels = channel * 1000000
        self.instance.write('%sVSG%s;FREQ:cent %d' % (mw, iq_rout, channels))
        self.instance.write('%sVSG%s;SRAT 160000000' % (mw, iq_rout))
        # loss = abs(loss)
        # logger.info(type(start), type(loss))
        # rlevel = start + loss
        rlevel = start
        logger.info(rlevel)
        time.sleep(sleep_time)
        self.instance.write('%sVSG%s;POW:lev %d' % (mw, iq_rout, rlevel))
        self.instance.write('VSG1;POW:STAT ON')
        self.instance.write('VSG1;MOD:STAT ON')
        self.instance.write('VSG1;WAVE:EXEC OFF')
        self.instance.write('VSG1;WLIS:COUN %d' % int(rx_packets))
        # wave = 'OFDM-6'
        if mode == '11b' and rates == '1':
            wave = 'DSSS-1L'
            vsg_delay = sleep_time + int(rx_packets) / 100 + 5 - int(rates) * int(rx_packets) / 1000
        elif mode == '11b' and rates == '2':
            wave = 'DSSS-2L'
            vsg_delay = sleep_time + int(rx_packets) / 100 - int(rates) * int(rx_packets) / 1000
        elif mode == '11b' and rates == '5.5':
            wave = 'CCK-5_5S'
            vsg_delay = sleep_time + int(rx_packets) / 100 - int(float(rates) * int(rx_packets) / 1000)
        elif mode == '11b' and rates == '11':
            wave = 'CCK-11S'
            vsg_delay = sleep_time + int(rx_packets) / 100 - int(rates) * int(rx_packets) / 1800
        elif mode == '11g' or mode == '11a':
            wave = 'OFDM-' + rates
            vsg_delay = sleep_time + int(rx_packets) / 100 - int(rates) * int(rx_packets) / 6000
        elif mode == '11n':
            wave = 'HT' + bw + '_MCS' + rates
            vsg_delay = sleep_time + int(rx_packets) / 100 - int(rates) * int(rx_packets) / 1000
        elif mode == '11ac':
            wave = '11AC_VHT' + bw + '_S1_MCS' + rates
            vsg_delay = sleep_time + int(rx_packets) / 100 - int(rates) * int(rx_packets) / 1000
        # logger.info(wave)
        logger.info('Delay Time:' + str(vsg_delay))
        self.instance.write('VSG1; WAVE:LOAD "/user/WiFi_%s.iqvsg"' % wave)
        self.instance.write('VSG1 ;wave:exec off')
        self.instance.write('WLIST:WSEG1:DATA "/user/WiFi_%s.iqvsg"' % wave)
        self.instance.write('wlist:wseg1:save')
        self.instance.write('WLIST:COUNT:ENABLE WSEG1')
        self.instance.write('WAVE:EXEC ON, WSEG1')
        time.sleep(vsg_delay)
        self.instance.write('VSG1 ;WAVE:EXEC OFF')
        self.instance.write('WLIST:COUNT:DISABLE WSEG1')


class dut():
    def __init__(self):
        self.tn = telnetlib.Telnet()
        self.tn.set_debuglevel(log_switch)

    def close(self):
        if self.tn is not None:
            self.tn.close()
            self.tn = None

    def login(self, host, username, password):
        self.tn.open(host, port=23)
        time.sleep(1)
        command_result = self.tn.read_very_eager().decode('ascii')
        logger.debug(command_result)
        self.tn.read_until(b'Login: ', timeout=1)
        self.tn.write(username.encode('ascii') + b'\r\n')
        self.tn.read_until(b'Password:', timeout=1)
        self.tn.write(password.encode('ascii') + b'\r\n')

        time.sleep(1)
        command_result = self.tn.read_very_eager().decode('ascii')
        logger.debug(command_result)
        if 'wrong' not in command_result:
            logging.info('%s Sign up' % host)
            return True
        else:
            logging.warning('%s Login Fail' % host)
            return False

    def init(self, str1, str2, str3):
        self.tn.read_until(b'WAP>', timeout=1)
        self.tn.write(str1.encode('ascii') + b'\r\n')

        self.tn.read_until(b'SU_WAP>', timeout=1)
        self.tn.write(str2.encode('ascii') + b'\r\n')

        self.tn.read_until(b'WAP(Dopra Linux) # ', timeout=1)
        self.tn.write(str3.encode('ascii') + b'\r\n')

        self.tn.read_until(b'WAP(Dopra Linux) # ', timeout=1)
        self.tn.write(b'Config_Set_WifiChip_[AdjustPpm] chip <chip_index> ppm <ppm>' + b'\r\n')

    def ex_command(self, command):
        self.tn.read_until(b'WAP(Dopra Linux) # ', timeout=1)
        self.tn.write(command.encode('ascii') + b'\r\n')
        time.sleep(2)
        command_result = self.tn.read_very_eager().decode('ascii')
        logging.info('\n%s' % command_result)

    def tx(self):
        logger.debug(channel)
        if bw == '40':
            channels = int(channel) - 10
            logger.debug(channel)
        else:
            channels = channel
        if int(channel) < 5000:
            channels = '{:.0f}'.format((int(channels) - 2407) / 5)  # 2407+5*1=2412
        else:
            channels = '{:.0f}'.format((int(channels) - 5000) / 5)  # 2407+5*1=2412
        if chain == '0':
            chains = '01'
        else:
            chains = '10'
        # tx
        self.tn.write(b'\r\n')
        self.tn.read_until(b'WAP(Dopra Linux) # ', timeout=1)
        self.tn.write(b'Config_Set_WifiChip_[TxParamSet] chip %s mode %s bw %s freq %s rate %s txch %s' %
                      (band.encode('ascii'), mode.encode('ascii'), bw.encode('ascii'), str(channels).encode('ascii'),
                       rate.encode('ascii'), chains.encode('ascii')) + b'\r\n')
        self.tn.read_until(b'WAP(Dopra Linux) # ', timeout=1)
        self.tn.write(b'Config_Set_WifiChip_[Switch] chip %s direction <direction> txrxSwitch <txrxSwitch>' % band.encode('ascii') + b'\r\n')
        logger.info('TX COMMANDS DONE')

    def get_paras(self):
        # self.tn.read_until(b'WAP(Dopra Linux) # ', timeout=1)
        # self.tn.write(b'hipriv.sh "vap%s set_tx_pow rf_reg_ctl 0"' % band.encode('ascii') + b'\r\n')
        self.tn.read_until(b'WAP(Dopra Linux) # ', timeout=1)
        self.tn.write(b'hipriv.sh "vap%s set_tx_pow rf_reg_ctl 1"' % band.encode('ascii') + b'\r\n')
        self.tn.read_until(b'WAP(Dopra Linux) # ', timeout=1)
        self.tn.write(b'hipriv.sh "vap%s reginfo soc 0x200380%s 0x200380%s";dmesg -c'
                      % (band.encode('ascii'), channel_groups.encode('ascii'), channel_groups.encode('ascii')) + b'\r\n')
        # self.tn.read_until(b'WAP(Dopra Linux) # ', timeout=1)
        # command_result = self.tn.read_very_eager().decode('ascii')
        command_result = self.tn.read_until(b'value=0x\w+\r\r\nWAP(Dopra Linux) # ', timeout=2)
        command_result = re.search(b'value=0x\w+', command_result)
        default_value = re.sub('value=0x', '', command_result.group().decode('ascii'))
        logger.info('Default Value(HEX):', default_value)
        default_value = int(default_value, 16)
        return default_value
        # logger.debug(default_value)
        # pwr_para = default_value + 2
        # pwr_paras = h#ex(pwr_para)
        ##logger.debug(pwr_paras)
        # logger.debug(channel_groups, pwr_paras)

    def adjust_power(self):
        self.tn.read_until(b'WAP(Dopra Linux) # ', timeout=1)
        self.tn.write(b'Config_Test_WifiChip_[TxPower] chip %s txpower %s"'
                      % (band.encode('ascii'), pwr_paras.encode('ascii'))+ b'\r\n')
        command_result = self.tn.read_until(b'value=0x\w+\r\r\nWAP(Dopra Linux) # ', timeout=1)
        command_result = re.search(b'value=0x\w+', command_result)
        default_value = re.sub('value=0x', '', command_result.group().decode('ascii'))
        logger.debug('Check Value(HEX):', default_value)
        gain = default_value
        return gain

    def set_default(self):
        self.tn.read_until(b'WAP(Dopra Linux) # ', timeout=1)
        self.tn.write(b'Config_Set_WifiChip_[Init]' % (band.encode('ascii'), band.encode('ascii')) + b'\r\n')

    def rx(self):
        if int(channel) < 5000:
            channels = '{:.0f}'.format((int(channel) - 2407) / 5)
        else:
            channels = '{:.0f}'.format((int(channel) - 5000) / 5)
        # logger.debug(channel)
        if chain == '0':
            chains = '01'
        else:
            chains = '10'
        # logger.debug('chain set',chain,type(chain),chains)
        if mode == '11b' or mode == '11g' or mode == '11a':
            rates = 'rate'
        else:
            rates = 'mcs'
        self.tn.read_until(b'WAP(Dopra Linux) # ', timeout=1)
        self.tn.write(b'Config_Test_WifiChip_[RxParamSet] chip %s mode %s bw %s freq %s rxch %s' %
                      (band.encode('ascii'), mode.encode('ascii'), bw.encode('ascii'), str(channels).encode('ascii'),
                       chains.encode('ascii')) + b'\r\n')
        logger.info('RX COMMANDS DONE')

    def get_statistics(self):
        self.tn.read_until(b'WAP(Dopra Linux) # ', timeout=1)
        self.tn.write(b'Query_Test_WifiChip_[GetRxInfo] chip %s' % band.encode('ascii') + b'\r\n')
        command_result = self.tn.read_until(b'rssi', timeout=2)
        logger.debug(command_result)
        command_result = re.search(b'succ:\w+', command_result)
        logger.debug(command_result)
        logger.debug(command_result.group())
        rx_good = re.sub('succ:', '', command_result.group().decode('ascii'))
        PER_value = int(rx_packets) - int(rx_good)
        logger.info('Packets:', rx_packets, 'PER:', PER_value)
        return PER_value


if __name__ == '__main__':
    init(autoreset=True)

    Number = input("SN:")
    sn = "SN_" + Number

    directory_r = os.path.exists(r'./Result')
    if directory_r is False:
        os.makedirs('Result')

    directory_l = os.path.exists(r'./log')
    if directory_l is False:
        os.makedirs('log')

    dut_file = csv.reader(open('./config.csv'))
    for rows in dut_file:
        if rows[0] == 'log_enable':
            le = rows[1]
            log_switch = int(le)

            logger.setLevel(logging.DEBUG)  # logger的总开关，只有大于Debug的日志才能被logger对象处理

            # 第二步，创建一个handler，用于写入日志文件
            file_handler = logging.FileHandler('./log/log.txt', mode='w')
            file_handler.setLevel(logging.DEBUG)  # 输出到file的log等级的开关
            # 创建该handler的formatter
            file_handler.setFormatter(
                logging.Formatter(
                    fmt='%(asctime)s - %(filename)s[line:%(lineno)d] - %(levelname)s: %(message)s',
                    datefmt='%Y-%m-%d %H:%M:%S')
            )
            # 添加handler到logger中
            logger.addHandler(file_handler)

            # 第三步，创建一个handler，用于输出到控制台
            console_handler = logging.StreamHandler()
            if log_switch == 1:
                console_handler.setLevel(logging.DEBUG)  # 输出到控制台的log等级的开关
            else:
                console_handler.setLevel(logging.INFO)  # 输出到控制台的log等级的开关
            # 创建该handler的formatter
            console_handler.setFormatter(
                logging.Formatter(
                    fmt='%(asctime)s - %(levelname)s: %(message)s',
                    datefmt='%Y-%m-%d %H:%M:%S')
            )
            logger.addHandler(console_handler)
            logger.info('logger.debug DEBUG Log: ' + le)

        elif rows[0] == 'DUT_IP':
            dut_ip = rows[1]
            logger.info('DUT IP: ', dut_ip)
        elif rows[0] == 'username':
            user = rows[1]
            logger.info('DUT Username: ', user)
        elif rows[0] == 'password':
            pwd = rows[1]
            logger.info('DUT Password: ', pwd)
        elif rows[0] == 'addition1':
            add1 = rows[1]
            logger.info('Addition Parameters: ', add1)
        elif rows[0] == 'addition2':
            add2 = rows[1]
            logger.info('Addition Parameters: ', add2)
        elif rows[0] == 'addition3':
            add3 = rows[1]
            logger.info('Addition Parameters: ', add3)
        elif rows[0] == 'addition4':
            add4 = rows[1]
            logger.info('Addition Parameters: ', add4)
        elif rows[0] == 'log_enable':
            le = rows[1]
            logger.info('logger.debug Log: ', le)
        elif rows[0] == 'sleep_time':
            st = rows[1]
            logger.info('Sleep time: ', st)
        elif rows[0] == 'IQ_IP':
            iq_ip = rows[1]
            logger.info('IQ time: ', iq_ip)
        elif rows[0] == 'IQ_Port':
            iq_port = rows[1]
            logger.info('IQ PORT: ', iq_port)
        elif rows[0] == 'IQ_Rout':
            iq_rout = rows[1]
            logger.info('IQ ROUT: ', iq_rout)
        elif rows[0] == 'RX_Packets':
            rx_packets = rows[1]
            logger.info('RX Packets: ', rx_packets)
        elif rows[0] == 'RX_Dynamic':
            rx_dynamic = rows[1]
            logger.info('RX Dynamic: ' + rx_dynamic)
        elif rows[0] == 'Super_Mode':
            super_mode = rows[1]
            logger.info('Super Mode: ', super_mode)
        elif rows[0] == '2G_ID':
            id_2g = rows[1]
            logger.info('2G ID: ', id_2g)
        elif rows[0] == '5G_Low_ID':
            id_5g_l = rows[1]
            logger.info('5G Low ID: ', id_5g_l)
        elif rows[0] == '5G_High_ID':
            id_5g_h = rows[1]
            logger.info('5G High ID: ', id_5g_h)
        elif rows[0] == '11a/g_Channel_Estimation':
            ag_Channel_Estimation = rows[1]
            logger.info('11a/g_Channel_Estimation: ', ag_Channel_Estimation)
        elif rows[0] == '11n_Channel_Estimation':
            n_Channel_Estimation = rows[1]
            logger.info('11n_Channel_Estimation: ', n_Channel_Estimation)
        elif rows[0] == '11ac_Channel_Estimation':
            ac_Channel_Estimation = rows[1]
            logger.info('11ac_Channel_Estimation: ', ac_Channel_Estimation)
        elif rows[0] == '11ax_Channel_Estimation':
            ax_Channel_Estimation = rows[1]
            logger.info('11ax_Channel_Estimation: ', ax_Channel_Estimation)
        elif rows[0] == 'accuracy_limit_left':
            accuracy_limit_left = rows[1]
            logger.info('accuracy_limit_left: ', accuracy_limit_left)
        elif rows[0] == 'accuracy_limit_right':
            accuracy_limit_right = rows[1]
            logger.info('accuracy_limit_right: ', accuracy_limit_right)

    log_switch = int(le)
    sleep_time = float(st)
    # INIT IQ
    iq = IQxel(iq_ip)
    iq_model = iq.read_idn
    iq.reset()
    iq.set_pathloss()

    # INIT DUT
    dt = dut()
    dt.login(dut_ip, user, pwd)
    dt.init(add1, add2, add3, add4)
    # dt.ex_command('reboot')
    # logger.debug('Reboot DUT, Please wait...')
    # time.sleep(60#)
    # dt = dut()
    # dt.login(dut_ip, user, pwd)
    # dt.init(add1, add2, add3)

    # define

    target_pwr_1M, target_pwr_2M, target_pwr_5_5M, target_pwr_11M, target_pwr_6M, target_pwr_9M, \
    target_pwr_12M, target_pwr_18M, target_pwr_24M, target_pwr_36M, target_pwr_48M, target_pwr_54M, \
    target_pwr_HT20_MCS0, target_pwr_HT20_MCS1, target_pwr_HT20_MCS2, \
    target_pwr_HT20_MCS3, target_pwr_HT20_MCS4, target_pwr_HT20_MCS5, target_pwr_HT20_MCS6, \
    target_pwr_HT20_MCS7, target_pwr_HT40_MCS0, target_pwr_HT40_MCS1, target_pwr_HT40_MCS2, \
    target_pwr_HT40_MCS3, target_pwr_HT40_MCS4, target_pwr_HT40_MCS5, target_pwr_HT40_MCS6, \
    target_pwr_HT40_MCS7, target_pwr_VHT20_MCS0, target_pwr_VHT20_MCS1, target_pwr_VHT20_MCS2, \
    target_pwr_VHT20_MCS3, target_pwr_VHT20_MCS4, target_pwr_VHT20_MCS5, target_pwr_VHT20_MCS6, \
    target_pwr_VHT20_MCS7, target_pwr_VHT20_MCS8, target_pwr_VHT40_MCS0, target_pwr_VHT40_MCS1, \
    target_pwr_VHT40_MCS2, target_pwr_VHT40_MCS3, target_pwr_VHT40_MCS4, target_pwr_VHT40_MCS5, \
    target_pwr_VHT40_MCS6, target_pwr_VHT40_MCS7, target_pwr_VHT40_MCS8, target_pwr_VHT40_MCS9, \
    target_pwr_VHT80_MCS0, target_pwr_VHT80_MCS1, target_pwr_VHT80_MCS2, target_pwr_VHT80_MCS3, \
    target_pwr_VHT80_MCS4, target_pwr_VHT80_MCS5, target_pwr_VHT80_MCS6, target_pwr_VHT80_MCS7, \
    target_pwr_VHT80_MCS8, target_pwr_VHT80_MCS9, \
    target_pwr_VHT160_MCS0, target_pwr_VHT160_MCS1, target_pwr_VHT160_MCS2, target_pwr_VHT160_MCS3, \
    target_pwr_VHT160_MCS4, target_pwr_VHT160_MCS5, target_pwr_VHT160_MCS6, target_pwr_VHT160_MCS7, \
    target_pwr_VHT160_MCS8, target_pwr_VHT160_MCS9, target_pwr_HE20_HE0, target_pwr_HE20_HE1, target_pwr_HE20_HE2, \
    target_pwr_HE20_HE3, target_pwr_HE20_HE4, target_pwr_HE20_HE5, \
    target_pwr_HE20_HE6, target_pwr_HE20_HE7, target_pwr_HE20_HE8, target_pwr_HE20_HE9, target_pwr_HE20_HE10, \
    target_pwr_HE20_HE11, target_pwr_HE40_HE0, target_pwr_HE40_HE1, target_pwr_HE40_HE2, target_pwr_HE40_HE3, \
    target_pwr_HE40_HE4, target_pwr_HE40_HE5, target_pwr_HE40_HE6, target_pwr_HE40_HE7, target_pwr_HE40_HE8, \
    target_pwr_HE40_HE9, target_pwr_HE40_HE10, target_pwr_HE40_HE11, target_pwr_HE80_HE0, target_pwr_HE80_HE1, \
    target_pwr_HE80_HE2, target_pwr_HE80_HE3, target_pwr_HE80_HE4, target_pwr_HE80_HE5, target_pwr_HE80_HE6, \
    target_pwr_HE80_HE7, target_pwr_HE80_HE8, target_pwr_HE80_HE9, target_pwr_HE80_HE10, target_pwr_HE80_HE11, \
    target_pwr_HE160_HE0, target_pwr_HE160_HE1, target_pwr_HE160_HE2, target_pwr_HE160_HE3, target_pwr_HE160_HE4, \
    target_pwr_HE160_HE5, target_pwr_HE160_HE6, target_pwr_HE160_HE7, target_pwr_HE160_HE8, target_pwr_HE160_HE9, \
    target_pwr_HE160_HE10, target_pwr_HE160_HE11 = [None] * 115

    target_1M_EVM, target_2M_EVM, target_5_5M_EVM, target_11M_EVM, target_6M_EVM, target_9M_EVM, target_12M_EVM, \
    target_18M_EVM, target_24M_EVM, target_36M_EVM, target_48M_EVM, target_54M_EVM, target_HT20_MCS0_EVM, \
    target_HT20_MCS1_EVM, target_HT20_MCS2_EVM, target_HT20_MCS3_EVM, target_HT20_MCS4_EVM, target_HT20_MCS5_EVM, \
    target_HT20_MCS6_EVM, target_HT20_MCS7_EVM, target_HT40_MCS0_EVM, target_HT40_MCS1_EVM, target_HT40_MCS2_EVM, \
    target_HT40_MCS3_EVM, target_HT40_MCS4_EVM, target_HT40_MCS5_EVM, target_HT40_MCS6_EVM, target_HT40_MCS7_EVM, \
    target_VHT20_MCS0_EVM, target_VHT20_MCS1_EVM, target_VHT20_MCS2_EVM, target_VHT20_MCS3_EVM, target_VHT20_MCS4_EVM, \
    target_VHT20_MCS5_EVM, target_VHT20_MCS6_EVM, target_VHT20_MCS7_EVM, target_VHT20_MCS8_EVM, target_VHT40_MCS0_EVM, \
    target_VHT40_MCS1_EVM, target_VHT40_MCS2_EVM, target_VHT40_MCS3_EVM, target_VHT40_MCS4_EVM, target_VHT40_MCS5_EVM, \
    target_VHT40_MCS6_EVM, target_VHT40_MCS7_EVM, target_VHT40_MCS8_EVM, target_VHT40_MCS9_EVM, target_VHT80_MCS0_EVM, \
    target_VHT80_MCS1_EVM, target_VHT80_MCS2_EVM, target_VHT80_MCS3_EVM, target_VHT80_MCS4_EVM, target_VHT80_MCS5_EVM, \
    target_VHT80_MCS6_EVM, target_VHT80_MCS7_EVM, target_VHT80_MCS8_EVM, target_VHT80_MCS9_EVM, \
    target_VHT160_MCS0_EVM, target_VHT160_MCS1_EVM, target_VHT160_MCS2_EVM, target_VHT160_MCS3_EVM, \
    target_VHT160_MCS4_EVM, target_VHT160_MCS5_EVM, target_VHT160_MCS6_EVM, target_VHT160_MCS7_EVM, \
    target_VHT160_MCS8_EVM, target_VHT160_MCS9_EVM, target_HE20_HE0_EVM, target_HE20_HE1_EVM, target_HE20_HE2_EVM, \
    target_HE20_HE3_EVM, target_HE20_HE4_EVM, target_HE20_HE5_EVM, \
    target_HE20_HE6_EVM, target_HE20_HE7_EVM, target_HE20_HE8_EVM, target_HE20_HE9_EVM, target_HE20_HE10_EVM, \
    target_HE20_HE11_EVM, target_HE40_HE0_EVM, target_HE40_HE1_EVM, target_HE40_HE2_EVM, target_HE40_HE3_EVM, \
    target_HE40_HE4_EVM, target_HE40_HE5_EVM, target_HE40_HE6_EVM, target_HE40_HE7_EVM, target_HE40_HE8_EVM, \
    target_HE40_HE9_EVM, target_HE40_HE10_EVM, target_HE40_HE11_EVM, target_HE80_HE0_EVM, target_HE80_HE1_EVM, \
    target_HE80_HE2_EVM, target_HE80_HE3_EVM, target_HE80_HE4_EVM, target_HE80_HE5_EVM, target_HE80_HE6_EVM, \
    target_HE80_HE7_EVM, target_HE80_HE8_EVM, target_HE80_HE9_EVM, target_HE80_HE10_EVM, target_HE80_HE11_EVM, \
    target_HE160_HE0_EVM, target_HE160_HE1_EVM, target_HE160_HE2_EVM, target_HE160_HE3_EVM, target_HE160_HE4_EVM, \
    target_HE160_HE5_EVM, target_HE160_HE6_EVM, target_HE160_HE7_EVM, target_HE160_HE8_EVM, target_HE160_HE9_EVM, \
    target_HE160_HE10_EVM, target_HE160_HE11_EVM = [None] * 115

    target_1M_sens, target_2M_sens, target_5_5M_sens, target_11M_sens, target_6M_sens, target_9M_sens, target_12M_sens, \
    target_18M_sens, target_24M_sens, target_36M_sens, target_48M_sens, target_54M_sens, target_HT20_MCS0_sens, \
    target_HT20_MCS1_sens, target_HT20_MCS2_sens, target_HT20_MCS3_sens, target_HT20_MCS4_sens, target_HT20_MCS5_sens, \
    target_HT20_MCS6_sens, target_HT20_MCS7_sens, target_HT40_MCS0_sens, target_HT40_MCS1_sens, target_HT40_MCS2_sens, \
    target_HT40_MCS3_sens, target_HT40_MCS4_sens, target_HT40_MCS5_sens, target_HT40_MCS6_sens, target_HT40_MCS7_sens, \
    target_VHT20_MCS0_sens, target_VHT20_MCS1_sens, target_VHT20_MCS2_sens, target_VHT20_MCS3_sens, target_VHT20_MCS4_sens, \
    target_VHT20_MCS5_sens, target_VHT20_MCS6_sens, target_VHT20_MCS7_sens, target_VHT20_MCS8_sens, target_VHT40_MCS0_sens, \
    target_VHT40_MCS1_sens, target_VHT40_MCS2_sens, target_VHT40_MCS3_sens, target_VHT40_MCS4_sens, target_VHT40_MCS5_sens, \
    target_VHT40_MCS6_sens, target_VHT40_MCS7_sens, target_VHT40_MCS8_sens, target_VHT40_MCS9_sens, target_VHT80_MCS0_sens, \
    target_VHT80_MCS1_sens, target_VHT80_MCS2_sens, target_VHT80_MCS3_sens, target_VHT80_MCS4_sens, target_VHT80_MCS5_sens, \
    target_VHT80_MCS6_sens, target_VHT80_MCS7_sens, target_VHT80_MCS8_sens, target_VHT80_MCS9_sens, \
    target_VHT160_MCS0_sens, target_VHT160_MCS1_sens, target_VHT160_MCS2_sens, target_VHT160_MCS3_sens, \
    target_VHT160_MCS4_sens, target_VHT160_MCS5_sens, target_VHT160_MCS6_sens, target_VHT160_MCS7_sens, \
    target_VHT160_MCS8_sens, target_VHT160_MCS9_sens, target_HE20_HE0_sens, target_HE20_HE1_sens, target_HE20_HE2_sens, \
    target_HE20_HE3_sens, target_HE20_HE4_sens, target_HE20_HE5_sens, \
    target_HE20_HE6_sens, target_HE20_HE7_sens, target_HE20_HE8_sens, target_HE20_HE9_sens, target_HE20_HE10_sens, \
    target_HE20_HE11_sens, target_HE40_HE0_sens, target_HE40_HE1_sens, target_HE40_HE2_sens, target_HE40_HE3_sens, \
    target_HE40_HE4_sens, target_HE40_HE5_sens, target_HE40_HE6_sens, target_HE40_HE7_sens, target_HE40_HE8_sens, \
    target_HE40_HE9_sens, target_HE40_HE10_sens, target_HE40_HE11_sens, target_HE80_HE0_sens, target_HE80_HE1_sens, \
    target_HE80_HE2_sens, target_HE80_HE3_sens, target_HE80_HE4_sens, target_HE80_HE5_sens, target_HE80_HE6_sens, \
    target_HE80_HE7_sens, target_HE80_HE8_sens, target_HE80_HE9_sens, target_HE80_HE10_sens, target_HE80_HE11_sens, \
    target_HE160_HE0_sens, target_HE160_HE1_sens, target_HE160_HE2_sens, target_HE160_HE3_sens, target_HE160_HE4_sens, \
    target_HE160_HE5_sens, target_HE160_HE6_sens, target_HE160_HE7_sens, target_HE160_HE8_sens, target_HE160_HE9_sens, \
    target_HE160_HE10_sens, target_HE160_HE11_sens = [None] * 115

    # GEN Report
    now_time = datetime.datetime.now()
    # logger.debug(now_time)
    now_time = str(now_time)
    now_time = now_time.split()
    # logger.debug(now_time)
    day_time = re.sub('-', '', now_time[0])
    now_time = now_time[1].split('.')
    # logger.debug(day_time)
    # logger.debug(now_time)
    now_time = re.sub(':', '', now_time[0])
    now_time = day_time + now_time
    # logger.debug(now_time)
    tx_result_name = sn + '_' + 'TX_Result' + '_' + now_time + '.csv'
    rx_result_name = sn + '_' + 'RX_Result' + '_' + now_time + '.csv'
    directory = os.path.exists(r'./Result')
    if directory is False:
        os.makedirs('Result')

    with open('./Result/' + tx_result_name, 'w', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(['FREQ', 'DATA_RATE', 'CHAIN', 'TX_POWER', 'POWER', 'GAIN', 'LIMIT',
                         'RESULT', 'EVM', 'LIMIT', 'RESULT', 'FREQ_ERROR', 'LIMIT', 'RESULT',
                         'LOLEAKAGE', 'LIMIT', 'RESULT', 'OBW', 'LIMIT', 'RESULT',
                         'MASK', 'LIMIT', 'RESULT', 'FLATNESS', 'LIMIT', 'RESULT',
                         'RAMPONTIME', 'LIMIT', 'RESULT', 'RAMPOFFTIME', 'LIMIT', 'RESULT'])

    with open('./Result/' + rx_result_name, 'w', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(['FREQ', 'DATA_RATE', 'CHAIN', 'SENSITIVITY', 'RX PACKETS', 'RESULT'])

    if rx_dynamic == '1':
        rx_dynamic_result = sn + '_' + 'RX_Dynamic' + '_' + now_time + '.csv'
        with open('./Result/' + rx_dynamic_result, 'w', newline='') as f:
            writer = csv.writer(f)
            writer.writerow(['FREQ', 'DATA_RATE', 'CHAIN', 'VSG_POWER', 'PER', 'RESULT'])
    # TEST FLOW
    filename = 'TEST_FLOW.txt'
    f = open(filename)
    result = list()
    for line in f.readlines():
        #logger.debug(len(line))
        if len(line) < 30 or line.startswith('//'):
            continue
            pass
        else:
            line = line.strip()
            line = line.split()
            # logger.debug(line)
            # logger.debug('Channel:', line[1], 'Rate:', line[2], 'Chain:', line[3])
            item = line[0]
            item = re.sub('IQ_WIFI_TEST_', '', item)
            logger.debug(item)
            channel = line[1]
            rate = line[2]
            chain = line[3]
            if item == 'RX':
                start = int(line[4])
                stop = int(line[5])
            # mode,bw,rates
            if rate == '1M' or rate == '2M' or rate == '5.5M' or rate == '11M':
                mode = '11b'
                bw = '20'
                rates = re.sub('M', '', rate)
                per_spec = float(rx_packets) * 0.08
            elif rate == '6M' or rate == '9M' or rate == '12M' or rate == '18M' \
                    or rate == '24M' or rate == '36M' or rate == '48M' or rate == '54M':
                if int(channel) < 5000:
                    mode = '11g'
                else:
                    mode = '11a'
                bw = '20'
                rates = re.sub('M', '', rate)
                per_spec = float(rx_packets) * 0.1
            elif rate == 'HT20-MCS0' or rate == 'HT20-MCS1' or rate == 'HT20-MCS2' or rate == 'HT20-MCS3' \
                    or rate == 'HT20-MCS4' or rate == 'HT20-MCS5' or rate == 'HT20-MCS6' or rate == 'HT20-MCS7':
                mode = '11ng20'
                bw = '20'
                rates = re.sub('HT20-MCS', '', rate)
                per_spec = float(rx_packets) * 0.1
            elif rate == 'HT40-MCS0' or rate == 'HT40-MCS1' or rate == 'HT40-MCS2' or rate == 'HT40-MCS3' \
                    or rate == 'HT40-MCS4' or rate == 'HT40-MCS5' or rate == 'HT40-MCS6' or rate == 'HT40-MCS7':
                mode = '11ng40plus'
                bw = '40'
                rates = re.sub('HT40-MCS', '', rate)
                per_spec = float(rx_packets) * 0.1
                #channel = int(channel) - 10
            elif rate == 'VHT20-MCS0' or rate == 'VHT20-MCS1' or rate == 'VHT20-MCS2' or rate == 'VHT20-MCS3' \
                    or rate == 'VHT20-MCS4' or rate == 'VHT20-MCS5' or rate == 'VHT20-MCS6' or rate == 'VHT20-MCS7' or \
                    rate == 'VHT20-MCS8':
                mode = '11ac'
                bw = '20'
                rates = re.sub('VHT20-MCS', '', rate)
                per_spec = float(rx_packets) * 0.1
            elif rate == 'VHT40-MCS0' or rate == 'VHT40-MCS1' or rate == 'VHT40-MCS2' or rate == 'VHT40-MCS3' \
                    or rate == 'VHT40-MCS4' or rate == 'VHT40-MCS5' or rate == 'VHT40-MCS6' or rate == 'VHT40-MCS7' or \
                    rate == 'VHT40-MCS8' or rate == 'VHT40-MCS9':
                mode = '11ac'
                bw = '40'
                rates = re.sub('VHT40-MCS', '', rate)
                per_spec = float(rx_packets) * 0.1
                #channel = int(channel) - 10
            elif rate == 'VHT80-MCS0' or rate == 'VHT80-MCS1' or rate == 'VHT80-MCS2' or rate == 'VHT80-MCS3' \
                    or rate == 'VHT80-MCS4' or rate == 'VHT80-MCS5' or rate == 'VHT80-MCS6' or rate == 'VHT80-MCS7' or \
                    rate == 'VHT80-MCS8' or rate == 'VHT80-MCS9':
                mode = '11ac'
                bw = '80'
                rates = re.sub('VHT80-MCS', '', rate)
                per_spec = float(rx_packets) * 0.1
            elif rate == 'VHT160-MCS0' or rate == 'VHT160-MCS1' or rate == 'VHT160-MCS2' or rate == 'VHT160-MCS3' \
                    or rate == 'VHT160-MCS4' or rate == 'VHT160-MCS5' or rate == 'VHT160-MCS6' or rate == 'VHT160-MCS7' or \
                    rate == 'VHT160-MCS8' or rate == 'VHT160-MCS9':
                mode = '11ac'
                bw = '160'
                rates = re.sub('VHT160-MCS', '', rate)
                per_spec = float(rx_packets) * 0.1
            elif rate == 'HE20-HE0' or rate == 'HE20-HE1' or rate == 'HE20-HE2' or rate == 'HE20-HE3' \
                    or rate == 'HE20-HE4' or rate == 'HE20-HE5' or rate == 'HE20-HE6' or rate == 'HE20-HE7' or \
                    rate == 'HE20-HE8' or rate == 'HE20-HE9' or rate == 'HE20-HE10' or rate == 'HE20-HE11':
                mode = '11ax'
                bw = '20'
                rates = re.sub('HE20-HE', '', rate)
                per_spec = float(rx_packets) * 0.1
            elif rate == 'HE40-HE0' or rate == 'HE40-HE1' or rate == 'HE40-HE2' or rate == 'HE40-HE3' \
                    or rate == 'HE40-HE4' or rate == 'HE40-HE5' or rate == 'HE40-HE6' or rate == 'HE40-HE7' or \
                    rate == 'HE40-HE8' or rate == 'HE40-HE9' or rate == 'HE40-HE10' or rate == 'HE40-HE11':
                mode = '11ax'
                bw = '40'
                rates = re.sub('HE40-HE', '', rate)
                per_spec = float(rx_packets) * 0.1
                #channel = int(channel) - 10
            elif rate == 'HE80-HE0' or rate == 'HE80-HE1' or rate == 'HE80-HE2' or rate == 'HE80-HE3' \
                    or rate == 'HE80-HE4' or rate == 'HE80-HE5' or rate == 'HE80-HE6' or rate == 'HE80-HE7' or \
                    rate == 'HE80-HE8' or rate == 'HE80-HE9' or rate == 'HE80-HE10' or rate == 'HE80-HE11':
                mode = '11ax'
                bw = '80'
                rates = re.sub('HE80-HE', '', rate)
                per_spec = float(rx_packets) * 0.1
            elif rate == 'HE160-HE0' or rate == 'HE160-HE1' or rate == 'HE160-HE2' or rate == 'HE160-HE3' \
                    or rate == 'HE160-HE4' or rate == 'HE160-HE5' or rate == 'HE160-HE6' or rate == 'HE160-HE7' or \
                    rate == 'HE160-HE8' or rate == 'HE160-HE9' or rate == 'HE160-HE10' or rate == 'HE160-HE11':
                mode = '11ax'
                bw = '160'
                rates = re.sub('HE160-HE', '', rate)
                per_spec = float(rx_packets) * 0.1
            # chain
            chain = re.sub('CHAIN', '', chain)
            # power para
            if int(channel) < 2437 and chain == '0':
                channel_groups = '98'
            elif int(channel) < 2437 and chain == '1':
                channel_groups = 'a8'
            elif 2437 < int(channel) < 2462 and chain == '0':
                channel_groups = '94'
            elif 2437 < int(channel) < 2462 and chain == '1':
                channel_groups = 'a4'
            elif int(channel) >= 2462 and chain == '0':
                channel_groups = '90'
            elif int(channel) >= 2462 and chain == '1':
                channel_groups = 'a0'
            else:
                logger.info('5g need add')
            # read spec
            if int(channel) < 5000:
                # INIT SPEC
                # 2.4g

                spec_file_2g = load_workbook('./spec_2g.xlsx')
                # logger.debug(spec_file.sheetnames)
                sheet_2g = spec_file_2g['Sheet1']
                rows_2g = []
                ratelist_2g = ['1M', '2M', '5_5M', '11M', '6M', '9M', '12M', '18M', '24M', '36M', '48M', '54M',
                               'HT20_MCS0', 'HT20_MCS1', 'HT20_MCS2', 'HT20_MCS3', 'HT20_MCS4', 'HT20_MCS5',
                               'HT20_MCS6', 'HT20_MCS7', 'HT40_MCS0', 'HT40_MCS1', 'HT40_MCS2', 'HT40_MCS3',
                               'HT40_MCS4', 'HT40_MCS5', 'HT40_MCS6', 'HT40_MCS7', 'VHT20_MCS0', 'VHT20_MCS1',
                               'VHT20_MCS2', 'VHT20_MCS3', 'VHT20_MCS4', 'VHT20_MCS5', 'VHT20_MCS6', 'VHT20_MCS7',
                               'VHT20_MCS8', 'VHT40_MCS0', 'VHT40_MCS1', 'VHT40_MCS2', 'VHT40_MCS3', 'VHT40_MCS4',
                               'VHT40_MCS5', 'VHT40_MCS6', 'VHT40_MCS7', 'VHT40_MCS8', 'VHT40_MCS9', 'HE20_HE0',
                               'HE20_HE1', 'HE20_HE2', 'HE20_HE3', 'HE20_HE4', 'HE20_HE5', 'HE20_HE6',
                               'HE20_HE7', 'HE20_HE8', 'HE20_HE9', 'HE20_HE10', 'HE20_HE11', 'HE40_HE0',
                               'HE40_HE1', 'HE40_HE2', 'HE40_HE3', 'HE40_HE4', 'HE40_HE5', 'HE40_HE6',
                               'HE40_HE7', 'HE40_HE8', 'HE40_HE9', 'HE40_HE10', 'HE40_HE11']
                ratelist_evm_2g = ['1M_EVM', '2M_EVM', '5_5M_EVM', '11M_EVM', '6M_EVM', '9M_EVM', '12M_EVM', '18M_EVM',
                                   '24M_EVM', '36M_EVM', '48M_EVM', '54M_EVM', 'HT20_MCS0_EVM', 'HT20_MCS1_EVM',
                                   'HT20_MCS2_EVM', 'HT20_MCS3_EVM', 'HT20_MCS4_EVM', 'HT20_MCS5_EVM', 'HT20_MCS6_EVM',
                                   'HT20_MCS7_EVM', 'HT40_MCS0_EVM', 'HT40_MCS1_EVM', 'HT40_MCS2_EVM', 'HT40_MCS3_EVM',
                                   'HT40_MCS4_EVM', 'HT40_MCS5_EVM', 'HT40_MCS6_EVM', 'HT40_MCS7_EVM', 'VHT20_MCS0_EVM',
                                   'VHT20_MCS1_EVM', 'VHT20_MCS2_EVM', 'VHT20_MCS3_EVM', 'VHT20_MCS4_EVM',
                                   'VHT20_MCS5_EVM', 'VHT20_MCS6_EVM', 'VHT20_MCS7_EVM', 'VHT20_MCS8_EVM',
                                   'VHT40_MCS0_EVM', 'VHT40_MCS1_EVM', 'VHT40_MCS2_EVM', 'VHT40_MCS3_EVM',
                                   'VHT40_MCS4_EVM', 'VHT40_MCS5_EVM', 'VHT40_MCS6_EVM', 'VHT40_MCS7_EVM',
                                   'VHT40_MCS8_EVM', 'VHT40_MCS9_EVM', 'HE20_HE0_EVM', 'HE20_HE1_EVM',
                                   'HE20_HE2_EVM', 'HE20_HE3_EVM', 'HE20_HE4_EVM', 'HE20_HE5_EVM', 'HE20_HE6_EVM',
                                   'HE20_HE7_EVM', 'HE20_HE8_EVM', 'HE20_HE9_EVM', 'HE20_HE10_EVM',
                                   'HE20_HE11_EVM', 'HE40_HE0_EVM', 'HE40_HE1_EVM', 'HE40_HE2_EVM',
                                   'HE40_HE3_EVM', 'HE40_HE4_EVM', 'HE40_HE5_EVM', 'HE40_HE6_EVM',
                                   'HE40_HE7_EVM', 'HE40_HE8_EVM', 'HE40_HE9_EVM', 'HE40_HE10_EVM',
                                   'HE40_HE11_EVM']

                for row_2g in sheet_2g:
                    rows_2g.append(row_2g)
                    # logger.debug(rows)
                for r in range(sheet_2g.max_row):
                    for c in range(sheet_2g.max_column):
                        # logger.debug(rows[r][c].value)
                        rows_2g[r][c].value = str(rows_2g[r][c].value).strip()
                        rs = r + 1
                        cs = c + 1
                        if rows_2g[r][c].value == 'POWER_ACCURACY':
                            spec_pwr = abs(rows_2g[r][cs].value)
                        elif rows_2g[r][c].value == 'EVM_MARGIN':
                            evm_margin = abs(rows_2g[r][cs].value)
                        elif rows_2g[r][c].value == 'Symbol_Clock_Error':
                            spec_symbol_clock_error = abs(rows_2g[r][cs].value)
                        elif rows_2g[r][c].value == 'XCAP':
                            xcap = abs(rows_2g[r][cs].value)
                        elif rows_2g[r][c].value == 'LO_Leakage':
                            spec_lo_leakage = -abs(rows_2g[r][cs].value)
                        elif rows_2g[r][c].value == 'MASK':
                            spec_mask = abs(rows_2g[r][cs].value)
                        elif rows_2g[r][c].value == 'OBW_20M':
                            spec_obw_20M = rows_2g[r][cs].value
                        elif rows_2g[r][c].value == 'OBW_40M':
                            spec_obw_40M = rows_2g[r][cs].value
                        for x in ratelist_2g:
                            if rows_2g[r][c].value == x + '_target':
                                exec('target_pwr_%s=%d' % (x, rows_2g[rs][c].value))
                                break
                        for i in ratelist_evm_2g:
                            if rows_2g[r][c].value == i + '_target':
                                exec('target_%s=%d' % (i, rows_2g[rs][c].value))
                                break
                        for j in ratelist_2g:
                            if rows_2g[r][c].value == j + '_sens':
                                exec('target_%s_sens=%d' % (j, rows_2g[rs][c].value))
                                break
                        spec_obw_80M = spec_obw_160M = None
            else:
                # 5g
                spec_file_5g = load_workbook('./spec_5g.xlsx')
                # logger.debug(spec_file.sheetnames)
                sheet_5g = spec_file_5g['Sheet1']
                rows_5g = []
                ratelist_5g = ['6M', '9M', '12M', '18M', '24M', '36M', '48M', '54M', 'HT20_MCS0', 'HT20_MCS1',
                               'HT20_MCS2', 'HT20_MCS3', 'HT20_MCS4', 'HT20_MCS5', 'HT20_MCS6', 'HT20_MCS7',
                               'HT40_MCS0', 'HT40_MCS1', 'HT40_MCS2', 'HT40_MCS3', 'HT40_MCS4', 'HT40_MCS5',
                               'HT40_MCS6', 'HT40_MCS7', 'VHT20_MCS0', 'VHT20_MCS1', 'VHT20_MCS2', 'VHT20_MCS3',
                               'VHT20_MCS4', 'VHT20_MCS5', 'VHT20_MCS6', 'VHT20_MCS7', 'VHT20_MCS8', 'VHT40_MCS0',
                               'VHT40_MCS1', 'VHT40_MCS2', 'VHT40_MCS3', 'VHT40_MCS4', 'VHT40_MCS5', 'VHT40_MCS6',
                               'VHT40_MCS7', 'VHT40_MCS8', 'VHT40_MCS9', 'VHT80_MCS0', 'VHT80_MCS1', 'VHT80_MCS2',
                               'VHT80_MCS3', 'VHT80_MCS4', 'VHT80_MCS5', 'VHT80_MCS6', 'VHT80_MCS7', 'VHT80_MCS8',
                               'VHT80_MCS9', 'VHT160_MCS0', 'VHT160_MCS1', 'VHT160_MCS2',
                               'VHT160_MCS3', 'VHT160_MCS4', 'VHT160_MCS5', 'VHT160_MCS6', 'VHT160_MCS7', 'VHT160_MCS8',
                               'VHT160_MCS9', 'HE20_HE0', 'HE20_HE1', 'HE20_HE2', 'HE20_HE3', 'HE20_HE4',
                               'HE20_HE5', 'HE20_HE6', 'HE20_HE7', 'HE20_HE8', 'HE20_HE9', 'HE20_HE10',
                               'HE20_HE11', 'HE40_HE0', 'HE40_HE1', 'HE40_HE2', 'HE40_HE3', 'HE40_HE4',
                               'HE40_HE5', 'HE40_HE6', 'HE40_HE7', 'HE40_HE8', 'HE40_HE9', 'HE40_HE10',
                               'HE40_HE11', 'HE80_HE0', 'HE80_HE1', 'HE80_HE2', 'HE80_HE3', 'HE80_HE4',
                               'HE80_HE5', 'HE80_HE6', 'HE80_HE7', 'HE80_HE8', 'HE80_HE9', 'HE80_HE10',
                               'HE80_HE11', 'HE160_HE0', 'HE160_HE1', 'HE160_HE2', 'HE160_HE3', 'HE160_HE4',
                               'HE160_HE5', 'HE160_HE6', 'HE160_HE7', 'HE160_HE8', 'HE160_HE9', 'HE160_HE10',
                               'HE160_HE11']
                ratelist_evm_5g = ['6M_EVM', '9M_EVM', '12M_EVM', '18M_EVM', '24M_EVM', '36M_EVM', '48M_EVM',
                                   '54M_EVM', 'HT20_MCS0_EVM', 'HT20_MCS1_EVM', 'HT20_MCS2_EVM', 'HT20_MCS3_EVM',
                                   'HT20_MCS4_EVM', 'HT20_MCS5_EVM', 'HT20_MCS6_EVM', 'HT20_MCS7_EVM', 'HT40_MCS0_EVM',
                                   'HT40_MCS1_EVM', 'HT40_MCS2_EVM', 'HT40_MCS3_EVM', 'HT40_MCS4_EVM', 'HT40_MCS5_EVM',
                                   'HT40_MCS6_EVM', 'HT40_MCS7_EVM', 'VHT20_MCS0_EVM', 'VHT20_MCS1_EVM',
                                   'VHT20_MCS2_EVM', 'VHT20_MCS3_EVM', 'VHT20_MCS4_EVM', 'VHT20_MCS5_EVM',
                                   'VHT20_MCS6_EVM', 'VHT20_MCS7_EVM', 'VHT20_MCS8_EVM', 'VHT40_MCS0_EVM',
                                   'VHT40_MCS1_EVM', 'VHT40_MCS2_EVM', 'VHT40_MCS3_EVM', 'VHT40_MCS4_EVM',
                                   'VHT40_MCS5_EVM', 'VHT40_MCS6_EVM', 'VHT40_MCS7_EVM', 'VHT40_MCS8_EVM',
                                   'VHT40_MCS9_EVM', 'VHT80_MCS0_EVM', 'VHT80_MCS1_EVM', 'VHT80_MCS2_EVM',
                                   'VHT80_MCS3_EVM', 'VHT80_MCS4_EVM', 'VHT80_MCS5_EVM', 'VHT80_MCS6_EVM',
                                   'VHT80_MCS7_EVM', 'VHT80_MCS8_EVM', 'VHT80_MCS9_EVM', 'VHT160_MCS0_EVM',
                                   'VHT160_MCS1_EVM', 'VHT160_MCS2_EVM', 'VHT160_MCS3_EVM', 'VHT160_MCS4_EVM',
                                   'VHT160_MCS5_EVM', 'VHT160_MCS6_EVM', 'VHT160_MCS7_EVM', 'VHT160_MCS8_EVM',
                                   'VHT160_MCS9_EVM', 'HE20_HE0_EVM', 'HE20_HE1_EVM', 'HE20_HE2_EVM',
                                   'HE20_HE3_EVM', 'HE20_HE4_EVM', 'HE20_HE5_EVM', 'HE20_HE6_EVM',
                                   'HE20_HE7_EVM', 'HE20_HE8_EVM', 'HE20_HE9_EVM', 'HE20_HE10_EVM',
                                   'HE20_HE11_EVM', 'HE40_HE0_EVM', 'HE40_HE1_EVM', 'HE40_HE2_EVM',
                                   'HE40_HE3_EVM', 'HE40_HE4_EVM', 'HE40_HE5_EVM', 'HE40_HE6_EVM',
                                   'HE40_HE7_EVM', 'HE40_HE8_EVM', 'HE40_HE9_EVM', 'HE40_HE10_EVM',
                                   'HE40_HE11_EVM', 'HE80_HE0_EVM', 'HE80_HE1_EVM', 'HE80_HE2_EVM',
                                   'HE80_HE3_EVM', 'HE80_HE4_EVM', 'HE80_HE5_EVM', 'HE80_HE6_EVM',
                                   'HE80_HE7_EVM', 'HE80_HE8_EVM', 'HE80_HE9_EVM', 'HE80_HE10_EVM',
                                   'HE80_HE11_EVM', 'HE160_HE0_EVM', 'HE160_HE1_EVM', 'HE160_HE2_EVM',
                                   'HE160_HE3_EVM', 'HE160_HE4_EVM', 'HE160_HE5_EVM', 'HE160_HE6_EVM',
                                   'HE160_HE7_EVM', 'HE160_HE8_EVM', 'HE160_HE9_EVM', 'HE160_HE10_EVM',
                                   'HE160_HE11_EVM']
                for row_5g in sheet_5g:
                    rows_5g.append(row_5g)
                    # logger.debug(rows)
                for rr in range(sheet_5g.max_row):
                    for cc in range(sheet_5g.max_column):
                        # logger.debug(rows[r][c].value)
                        rows_5g[rr][cc].value = str(rows_5g[rr][cc].value).strip()
                        rrs = rr + 1
                        ccs = cc + 1
                        if rows_5g[rr][cc].value == 'POWER_ACCURACY':
                            spec_pwr = abs(rows_5g[rr][ccs].value)
                        elif rows_5g[rr][cc].value == 'EVM_MARGIN':
                            evm_margin = abs(rows_5g[rr][ccs].value)
                        elif rows_5g[rr][cc].value == 'Symbol_Clock_Error':
                            spec_symbol_clock_error = abs(rows_5g[rr][ccs].value)
                        elif rows_5g[rr][cc].value == 'XCAP':
                            xcap = abs(rows_5g[rr][ccs].value)
                        elif rows_5g[rr][cc].value == 'LO_Leakage':
                            spec_lo_leakage = -abs(rows_5g[rr][ccs].value)
                        elif rows_5g[rr][cc].value == 'MASK':
                            spec_mask = abs(rows_5g[rr][ccs].value)
                        elif rows_5g[rr][cc].value == 'OBW_20M':
                            spec_obw_20M = rows_5g[rr][ccs].value
                        elif rows_5g[rr][cc].value == 'OBW_40M':
                            spec_obw_40M = rows_5g[rr][ccs].value
                        elif rows_5g[rr][cc].value == 'OBW_80M':
                            spec_obw_80M = rows_5g[rr][ccs].value
                        elif rows_5g[rr][cc].value == 'OBW_160M':
                            spec_obw_160M = rows_5g[rr][ccs].value
                        for xx in ratelist_5g:
                            if rows_5g[rr][cc].value == xx + '_target':
                                exec('target_pwr_%s=%d' % (xx, rows_5g[rrs][cc].value))
                                break
                        for ii in ratelist_evm_5g:
                            if rows_5g[rr][cc].value == ii + '_target':
                                exec('target_%s=%d' % (ii, rows_5g[rrs][cc].value))
                                break
                        for jj in ratelist_5g:
                            if rows_5g[rr][cc].value == jj + '_sens':
                                exec('target_%s_sens=%d' % (jj, rows_5g[rrs][cc].value))
                                break
            # targetpower = 0
            # spec_evm = 0

            logger.info('*************************************************************')
            logger.info('Mode: '+ mode + ' Channel: ' + channel+' BW: '+ bw+ ' Rate: '+ rate + ' Chain: '+ chain)
            if int(channel) < 5000:
                band = id_2g
            elif int(channel) < 5500:
                band = id_5g_l
            else:
                band = id_5g_h
            if item == 'TX':
                adjust_result = result_evm = result_symbol_clock_error = result_lo_leakage = result_mask = \
                    result_flatness = 'Pass'
                # dt.set_default()
                dt.tx()
                # RESULT
                rate_t = re.sub('-', '_', rate)
                rate_t = re.sub('\.', '_', rate_t)
                targetpower = eval('target_pwr_' + rate_t)
                rate_e = rate_t + '_EVM'
                spec_evm = eval('target_' + rate_e)
                iq.set_port(targetpower)
                iq.analysis()
                pwr_len, txq_len, data_pwr, data_txq = iq.get_status()
                avg_power, result_evm, result_symbol_clock_error, result_lo_leakage, result_mask = \
                    iq.get_data(pwr_len, txq_len, data_pwr, data_txq, mode, channel, rate, chain, tx_result_name,
                                targetpower, spec_pwr, gain, spec_evm, evm_margin, spec_symbol_clock_error,
                                spec_lo_leakage, spec_mask, spec_obw_20M, spec_obw_40M, spec_obw_80M)
                if super_mode == '1':
                    # if avg_power == 'NA' or result_evm == 'NA':
                    if avg_power == 'NA' or float(avg_power) > 99.000:
                        logger.info(Fore.RED + 'Error!' + Style.RESET_ALL)
                    else:
                        # accuracy_limit_left = 0.5
                        # accuracy_limit_right = 0.5
                        power_accuracy_left = float(targetpower) - float(spec_pwr) + float(accuracy_limit_left)
                        power_accuracy_right = float(targetpower) + float(spec_pwr) - float(accuracy_limit_right)
                        max_power = float(targetpower + spec_pwr)
                        delta_power = float(avg_power) - float(targetpower)
                        delta_power = float('{:.3f}'.format(delta_power))
                        logger.debug('Default Measurer Power: ' + avg_power)
                        logger.debug('Target Power: ' + targetpower)
                        logger.debug('Delta Power: ' + delta_power)
                        # power is nomal but some is fail
                        if targetpower <= avg_power <= max_power:
                            if result_evm == 'Fail' or result_symbol_clock_error == 'Fail' \
                                    or result_lo_leakage == 'Fail' or result_mask == 'Fail':
                                logger.info(Fore.RED + 'TX\'s quality are failed!' + Style.RESET_ALL)
                            else:
                                logger.info('Get Good Result!')
                        # power is not nomal
                        else:
                            low_power_counts = 0
                            setup = 2
                            setups = 2
                            logger.debug('Step:' + setup)
                            init_value = dt.get_paras()
                            pwr_paras = init_value + setup
                            pwr_paras = hex(pwr_paras)
                            pwr_paras = re.sub('0x', '', pwr_paras)
                            get_power = []
                            get_delta_power = []
                            c = 1
                            get_power.append(avg_power)
                            logger.debug('Measurer Power List: '+get_power)
                            get_delta_power.append(delta_power)
                            logger.debug('Power Deviation List: ' + get_delta_power)
                            logger.debug('Adjust Counts:' + c)
                            # ADJUST
                            min_adj = avg_power - targetpower
                            max_adj = avg_power - max_power
                            while min_adj < 0 or max_adj > 0:
                                logger.info(Fore.GREEN + 'Adjust...' + Style.RESET_ALL)
                                if avg_power > max_power:
                                    adj = -1
                                else:
                                    adj = 1
                                c = c + 1
                                logger.debug('NEW VALUE(HEX):', channel_groups, pwr_paras)
                                gain = dt.adjust_power()
                                iq.analysis()
                                pwr_len, txq_len, data_pwr, data_txq = iq.get_status()
                                avg_power, result_evm, result_symbol_clock_error, result_lo_leakage, result_mask = \
                                    iq.get_data(pwr_len, txq_len, data_pwr, data_txq, mode, channel, rate, chain,
                                                tx_result_name, targetpower, spec_pwr, gain, spec_evm, evm_margin,
                                                spec_symbol_clock_error, spec_lo_leakage, spec_mask, spec_obw_20M,
                                                spec_obw_40M, spec_obw_80M)
                                get_power.append(avg_power)
                                logger.debug('Measurer Power List: ' + get_power)
                                delta_power = float(avg_power) - targetpower
                                delta_power = float('{:.3f}'.format(delta_power))
                                avg_power = float(avg_power)
                                min_adj = avg_power - targetpower
                                max_adj = avg_power - max_power
                                get_delta_power.append(delta_power)
                                logger.debug('Power Deviation List: ' + get_delta_power)
                                logger.debug('Adjust Counts: ' + c)
                                adjust_status = get_delta_power[c - 1] - get_delta_power[c - 2]
                                adjust_status = float('{:.3f}'.format(adjust_status))
                                logger.debug('Power Added: ' + adjust_status)
                                if adjust_status < 0.1:
                                    adjust_result = 'Pass'
                                    setup = 2 * setups * adj
                                    low_power_counts = low_power_counts + 1
                                elif adjust_status > 0.1:
                                    adjust_status = 'Pass'
                                    setup = 2 * adj
                                # logger.debug(low_power_counts)
                                if low_power_counts > 5:
                                    logger.info(Fore.RED + 'Power added was too small, Power adjust stop' + Style.RESET_ALL)
                                    adjust_result = 'Fail'
                                else:
                                    logger.debug('Step: ' + setup)
                                    pwr_paras = int(pwr_paras, 16)
                                    pwr_paras = int(pwr_paras) + setup
                                    pwr_paras = hex(pwr_paras)
                                    pwr_paras = re.sub('0x', '', pwr_paras)
                                    logger.debug(result_evm + result_symbol_clock_error + result_lo_leakage + result_mask)
                                    logger.info('*************************************************************')
            else:
                per_list = []
                sens_list = []
                logger.info('Start: ' + str(start) + 'Stop: ' + str(stop))
                # loss = iq.read_pathloss(channel, chain)
                dt.rx(mode, channel, bw, rates, chain)
                # logger.debug(channel)
                iq.vsg()
                per = dt.get_statistics()
                per = int(per)
                per_list.append(per)
                sens_list.append(start)
                while start > stop and per <= int(per_spec):
                    start = start - 1
                    # dt.rx(mode, channel, bw, rates, chain)
                    iq.vsg()
                    time.sleep(sleep_time)
                    per = dt.get_statistics()
                    per = int(per)
                    per_list.append(per)
                    sens_list.append(start)
                    if rx_dynamic == '1':
                        if per <= per_spec:
                            per_result = 'Pass'
                        else:
                            per_result = 'Fail'
                        with open('./Result/' + rx_dynamic_result, 'a+', newline='') as f2:
                            writer2 = csv.writer(f2)
                            writer2.writerow([channel, rate, chain, start, per, per_result])
                logger.debug(len(per_list), str(per_list))
                logger.debug(len(sens_list), str(sens_list))
                per = per_list[len(per_list) - 2]
                sens = sens_list[len(sens_list) - 2]
                logger.info('Sensitivity: ' + str(sens))
                rate_t = re.sub('-', '_', rate)
                rate_t = re.sub('\.', '_', rate_t)
                rate_e = rate_t + '_sens'
                sens_spec = eval('target_' + rate_e)
                if sens > sens_spec:
                    result = 'Fail'
                else:
                    result = 'Pass'
                with open('./Result/' + rx_result_name, 'a+', newline='') as f2:
                    writer2 = csv.writer(f2)
                    writer2.writerow([channel, rate, chain, sens, per, result])
                logger.info('*************************************************************')
    logger.info('************************TEST DONE****************************')
    dt.close()
    iq.close()
