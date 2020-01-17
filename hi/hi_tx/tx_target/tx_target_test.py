# -*- coding: utf-8 -*-
# @Time    : 2019/1/18 13:23
# @Author  : Ethan
# @FileName: tx_test.py

from __future__ import division
from openpyxl import load_workbook
import visa
import telnetlib
import pandas as pd
import time
import datetime
from colorama import init, Fore, Style
import csv
import re
import logging


class IQxel():
    def __init__(self, ip, visaDLL=None, *args):
        self.ip = ip
        # self.visaDLL = 'C:\windows\system32\visa64.dll' if visaDLL is None else visaDLL
        self.address = 'TCPIP0::%s::hislip0::INSTR' % self.ip
        # self.resourceManager = visa.ResourceManager(self.visaDLL)
        self.resourceManager = visa.ResourceManager()

    def open(self):
        self.instance = self.resourceManager.open_resource(self.address)

    def close(self):
        if self.instance is not None:
            self.instance.close()
            self.instance = None

    def __formats__(self, format_spec):
        self.format_spec = format_spec
        format_spec = float(format_spec)
        format_spec = '{:.3f}'.format(format_spec)
        return format_spec

    def __log__(self, switch, log):
        if switch == 1:
            print(log)
        else:
            # print('no log')
            pass

    def __isset__(self, v):
        try:
            type(eval(v))
        except:
            return 0
        else:
            return 1

    def save_image(self, imagname, fmt):
        assert fmt in ['jpg', 'png'], 'Invalid postfix of image'
        print('MMEM:STOR:IMAG "%s.%s"' % (imagname, fmt))
        self.instance.write('MMEM:STOR:IMAG "%s.%s"' % (imagname, fmt))

    def reset(self):
        self.instance.write('*RST')

    def read_idn(self):
        idn = self.instance.query('*IDN?')
        print(idn)
        # return idn

    def init(self):
        self.instance.write('ROUT1;PORT:RES RF1,OFF')
        self.instance.write('ROUT1;PORT:RES RF2,OFF')

        # print('iq init...')

    def read_pathloss(self, get_freqs, get_chains):
        self.get_freqs = get_freqs
        self.get_chains = get_chains
        pathloss = pd.read_csv('./pathloss.csv')
        self.__log__(log_switch, pathloss)
        freqlist = pathloss['Frequency']
        # print(freqlist)
        lens = len(freqlist)
        # print(lens)
        for x in range(lens):
            # print(x)
            # print(pathloss.loc[x,'Frequency'])
            if get_freqs == pathloss.loc[x, 'Frequency']:
                loss = pathloss.loc[x]  # get frequency loss list
                print(loss)
                loss = loss[get_chains]  # get chain loss value
                print('Channel:', get_freqs, 'Chain:', get_chains, 'pathloss:', loss)
                break
            else:
                pass
        return loss
        # freq = freqlist[0]
        # print (freq)
        # chaina = pathloss['1']
        # print(chaina)
        # chaina_2412 = chaina[0]
        # print (chaina_2412)
        # chainb = pathloss['2']
        # print(chainb)
        # chainb_2412 = chaina[0]
        # print(chaina_2412)
        # print(pathloss.loc[1, ['Frequency', '1','2','3','4']])
        # print(pathloss.loc[1, 'Frequency'])
        # print(pathloss.loc[1,'1'])

    def set_pathloss(self):
        pathloss = pd.read_csv('./pathloss.csv')
        self.__log__(log_switch, pathloss)
        freqlist = pathloss['Frequency']
        # print(freqlist)
        lens = len(freqlist)
        # print(lens)
        # creat loss table
        self.instance.write('MEM:TABLE "1"; MEM:TABLE:DEFINE "FREQ,LOSS"')
        for x in range(lens):
            channel = pathloss.loc[x, 'Frequency']
            loss = pathloss.loc[x, '1']
            # print(loss)
            self.instance.write('MEM:TABLE "1";MEMory:TABLe:INSert:POINt %f MHz,%f' % (channel, loss))
        self.instance.write('MEMory:TABLe "1";MEMory:TABLe:STORe')
        self.instance.write('MEM:TABLE "2"; MEM:TABLE:DEFINE "FREQ,LOSS"')
        for x in range(lens):
            channel = pathloss.loc[x, 'Frequency']
            loss = pathloss.loc[x, '2']
            # print(loss)
            self.instance.write('MEM:TABLE "2";MEMory:TABLe:INSert:POINt %f MHz,%f' % (channel, loss))
        self.instance.write('MEMory:TABLe "2";MEMory:TABLe:STORe')
        self.instance.write('MEM:TABLE "3"; MEM:TABLE:DEFINE "FREQ,LOSS"')
        for x in range(lens):
            channel = pathloss.loc[x, 'Frequency']
            loss = pathloss.loc[x, '3']
            # print(loss)
            self.instance.write('MEM:TABLE "3";MEMory:TABLe:INSert:POINt %f MHz,%f' % (channel, loss))
        self.instance.write('MEMory:TABLe "3";MEMory:TABLe:STORe')
        self.instance.write('MEM:TABLE "4"; MEM:TABLE:DEFINE "FREQ,LOSS"')
        for x in range(lens):
            channel = pathloss.loc[x, 'Frequency']
            loss = pathloss.loc[x, '4']
            # print(loss)
            self.instance.write('MEM:TABLE "4";MEMory:TABLe:INSert:POINt %f MHz,%f' % (channel, loss))
        self.instance.write('MEMory:TABLe "4";MEMory:TABLe:STORe')

    def read_data(self):
        loss_path = 1
        idns = self.instance.query('*IDN?')
        # print(idns)
        iq_port_model = iq_port.isdigit()
        # print(iq_port_model)
        iq_model = idns.split(',')
        # print(iq_model)
        iq_model = iq_model[1]
        # print(iq_model)
        if iq_model == 'IQXEL' and iq_port_model is True:
            mw = ''
        elif iq_model == 'IQXEL-M' and iq_port_model is False:
            mw = ''
        elif iq_model == 'IQXEL-MW' and iq_port_model is False:
            mw = 'M'
        elif chain == '0':
            loss_path = 1
        elif chain == '1':
            loss_path = 2
        self.instance.write('%sVSA%s;RFC:USE "%s",RF%s' % (mw, iq_rout, loss_path, iq_port))
        self.instance.write('%sVSA%s;RFC:STAT  ON,RF%s' % (mw, iq_rout, iq_port))
        # rint(mw, iq_rout, iq_port, iq_rout)
        self.instance.write('%sROUT%s;PORT:RES RF%s,VSA%s' % (mw, iq_rout, iq_port, iq_rout))
        self.instance.write('CHAN1;WIFI')
        channels = channel * 1000000
        self.instance.write('%sVSA%s;FREQ:cent %d' % (mw, iq_rout, channels))
        self.instance.write('%sVSA%s;SRAT 160000000' % (mw, iq_rout))
        self.instance.write('%sVSA%s ;RLEVel:AUTO' % (mw, iq_rout))
        # rlevel = target_power + 15
        # print(rlevel)
        # time.sleep(sleep_time)
        # self.instance.write('%sVSA%s;RLEV %d;*wai;*opc?' % (mw, iq_rout, rlevel))

        # self.__log__(log_switch, chain)

        if mode == '11b':
            # print('DSSS')
            self.instance.write('CHAN1;WIFI;CONF:STAN DSSS')
            self.instance.write('VSA%s;CAPT:TIME 0.03' % iq_rout)
            self.instance.write('CHAN1')
            self.instance.write('VSA%s ;init' % iq_rout)
            self.instance.write('WIFI')
            self.instance.write('calc:pow 0, 2')
            self.instance.write('calc:txq 0, 2')
            self.instance.write('calc:ccdf 0, 2')
            self.instance.write('calc:ramp 0, 2')
            self.instance.write('calc:spec 0, 2')
            # print(rec)
            time.sleep(1)
            # power
            data = self.instance.query('WIFI;FETC:SEGM:POW:AVER?')
            self.__log__(log_switch, data)
        else:
            # print('OFDM')
            self.instance.write('CHAN1;WIFI;CONF:STAN OFDM')
            self.instance.write('CHAN1;WIFI;CONF:OFDM:CEST DATA')
            self.instance.write('VSA%s;CAPT:TIME 0.005' % iq_rout)
            self.instance.write('CHAN1')
            self.instance.write('VSA%s ;init' % iq_rout)
            self.instance.write('WIFI')
            self.instance.write('calc:pow 0, 2')
            self.instance.write('calc:txq 0, 2')
            self.instance.write('calc:ccdf 0, 2')
            self.instance.write('calc:ramp 0, 2')
            self.instance.write('calc:spec 0, 2')
            data = self.instance.query('WIFI;FETC:SEGM:POW:AVER?')
            self.__log__(log_switch, data)

    def get_data(self, mode, channel, rate, chain, result_name, target_power, spec_pwr, gain, spec_evm, evm_margin,
                 spec_symbol_clock_error, spec_lo_leakage, spec_mask, spec_obw_20M, spec_obw_40M):
        if mode == '11b':
            data = self.instance.query('WIFI;FETC:SEGM:POW:AVER?')
            self.__log__(log_switch, data)
            data_len = len(data)
            self.__log__(log_switch, data_len)
            data = data.split(',')
            if data_len < 3:
                print('Error', data)
                avg_power = 'NA'
                result_pwr = 'NA'
            else:
                # self.__log__(log_switch, data)
                avg_power = data[1]
                # print(avg_power)
                avg_power = self.__formats__(avg_power)
                delta_pwr = abs(float(avg_power) - target_power)
                # print(delta_pwr)
                # spec_pwr = 2
                if delta_pwr > spec_pwr:
                    print('Power:               ', Fore.RED + avg_power + Style.RESET_ALL, 'dBm')
                    result_pwr = 'Fail'
                else:
                    print('Power:               ', Fore.BLUE + avg_power + Style.RESET_ALL, 'dBm')
                    result_pwr = 'Pass'
            # txquality
            data = self.instance.query('WIFI;FETC:SEGM:TXQ:DSSS:AVER?')
            time.sleep(sleep_time)
            self.__log__(log_switch, data)
            data_len = len(data)
            self.__log__(log_switch, data_len)
            data = data.split(',')
            if data_len < 100:
                print('Error', data)
                avg_evm = 'NA'
                result_evm = 'NA'
            else:
                avg_evm = data[1]
                avg_evm = self.__formats__(avg_evm)
                peak_evm = data[2]
                peak_evm = self.__formats__(peak_evm)
                freq_error = data[4]
                freq_error = self.__formats__(freq_error)
                peak_freq_error = data[5]
                peak_freq_error = self.__formats__(peak_freq_error)
                symbol_clock_error = data[6]
                symbol_clock_error = self.__formats__(symbol_clock_error)
                lo_leakage = data[7]
                lo_leakage = self.__formats__(lo_leakage)
                # rate = rate + 'EVM'
                spec_evm = spec_evm - evm_margin
                if float(avg_evm) > spec_evm:
                    print('EVM:                 ', Fore.RED + avg_evm + Style.RESET_ALL, 'dB')
                    result_evm = 'Fail'
                else:
                    print('EVM:                 ', Fore.BLUE + avg_evm + Style.RESET_ALL, 'dB')
                    result_evm = 'Pass'
                print('EVM Peak:            ', peak_evm, 'dB')
                print('Frequency Error:     ', freq_error, 'kHz')
                print('Frequency Error Peak:', peak_freq_error, 'kHz')
                # spec_symbol_clock_error = eval('target_evm_'+rate)
                if abs(float(symbol_clock_error)) > abs(float(spec_symbol_clock_error)):
                    print('Symbol Clock Error:  ', Fore.RED + symbol_clock_error + Style.RESET_ALL, 'ppm')
                    result_symbol_clock_error = 'Fail'
                else:
                    print('Symbol Clock Error:  ', Fore.BLUE + symbol_clock_error + Style.RESET_ALL, 'ppm')
                    result_symbol_clock_error = 'Pass'
                # spec_lo_leakage = -15
                if abs(float(lo_leakage)) < abs(float(spec_lo_leakage)):
                    print('LO Leakage:          ', Fore.RED + lo_leakage + Style.RESET_ALL, 'dB')
                    result_lo_leakage = 'Fail'
                else:
                    print('LO Leakage:          ', Fore.BLUE + lo_leakage + Style.RESET_ALL, 'dB')
                    result_lo_leakage = 'Pass'
            # mask
            data = self.instance.query('WIFI;FETC:SEGM:SPEC:AVER:VIOL?')
            status = self.instance.query('*opc?')
            status = status.split(',')
            status = status[0]
            self.__log__(log_switch, status)
            while status == '0':
                print('...')
                status = self.instance.query('*opc?')
                status = status.split(',')
                status = status[0]
            time.sleep(sleep_time)
            self.__log__(log_switch, data)
            data_len = len(data)
            self.__log__(log_switch, data_len)
            data = data.split(',')
            if data_len < 3:
                print('Error', data)
            else:
                # self.__log__(log_switch, data)
                mask = data[1]
                mask = self.__formats__(mask)
                # spec_mask = 5.12
                if abs(float(mask)) > spec_mask:
                    print('Mask:                ', Fore.RED + mask + Style.RESET_ALL, '%')
                    result_mask = 'Fail'
                else:
                    print('Mask:                ', Fore.BLUE + mask + Style.RESET_ALL, '%')
                    result_mask = 'Pass'
            # OBW
            data = self.instance.query('WIFI;FETC:SEGM:SPEC:AVER:OBW?')
            status = self.instance.query('*opc?')
            self.__log__(log_switch, status)
            status = status.split(',')
            status = status[0]
            while status == '0':
                print('...')
                status = self.instance.query('*opc?')
                status = status.split(',')
                status = status[0]
            time.sleep(sleep_time)
            self.__log__(log_switch, data)
            data_len = len(data)
            self.__log__(log_switch, data_len)
            data = data.split(',')
            if data_len < 3:
                print('Error', data)
            else:
                # self.__log__(log_switch, data)
                obw = data[1]
                obw = self.__formats__(obw)
                obw = round(float(obw) / 1000000, 3)
                spec_obw = spec_obw_20M
                if obw > spec_obw:
                    print('OBW:                 ', Fore.RED + str(obw) + Style.RESET_ALL, 'MHz')
                    result_obw = 'Fail'
                else:
                    print('OBW:                 ', Fore.BLUE + str(obw) + Style.RESET_ALL, 'MHz')
                    result_obw = 'Pass'
            # RAMP
            # on time
            data = self.instance.query('WIFI;FETC:SEGM:RAMP:ON:TRIS?')
            status = self.instance.query('*opc?')
            self.__log__(log_switch, status)
            status = status.split(',')
            status = status[0]
            while status == '0':
                print('...')
                status = self.instance.query('*opc?')
                status = status.split(',')
                status = status[0]
            time.sleep(sleep_time)
            self.__log__(log_switch, data)
            data_len = len(data)
            self.__log__(log_switch, data_len)
            data = data.split(',')
            if data_len < 3:
                print('Error', data)
            else:
                # self.__log__(log_switch, data)
                ramp_on_time = data[1]
                ramp_on_time = float(ramp_on_time) * 1000000
                ramp_on_time = self.__formats__(ramp_on_time)
                spec_ramp_on_time = 2.0
                if float(ramp_on_time) > spec_ramp_on_time:
                    print('Ramp On Time:        ', Fore.RED + ramp_on_time + Style.RESET_ALL, 'us')
                    result_ramp_on_time = 'Fail'
                else:
                    print('Ramp On Time:        ', Fore.BLUE + ramp_on_time + Style.RESET_ALL, 'us')
                    result_ramp_on_time = 'Pass'
            # off time
            data = self.instance.query('WIFI;FETC:SEGM:RAMP:OFF:TRIS?')
            status = self.instance.query('*opc?')
            self.__log__(log_switch, status)
            status = status.split(',')
            status = status[0]
            while status == '0':
                print('...')
                status = self.instance.query('*opc?')
                status = status.split(',')
                status = status[0]
            time.sleep(sleep_time)
            self.__log__(log_switch, data)
            data_len = len(data)
            self.__log__(log_switch, data_len)
            data = data.split(',')
            if data_len < 3:
                print('Error', data)
            else:
                # self.__log__(log_switch, data)
                ramp_off_time = data[1]
                ramp_off_time = float(ramp_off_time) * 1000000
                ramp_off_time = self.__formats__(ramp_off_time)
                spec_ramp_off_time = 2.0
                if float(ramp_off_time) > spec_ramp_off_time:
                    print('Ramp Off Time:       ', Fore.RED + ramp_off_time + Style.RESET_ALL, 'us')
                    result_ramp_off_time = 'Fail'
                else:
                    print('Ramp Off Time:       ', Fore.BLUE + ramp_off_time + Style.RESET_ALL, 'us')
                    result_ramp_off_time = 'Pass'
            flatness = 'NA'
            spec_flatness = 'NA'
            result_flatness = 'NA'

        else:
            # print('OFDM')
            data = self.instance.query('WIFI;FETC:SEGM:POW:AVER?')
            time.sleep(sleep_time)
            self.__log__(log_switch, data)
            data_len = len(data)
            self.__log__(log_switch, data_len)
            data = data.split(',')
            if data[0] != '0' or data_len < 3:
                print('Error', data)
                avg_power = 'NA'
                result_pwr = 'NA'
            else:
                self.__log__(log_switch, data)
                avg_power = data[1]
                avg_power = self.__formats__(avg_power)
                delta_pwr = abs(float(avg_power) - target_power)
                # print(delta_pwr)
                ##spec_pwr = 2
                if delta_pwr > spec_pwr:
                    print('Power:               ', Fore.RED + avg_power + Style.RESET_ALL, 'dBm')
                    result_pwr = 'Fail'
                else:
                    print('Power:               ', Fore.BLUE + avg_power + Style.RESET_ALL, 'dBm')
                    result_pwr = 'Pass'
            # txquality
            data = self.instance.query('WIFI;FETC:SEGM:TXQ:OFDM:AVER?')
            time.sleep(sleep_time)
            self.__log__(log_switch, data)
            data_len = len(data)
            self.__log__(log_switch, data_len)
            data = data.split(',')
            if data_len < 100:
                print('Error', data)
                avg_evm = 'NA'
                result_evm = 'NA'
            else:
                self.__log__(log_switch, data)
                avg_evm = data[1]
                avg_evm = self.__formats__(avg_evm)
                freq_error = data[3]
                freq_error = self.__formats__(freq_error)
                symbol_clock_error = data[4]
                symbol_clock_error = self.__formats__(symbol_clock_error)
                lo_leakage = data[5]
                lo_leakage = self.__formats__(lo_leakage)
                spec_evm = spec_evm - abs(evm_margin)
                # print(spec_evm)
                if float(avg_evm) > spec_evm:
                    print('EVM:                 ', Fore.RED + avg_evm + Style.RESET_ALL, 'dB')
                    result_evm = 'Fail'
                else:
                    print('EVM:                 ', Fore.BLUE + avg_evm + Style.RESET_ALL, 'dB')
                    result_evm = 'Pass'
                print('Frequency Error:     ', freq_error, 'kHz')
                # spec_symbol_clock_error = 10.0
                if abs(float(symbol_clock_error)) > abs(float(spec_symbol_clock_error)):
                    print('Symbol Clock Error:  ', Fore.RED + symbol_clock_error + Style.RESET_ALL, 'ppm')
                    result_symbol_clock_error = 'Fail'
                else:
                    print('Symbol Clock Error:  ', Fore.BLUE + symbol_clock_error + Style.RESET_ALL, 'ppm')
                    result_symbol_clock_error = 'Pass'
                # spec_lo_leakage = -15
                if abs(float(lo_leakage)) < abs(float(spec_lo_leakage)):
                    print('LO Leakage:          ', Fore.RED + lo_leakage + Style.RESET_ALL, 'dB')
                    result_lo_leakage = 'Fail'
                else:
                    print('LO Leakage:          ', Fore.BLUE + lo_leakage + Style.RESET_ALL, 'dB')
                    result_lo_leakage = 'Pass'
            # mask
            data = self.instance.query('WIFI;FETC:SEGM:SPEC:AVER:VIOL?')
            status = self.instance.query('*opc?')
            self.__log__(log_switch, status)
            status = status.split(',')
            status = status[0]
            while status == '0':
                print('...')
                status = self.instance.query('*opc?')
                status = status.split(',')
                status = status[0]
            time.sleep(sleep_time)
            self.__log__(log_switch, data)
            data_len = len(data)
            self.__log__(log_switch, data_len)
            data = data.split(',')
            if data_len < 3:
                print('Error', data)
            else:
                # self.__log__(log_switch, data)
                mask = data[1]
                mask = self.__formats__(mask)
                # spec_mask = 5.12
                if abs(float(mask)) > spec_mask:
                    print('Mask:                ', Fore.RED + mask + Style.RESET_ALL, '%')
                    result_mask = 'Fail'
                else:
                    print('Mask:                ', Fore.BLUE + mask + Style.RESET_ALL, '%')
                    result_mask = 'Pass'
            # OBW
            data = self.instance.query('WIFI;FETC:SEGM:SPEC:AVER:OBW?')
            status = self.instance.query('*opc?')
            self.__log__(log_switch, status)
            status = status.split(',')
            status = status[0]
            while status == '0':
                print('...')
                status = self.instance.query('*opc?')
                status = status.split(',')
                status = status[0]
            time.sleep(sleep_time)
            self.__log__(log_switch, data)
            data_len = len(data)
            self.__log__(log_switch, data_len)
            data = data.split(',')
            if data_len < 3:
                print('Error', data)
            else:
                # self.__log__(log_switch, data)
                obw = data[1]
                obw = self.__formats__(obw)
                obw = round(float(obw) / 1000000, 3)
                if mode == '11ng40plus':
                    spec_obw = spec_obw_40M
                else:
                    spec_obw = spec_obw_20M
                if obw > spec_obw:
                    print('OBW:                 ', Fore.RED + str(obw) + Style.RESET_ALL, 'MHz')
                    result_obw = 'Fail'
                else:
                    print('OBW:                 ', Fore.BLUE + str(obw) + Style.RESET_ALL, 'MHz')
                    result_obw = 'Pass'
            # flatness
            # datas = self.instance.query('FETC:SEGM1:OFDM:SFL:SIGN1:AVER?')
            # time.sleep(sleep_time)
            # time.sleep(10)
            # print('flatness', datas)
            data = self.instance.query('WIFI;FETC:SEGM:OFDM:SFL:AVER:CHEC?')
            status = self.instance.query('*opc?')
            self.__log__(log_switch, status)
            status = status.split(',')
            status = status[0]
            while status == '0':
                print('...')
                status = self.instance.query('*opc?')
                status = status.split(',')
                status = status[0]
            time.sleep(sleep_time)
            self.__log__(log_switch, data)
            data_len = len(data)
            self.__log__(log_switch, data_len)
            data = data.split(',')
            if data_len < 3:
                print('Error', data)
            else:
                # self.__log__(log_switch, data)
                flatness = data[0]
                spec_flatness = 0
                if int(flatness) == spec_flatness:
                    print('Flatness:            ', Fore.BLUE + flatness + Style.RESET_ALL)
                    result_flatness = 'Pass'
                else:
                    print('Flatness:            ', Fore.RED + flatness + Style.RESET_ALL)
                    result_flatness = 'Fail'
            ramp_on_time = 'NA'
            spec_ramp_on_time = 'NA'
            result_ramp_on_time = 'NA'
            ramp_off_time = 'NA'
            spec_ramp_off_time = 'NA'
            result_ramp_off_time = 'NA'

        # if mode == '11b' or mode == '11g':
        #    rates = 'M'
        #    rate = rate+rates
        # else:
        #    rates = 'MCS'
        #    rate = rates+rate
        # rate = mode+'_'+rate
        avg_power = float(avg_power)
        if result_pwr == 'Fail' or result_evm == 'Fail' or result_symbol_clock_error == 'Fail' or \
                result_lo_leakage == 'Fail' or result_mask == 'Fail':
            pass
        elif avg_power < targetpower:
            pass
        else:
            with open('./Result/' + result_name, 'a+', newline='') as f2:
                writer2 = csv.writer(f2)
                writer2.writerow([channel, rate, chain, target_power, avg_power, gain, spec_pwr, result_pwr,
                                  avg_evm, spec_evm, result_evm, symbol_clock_error, spec_symbol_clock_error,
                                  result_symbol_clock_error, lo_leakage, spec_lo_leakage, result_lo_leakage, obw,
                                  spec_obw,
                                  result_obw, mask, spec_mask, result_mask, flatness, spec_flatness, result_flatness,
                                  ramp_on_time, spec_ramp_on_time, result_ramp_on_time, ramp_off_time,
                                  spec_ramp_off_time,
                                  result_ramp_off_time])

        return avg_power, result_evm, result_symbol_clock_error, result_lo_leakage, result_mask

    def vsg(self):
        loss_path = 1
        idns = self.instance.query('*IDN?')
        # print(idns)
        iq_port_model = iq_port.isdigit()
        # print(iq_port_model)
        iq_model = idns.split(',')
        # print(iq_model)
        iq_model = iq_model[1]
        # print(iq_model)
        if iq_model == 'IQXEL' and iq_port_model is True:
            mw = ''
        elif iq_model == 'IQXEL-M' and iq_port_model is False:
            mw = ''
        elif iq_model == 'IQXEL-MW' and iq_port_model is False:
            mw = 'M'
        elif chain == '0':
            loss_path = '1'
        elif chain == '1':
            loss_path = '2'
        elif chain == '3':
            loss_path = '3'
        self.instance.write('%sVSG%s;RFC:USE "%s",RF%s' % (mw, iq_rout, loss_path, iq_port))
        self.instance.write('%sVSG%s;RFC:STAT  ON,RF%s' % (mw, iq_rout, iq_port))
        # rint(mw, iq_rout, iq_port, iq_rout)
        self.instance.write('%sROUT%s;PORT:RES RF%s,VSG%s' % (mw, iq_rout, iq_port, iq_rout))
        self.instance.write('CHAN1;WIFI')
        channels = int(channel * 1000000)
        self.instance.write('%sVSG%s;FREQ:cent %d' % (mw, iq_rout, channels))
        self.instance.write('%sVSG%s;SRAT 160000000' % (mw, iq_rout))
        # loss = abs(loss)
        # print(type(start), type(loss))
        # rlevel = start + loss
        rlevel = start
        # print(start, rlevel)
        time.sleep(sleep_time)
        self.instance.write('%sVSG%s;POW:lev %d' % (mw, iq_rout, rlevel))
        self.instance.write('VSG1;POW:STAT ON')
        self.instance.write('VSG1;MOD:STAT ON')
        self.instance.write('VSG1;WAVE:EXEC OFF')
        self.instance.write('VSG1;WLIS:COUN %d' % int(rx_packets))
        # wave = 'OFDM-6'
        if mode == '11b' and rates == '1':
            wave = 'DSSS-1L'
            vsg_delay = sleep_time + int(rx_packets) / 100 + 5 - int(rates) * int(rx_packets) / 1000
        elif mode == '11b' and rates == '2':
            wave = 'DSSS-2L'
            vsg_delay = sleep_time + int(rx_packets) / 100 - int(rates) * int(rx_packets) / 1000
        elif mode == '11b' and rates == '5.5':
            wave = 'CCK-5_5S'
            vsg_delay = sleep_time + int(rx_packets) / 100 - int(float(rates) * int(rx_packets) / 1000)
        elif mode == '11b' and rates == '11':
            wave = 'CCK-11S'
            vsg_delay = sleep_time + int(rx_packets) / 100 - int(rates) * int(rx_packets) / 1800
        elif mode == '11g' or mode == '11a':
            wave = 'OFDM-' + rates
            vsg_delay = sleep_time + int(rx_packets) / 100 - int(rates) * int(rx_packets) / 6000
        elif mode == '11n':
            wave = 'HT' + bw + '_MCS' + rates
            vsg_delay = sleep_time + int(rx_packets) / 100 - int(rates) * int(rx_packets) / 1000
        elif mode == '11ac':
            wave = '11AC_VHT' + bw + '_S1_MCS' + rates
            vsg_delay = sleep_time + int(rx_packets) / 100 - int(rates) * int(rx_packets) / 1000
        # print(wave)
        print('Delay Time:', vsg_delay)
        self.instance.write('VSG1; WAVE:LOAD "/user/WiFi_%s.iqvsg"' % wave)
        self.instance.write('VSG1 ;wave:exec off')
        self.instance.write('WLIST:WSEG1:DATA "/user/WiFi_%s.iqvsg"' % wave)
        self.instance.write('wlist:wseg1:save')
        self.instance.write('WLIST:COUNT:ENABLE WSEG1')
        self.instance.write('WAVE:EXEC ON, WSEG1')
        time.sleep(vsg_delay)
        self.instance.write('VSG1 ;WAVE:EXEC OFF')
        self.instance.write('WLIST:COUNT:DISABLE WSEG1')


class dut():
    def __init__(self):
        self.tn = telnetlib.Telnet()
        self.tn.set_debuglevel(log_switch)

    def close(self):
        if self.tn is not None:
            self.tn.close()
            self.tn = None

    def login(self, host, username, password):
        self.tn.open(host, port=23)
        time.sleep(1)
        command_result = self.tn.read_very_eager().decode('ascii')
        print(command_result)
        self.tn.read_until(b'Login: ', timeout=1)
        self.tn.write(username.encode('ascii') + b'\n')
        self.tn.read_until(b'Password:', timeout=1)
        self.tn.write(password.encode('ascii') + b'\n')

        time.sleep(1)
        command_result = self.tn.read_very_eager().decode('ascii')
        print(command_result)
        if 'wrong' not in command_result:
            logging.info('%s Sign up' % host)
            return True
        else:
            logging.warning('%s Login Fail' % host)
            return False

    def init(self, str1, str2, str3):
        self.tn.read_until(b'WAP>', timeout=1)
        self.tn.write(str1.encode('ascii') + b'\n')

        self.tn.read_until(b'SU_WAP>', timeout=1)
        self.tn.write(str2.encode('ascii') + b'\n')

        self.tn.read_until(b'WAP(Dopra Linux) # ', timeout=1)
        self.tn.write(str3.encode('ascii') + b'\n')

        self.tn.read_until(b'WAP(Dopra Linux) # ', timeout=1)
        self.tn.write(b'iwpriv Hisilicon0 adjust_ppm -10' + b'\n')

        # command_result = self.tn.read_very_eager().decode('ascii')
        # logging.info('\n%s' % command_result)

    def ex_command(self, command):
        self.tn.read_until(b'WAP(Dopra Linux) # ', timeout=1)
        self.tn.write(command.encode('ascii') + b'\n')
        time.sleep(2)
        command_result = self.tn.read_very_eager().decode('ascii')
        logging.info('\n%s' % command_result)

    def tx(self, mode, channel, bw, rate, chain):
        if int(channel) < 5000:
            band = id_2g
        else:
            band = id_5g
        # for 2.4g
        self.tn.read_until(b'WAP(Dopra Linux) # ', timeout=1)
        self.tn.write(b'\r\n')
        self.tn.read_until(b'WAP(Dopra Linux) # ', timeout=1)
        self.tn.write(b'ifconfig vap%s down' % band.encode('ascii') + b'\n')
        self.tn.read_until(b'WAP(Dopra Linux) # ', timeout=1)
        self.tn.write(b'iwpriv vap%s setessid Hi24g' % band.encode('ascii') + b'\n')
        self.tn.read_until(b'WAP(Dopra Linux) # ', timeout=1)
        self.tn.write(b'iwpriv vap%s mode %s' % (band.encode('ascii'), mode.encode('ascii')) + b'\n')
        # print(channel)
        if bw == '40':
            channels = int(channel) - 10
            # print(channel)
        else:
            channels = channel
        channels = '{:.0f}'.format((int(channels) - 2407) / 5)  # 2407+5*1=2412
        # print(channel)'{:.3f}'.format(delta_power)
        channels = str(channels)
        # print(channel)
        self.tn.read_until(b'WAP(Dopra Linux) # ', timeout=1)
        self.tn.write(b'iwpriv vap%s channel %s' % (band.encode('ascii'), channels.encode('ascii')) + b'\n')
        self.tn.read_until(b'WAP(Dopra Linux) # ', timeout=1)
        self.tn.write(b'iwpriv vap%s privflag 1' % band.encode('ascii') + b'\n')
        self.tn.read_until(b'WAP(Dopra Linux) # ', timeout=1)
        self.tn.write(b'iwpriv vap%s al_rx 0' % band.encode('ascii') + b'\n')
        self.tn.read_until(b'WAP(Dopra Linux) # ', timeout=1)
        self.tn.write(b'hipriv.sh "vap%s 2040bss_enable 0"' % band.encode('ascii') + b'\n')
        self.tn.read_until(b'WAP(Dopra Linux) # ', timeout=1)
        self.tn.write(b'hipriv.sh "vap%s radartool enable 0"' % band.encode('ascii') + b'\n')
        self.tn.read_until(b'WAP(Dopra Linux) # ', timeout=1)
        self.tn.write(b'hipriv.sh "vap%s acs sw 0"' % band.encode('ascii') + b'\n')
        self.tn.read_until(b'WAP(Dopra Linux) # ', timeout=1)
        self.tn.write(b'ifconfig vap%s up' % band.encode('ascii') + b'\n')
        self.tn.read_until(b'WAP(Dopra Linux) # ', timeout=1)
        self.tn.write(b'iwpriv vap%s al_tx 0' % band.encode('ascii') + b'\n')
        self.tn.read_until(b'WAP(Dopra Linux) # ', timeout=1)
        # self.tn.write(b'hipriv.sh "vap%s add_user 07:06:05:04:03:02 1"\r')
        self.tn.write(b'iwpriv vap%s bw %s' % (band.encode('ascii'), bw.encode('ascii')) + b'\n')
        if mode == '11b' or mode == '11g':
            rates = 'rate'
        else:
            rates = 'mcs'
        self.tn.read_until(b'WAP(Dopra Linux) # ', timeout=1)
        self.tn.write(
            b'iwpriv vap%s %s %s' % (band.encode('ascii'), rates.encode('ascii'), rate.encode('ascii')) + b'\n')
        if chain == '0':
            chains = '01'
        else:
            chains = '10'
        # print('chain set',chain,type(chain),chains)
        self.tn.read_until(b'WAP(Dopra Linux) # ', timeout=1)
        self.tn.write(b'iwpriv vap%s txch 00%s' % (band.encode('ascii'), chains.encode('ascii')) + b'\n')
        self.tn.read_until(b'WAP(Dopra Linux) # ', timeout=1)
        self.tn.write(b'iwpriv vap%s al_tx "1 2 1000"' % band.encode('ascii') + b'\n')
        self.tn.read_until(b'WAP(Dopra Linux) # ', timeout=1)
        self.tn.write(b'hipriv.sh "vap%s set_tx_pow rf_reg_ctl 0"' % band.encode('ascii') + b'\n')
        print('TX COMMANDS DONE')

    def get_paras(self):
        if int(channel) < 5000:
            band = id_2g
        else:
            band = id_5g
        # self.tn.read_until(b'WAP(Dopra Linux) # ', timeout=1)
        # self.tn.write(b'hipriv.sh "vap%s set_tx_pow rf_reg_ctl 0"' % band.encode('ascii') + b'\n')
        self.tn.read_until(b'WAP(Dopra Linux) # ', timeout=1)
        self.tn.write(b'hipriv.sh "vap%s set_tx_pow rf_reg_ctl 1"' % band.encode('ascii') + b'\n')
        self.tn.read_until(b'WAP(Dopra Linux) # ', timeout=1)
        self.tn.write(b'hipriv.sh "vap%s reginfo soc 0x200380%s 0x200380%s";dmesg -c'
                      % (band.encode('ascii'), channel_groups.encode('ascii'), channel_groups.encode('ascii')) + b'\n')
        # self.tn.read_until(b'WAP(Dopra Linux) # ', timeout=1)
        # command_result = self.tn.read_very_eager().decode('ascii')
        command_result = self.tn.read_until(b'value=0x\w+\r\r\nWAP(Dopra Linux) # ', timeout=2)
        command_result = re.search(b'value=0x\w+', command_result)
        default_value = re.sub('value=0x', '', command_result.group().decode('ascii'))
        print('Default Value(HEX):', default_value)
        default_value = int(default_value, 16)
        return default_value
        # print(default_value)
        # pwr_para = default_value + 2
        # pwr_paras = h#ex(pwr_para)
        ##print(pwr_paras)
        # print(channel_groups, pwr_paras)

    def adjust_power(self):
        if int(channel) < 5000:
            band = id_2g
        else:
            band = id_5g
        self.tn.read_until(b'WAP(Dopra Linux) # ', timeout=1)
        self.tn.write(b'hipriv.sh "vap%s regwrite soc 0x200380%s 0x%s"'
                      % (band.encode('ascii'), channel_groups.encode('ascii'), pwr_paras.encode('ascii'))+ b'\n')
        self.tn.read_until(b'WAP(Dopra Linux) # ', timeout=1)
        self.tn.write(b'hipriv.sh "vap%s reginfo soc 0x200380%s 0x200380%s";dmesg -c'
                      % (band.encode('ascii'), channel_groups.encode('ascii'), channel_groups.encode('ascii')) + b'\n')
        command_result = self.tn.read_until(b'value=0x\w+\r\r\nWAP(Dopra Linux) # ', timeout=1)
        command_result = re.search(b'value=0x\w+', command_result)
        default_value = re.sub('value=0x', '', command_result.group().decode('ascii'))
        print('Check Value(HEX):', default_value)
        gain = default_value
        return gain

    def set_default(self):
        if int(channel) < 5000:
            band = id_2g
        else:
            band = id_5g
        self.tn.read_until(b'WAP(Dopra Linux) # ', timeout=1)
        self.tn.write(b'exit' + b'\n')
        self.tn.read_until(b'SU_WAP>', timeout=1)
        self.tn.write(b'wifi calibrate test parameter write wifichip 1 type power len 77 value default' + b'\n')
        time.sleep(2)
        self.tn.read_until(b'SU_WAP>', timeout=1)
        self.tn.write(b'shell' + b'\n')
        self.tn.read_until(b'WAP(Dopra Linux) # ', timeout=1)
        self.tn.write(b'wifi_equipment.sh init chip 1' + b'\n')
        time.sleep(2)
        self.tn.read_until(b'WAP(Dopra Linux) # ', timeout=1)
        self.tn.write(
            b'hipriv.sh "vap%s set_tx_pow rf_reg_ctl 0"' % (band.encode('ascii'), band.encode('ascii'))  + b'\n')

    def rx(self, mode, channel, bw, rate, chain):
        if int(channel) < 5000:
            band = id_2g
        else:
            band = id_5g
        # for 2.4g
        self.tn.read_until(b'WAP(Dopra Linux) # ', timeout=1)
        self.tn.write(b'\r\n')
        self.tn.read_until(b'WAP(Dopra Linux) # ', timeout=1)
        self.tn.write(b'iwpriv vap%s al_tx 0' % band.encode('ascii') + b'\n')
        self.tn.read_until(b'WAP(Dopra Linux) # ', timeout=1)
        self.tn.write(b'iwpriv vap%s al_rx 0' % band.encode('ascii') + b'\n')
        self.tn.read_until(b'WAP(Dopra Linux) # ', timeout=1)
        self.tn.write(b'ifconfig vap%s down' % band.encode('ascii') + b'\n')
        self.tn.read_until(b'WAP(Dopra Linux) # ', timeout=1)
        self.tn.write(b'ifconfig vap%s hw ether 00:E0:52:22:22:14' % band.encode('ascii') + b'\n')
        self.tn.read_until(b'WAP(Dopra Linux) # ', timeout=1)
        self.tn.write(b'ifconfig vap%s al_rx 1' % band.encode('ascii') + b'\n')
        self.tn.read_until(b'WAP(Dopra Linux) # ', timeout=1)
        self.tn.write(b'ifconfig vap%s setessid HS8145V_HW_TEST_2G ' % band.encode('ascii') + b'\n')
        self.tn.read_until(b'WAP(Dopra Linux) # ', timeout=5)
        self.tn.write(b'iwpriv vap%s mode %s' % (band.encode('ascii'), mode.encode('ascii')) + b'\n')
        self.tn.read_until(b'WAP(Dopra Linux) # ', timeout=1)
        self.tn.write(b'iwpriv vap%s bw %s' % (band.encode('ascii'), bw.encode('ascii')) + b'\n')
        if int(channel) < 5000:
            channels = '{:.0f}'.format((int(channel) - 2407) / 5)
        else:
            channels = '{:.0f}'.format((int(channel) - 5000) / 5)
        channels = str(channels)
        # print(channel)
        self.tn.read_until(b'WAP(Dopra Linux) # ', timeout=1)
        self.tn.write(b'iwpriv vap%s channel %s' % (band.encode('ascii'), channels.encode('ascii')) + b'\n')
        if chain == '0':
            chains = '01'
        else:
            chains = '10'
        # print('chain set',chain,type(chain),chains)
        self.tn.read_until(b'WAP(Dopra Linux) # ', timeout=1)
        self.tn.write(b'iwpriv vap%s rxch 00%s' % (band.encode('ascii'), chains.encode('ascii')) + b'\n')
        if mode == '11b' or mode == '11g':
            rates = 'rate'
        else:
            rates = 'mcs'
        self.tn.read_until(b'WAP(Dopra Linux) # ', timeout=1)
        self.tn.write(b'iwpriv vap%s %s %s' % (band.encode('ascii'), rates.encode('ascii'), rate.encode('ascii')) + b'\n')
        self.tn.read_until(b'WAP(Dopra Linux) # ', timeout=1)
        self.tn.write(b'ifconfig vap%s up' % band.encode('ascii') + b'\n')
        print('RX COMMANDS DONE')

    def get_statistics(self):
        if int(channel) < 5000:
            band = id_2g
        else:
            band = id_5g
        self.tn.read_until(b'WAP(Dopra Linux) # ', timeout=1)
        self.tn.write(b'hipriv.sh "vap%s rx_fcs_info 1 2";dmesg -c' % band.encode('ascii') + b'\n')
        command_result = self.tn.read_until(b'rssi', timeout=2)
        # print(command_result)
        command_result = re.search(b'succ:\w+', command_result)
        # print(command_result)
        # print(command_result.group())
        PER_value = re.sub('succ:', '', command_result.group().decode('ascii'))
        print('Packets:', rx_packets, 'PER:', PER_value)
        return PER_value


if __name__ == '__main__':
    init(autoreset=True)

    Number = input("SN:")
    sn = "SN_" + Number

    dut_file = csv.reader(open('./config.csv'))
    for rows in dut_file:
        if rows[0] == 'DUT_IP':
            dut_ip = rows[1]
            print('DUT IP: ', dut_ip)
        elif rows[0] == 'username':
            user = rows[1]
            print('DUT Username: ', user)
        elif rows[0] == 'password':
            pwd = rows[1]
            print('DUT Password: ', pwd)
        elif rows[0] == 'addition1':
            add1 = rows[1]
            print('Addition Parameters: ', add1)
        elif rows[0] == 'addition2':
            add2 = rows[1]
            print('Addition Parameters: ', add2)
        elif rows[0] == 'addition3':
            add3 = rows[1]
            print('Addition Parameters: ', add3)
        elif rows[0] == 'log_enable':
            le = rows[1]
            print('Print Log: ', le)
        elif rows[0] == 'sleep_time':
            st = rows[1]
            print('Sleep time: ', st)
        elif rows[0] == 'IQ_IP':
            iq_ip = rows[1]
            print('IQ time: ', iq_ip)
        elif rows[0] == 'IQ_Port':
            iq_port = rows[1]
            print('IQ PORT: ', iq_port)
        elif rows[0] == 'IQ_Rout':
            iq_rout = rows[1]
            print('IQ ROUT: ', iq_rout)
        elif rows[0] == 'RX_Packets':
            rx_packets = rows[1]
            print('RX Packets: ', rx_packets)
        elif rows[0] == '2G_ID':
            id_2g = rows[1]
            print('2G ID: ', id_2g)
        elif rows[0] == '5G_ID':
            id_5g = rows[1]
            print('5G ID: ', id_5g)

    log_switch = int(le)
    sleep_time = float(st)
    # INIT IQ
    iq = IQxel(iq_ip)
    iq.open()
    iq.read_idn()
    iq.reset()
    iq.set_pathloss()

    # INIT DUT
    dt = dut()
    dt.login(dut_ip, user, pwd)
    dt.init(add1, add2, add3)
    # dt.ex_command('reboot')
    # print('Reboot DUT, Please wait...')
    # time.sleep(60#)
    # dt = dut()
    # dt.login(dut_ip, user, pwd)
    # dt.init(add1, add2, add3)

    # INIT SPEC
    spec_file_2g = load_workbook('./spec_2g.xlsx')
    # print(spec_file_2g.sheetnames)
    sheet_2g = spec_file_2g['Sheet1']
    rows_2g = []
    ratelist_2g = ['1M', '2M', '5_5M', '11M', '6M', '9M', '12M', '18M', '24M', '36M', '48M', '54M', 'HT20_MCS0',
                   'HT20_MCS1',
                   'HT20_MCS2', 'HT20_MCS3', 'HT20_MCS4', 'HT20_MCS5', 'HT20_MCS6', 'HT20_MCS7', 'HT40_MCS0',
                   'HT40_MCS1',
                   'HT40_MCS2', 'HT40_MCS3', 'HT40_MCS4', 'HT40_MCS5', 'HT40_MCS6', 'HT40_MCS7']
    ratelist_evm_2g = ['1M_EVM', '2M_EVM', '5_5M_EVM', '11M_EVM', '6M_EVM', '9M_EVM', '12M_EVM', '18M_EVM', '24M_EVM',
                       '36M_EVM', '48M_EVM', '54M_EVM', 'HT20_MCS0_EVM', 'HT20_MCS1_EVM', 'HT20_MCS2_EVM',
                       'HT20_MCS3_EVM',
                       'HT20_MCS4_EVM', 'HT20_MCS5_EVM', 'HT20_MCS6_EVM', 'HT20_MCS7_EVM', 'HT40_MCS0_EVM',
                       'HT40_MCS1_EVM',
                       'HT40_MCS2_EVM', 'HT40_MCS3_EVM', 'HT40_MCS4_EVM', 'HT40_MCS5_EVM', 'HT40_MCS6_EVM',
                       'HT40_MCS7_EVM']
    target_pwr_1M, target_pwr_2M, target_pwr_5_5M, target_pwr_11M, target_pwr_6M, target_pwr_9M, target_pwr_12M, \
    target_pwr_18M, target_pwr_24M, target_pwr_36M, target_pwr_48M, target_pwr_54M, target_pwr_HT20_MCS0, \
    target_pwr_HT20_MCS1, target_pwr_HT20_MCS2, target_pwr_HT20_MCS3, target_pwr_HT20_MCS4, target_pwr_HT20_MCS5, \
    target_pwr_HT20_MCS6, target_pwr_HT20_MCS7, target_pwr_HT40_MCS0, target_pwr_HT40_MCS1, target_pwr_HT40_MCS2, \
    target_pwr_HT40_MCS3, target_pwr_HT40_MCS4, target_pwr_HT40_MCS5, target_pwr_HT40_MCS6, target_pwr_HT40_MCS7 \
        = [None] * 28
    target_1M_EVM, target_2M_EVM, target_5_5M_EVM, target_11M_EVM, target_6M_EVM, target_9M_EVM, target_12M_EVM, \
    target_18M_EVM, target_24M_EVM, target_36M_EVM, target_48M_EVM, target_54M_EVM, target_HT20_MCS0_EVM, \
    target_HT20_MCS1_EVM, target_HT20_MCS2_EVM, target_HT20_MCS3_EVM, target_HT20_MCS4_EVM, target_HT20_MCS5_EVM, \
    target_HT20_MCS6_EVM, target_HT20_MCS7_EVM, target_HT40_MCS0_EVM, target_HT40_MCS1_EVM, target_HT40_MCS2_EVM, \
    target_HT40_MCS3_EVM, target_HT40_MCS4_EVM, target_HT40_MCS5_EVM, target_HT40_MCS6_EVM, target_HT40_MCS7_EVM \
        = [None] * 28
    for row in sheet_2g:
        rows_2g.append(row)
        # print(rows_2g)
    for r in range(sheet_2g.max_row):
        for c in range(sheet_2g.max_column):
            # print(rows_2g[r][c].value)
            rows_2g[r][c].value = str(rows_2g[r][c].value).strip()
            rs = r + 1
            cs = c + 1
            if rows_2g[r][c].value == 'POWER_ACCURACY':
                spec_pwr = abs(rows_2g[r][cs].value)
            elif rows_2g[r][c].value == 'EVM_MARGIN':
                evm_margin = abs(rows_2g[r][cs].value)
            elif rows_2g[r][c].value == 'Symbol_Clock_Error':
                spec_symbol_clock_error = abs(rows_2g[r][cs].value)
            elif rows_2g[r][c].value == 'LO_Leakage':
                spec_lo_leakage = abs(rows_2g[r][cs].value)
            elif rows_2g[r][c].value == 'MASK':
                spec_mask = abs(rows_2g[r][cs].value)
            elif rows_2g[r][c].value == 'OBW_20M':
                spec_obw_20M = rows_2g[r][cs].value
            elif rows_2g[r][c].value == 'OBW_40M':
                spec_obw_40M = rows_2g[r][cs].value
            for x in ratelist_2g:
                if rows_2g[r][c].value == x + '_target':
                    exec('target_pwr_%s=%d' % (x, rows_2g[rs][c].value))
                    break
            for i in ratelist_evm_2g:
                if rows_2g[r][c].value == i + '_target':
                    exec('target_%s=%d' % (i, rows_2g[rs][c].value))
                    break
            for j in ratelist_2g:
                if rows_2g[r][c].value == j + '_sens':
                    exec('sens_%s=%d' % (j, rows_2g[rs][c].value))
                    break
    targetpower = 0
    spec_evm = 0
    # get pwr_paras

    # GEN Report
    now_time = datetime.datetime.now()
    # print(now_time)
    now_time = str(now_time)
    now_time = now_time.split()
    # print(now_time)
    day_time = re.sub('-', '', now_time[0])
    now_time = now_time[1].split('.')
    # print(day_time)
    # print(now_time)
    now_time = re.sub(':', '', now_time[0])
    now_time = day_time + now_time
    # print(now_time)
    tx_result_name = sn + '_' + 'TX_Result' + '_' + now_time + '.csv'
    rx_result_name = sn + '_' + 'RX_Result' + '_' + now_time + '.csv'
    with open('./Result/' + tx_result_name , 'w', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(['FREQ', 'DATA_RATE', 'CHAIN', 'TX_POWER', 'POWER', 'GAIN', 'LIMIT',
                         'RESULT', 'EVM', 'LIMIT', 'RESULT', 'FREQ_ERROR', 'LIMIT', 'RESULT',
                         'LOLEAKAGE', 'LIMIT', 'RESULT', 'OBW', 'LIMIT', 'RESULT',
                         'MASK', 'LIMIT', 'RESULT', 'FLATNESS', 'LIMIT', 'RESULT',
                         'RAMPONTIME', 'LIMIT', 'RESULT', 'RAMPOFFTIME', 'LIMIT', 'RESULT'])
    with open('./Result/' + rx_result_name, 'w', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(['FREQ', 'DATA_RATE', 'CHAIN', 'SENSITIVITY', 'PER', 'RESULT'])

    # TEST FLOW
    filename = 'TEST_FLOW.txt'
    f = open(filename)
    result = list()
    for line in f.readlines():
        if not len(line) or line.startswith('//'):
            continue
            pass
        else:
            line = line.strip()
            line = line.split()
            # print(line)
            # print('Channel:',line[1],'Rate:',line[2],'Chain:',line[3])
            item = re.sub('IQ_WIFI_TEST_', '', line[0])
            channel = line[1]
            rate = line[2]
            chain = line[3]
            if item == 'RX':
                start = int(line[4])
                stop = int(line[5])
            # mode,bw,rates
            if rate == '1M' or rate == '2M' or rate == '5.5M' or rate == '11M':
                mode = '11b'
                bw = '20'
                rates = re.sub('M', '', rate)
                per_spec = float(rx_packets) * 0.92
            elif rate == '6M' or rate == '9M' or rate == '12M' or rate == '18M' \
                    or rate == '24M' or rate == '36M' or rate == '48M' or rate == '54M':
                if int(channel) < 5000:
                    mode = '11g'
                else:
                    mode = '11a'
                bw = '20'
                rates = re.sub('M', '', rate)
                per_spec = float(rx_packets) * 0.9
            elif rate == 'HT20-MCS0' or rate == 'HT20-MCS1' or rate == 'HT20-MCS2' or rate == 'HT20-MCS3' \
                    or rate == 'HT20-MCS4' or rate == 'HT20-MCS5' or rate == 'HT20-MCS6' or rate == 'HT20-MCS7':
                mode = '11ng20'
                bw = '20'
                rates = re.sub('HT20-MCS', '', rate)
                per_spec = float(rx_packets) * 0.9
            elif rate == 'HT40-MCS0' or rate == 'HT40-MCS1' or rate == 'HT40-MCS2' or rate == 'HT40-MCS3' \
                    or rate == 'HT40-MCS4' or rate == 'HT40-MCS5' or rate == 'HT40-MCS6' or rate == 'HT40-MCS7':
                mode = '11ng40plus'
                bw = '40'
                rates = re.sub('HT40-MCS', '', rate)
                per_spec = float(rx_packets) * 0.9
            # chain
            if chain == 'CHAIN0':
                chain = re.sub('CHAIN', '', chain)
            elif chain == 'CHAIN1':
                chain = re.sub('CHAIN', '', chain)
            elif chain == 'CHAIN2':
                chain = re.sub('CHAIN', '', chain)
            else:
                chain = re.sub('CHAIN', '', chain)
            # power para
            if int(channel) < 2437 and chain == '0':
                channel_groups = '98'
            elif int(channel) < 2437 and chain == '1':
                channel_groups = 'a8'
            elif 2437 < int(channel) < 2462 and chain == '0':
                channel_groups = '94'
            elif 2437 < int(channel) < 2462 and chain == '1':
                channel_groups = 'a4'
            elif int(channel) >= 2462 and chain == '0':
                channel_groups = '90'
            elif int(channel) >= 2462 and chain == '1':
                channel_groups = 'a0'
            if item == 'TX':
                print('*************************************************************')
                print('Mode:', mode, 'Channel:', channel, 'BW:', bw, 'Rate:', rate, 'Chain:', chain)
                adjust_result = result_evm = result_symbol_clock_error = result_lo_leakage = result_mask = result_flatness = \
                    'Pass'
                avg_power = targetpower
                gain = 0
                # TX
                channel = int(channel)
                bw = str(bw)
                rates = str(rates)
                # dt.set_default()
                dt.tx(mode, channel, bw, rates, chain)
                init_value = dt.get_paras()
                gain = init_value
                print('INIT_VALUE(INT):', init_value)

                # RESULT
                rate_t = re.sub('-', '_', rate)
                rate_t = re.sub('\.', '_', rate_t)
                targetpower = eval('target_pwr_' + rate_t)
                rate_e = rate_t + '_EVM'
                spec_evm = eval('target_' + rate_e)
                channel = int(channel)
                chain = int(chain)
                iq.read_data()
                avg_power, result_evm, result_symbol_clock_error, result_lo_leakage, result_mask = \
                    iq.get_data(mode, rate, channel, chain, tx_result_name , targetpower, spec_pwr, gain, spec_evm,
                                evm_margin,
                                spec_symbol_clock_error, spec_lo_leakage, spec_mask, spec_obw_20M, spec_obw_40M)
                if avg_power == 'NA' or result_evm == 'NA':
                    # avg_power = avg_evm = symbol_clock_error = lo_leakage = result = result_pwr = result_evm = \
                    #    result_symbol_clock_error = result_lo_leakage = result_mask = 'NA'
                    print(Fore.RED + 'No Result!' + Style.RESET_ALL)
                    continue
                else:
                    avg_power = float(avg_power)
                    targetpower = float(targetpower)
                    max_power = float(targetpower + spec_pwr)
                    delta_power = avg_power - targetpower
                    delta_power = '{:.3f}'.format(delta_power)
                    delta_power = float(delta_power)
                    print('Default Measurer Power:', avg_power, 'Target Power:', targetpower, 'Delta Power:',
                          delta_power)
                    # power is nomal but some is fail
                    if targetpower <= avg_power <= max_power:
                        if result_evm == 'Fail' or result_symbol_clock_error == 'Fail' or result_lo_leakage == 'Fail' \
                                or result_mask == 'Fail':
                            print(Fore.RED + 'TX\'s quality are failed!' + Style.RESET_ALL)
                        else:
                            print('Get Good Result!')
                    # power is not nomal
                    else:
                        low_power_counts = 0
                        setup = 2
                        setups = 2
                        print('Step:', setup)
                        pwr_paras = init_value + setup
                        pwr_paras = hex(pwr_paras)
                        pwr_paras = re.sub('0x', '', pwr_paras)
                        get_power = []
                        get_delta_power = []
                        c = 1
                        get_power.append(avg_power)
                        print('Measurer Power List:', get_power)
                        get_delta_power.append(delta_power)
                        print('Power Deviation List:', get_delta_power)
                        print('Adjust Counts:', c)
                        # ADJUST
                        min_adj = avg_power - targetpower
                        max_adj = avg_power - max_power
                        while min_adj < 0 or max_adj > 0:
                            print(Fore.GREEN + 'Adjust...' + Style.RESET_ALL)
                            if avg_power > max_power:
                                adj = -1
                            else:
                                adj = 1
                            c = c + 1
                            print('NEW VALUE(HEX):', channel_groups, pwr_paras)
                            gain = dt.adjust_power()
                            iq.read_data()
                            avg_power, result_evm, result_symbol_clock_error, result_lo_leakage, result_mask = \
                                iq.get_data(mode, rate, channel, chain, tx_result_name , targetpower, spec_pwr, gain,
                                            spec_evm,
                                            evm_margin, spec_symbol_clock_error, spec_lo_leakage, spec_mask,
                                            spec_obw_20M,
                                            spec_obw_40M)
                            get_power.append(avg_power)
                            print('Measurer Power List:', get_power)
                            delta_power = float(avg_power) - targetpower
                            delta_power = '{:.3f}'.format(delta_power)
                            delta_power = float(delta_power)
                            avg_power = float(avg_power)
                            min_adj = avg_power - targetpower
                            max_adj = avg_power - max_power
                            get_delta_power.append(delta_power)
                            print('Power Deviation List:', get_delta_power)
                            print('Adjust Counts:', c)
                            adjust_status = get_delta_power[c - 1] - get_delta_power[c - 2]
                            adjust_status = float('{:.3f}'.format(adjust_status))
                            print('Power Added:', adjust_status)
                            if adjust_status < 0.1:
                                adjust_result = 'Pass'
                                setup = 2 * setups * adj
                                low_power_counts = low_power_counts + 1
                            elif adjust_status > 0.1:
                                adjust_status = 'Pass'
                                setup = 2 * adj
                            # print(low_power_counts)
                            if low_power_counts > 5:
                                print(Fore.RED + 'Power added was too small, Power adjust stop' + Style.RESET_ALL)
                                adjust_result = 'Fail'
                            else:
                                print('Step:', setup)
                                pwr_paras = int(pwr_paras, 16)
                                pwr_paras = int(pwr_paras) + setup
                                pwr_paras = hex(pwr_paras)
                                pwr_paras = re.sub('0x', '', pwr_paras)
                                print(result_evm, result_symbol_clock_error, result_lo_leakage, result_mask)
                                print('*************************************************************')
            else:
                per_list = []
                sens_list = []
                print('Start: ' + str(start) + 'Stop: ' + str(stop))
                # loss = iq.read_pathloss(channel, chain)
                dt.rx(mode, channel, bw, rates, chain)
                iq.vsg()
                per = dt.get_statistics()
                per = int(per)
                per_list.append(per)
                sens_list.append(start)
                while start > stop and per >= int(per_spec):
                    start = start - 1
                    # dt.rx(mode, channel, bw, rates, chain)
                    iq.vsg()
                    time.sleep(sleep_time)
                    per = dt.get_statistics()
                    per = int(per)
                    per_list.append(per)
                    sens_list.append(start)
                print(len(per_list) + str(per_list))
                print(len(sens_list) + str(sens_list))
                per = per_list[len(per_list) - 2]
                sens = sens_list[len(sens_list) - 2]
                if per < per_spec:
                    result = 'Fail'
                else:
                    result = 'Pass'
                with open('./Result/' + rx_result_name, 'a+', newline='') as f2:
                    writer2 = csv.writer(f2)
                    writer2.writerow([channel, rate, chain, sens, per, result])
                print('*************************************************************')
    print('************************TEST DONE****************************')
    dt.close()
    iq.close()
